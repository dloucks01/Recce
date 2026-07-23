"""High-fidelity integration tests: the whole workflow end-to-end, with a hard
focus on correctness of the spreadsheet - that the RIGHT fields land on the RIGHT
IP row, that per-IP tracking never bleeds across hosts, and that re-scans/updates
preserve everything.

These drive the real parser -> store -> workbook writer -> read-back -> report
paths (no nmap needed) against the bundled sample scan, whose four hosts each
have a distinct fingerprint:

    10.0.10.10  dc01.corp.local  Windows Server 2019  88,389,445,3389  ms17-010
    10.0.10.25  ws01.corp.local  Windows 10 21H2      135,445,3389     (no vulns)
    10.0.20.5   web01            Linux 5.4            22,80,443        4 vulns
    10.0.20.6   web02            Linux 5.4            22,80,21,3306    ftp-anon
"""

import contextlib
import io
import os
import shutil
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from recce import ad, parser, xlsx
from recce import tracking as tr
from recce.models import Host, Port, Vuln
from recce.report_excel import (build_workbook, read_key_order,
                                 read_workbook_edits, read_workbook_tracking,
                                 update_workbook, STATUS_WIP, STATUS_TODO)
from recce.store import Store
from recce.targets import _subnet_of

SAMPLE = os.path.join(os.path.dirname(parser.__file__), "sample_scan.xml")

# Ground-truth facts, keyed by IP, for cross-checking the spreadsheet.
FACTS = {
    "10.0.10.10": {"host": "dc01.corp.local", "os": "Windows Server 2019",
                   "ports": [88, 389, 445, 3389], "nvulns": 1},
    "10.0.10.25": {"host": "ws01.corp.local", "os": "Windows 10",
                   "ports": [135, 445, 3389], "nvulns": 0},
    "10.0.20.5": {"host": "web01", "os": "Linux",
                  "ports": [22, 80, 443], "nvulns": 4},
    "10.0.20.6": {"host": "web02", "os": "Linux",
                  "ports": [21, 22, 80, 3306], "nvulns": 1},
}


def sample_hosts():
    hosts = parser.parse_nmap_xml(SAMPLE)
    for h in hosts:
        h.subnet = _subnet_of(h.ip)
        h.enumerated = True
    ad.analyze_hosts(hosts)
    return hosts


def rows_by_ip(sheets, title):
    """Return (header, {ip: [row-as-dict, ...]}) for a sheet with an IP column.

    Skips collapsible group-header band rows (they carry a label in the IP column
    but no Key), so callers only see real data rows keyed by a bare IP."""
    rows = sheets[title]
    hdr = rows[0]
    ipc = hdr.index("IP")
    kidx = hdr.index("Key") if "Key" in hdr else None
    out: dict = {}
    for r in rows[1:]:
        if kidx is not None and (len(r) <= kidx or not r[kidx]):
            continue                       # group-header band row - not data
        if len(r) > ipc and r[ipc]:
            out.setdefault(str(r[ipc]), []).append(dict(zip(hdr, r)))
    return hdr, out


class ChecklistPerIpFidelityTest(unittest.TestCase):
    """Every host's Checklist row carries ITS OWN facts - no cross-IP bleed."""

    def setUp(self):
        self.d = tempfile.mkdtemp()
        self.out = os.path.join(self.d, "wb.xlsx")
        build_workbook(sample_hosts(), self.out)
        self.sheets = xlsx.read_sheets(self.out)

    def test_each_ip_row_has_its_own_identity(self):
        _hdr, by_ip = rows_by_ip(self.sheets, "Checklist")
        self.assertEqual(set(by_ip), set(FACTS))
        for ip, facts in FACTS.items():
            self.assertEqual(len(by_ip[ip]), 1, f"{ip} should be exactly one row")
            row = by_ip[ip][0]
            self.assertEqual(row["Hostname"], facts["host"])
            self.assertIn(facts["os"].split()[0], row["OS"])
            self.assertEqual(str(row["# Vulns"]), str(facts["nvulns"]))
            # Open-ports cell lists exactly this host's ports, in order.
            self.assertEqual(row["Open ports"],
                             ", ".join(str(p) for p in facts["ports"]))

    def test_dc_role_only_on_the_dc_row(self):
        _hdr, by_ip = rows_by_ip(self.sheets, "Checklist")
        self.assertIn("Domain Controller", by_ip["10.0.10.10"][0]["Roles"])
        for ip in ("10.0.20.5", "10.0.20.6"):
            self.assertNotIn("Domain Controller", by_ip[ip][0]["Roles"] or "")

    def test_surface_steps_match_host_type(self):
        _hdr, by_ip = rows_by_ip(self.sheets, "Checklist")
        # DC: AD applies (☐, manual), Web/DB do not (—).
        dc = by_ip["10.0.10.10"][0]
        self.assertEqual(dc["AD"], xlsx.CHECK_OFF)
        self.assertEqual(dc["Web"], tr.STEP_NA)
        self.assertEqual(dc["DB"], tr.STEP_NA)
        # web01: Web applies, AD does not, DB does not.
        web = by_ip["10.0.20.5"][0]
        self.assertEqual(web["Web"], xlsx.CHECK_OFF)
        self.assertEqual(web["AD"], tr.STEP_NA)
        self.assertEqual(web["DB"], tr.STEP_NA)
        # web02: has MySQL -> DB applies.
        self.assertEqual(by_ip["10.0.20.6"][0]["DB"], xlsx.CHECK_OFF)


class ServicesPerIpFidelityTest(unittest.TestCase):
    def setUp(self):
        self.d = tempfile.mkdtemp()
        self.out = os.path.join(self.d, "wb.xlsx")
        build_workbook(sample_hosts(), self.out)
        self.sheets = xlsx.read_sheets(self.out)

    def test_every_service_row_maps_port_to_correct_ip(self):
        hdr, by_ip = rows_by_ip(self.sheets, "Services")
        for ip, facts in FACTS.items():
            got = sorted(int(r["Port"]) for r in by_ip.get(ip, []))
            self.assertEqual(got, sorted(facts["ports"]),
                             f"{ip} services should be exactly its own ports")
        # The FTP service (21) exists on web02 ONLY.
        ftp_rows = [r for rs in by_ip.values() for r in rs if str(r["Port"]) == "21"]
        self.assertEqual(len(ftp_rows), 1)
        self.assertEqual(ftp_rows[0]["IP"], "10.0.20.6")
        self.assertIn("ftp", ftp_rows[0]["Service"].lower())

    def test_hostname_rides_in_the_collapsible_ip_band(self):
        # Hostname is no longer a per-row column; it appears once in each host's
        # collapsible band (IP · hostname · N ports), not repeated on every port row.
        rows = self.sheets["Services"]
        self.assertNotIn("Hostname", rows[0])
        ipc = rows[0].index("IP")
        bands = " ".join(str(r[ipc]) for r in rows[1:] if "·" in str(r[ipc]))
        for ip, facts in FACTS.items():
            if facts["host"]:
                self.assertIn(facts["host"], bands)


class VulnerabilitiesPerIpFidelityTest(unittest.TestCase):
    def setUp(self):
        self.d = tempfile.mkdtemp()
        self.out = os.path.join(self.d, "wb.xlsx")
        build_workbook(sample_hosts(), self.out)
        self.sheets = xlsx.read_sheets(self.out)

    def test_findings_attributed_to_correct_ip(self):
        _hdr, by_ip = rows_by_ip(self.sheets, "Vulnerabilities")
        # ms17-010 is on the DC only.
        dc_finds = " ".join(r["Finding"] for r in by_ip.get("10.0.10.10", []))
        self.assertIn("ms17-010", dc_finds)
        # ...and NOT attributed to any other host.
        for ip in ("10.0.20.5", "10.0.20.6", "10.0.10.25"):
            self.assertNotIn("ms17-010",
                             " ".join(r["Finding"] for r in by_ip.get(ip, [])))
        # ftp-anon is web02 only.
        self.assertIn("FTP", " ".join(r["Finding"] for r in by_ip.get("10.0.20.6", [])))

    def test_grouped_by_host_no_hostname_col_and_full_details(self):
        from recce.report_excel import build_workbook
        from recce.models import Host, Port, Vuln
        rows = self.sheets["Vulnerabilities"]
        hdr = rows[0]
        self.assertNotIn("Hostname", hdr)                 # Hostname moved to the band
        ipc = hdr.index("IP")
        # A collapsible per-host band exists (IP · hostname · N findings · worst ...).
        bands = [str(r[ipc]) for r in rows[1:] if "finding" in str(r[ipc])]
        self.assertTrue(bands)
        self.assertTrue(any("worst:" in b for b in bands))
        # Details is shown IN FULL (wrapped), never truncated with an ellipsis.
        big = "A" * 1200
        h = Host(ip="10.9.9.9", subnet="10.9.9.0/24", state="up", enumerated=True,
                 ports=[Port(portid=445, service="smb")],
                 vulns=[Vuln(ip="10.9.9.9", port=445, protocol="tcp", script_id="x",
                             title="f", severity="high", source="nse", state="finding",
                             output=big)])
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook([h], out)
            vr = xlsx.read_sheets(out)["Vulnerabilities"]
        dc = vr[0].index("Details")
        detail = next(str(r[dc]) for r in vr[1:] if len(r) > dc and "A" in str(r[dc]))
        self.assertEqual(detail, big)                     # full, untruncated
        self.assertNotIn("…", detail)

    def test_exploit_column_proven_vs_candidate(self):
        _hdr, by_ip = rows_by_ip(self.sheets, "Vulnerabilities")
        self.assertIn("Exploit", _hdr)
        self.assertNotIn("Proven exploit", _hdr)
        # The DC's ms17-010 finding carries the proven EternalBlue exploit (curated).
        dc = by_ip["10.0.10.10"]
        ms17 = next(r for r in dc if "ms17-010" in r["Finding"])
        self.assertIn("eternalblue", ms17["Exploit"].lower())
        # A potential/advisory finding never claims an exploit.
        for r in dc:
            if r["Conf."] == "potential":
                self.assertEqual(r["Exploit"], "")
        # A config/hardening finding (weak TLS cipher/protocol, missing header)
        # never gets a PROVEN exploit, even if a CVE leaked into its output.
        for rs in by_ip.values():
            for r in rs:
                f = r["Finding"].lower()
                if any(k in f for k in ("weak", "cipher", "tlsv1", "missing", "header")):
                    self.assertFalse(r["Exploit"].startswith(("Metasploit", "impacket")),
                                     f"hardening finding wrongly proven: {r['Finding']}")
        # searchsploit hits are shown as CANDIDATES to verify, never as proof.
        for rs in by_ip.values():
            for r in rs:
                if r["Exploit"].startswith("candidate"):
                    self.assertIn("verify", r["Exploit"].lower())
        # Overview 'curated exploit' tile counts only the curated (non-candidate) ones.
        n_proven = sum(1 for rs in by_ip.values() for r in rs
                       if r["Exploit"] and not r["Exploit"].startswith("candidate"))
        ov = ["|".join(str(c) for c in r) for r in self.sheets["Overview"]]
        self.assertTrue(any(f"Findings with a curated exploit|{n_proven}" in t for t in ov))
        self.assertEqual(by_ip.get("10.0.10.25", []), [])

    def test_vuln_row_counts_match_per_host(self):
        _hdr, by_ip = rows_by_ip(self.sheets, "Vulnerabilities")
        for ip, facts in FACTS.items():
            self.assertEqual(len(by_ip.get(ip, [])), facts["nvulns"], ip)


class TrackingIsolationTest(unittest.TestCase):
    """Ticking a box for one IP must never touch another IP's tracking."""

    def test_reviewing_one_host_isolates_to_that_key(self):
        with tempfile.TemporaryDirectory() as d:
            store = Store(os.path.join(d, "s.sqlite"))
            for h in sample_hosts():
                store.upsert_host(h)
            store.set_reviewed(tr.host_key("10.0.20.5"), True, notes="looked at web01")
            got = store.get_tracking()
            self.assertTrue(got[tr.host_key("10.0.20.5")][0])
            for ip in ("10.0.10.10", "10.0.10.25", "10.0.20.6"):
                self.assertNotIn(tr.host_key(ip), got)
            store.close()

    def test_step_and_status_keys_are_per_ip(self):
        # Distinct hosts, same open port -> distinct svc/step keys, no collision.
        a = tr.svc_key("10.0.20.5", "tcp", 80)
        b = tr.svc_key("10.0.20.6", "tcp", 80)
        self.assertNotEqual(a, b)
        self.assertNotEqual(tr.step_key("vuln", "10.0.20.5"),
                            tr.step_key("vuln", "10.0.20.6"))
        self.assertNotEqual(tr.vuln_key("10.0.20.5", 80, "http-x"),
                            tr.vuln_key("10.0.20.6", 80, "http-x"))

    def test_workbook_reviewed_readback_targets_the_edited_ip(self):
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            hosts = sample_hosts()
            # Operator ticks Reviewed for web02 (10.0.20.6) only.
            build_workbook(hosts, out,
                           tracking={tr.host_key("10.0.20.6"): (True, "done")})
            back = read_workbook_tracking(out)
            self.assertTrue(back[tr.host_key("10.0.20.6")][0])
            for ip in ("10.0.10.10", "10.0.10.25", "10.0.20.5"):
                self.assertFalse(back.get(tr.host_key(ip), (False, ""))[0])

    def test_per_port_status_isolates_to_one_service_row(self):
        # Every service row carries a Status (default "Not started"); only the one
        # we set should read back as In-progress - no other row is elevated.
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            k = tr.svc_key("10.0.20.6", "tcp", 21)
            build_workbook(sample_hosts(), out, statuses={k: STATUS_WIP})
            _edits, statuses = read_workbook_edits(out)
            self.assertEqual(statuses.get(k), STATUS_WIP)
            others = {kk: v for kk, v in statuses.items() if kk != k}
            self.assertTrue(others, "other service rows should still be present")
            self.assertTrue(all(v == STATUS_TODO for v in others.values()),
                            "no other port should be marked in-progress/done")


class MergeRescanFidelityTest(unittest.TestCase):
    """Re-scanning a host merges into the right IP and leaves others untouched."""

    def test_rescan_merges_ports_flags_vulns_without_touching_other_hosts(self):
        with tempfile.TemporaryDirectory() as d:
            store = Store(os.path.join(d, "s.sqlite"))
            a1 = Host(ip="10.0.0.5", subnet="10.0.0.0/24", enumerated=True,
                      hostnames=["a"], ports=[Port(portid=80, service="http")])
            b = Host(ip="10.0.0.9", subnet="10.0.0.0/24", enumerated=True,
                     hostnames=["b"], ports=[Port(portid=22, service="ssh")])
            store.upsert_host(a1)
            store.upsert_host(b)
            # Re-scan A: new port + a vuln + db flag; same vuln twice to test dedup.
            v = Vuln(ip="10.0.0.5", port=443, protocol="tcp", script_id="ssl-x",
                     title="weak tls", severity="medium")
            a2 = Host(ip="10.0.0.5", subnet="10.0.0.0/24", db_scanned=True,
                      ports=[Port(portid=443, service="https")], vulns=[v, v])
            store.upsert_host(a2)

            A = store.get_host("10.0.0.5")
            self.assertEqual(sorted(p.portid for p in A.ports), [80, 443])
            self.assertTrue(A.enumerated)          # preserved from first scan
            self.assertTrue(A.db_scanned)           # merged from rescan
            self.assertEqual(len(A.vulns), 1)       # deduped
            # B is completely untouched.
            B = store.get_host("10.0.0.9")
            self.assertEqual([p.portid for p in B.ports], [22])
            self.assertEqual(B.vulns, [])
            store.close()

    def test_regenerate_preserves_row_order_and_appends_new_host(self):
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            hosts = sample_hosts()
            build_workbook(hosts, out)
            order_before = read_key_order(out)["Checklist"]
            # A new host is discovered on a later scan.
            newh = Host(ip="10.0.20.99", subnet="10.0.20.0/24", enumerated=True,
                        hostnames=["new"], ports=[Port(portid=8080, service="http")])
            update_workbook(out, hosts + [newh])
            order_after = read_key_order(out)["Checklist"]
            # Existing rows keep their exact positions; the new one is appended.
            self.assertEqual(order_after[:len(order_before)], order_before)
            self.assertEqual(order_after[-1], tr.host_key("10.0.20.99"))


class CoverageMathFidelityTest(unittest.TestCase):
    def test_marking_one_host_counts_once_in_the_right_subnet(self):
        hosts = sample_hosts()
        tracking = {tr.host_key("10.0.10.10"): (True, "")}
        cov = tr.compute_coverage(hosts, tracking)
        self.assertEqual(cov["hosts"]["total"], 4)
        self.assertEqual(cov["hosts"]["done"], 1)
        sc = tr.subnet_coverage(hosts, tracking)
        self.assertEqual(sc["10.0.10.0/24"]["done"], 1)   # the DC's subnet
        self.assertEqual(sc["10.0.10.0/24"]["total"], 2)
        self.assertEqual(sc["10.0.20.0/24"]["done"], 0)   # other subnet unaffected

    def test_service_coverage_counts_all_open_ports(self):
        hosts = sample_hosts()
        keys = tr.item_keys(hosts)
        total_ports = sum(len(h.open_ports) for h in hosts)
        self.assertEqual(len(keys["services"]), total_ports)
        # Mark exactly one service done -> coverage done == 1.
        k = tr.svc_key("10.0.20.6", "tcp", 21)
        cov = tr.compute_coverage(hosts, {k: (True, "")})
        self.assertEqual(cov["services"]["done"], 1)

    def test_overview_phase_table_honors_operator_override(self):
        """Regression: the Overview per-subnet phase table must reflect an
        operator un-tick the same way the Checklist does, or the two diverge."""
        import openpyxl
        from recce.report_excel import build_workbook

        def enum_cell(path):
            ov = openpyxl.load_workbook(path)["Overview"]
            for row in ov.iter_rows(values_only=True):
                if row and row[0] == "10.0.0.0/24":
                    return row[3]   # "Enumerated" column
        hosts = [Host(ip="10.0.0.5", subnet="10.0.0.0/24", enumerated=True, state="up",
                      ports=[Port(portid=80, service="http", state="open")]),
                 Host(ip="10.0.0.6", subnet="10.0.0.0/24", enumerated=True, state="up",
                      ports=[Port(portid=80, service="http", state="open")])]
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "wb.xlsx")
            build_workbook(hosts, p, tracking={})
            self.assertEqual(enum_cell(p), "2/2")
            build_workbook(hosts, p,
                           tracking={tr.step_key("enum", "10.0.0.6"): (False, "redo")})
            self.assertEqual(enum_cell(p), "1/2")

    def test_accounts_differing_only_by_rid_dont_collide(self):
        """Regression: the store keeps accounts distinct by rid, so acct_key must
        include rid or two such accounts collapse to one row + undercount."""
        from recce.models import Account
        a = Account(ip="10.0.0.5", source="ldap", kind="user", name="svc", domain="corp", rid="1103")
        b = Account(ip="10.0.0.5", source="ldap", kind="user", name="svc", domain="corp", rid="1104")
        ka = tr.acct_key(a.source, a.kind, a.domain, a.name, a.rid)
        kb = tr.acct_key(b.source, b.kind, b.domain, b.name, b.rid)
        self.assertNotEqual(ka, kb)
        # A rid-less account keeps its historical (colon-free) key.
        self.assertEqual(tr.acct_key("ldap", "share", "corp", "SYSVOL"),
                         "acct:ldap:share:corp:SYSVOL")
        h = Host(ip="10.0.0.5", accounts=[a, b])
        self.assertEqual(len(tr.item_keys([h])["accounts"]), 2)


class WriteupPerIpFidelityTest(unittest.TestCase):
    def test_grouped_finding_lists_only_the_affected_ip(self):
        from recce.report_docx import group_findings
        findings = group_findings(sample_hosts())
        ms17 = next(f for f in findings if "ms17-010" in f.title.lower())
        self.assertEqual(sorted({a[0] for a in ms17.affected}), ["10.0.10.10"])
        ftp = next(f for f in findings if "ftp" in f.title.lower())
        self.assertEqual(sorted({a[0] for a in ftp.affected}), ["10.0.20.6"])

    def test_shared_finding_across_hosts_lists_all_affected(self):
        # Two hosts with the same finding title -> one write-up, both IPs.
        from recce.report_docx import group_findings
        hosts = [
            Host(ip="10.0.0.1", ports=[Port(portid=443, service="https")],
                 vulns=[Vuln(ip="10.0.0.1", port=443, protocol="tcp",
                             script_id="ssl", title="Weak TLS", severity="medium")]),
            Host(ip="10.0.0.2", ports=[Port(portid=443, service="https")],
                 vulns=[Vuln(ip="10.0.0.2", port=443, protocol="tcp",
                             script_id="ssl", title="Weak TLS", severity="medium")]),
        ]
        f = group_findings(hosts)[0]
        self.assertEqual(sorted({a[0] for a in f.affected}),
                         ["10.0.0.1", "10.0.0.2"])

    def test_generated_docx_contains_correct_ip_only(self):
        from recce.report_docx import build_writeups
        import zipfile
        import xml.etree.ElementTree as ET
        W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "w")
            build_writeups(sample_hosts(), out)
            fn = next(p for p in os.listdir(out) if "ms17" in p)
            root = ET.fromstring(zipfile.ZipFile(os.path.join(out, fn))
                                 .read("word/document.xml"))
            text = "\n".join("".join(t.text or "" for t in p.iter(f"{W}t"))
                             for p in root.iter(f"{W}p"))
            self.assertIn("10.0.10.10", text)          # the affected host
            self.assertNotIn("10.0.20.5", text)         # not an unrelated host


class FullCliRoundTripTest(unittest.TestCase):
    """Drive the real cli report/import functions, as the commands do."""

    def test_report_then_operator_edit_then_import_persists_per_ip(self):
        from recce import cli
        with tempfile.TemporaryDirectory() as d:
            paths = cli._open_paths(d)
            store = Store(paths["db"])
            store.set_meta("engagement", "roundtrip")
            for h in sample_hosts():
                store.upsert_host(h)
            # 1. Generate all reports from the datastore (like `report`).
            cli._generate_reports(store, paths, "roundtrip", quiet=True)
            self.assertTrue(os.path.exists(paths["xlsx"]))
            # 2. Operator edits the workbook: tick Reviewed for web01 only.
            build_workbook(store.all_hosts(), paths["xlsx"],
                           tracking={tr.host_key("10.0.20.5"): (True, "reviewed")},
                           order_map=read_key_order(paths["xlsx"]))
            # 3. Import edits back (like the start of any command).
            cli._import_excel_tracking(store, paths)
            got = store.get_tracking()
            self.assertTrue(got[tr.host_key("10.0.20.5")][0])
            self.assertEqual(got[tr.host_key("10.0.20.5")][1], "reviewed")
            # No other host got reviewed.
            for ip in ("10.0.10.10", "10.0.10.25", "10.0.20.6"):
                self.assertFalse(got.get(tr.host_key(ip), (False, ""))[0])
            # 4. Regenerate -> the tick survives and stays on the right IP.
            cli._generate_reports(store, paths, "roundtrip", quiet=True)
            back = read_workbook_tracking(paths["xlsx"])
            self.assertTrue(back[tr.host_key("10.0.20.5")][0])
            store.close()

    def test_manual_step_override_survives_regeneration(self):
        from recce import cli
        with tempfile.TemporaryDirectory() as d:
            paths = cli._open_paths(d)
            store = Store(paths["db"])
            for h in sample_hosts():
                store.upsert_host(h)
            cli._generate_reports(store, paths, "t", quiet=True)
            # Operator ticks the manual 'Access' box for the DC only.
            key = tr.step_key("access", "10.0.10.10")
            build_workbook(store.all_hosts(), paths["xlsx"],
                           tracking={key: (True, "")},
                           order_map=read_key_order(paths["xlsx"]))
            cli._import_excel_tracking(store, paths)
            self.assertTrue(store.get_tracking()[key][0])
            # Regenerate and confirm it is still checked on the DC row only.
            cli._generate_reports(store, paths, "t", quiet=True)
            sheets = xlsx.read_sheets(paths["xlsx"])
            _hdr, by_ip = rows_by_ip(sheets, "Checklist")
            self.assertEqual(by_ip["10.0.10.10"][0]["Access"], xlsx.CHECK_ON)
            self.assertEqual(by_ip["10.0.20.5"][0]["Access"], xlsx.CHECK_OFF)
            store.close()


class WorkbookStructureTest(unittest.TestCase):
    def test_all_sheets_present_and_key_column_hidden_consistent(self):
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(sample_hosts(), out)
            sheets = xlsx.read_sheets(out)
        # Always present.
        for name in ("Start Here", "Overview", "Checklist", "Services",
                     "Vulnerabilities", "Services by Product"):
            self.assertIn(name, sheets)
        # Present because the sample has this data (Exploits is skip-if-empty and
        # absent here since searchsploit didn't run).
        for name in ("Databases", "Active Directory", "AD Quick Wins",
                     "Users & Accounts", "Priv-Esc"):
            self.assertIn(name, sheets)
        self.assertNotIn("Exploits", sheets)   # no exploit data -> sheet omitted
        # Tracked sheets carry a Key column so read-back can find every row.
        for name in ("Checklist", "Services", "Vulnerabilities"):
            self.assertIn("Key", sheets[name][0])

    def test_openpyxl_can_open_the_workbook(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(sample_hosts(), out)
            wb = load_workbook(out)
            self.assertIn("Checklist", wb.sheetnames)

    def test_styling_freeze_gridlines_and_severity_contrast(self):
        """The polish pass: identity columns frozen, gridlines off, and critical
        severity is solid with white text (openpyxl reads the applied styles)."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(sample_hosts(), out)
            wb = load_workbook(out)
        self.assertEqual(wb["Checklist"].freeze_panes, "D2")   # header + 3 id cols
        self.assertEqual(wb["Services"].freeze_panes, "C2")
        self.assertFalse(wb["Checklist"].sheet_view.showGridLines)
        vs = wb["Vulnerabilities"]
        hdr = [c.value for c in vs[1]]
        sev_i = hdr.index("Severity") + 1
        crit = next(vs.cell(row=r, column=sev_i)
                    for r in range(2, vs.max_row + 1)
                    if vs.cell(row=r, column=sev_i).value == "CRITICAL")
        self.assertEqual(crit.fill.fgColor.rgb, "FFC00000")     # solid red
        self.assertEqual(crit.font.color.rgb, "FFFFFFFF")       # white text
        self.assertTrue(crit.font.bold)

    def test_raw_nse_sheet_scopes_host_and_port_scripts_per_host(self):
        from recce.models import Host, Port, Script
        hosts = [
            Host(ip="10.0.0.5", hostnames=["a"], ports=[Port(portid=445,
                 service="microsoft-ds", scripts=[Script(id="smb2-security-mode",
                 output="Message signing enabled but not required")])],
                 host_scripts=[Script(id="smb-os-discovery", output="OS: Windows")]),
            Host(ip="10.0.0.6", ports=[Port(portid=21, service="ftp",
                 scripts=[Script(id="ftp-anon", output="Anonymous FTP login allowed")])]),
        ]
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(hosts, out)
            sheets = xlsx.read_sheets(out)
        self.assertIn("Raw NSE", sheets)
        self.assertIn("Scope", sheets["Raw NSE"][0])          # unified Scope column

        _h, by_ip = rows_by_ip(sheets, "Raw NSE")
        rows5 = {r["Script"]: r for r in by_ip["10.0.0.5"]}
        # Host-level script -> Scope "host"; port script -> Scope "445".
        self.assertEqual(rows5["smb-os-discovery"]["Scope"], "host")
        self.assertEqual(rows5["smb-os-discovery"]["Output"], "OS: Windows")
        self.assertEqual(rows5["smb2-security-mode"]["Scope"], "445")
        # web02's per-port script lands on its own IP, no cross-host bleed.
        rows6 = {r["Script"]: r for r in by_ip["10.0.0.6"]}
        self.assertEqual(rows6["ftp-anon"]["Scope"], "21")
        self.assertNotIn("ftp-anon", rows5)

    def test_design_language_fonts_and_accent(self):
        """Machine data (IP/version) renders monospace with a teal IP accent;
        prose (Product) stays sans - the light HTML-preview design language."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(sample_hosts(), out)
            wb = load_workbook(out)
        sv = wb["Services"]
        hdr = [c.value for c in sv[1]]
        ki = hdr.index("Key") + 1                          # openpyxl cols are 1-based
        drow = next(r for r in range(2, sv.max_row + 1)
                    if sv.cell(row=r, column=ki).value)   # first real data row

        def cell(name):
            return sv.cell(row=drow, column=hdr.index(name) + 1)
        ip = cell("IP")
        self.assertEqual(ip.font.name, "Consolas")
        self.assertEqual(ip.font.color.rgb, "FF0E6E67")       # teal accent
        self.assertEqual(cell("Version").font.name, "Consolas")  # mono data
        self.assertEqual(cell("Product").font.name, "Calibri")   # prose stays sans
        self.assertEqual(sv.cell(row=1, column=1).fill.fgColor.rgb, "FF0E6E67")  # teal header

    def test_tab_colors_and_overview_nav_links(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(sample_hosts(), out)
            wb = load_workbook(out)
        # Tab colours group the tabs into role bands.
        self.assertIsNotNone(wb["Checklist"].sheet_properties.tabColor)
        self.assertIsNotNone(wb["Vulnerabilities"].sheet_properties.tabColor)
        # Overview is a nav hub: a jump bar + clickable totals link to real sheets.
        ov = wb["Overview"]
        targets = {c.hyperlink.location for row in ov.iter_rows()
                   for c in row if c.hyperlink}
        self.assertTrue(any("Checklist" in t for t in targets))
        self.assertTrue(any("Vulnerabilities" in t for t in targets))
        # Every link points at a sheet that actually exists (no dangling jumps).
        present = {f"'{n}'" for n in wb.sheetnames}
        for loc in targets:
            self.assertTrue(loc.split("!")[0] in present, f"dangling link: {loc}")

    def test_credentialed_access_matrix(self):
        from recce.models import Host, Port, Vuln

        def cv(t):
            return Vuln(ip="x", port=445, protocol="tcp", script_id="c",
                        title=t, severity="high", source="cred")
        hosts = [
            Host(ip="10.0.10.10", hostnames=["dc01"], os_family="Windows",
                 cred_enumerated=True, ports=[Port(portid=445, service="microsoft-ds")],
                 vulns=[cv("Local admin confirmed - privileged account"),
                        cv("Credential hashes dumped (5 accounts)")]),
            Host(ip="10.0.10.25", hostnames=["ws01"], os_family="Windows",
                 cred_enumerated=True, ports=[Port(portid=445, service="microsoft-ds")],
                 vulns=[cv("Local admin confirmed - user account")]),
        ]
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(hosts, out)
            rows = xlsx.read_sheets(out)["Overview"]
        # Find the matrix header + the two host rows.
        text = ["|".join(str(c) for c in r) for r in rows]
        self.assertTrue(any("access matrix" in t for t in text))
        dc = next(r for r in rows if r and str(r[0]).startswith("10.0.10.10"))
        ws = next(r for r in rows if r and str(r[0]).startswith("10.0.10.25"))
        # dc01: privileged account is admin + hashes dumped; user account is not admin.
        self.assertEqual([dc[2], dc[3], dc[4]], ["—", "✓", "✓"])
        # ws01: the LOW-PRIV user account is admin (over-privileged) -> flagged.
        self.assertEqual([ws[2], ws[3], ws[4]], ["✓", "—", "—"])

    def test_overview_host_index_deep_links_hit_correct_checklist_rows(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        from recce.report_excel import read_key_order
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(sample_hosts(), out)

            def _check(path):
                import re as _re
                wb = load_workbook(path)
                ck = wb["Checklist"]
                ipc = [c.value for c in ck[1]].index("IP") + 1
                # The IP column also carries the collapsible subnet-band labels; keep
                # only the rows whose IP cell is a bare IPv4 (the real host rows).
                ip_row = {}
                for r in range(2, ck.max_row + 1):
                    v = ck.cell(row=r, column=ipc).value
                    if isinstance(v, str) and _re.fullmatch(r"\d+\.\d+\.\d+\.\d+", v):
                        ip_row[v] = r
                ov = wb["Overview"]
                deep = {c.value: c.hyperlink.location
                        for row in ov.iter_rows() for c in row
                        if c.hyperlink and c.hyperlink.location.startswith(
                            "'Checklist'!A") and c.value in ip_row}
                # Every host is indexed, and each link targets its real row.
                self.assertEqual(set(deep), set(ip_row))
                for ip, loc in deep.items():
                    self.assertEqual(loc, f"'Checklist'!A{ip_row[ip]}")

            _check(out)
            # Regenerate preserving row order -> links must still be correct.
            build_workbook(sample_hosts(), out, order_map=read_key_order(out))
            _check(out)

    def test_step_headers_colour_auto_vs_manual(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        from recce import tracking as tr
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(sample_hosts(), out)
            ws = load_workbook(out)["Checklist"]
        hdr = [c.value for c in ws[1]]
        auto = {"Enumerated", "Vuln-scan", "Web", "DB", "Priv-esc"}
        manual = {"AD", "Access", "Creds", "Lateral"}
        for h in auto:
            c = ws.cell(row=1, column=hdr.index(h) + 1)
            self.assertEqual(c.fill.fgColor.rgb, "FF2E7D32", f"{h} should be auto-green")
        for h in manual:
            c = ws.cell(row=1, column=hdr.index(h) + 1)
            self.assertEqual(c.fill.fgColor.rgb, "FFC55A11", f"{h} should be manual-amber")
        # Sanity: the split matches the tracking module's source of truth.
        self.assertEqual(manual, {h for h, s in tr.STEP_COLUMNS.items()
                                  if s in tr.MANUAL_STEPS})

    def test_grouped_sheet_has_collapsible_host_bands(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(sample_hosts(), out)
            wb = load_workbook(out)
        sv = wb["Services"]
        self.assertEqual(sv.sheet_format.outlineLevelRow, 1)
        self.assertFalse(sv.sheet_properties.outlinePr.summaryBelow)  # header above
        # Detail rows are grouped (outline level 1); host-header rows are level 0.
        levels = {sv.row_dimensions[r].outlineLevel
                  for r in range(2, sv.max_row + 1)}
        self.assertIn(1, levels)                              # some grouped detail
        self.assertIn(0, levels)                              # some header/summary


class ScannerCommandTest(unittest.TestCase):
    """Verify the actual nmap command assembled for each phase (mock _run)."""

    def _capture(self, fn, *a, **k):
        import recce.scanner as s
        calls = []
        orig = s._run
        s._run = lambda cmd, timeout=None: (calls.append((cmd, timeout))
                                            or s.RunOutcome(returncode=0))
        try:
            fn(*a, **k)
        finally:
            s._run = orig
        return calls

    def test_full_port_scan_flags(self):
        import recce.scanner as s
        with tempfile.TemporaryDirectory() as d:
            calls = self._capture(s.full_port_scan, "1.2.3.4",
                                  os.path.join(d, "p.xml"), s.PROFILES["standard"])
        cmd = calls[0][0]
        self.assertIn("-p-", cmd)                 # full sweep by default
        self.assertIn("--host-timeout", cmd)
        self.assertIn("--max-retries", cmd)
        self.assertIn("1.2.3.4", cmd)
        self.assertIsNotNone(calls[0][1])         # subprocess timeout set

    def test_pn_scan_retries_for_completeness_bounded_by_host_timeout(self):
        # -Pn no longer drops to a single retry (that silently lost open ports on
        # any lossy link). It uses the profile's max_retries (default 3) so a
        # dropped SYN doesn't lose a port; dead IPs are bounded by --host-timeout.
        import copy
        import recce.scanner as s
        prof = copy.copy(s.PROFILES["standard"])
        prof.assume_up = True
        with tempfile.TemporaryDirectory() as d:
            calls = self._capture(s.full_port_scan, "1.2.3.4",
                                  os.path.join(d, "p.xml"), prof)
        cmd = calls[0][0]
        self.assertEqual(cmd[cmd.index("--max-retries") + 1], "3")   # completeness
        self.assertIn("--host-timeout", cmd)                         # bounds dead IPs
        self.assertIn("--min-rate", cmd)                             # packet floor

    def test_port_sweep_auto_retries_reliably_on_dropped_probes(self):
        """A rate-limiting network (nmap drops probes) must trigger an automatic
        congestion-adaptive re-scan - no --min-rate floor, more retries, -T3 -
        which is what actually finds the ports (the fast pass under-reports)."""
        import recce.scanner as s
        calls = []
        outs = iter([
            s.RunOutcome(returncode=0, stderr="Increasing send delay for 1.2.3.4 "
                         "from 0 to 5 due to 11 out of 11 dropped probes"),
            s.RunOutcome(returncode=0),
        ])
        orig = s._run
        s._run = lambda cmd, timeout=None: (calls.append(cmd) or next(outs))
        try:
            _, issue = s.full_port_scan("1.2.3.4", "/tmp/x.xml", s.ScanProfile())
        finally:
            s._run = orig
        self.assertEqual(len(calls), 2)                       # fast pass + reliable re-scan
        self.assertIn("--min-rate", calls[0])                 # fast pass keeps the floor
        self.assertNotIn("--min-rate", calls[1])              # reliable drops it
        self.assertEqual(calls[1][calls[1].index("--max-retries") + 1], "6")
        self.assertIn("-T3", calls[1])
        # bounded: the adaptive re-scan keeps the SAME --host-timeout as any host
        # (no silent extension), so it can't run for hours - it returns partial.
        self.assertIn("--host-timeout", calls[1])
        self.assertEqual(calls[1][calls[1].index("--host-timeout") + 1],
                         f"{s.ScanProfile().host_timeout}m")
        self.assertTrue(issue and issue.level == "warning")   # rate-limit surfaced

    def test_reliable_flag_drops_min_rate_from_first_pass(self):
        import recce.scanner as s
        calls = []
        orig = s._run
        s._run = lambda cmd, timeout=None: (calls.append(cmd)
                                            or s.RunOutcome(returncode=0))
        prof = s.ScanProfile(reliable=True)
        try:
            s.full_port_scan("1.2.3.4", "/tmp/y.xml", prof)
        finally:
            s._run = orig
        self.assertEqual(len(calls), 1)                       # no wasted fast pass
        self.assertNotIn("--min-rate", calls[0])
        self.assertEqual(calls[0][calls[0].index("--max-retries") + 1], "6")

    def test_clean_fast_pass_does_not_rescan(self):
        """No dropped probes -> a single scan, no wasteful reliable re-run."""
        import recce.scanner as s
        calls = []
        orig = s._run
        s._run = lambda cmd, timeout=None: (calls.append(cmd)
                                            or s.RunOutcome(returncode=0))
        try:
            s.full_port_scan("1.2.3.4", "/tmp/z.xml", s.ScanProfile())
        finally:
            s._run = orig
        self.assertEqual(len(calls), 1)

    def test_enum_scan_flags(self):
        import recce.scanner as s
        with tempfile.TemporaryDirectory() as d:
            calls = self._capture(s.enum_scan, "1.2.3.4", [80, 445],
                                  os.path.join(d, "e.xml"), s.PROFILES["standard"])
        cmd = calls[0][0]
        j = " ".join(cmd)
        self.assertIn("-sV", cmd)
        self.assertIn("--version-intensity", cmd)     # standard = intensity gate
        self.assertIn("--host-timeout", cmd)
        self.assertIn("smb-os-discovery", j)          # AD enrichment scripts
        self.assertIn("80,445", j)                    # exactly the ports given

    def test_vuln_scan_safe_vs_aggressive(self):
        import recce.scanner as s
        with tempfile.TemporaryDirectory() as d:
            safe = self._capture(s.vuln_scan, "1.2.3.4", [80],
                                 os.path.join(d, "v.xml"), s.PROFILES["standard"])
            agg = self._capture(s.vuln_scan, "1.2.3.4", [80],
                                os.path.join(d, "v.xml"), s.PROFILES["standard"],
                                aggressive=True)
        safe_j, agg_j = " ".join(safe[0][0]), " ".join(agg[0][0])
        self.assertIn("vuln and safe", safe_j)
        self.assertIn("--version-light", safe[0][0])   # not a full re-scan
        self.assertIn("vuln or vulners", agg_j)
        # KEY FIX: high-value detection scripts that nmap does NOT tag "safe"
        # (ms17-010, heartbleed, vsftpd backdoor) still run in the default scan -
        # no flag needed. --aggressive adds the full intrusive vuln category.
        for script in ("smb-vuln-ms17-010", "ssl-heartbleed", "ftp-vsftpd-backdoor"):
            self.assertIn(script, safe_j)
            self.assertIn(script, agg_j)

    def test_vuln_scan_fast_tier(self):
        import recce.scanner as s
        with tempfile.TemporaryDirectory() as d:
            fast = self._capture(s.vuln_scan, "1.2.3.4", [80, 445],
                                 os.path.join(d, "v.xml"), s.PROFILES["standard"],
                                 fast=True)
        fast_j = " ".join(fast[0][0])
        # --fast drops the broad category net + deep enum, keeps top-signal detection.
        self.assertNotIn("vuln and safe", fast_j)
        self.assertNotIn("http-enum", fast_j)          # deep-enum script excluded
        self.assertIn("smb-vuln-ms17-010", fast_j)     # top-signal detection kept
        self.assertIn("ssl-heartbleed", fast_j)
        self.assertIn("90s", fast_j)                   # lighter script-timeout

    def test_enum_deep_scripts_on_standard_not_quick(self):
        import recce.scanner as s
        with tempfile.TemporaryDirectory() as d:
            std = self._capture(s.enum_scan, "1.2.3.4", [80],
                                os.path.join(d, "e.xml"), s.PROFILES["standard"])
            quick = self._capture(s.enum_scan, "1.2.3.4", [80],
                                  os.path.join(d, "e2.xml"), s.PROFILES["quick"])
        # Deep service-enum scripts run in enum on standard, dropped on quick.
        self.assertIn("http-enum", " ".join(std[0][0]))
        self.assertNotIn("http-enum", " ".join(quick[0][0]))

    def test_version_all_profile_uses_version_all(self):
        import recce.scanner as s
        with tempfile.TemporaryDirectory() as d:
            calls = self._capture(s.enum_scan, "1.2.3.4", [80],
                                  os.path.join(d, "e.xml"), s.PROFILES["thorough"])
        self.assertIn("--version-all", calls[0][0])

    def test_no_ports_writes_empty_xml_and_no_scan(self):
        import recce.scanner as s
        with tempfile.TemporaryDirectory() as d:
            xmlp = os.path.join(d, "e.xml")
            calls = self._capture(s.enum_scan, "1.2.3.4", [], xmlp,
                                  s.PROFILES["standard"])
            self.assertEqual(calls, [])                # nothing scanned
            self.assertTrue(os.path.exists(xmlp))      # but a parseable stub exists
            self.assertEqual(parser.parse_nmap_xml(xmlp), [])


class MarkdownCsvFidelityTest(unittest.TestCase):
    def test_markdown_attributes_findings_to_correct_host(self):
        from recce.report_markdown import build_markdown
        with tempfile.TemporaryDirectory() as d:
            md = os.path.join(d, "r.md")
            build_markdown(sample_hosts(), md, title="Eng", domains=[])
            with open(md) as fh:
                text = fh.read()
        self.assertIn("Eng", text)
        self.assertIn("10.0.10.10", text)
        # The DC's finding is present and tied to the DC's IP line.
        dc_line = next(ln for ln in text.splitlines()
                       if "ms17-010" in ln)
        self.assertIn("10.0.10.10", dc_line)

    def test_csv_one_row_per_open_port_with_correct_ip(self):
        import csv as csvmod
        from recce.report_markdown import build_csv
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "s.csv")
            build_csv(sample_hosts(), p)
            with open(p) as fh:
                rows = list(csvmod.reader(fh))
        hdr, data = rows[0], rows[1:]
        self.assertEqual(len(data), sum(len(FACTS[ip]["ports"]) for ip in FACTS))
        ipc, portc = hdr.index("ip"), hdr.index("port")
        # The FTP port row (21) belongs to web02.
        ftp = [r for r in data if r[portc] == "21"]
        self.assertEqual(len(ftp), 1)
        self.assertEqual(ftp[0][ipc], "10.0.20.6")
        # Every row's port genuinely belongs to that row's IP.
        for r in data:
            self.assertIn(int(r[portc]), FACTS[r[ipc]]["ports"])


def _nmap_xml(ip, ports):
    """Build a minimal, parseable nmap XML for a host with the given ports.

    ports: list of dicts {port, service, product?, version?, scripts?:[(id,out)]}.
    """
    body = [f'<host><status state="up"/>'
            f'<address addr="{ip}" addrtype="ipv4"/><ports>']
    for p in ports:
        body.append(f'<port protocol="tcp" portid="{p["port"]}">'
                    f'<state state="open"/>')
        svc = f'<service name="{p.get("service", "")}"'
        if p.get("product"):
            svc += f' product="{p["product"]}"'
        if p.get("version"):
            svc += f' version="{p["version"]}"'
        body.append(svc + '/>')
        for sid, out in p.get("scripts", []):
            body.append(f'<script id="{sid}" output="{out}"/>')
        body.append('</port>')
    body.append('</ports></host>')
    return '<?xml version="1.0"?><nmaprun start="1">' + "".join(body) + '</nmaprun>'


def _fake_scan(out, ip, ports):
    """Write a canned nmap XML and return the (path, issue) tuple a scan fn does."""
    with open(out, "w") as fh:
        fh.write(_nmap_xml(ip, ports))
    return out, None


class ModelSerializationTest(unittest.TestCase):
    """Host/Domain survive the exact JSON round-trip the store uses (no field loss)."""

    def _rich_host(self):
        from recce.models import Account, Exploit, Script
        return Host(
            ip="10.0.0.5", subnet="10.0.0.0/24", hostnames=["h1", "h1.corp"],
            os_name="Linux 5.4", os_family="Linux", os_accuracy=95,
            roles=["Domain Controller"], ntlm={"domain": "CORP"},
            smb_signing="not required", enumerated=True, db_scanned=True,
            privesc_checked=True, cred_enumerated=True, notes="a note",
            ports=[Port(portid=445, service="microsoft-ds", product="Samba",
                        version="4.13", vuln_scanned=True,
                        scripts=[Script(id="smb-os", output="x", elements={"k": "v"})])],
            vulns=[Vuln(ip="10.0.0.5", port=445, protocol="tcp", script_id="v",
                        title="t", severity="high", ids=["CVE-2020-1"],
                        cwes=["CWE-78"], source="version-db", confidence="likely")],
            accounts=[Account(ip="10.0.0.5", source="smb", kind="user", name="a",
                              domain="CORP", rid="500", attrs={"spn": "x"})],
            exploits=[Exploit(ip="10.0.0.5", port=445, edb_id="123",
                              title="e", cves=["CVE-2020-1"])],
            host_scripts=[Script(id="hs", output="o")])

    def test_host_json_roundtrip_via_store_encoding(self):
        import json
        h = self._rich_host()
        h2 = Host.from_json(json.loads(json.dumps(h.to_json())))
        # Scalars + all progress flags.
        for f in ("ip", "subnet", "os_name", "os_family", "smb_signing", "notes",
                  "enumerated", "db_scanned", "privesc_checked", "cred_enumerated"):
            self.assertEqual(getattr(h2, f), getattr(h, f), f)
        self.assertEqual(h2.hostnames, h.hostnames)
        self.assertEqual(h2.roles, h.roles)
        self.assertEqual(h2.ntlm, h.ntlm)
        # Nested structures.
        self.assertEqual(h2.ports[0].scripts[0].elements, {"k": "v"})
        self.assertTrue(h2.ports[0].vuln_scanned)
        self.assertEqual(h2.vulns[0].cwes, ["CWE-78"])       # newest field survives
        self.assertEqual(h2.vulns[0].ids, ["CVE-2020-1"])
        self.assertEqual(h2.accounts[0].attrs, {"spn": "x"})
        self.assertEqual(h2.exploits[0].edb_id, "123")
        self.assertEqual(h2.host_scripts[0].id, "hs")

    def test_domain_json_roundtrip(self):
        from recce.models import Domain
        import json
        d = Domain(name="corp.local", netbios="CORP", dc_ips=["10.0.10.10"],
                   anonymous_bind=True, password_policy={"min": 7},
                   trusts=[{"name": "x"}], sources=["nse"])
        d2 = Domain.from_json(json.loads(json.dumps(d.to_json())))
        self.assertEqual(d2.name, "corp.local")
        self.assertEqual(d2.dc_ips, ["10.0.10.10"])
        self.assertTrue(d2.anonymous_bind)
        self.assertEqual(d2.password_policy, {"min": 7})


class PhaseWorkerTest(unittest.TestCase):
    """The real enum/vuln/db/privesc workers, with scanner mocked (no nmap)."""

    def _paths(self, d):
        from recce import cli
        return cli._open_paths(d)

    def test_enum_worker_folds_ports_flags_and_runs_vulndb(self):
        from recce import cli
        import recce.scanner as s
        orig = s.enum_scan
        s.enum_scan = lambda ip, ports, out, profile, creds=None: _fake_scan(
            out, ip, [{"port": 21, "service": "ftp", "product": "vsftpd",
                       "version": "2.3.4"}])
        try:
            with tempfile.TemporaryDirectory() as d:
                paths = self._paths(d)
                host, issues = cli._enum_worker(
                    "10.0.0.5", s.PROFILES["standard"], paths, None,
                    {"10.0.0.5": [21]}, {"10.0.0.5": "10.0.0.0/24"})
        finally:
            s.enum_scan = orig
        self.assertTrue(host.enumerated)
        self.assertEqual([p.portid for p in host.open_ports], [21])
        self.assertEqual(host.subnet, "10.0.0.0/24")
        # vulndb ran inside the worker and flagged the vsftpd backdoor.
        self.assertTrue(any("vsftpd 2.3.4 backdoor" in v.title for v in host.vulns))
        self.assertEqual(issues, [])

    def test_vuln_worker_merges_findings_and_marks_ports(self):
        from recce import cli
        import recce.scanner as s
        orig = s.vuln_scan
        s.vuln_scan = lambda ip, ports, out, profile, creds=None, aggressive=False, \
            fast=False: \
            _fake_scan(out, ip, [{"port": 80, "service": "http", "scripts": [
                ("http-vuln-x", "VULNERABLE: demo issue State: VULNERABLE")]}])
        try:
            with tempfile.TemporaryDirectory() as d:
                paths = self._paths(d)
                host = Host(ip="10.0.0.5", ports=[Port(portid=80, service="http",
                            state="open")])
                host, issues = cli._vuln_worker(
                    host, [80], s.PROFILES["standard"], paths, None,
                    aggressive=False, use_ss=False, use_probes=False)
        finally:
            s.vuln_scan = orig
        self.assertTrue(host.ports[0].vuln_scanned)          # port marked done
        self.assertTrue(any("http-vuln-x" in (v.script_id or "") for v in host.vulns))
        self.assertEqual(issues, [])

    def test_db_worker_sets_db_scanned(self):
        from recce import cli
        import recce.scanner as s
        orig = s.nse_scan
        s.nse_scan = lambda ip, ports, out, profile, scripts, creds=None: _fake_scan(
            out, ip, [{"port": 3306, "service": "mysql"}])
        try:
            with tempfile.TemporaryDirectory() as d:
                paths = self._paths(d)
                host = Host(ip="10.0.0.5", ports=[Port(portid=3306, service="mysql",
                            state="open")])
                host, issues = cli._db_worker(host, [3306], s.PROFILES["standard"],
                                              paths, None, aggressive=False,
                                              use_ss=False)
        finally:
            s.nse_scan = orig
        self.assertTrue(host.db_scanned)
        self.assertTrue(host.ports[0].vuln_scanned)
        self.assertEqual(issues, [])

    def test_privesc_worker_returns_host_and_issues(self):
        from recce import cli
        import recce.scanner as s
        orig = s.nse_scan
        s.nse_scan = lambda ip, ports, out, profile, scripts, creds=None: _fake_scan(
            out, ip, [{"port": 445, "service": "microsoft-ds"}])
        try:
            with tempfile.TemporaryDirectory() as d:
                paths = self._paths(d)
                host = Host(ip="10.0.0.9", os_family="Windows",
                            ports=[Port(portid=445, service="microsoft-ds",
                                        state="open")])
                host, issues = cli._privesc_worker(host, s.PROFILES["standard"],
                                                   paths, None, aggressive=False)
        finally:
            s.nse_scan = orig
        self.assertEqual(host.ip, "10.0.0.9")
        self.assertIsInstance(issues, list)


class EnvironmentAndTargetsTest(unittest.TestCase):
    def test_check_environment_requires_nmap_and_warns(self):
        import recce.scanner as s
        from recce.scanner import ScanProfile, ScannerError
        oh, orr = s._have, s._is_root
        try:
            s._have = lambda t: False        # nothing installed
            s._is_root = lambda: False
            with self.assertRaises(ScannerError):
                s.check_environment(ScanProfile())      # nmap missing -> raise
            s._have = lambda t: t == "nmap"             # only nmap present
            warns = s.check_environment(ScanProfile())
            self.assertTrue(any("root" in w.lower() for w in warns))
            # masscan requested but absent -> warn + fall back to nmap.
            prof = ScanProfile(scanner="masscan")
            warns = s.check_environment(prof)
            self.assertEqual(prof.scanner, "nmap")
            self.assertTrue(any("masscan" in w.lower() for w in warns))
        finally:
            s._have, s._is_root = oh, orr

    def test_load_targets_from_file_with_comments_and_cidr(self):
        from recce.targets import load_targets
        with tempfile.TemporaryDirectory() as d:
            f = os.path.join(d, "scope.txt")
            with open(f, "w") as fh:
                fh.write("10.0.0.1\n# a comment\n10.0.0.2   # trailing\n"
                         "10.0.1.0/30\n\n")
            hosts, sm = load_targets(["@" + f])
        self.assertIn("10.0.0.1", hosts)
        self.assertIn("10.0.0.2", hosts)
        self.assertIn("10.0.1.1", hosts)                 # expanded from the CIDR
        self.assertEqual(sm["10.0.1.1"], "10.0.1.0/30")  # CIDR line -> subnet label

    def test_missing_target_file_raises(self):
        from recce.targets import load_targets
        with self.assertRaises(FileNotFoundError):
            load_targets(["@/no/such/file.txt"])


class TargetingFormE2ETest(unittest.TestCase):
    """Drive the REAL `vulns` command end-to-end (parser -> selection -> workers ->
    store) and prove each targeting form selects EXACTLY the right stored hosts.

    This locks in the core promise that every phase works on a single IP, several
    IPs, a dash range, a whole subnet, an @file, or 'everything' - using the same
    CLI grammar, with no cross-host bleed. The scanner is mocked so no nmap runs;
    the selection layer under test is entirely real.

    Seeded scope (from the bundled sample): 10.0.10.10, 10.0.10.25 (10.0.10.0/24)
    and 10.0.20.5, 10.0.20.6 (10.0.20.0/24).
    """

    ALL = {"10.0.10.10", "10.0.10.25", "10.0.20.5", "10.0.20.6"}

    def _run_vulns(self, targets):
        """Run `vulns <targets>` against a freshly seeded store; return the set of
        IPs that actually got vuln-scanned (i.e. the hosts the phase selected)."""
        from recce import cli
        import recce.scanner as s

        d = tempfile.mkdtemp()
        self.addCleanup(lambda: shutil.rmtree(d, ignore_errors=True))
        paths = cli._open_paths(d)
        seed = Store(paths["db"])
        seed.set_meta("engagement", "e2e")
        for h in sample_hosts():
            seed.upsert_host(h)
        seed.close()

        def fake_vuln(ip, portids, out, profile, creds=None, aggressive=False,
                      fast=False):
            # Echo the requested ports back as an open-port XML the worker parses.
            return _fake_scan(out, ip, [{"port": p, "service": "tcp"}
                                        for p in portids])

        oc, ov, ou = s.check_environment, s.vuln_scan, s.udp_scan
        s.check_environment = lambda profile: []          # no nmap/root needed
        s.vuln_scan = fake_vuln
        s.udp_scan = lambda ip, out, profile: _fake_scan(out, ip, [])
        argv = ["vulns", *targets, "-o", d,
                "--no-searchsploit", "--no-probes", "--workers", "1"]
        try:
            with contextlib.redirect_stdout(io.StringIO()):
                rc = cli.main(argv)
        finally:
            s.check_environment, s.vuln_scan, s.udp_scan = oc, ov, ou
        self.assertEqual(rc, 0)

        store = Store(paths["db"])
        try:
            scanned = {h.ip for h in store.all_hosts()
                       if any(p.vuln_scanned for p in h.open_ports)}
            # Every seeded host must still be present (selection never drops rows).
            self.assertEqual({h.ip for h in store.all_hosts()}, self.ALL)
        finally:
            store.close()
        return scanned

    def test_single_ip(self):
        self.assertEqual(self._run_vulns(["10.0.10.10"]), {"10.0.10.10"})

    def test_several_ips(self):
        self.assertEqual(self._run_vulns(["10.0.10.10", "10.0.20.6"]),
                         {"10.0.10.10", "10.0.20.6"})

    def test_dash_range(self):
        # .10-.25 covers both 10.0.10.x hosts and nothing in 10.0.20.x.
        self.assertEqual(self._run_vulns(["10.0.10.10-25"]),
                         {"10.0.10.10", "10.0.10.25"})

    def test_range_excludes_outside(self):
        # A range that stops before .25 must NOT pick it up.
        self.assertEqual(self._run_vulns(["10.0.10.1-15"]), {"10.0.10.10"})

    def test_whole_subnet(self):
        self.assertEqual(self._run_vulns(["10.0.20.0/24"]),
                         {"10.0.20.5", "10.0.20.6"})

    def test_mixed_single_and_subnet(self):
        self.assertEqual(self._run_vulns(["10.0.10.25", "10.0.20.0/24"]),
                         {"10.0.10.25", "10.0.20.5", "10.0.20.6"})

    def test_at_file(self):
        d = tempfile.mkdtemp()
        self.addCleanup(lambda: shutil.rmtree(d, ignore_errors=True))
        f = os.path.join(d, "scope.txt")
        with open(f, "w") as fh:
            fh.write("10.0.10.10\n"
                     "10.0.20.0/24   # the linux subnet\n")
        self.assertEqual(self._run_vulns(["@" + f]),
                         {"10.0.10.10", "10.0.20.5", "10.0.20.6"})

    def test_empty_targets_selects_everything(self):
        self.assertEqual(self._run_vulns([]), self.ALL)


class UsabilityAndDiscoveryTest(unittest.TestCase):
    def test_pn_alias_parses(self):
        from recce import cli
        for argv in (["enum", "10.0.0.0/24", "-Pn"],
                     ["enum", "10.0.0.0/24", "--no-discovery"],
                     ["scan", "10.0.0.5", "-Pn"]):
            self.assertTrue(cli.build_arg_parser().parse_args(argv).no_discovery, argv)

    def test_bare_recce_prints_quickstart_not_error(self):
        from recce import cli
        buf = io.StringIO()
        with contextlib.redirect_stdout(buf):
            rc = cli.main([])
        self.assertEqual(rc, 0)                        # no argparse error exit
        out = buf.getvalue()
        self.assertIn("recce enum", out)
        self.assertIn("-Pn", out)                      # the ping-blocking hint

    def test_zero_discovery_auto_falls_back_to_pn(self):
        # The killer field bug: hosts that block ping got dropped -> zero ports.
        # Now 0 discovery responses must auto-fall-back to scanning all as up.
        from recce import cli
        import recce.scanner as s
        from recce.store import Store

        def empty_disc(tf, out):
            with open(out, "w") as fh:
                fh.write('<?xml version="1.0"?><nmaprun></nmaprun>')
            return out, None

        def fps(ip, out, profile):
            with open(out, "w") as fh:
                fh.write(f'<?xml version="1.0"?><nmaprun><host><status state="up"/>'
                         f'<address addr="{ip}" addrtype="ipv4"/><ports><port '
                         f'protocol="tcp" portid="445"><state state="open"/>'
                         f'<service name="microsoft-ds"/></port></ports></host></nmaprun>')
            return out, None

        def enum(ip, ports, out, profile, creds=None):
            return fps(ip, out, profile)

        saved = (s.check_environment, s.discover_hosts, s.full_port_scan, s.enum_scan)
        s.check_environment = lambda p: []
        s.discover_hosts, s.full_port_scan, s.enum_scan = empty_disc, fps, enum
        try:
            with tempfile.TemporaryDirectory() as d:
                buf = io.StringIO()
                with contextlib.redirect_stdout(buf):
                    rc = cli.main(["enum", "10.0.10.10", "10.0.10.11",
                                   "-o", d, "--workers", "1"])
                self.assertEqual(rc, 0)
                self.assertIn("Falling back to -Pn", buf.getvalue())
                hosts = Store(os.path.join(d, "results.sqlite")).all_hosts()
                # Both ping-blocking hosts still got enumerated with their ports.
                self.assertEqual(len(hosts), 2)
                self.assertTrue(all(h.open_ports for h in hosts))
        finally:
            (s.check_environment, s.discover_hosts,
             s.full_port_scan, s.enum_scan) = saved


class PhaseIdempotencyTest(unittest.TestCase):
    """Re-running a phase must NOT duplicate rows. Guards the core store-merge
    contract: hosts/vulns/accounts/exploits/issues dedupe on re-scan."""

    def _counts(self, db):
        s = Store(db)
        try:
            hosts = s.all_hosts()
            return {
                "hosts": len(hosts),
                "vulns": sum(len(h.vulns) for h in hosts),
                "accounts": sum(len(h.accounts) for h in hosts),
                "exploits": sum(len(h.exploits) for h in hosts),
                "issues": s.count_issues().get("total", 0),
            }
        finally:
            s.close()

    def test_rerunning_vulns_does_not_duplicate(self):
        from recce import cli
        import recce.scanner as s

        d = tempfile.mkdtemp()
        self.addCleanup(lambda: shutil.rmtree(d, ignore_errors=True))
        paths = cli._open_paths(d)
        seed = Store(paths["db"])
        seed.set_meta("engagement", "idem")
        for h in sample_hosts():
            seed.upsert_host(h)
        seed.close()

        def fake_vuln(ip, portids, out, profile, creds=None, aggressive=False,
                      fast=False):
            # Emit a real NSE finding + an issue every run, so a broken dedup WOULD
            # grow the counts.
            return _fake_scan(out, ip, [{"port": 80, "service": "http", "scripts": [
                ("http-vuln-x", "VULNERABLE: demo State: VULNERABLE")]}])

        oc, ov, ou = s.check_environment, s.vuln_scan, s.udp_scan
        s.check_environment = lambda profile: []
        s.vuln_scan = fake_vuln
        s.udp_scan = lambda ip, out, profile: _fake_scan(out, ip, [])
        argv = ["vulns", "-o", d, "--no-searchsploit", "--no-probes", "--workers", "1"]
        try:
            with contextlib.redirect_stdout(io.StringIO()):
                self.assertEqual(cli.main(argv), 0)
            first = self._counts(paths["db"])
            with contextlib.redirect_stdout(io.StringIO()):
                self.assertEqual(cli.main(argv), 0)          # SAME phase again
            second = self._counts(paths["db"])
        finally:
            s.check_environment, s.vuln_scan, s.udp_scan = oc, ov, ou
        self.assertEqual(first, second, f"re-run changed counts: {first} -> {second}")
        self.assertGreater(first["vulns"], 0)                # actually produced findings

    def test_store_merge_is_idempotent_for_all_collections(self):
        # Upserting the identical rich host twice must not grow any collection.
        from recce.models import Account, Exploit, Script
        d = tempfile.mkdtemp()
        self.addCleanup(lambda: shutil.rmtree(d, ignore_errors=True))
        s = Store(os.path.join(d, "r.sqlite"))
        h = Host(
            ip="10.0.0.5", os_family="Linux",
            ports=[Port(portid=445, state="open", service="smb",
                        scripts=[Script(id="smb-os", output="o")])],
            vulns=[Vuln(ip="10.0.0.5", port=445, protocol="tcp", script_id="v",
                        title="t", severity="high")],
            accounts=[Account(ip="10.0.0.5", source="nxc", kind="user", name="a",
                              domain="C", rid="500")],
            exploits=[Exploit(ip="10.0.0.5", port=445, edb_id="1", title="e")],
            host_scripts=[Script(id="hs", output="o")],
            local_findings=[{"category": "sudo", "vector": "NOPASSWD"}])
        s.upsert_host(h)
        s.upsert_host(Host.from_json(h.to_json()))          # identical, again
        got = s.get_host("10.0.0.5")
        s.close()
        self.assertEqual(len(got.vulns), 1)
        self.assertEqual(len(got.accounts), 1)
        self.assertEqual(len(got.exploits), 1)
        self.assertEqual(len(got.host_scripts), 1)
        self.assertEqual(len(got.local_findings), 1)
        self.assertEqual(len(got.ports), 1)


_LOOT_LINUX = """\
recce-enum  host=web01  user=www-data  Mon Jul 20 12:00:00 UTC 2026

==== System & kernel ====
    Linux web01 5.4.0-42-generic
[!] Old kernel (5.4.0) - run a local-exploit suggester offline

==== Sudo ====
[!] NOPASSWD sudo entries present -> check GTFOBins for the allowed binaries
[!] sudo grants (ALL) ALL -> full root

==== SUID / SGID / capabilities ====
[!] SUID /usr/bin/find - GTFOBins escalation candidate

==== How to exploit (reference for the [!] findings above) ====
  Sudo: NOPASSWD / (ALL) ALL
      sudo <binary> ; see GTFOBins
[!] THIS LINE MUST NOT BE INGESTED (it lives in the how-to section)

==== Writable files & PATH hijack ====
[!] /etc/shadow is READABLE by www-data -> crack hashes
"""

_LOOT_WIN = """\
recce-enum  host=DBSRV01  user=svc_sql  07/20/2026 12:00:00

==== current context ====
[!] Token holds SeImpersonate -> SYSTEM via Potato (GodPotato/PrintSpoofer)

==== AlwaysInstallElevated ====
[!] AlwaysInstallElevated is set (HKLM+HKCU) -> install a malicious MSI as SYSTEM
"""


_GNMAP = ("# Nmap 7.94 scan initiated\n"
          "Host: 10.0.20.6 (web02)\tStatus: Up\n"
          "Host: 10.0.20.6 (web02)\tPorts: 21/open/tcp//ftp//vsftpd 2.3.4/, "
          "22/open/tcp//ssh//OpenSSH 7.4/, 80/open/tcp//http//Apache httpd 2.4.49/"
          "\tIgnored State: closed (997)\n"
          "Host: 10.0.20.6 (web02)\tOS: Linux 5.4\n")


class NmapImportTest(unittest.TestCase):
    def test_split_product_version(self):
        from recce.parser import _split_product_version
        self.assertEqual(_split_product_version("OpenSSH 8.2p1 Ubuntu"),
                         ("OpenSSH", "8.2p1"))
        self.assertEqual(_split_product_version("Apache httpd 2.4.49"),
                         ("Apache httpd", "2.4.49"))
        self.assertEqual(_split_product_version("Microsoft Windows RPC"),
                         ("Microsoft Windows RPC", ""))

    def test_parse_gnmap(self):
        from recce import parser
        with tempfile.TemporaryDirectory() as d:
            f = os.path.join(d, "s.gnmap")
            with open(f, "w") as fh:
                fh.write(_GNMAP)
            hosts = parser.parse_gnmap(f)
            self.assertEqual(len(hosts), 1)
            h = hosts[0]
            self.assertEqual(h.ip, "10.0.20.6")
            self.assertIn("web02", h.hostnames)
            self.assertEqual({p.portid for p in h.ports}, {21, 22, 80})
            ftp = next(p for p in h.ports if p.portid == 21)
            self.assertEqual(ftp.product, "vsftpd")
            self.assertEqual(ftp.version, "2.3.4")
            self.assertEqual(h.os_family, "Linux")

    def test_parse_normal_text(self):
        from recce import parser
        normal = ("Nmap scan report for web02 (10.0.20.6)\n"
                  "Host is up (0.00042s latency).\n"
                  "PORT   STATE SERVICE VERSION\n"
                  "21/tcp open  ftp     vsftpd 2.3.4\n"
                  "80/tcp open  http    Apache httpd 2.4.49\n"
                  "445/tcp closed microsoft-ds\n")
        with tempfile.TemporaryDirectory() as d:
            f = os.path.join(d, "s.nmap")
            with open(f, "w") as fh:
                fh.write(normal)
            hosts = parser.parse_normal(f)
            self.assertEqual(len(hosts), 1)
            h = hosts[0]
            self.assertEqual(h.ip, "10.0.20.6")
            self.assertIn("web02", h.hostnames)
            self.assertEqual({p.portid for p in h.ports}, {21, 80})  # closed dropped
            ftp = next(p for p in h.ports if p.portid == 21)
            self.assertEqual((ftp.product, ftp.version), ("vsftpd", "2.3.4"))

    def test_parse_normal_bare_ip(self):
        from recce import parser
        with tempfile.TemporaryDirectory() as d:
            f = os.path.join(d, "s.nmap")
            with open(f, "w") as fh:
                fh.write("Nmap scan report for 10.0.0.9\n22/tcp open ssh\n")
            hosts = parser.parse_normal(f)
            self.assertEqual(hosts[0].ip, "10.0.0.9")
            self.assertEqual(hosts[0].hostnames, [])

    def test_parse_nmap_file_autodetects_all_formats(self):
        from recce import parser
        with tempfile.TemporaryDirectory() as d:
            # grepable content, no extension -> sniffed
            g = os.path.join(d, "noext_grep")
            with open(g, "w") as fh:
                fh.write(_GNMAP)
            self.assertTrue(parser.parse_nmap_file(g))
            # normal text, no extension -> sniffed
            n = os.path.join(d, "noext_normal")
            with open(n, "w") as fh:
                fh.write("Nmap scan report for 1.2.3.9\n80/tcp open http\n")
            self.assertEqual(parser.parse_nmap_file(n)[0].ip, "1.2.3.9")
            # xml
            x = os.path.join(d, "s.xml")
            with open(x, "w") as fh:
                fh.write('<?xml version="1.0"?><nmaprun><host><status state="up"/>'
                         '<address addr="1.2.3.4" addrtype="ipv4"/><ports>'
                         '<port protocol="tcp" portid="80"><state state="open"/>'
                         '<service name="http"/></port></ports></host></nmaprun>')
            self.assertEqual(parser.parse_nmap_file(x)[0].ip, "1.2.3.4")

    def test_masscan_xml_is_nmap_compatible(self):
        from recce import parser
        with tempfile.TemporaryDirectory() as d:
            f = os.path.join(d, "mass.xml")
            with open(f, "w") as fh:
                fh.write('<?xml version="1.0"?><nmaprun scanner="masscan"><host>'
                         '<address addr="10.0.30.9" addrtype="ipv4"/><ports>'
                         '<port protocol="tcp" portid="5432"><state state="open"/>'
                         '<service name="postgresql"/></port></ports></host></nmaprun>')
            hosts = parser.parse_nmap_file(f)
            self.assertEqual(hosts[0].ip, "10.0.30.9")
            self.assertEqual(hosts[0].open_ports[0].portid, 5432)

    def test_oa_directory_imports_once_prefers_xml(self):
        # A -oA set (base.xml + base.gnmap + base.nmap) must import once, from xml.
        from recce import cli
        with tempfile.TemporaryDirectory() as d:
            for ext, body in ((".gnmap", _GNMAP),
                              (".nmap", "Nmap scan report for x (10.0.20.6)\n"
                                        "21/tcp open ftp vsftpd 2.3.4\n"),
                              (".xml", '<?xml version="1.0"?><nmaprun><host>'
                                       '<status state="up"/><address addr="10.0.20.6" '
                                       'addrtype="ipv4"/><ports><port protocol="tcp" '
                                       'portid="21"><state state="open"/><service '
                                       'name="ftp" product="vsftpd" version="2.3.4"/>'
                                       '</port></ports></host></nmaprun>')):
                with open(os.path.join(d, "scan" + ext), "w") as fh:
                    fh.write(body)
            files = cli._collect_scan_files([d])
            self.assertEqual(len(files), 1)
            self.assertTrue(files[0].endswith(".xml"))

    def test_import_builds_workbook_with_checkmarks_and_findings(self):
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            f = os.path.join(d, "s.gnmap")
            with open(f, "w") as fh:
                fh.write(_GNMAP)
            eng = os.path.join(d, "eng")
            argv = ["import", f, "-o", eng, "--title", "T"]
            args = cli.build_arg_parser().parse_args(argv)
            with contextlib.redirect_stdout(io.StringIO()):
                self.assertEqual(args.func(args), 0)
            h = Store(os.path.join(eng, "results.sqlite")).get_host("10.0.20.6")
            self.assertTrue(h.enumerated)                  # checkmark set
            self.assertEqual(len(h.open_ports), 3)
            # offline version->CVE engine fired on the imported versions
            titles = " ".join(v.title for v in h.vulns)
            self.assertIn("vsftpd 2.3.4 backdoor", titles)
            self.assertIn("path traversal", titles.lower())
            self.assertTrue(os.path.exists(os.path.join(eng, "enumeration.xlsx")))

    def test_import_appends_subnets_and_merges_overlap(self):
        # Several scans (different subnets, then an overlapping host) must APPEND
        # new hosts and MERGE overlaps - never duplicate a host or a port.
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            eng = os.path.join(d, "eng")

            def imp(text):
                f = os.path.join(d, "s.gnmap")
                with open(f, "w") as fh:
                    fh.write(text)
                a = cli.build_arg_parser().parse_args(["import", f, "-o", eng])
                with contextlib.redirect_stdout(io.StringIO()):
                    a.func(a)

            imp("Host: 10.0.10.10 (dc01)\tPorts: 445/open/tcp//microsoft-ds//, "
                "3389/open/tcp//ms-wbt-server//\tIgnored State: closed\n"
                "Host: 10.0.10.25 (ws01)\tPorts: 445/open/tcp//microsoft-ds//"
                "\tIgnored State: closed\n")                       # subnet .10
            imp("Host: 10.0.20.5 (web01)\tPorts: 22/open/tcp//ssh//OpenSSH 8.2p1/"
                "\tIgnored State: closed\n")                       # subnet .20 (appended)
            imp("Host: 10.0.10.10 (dc01)\tPorts: 88/open/tcp//kerberos-sec//, "
                "445/open/tcp//microsoft-ds//\tIgnored State: closed\n")  # overlap
            s = Store(os.path.join(eng, "results.sqlite"))
            hosts = s.all_hosts()
            self.assertEqual(len(hosts), 3)                        # dc01 not duplicated
            self.assertEqual({h.subnet for h in hosts},
                             {"10.0.10.0/24", "10.0.20.0/24"})     # both subnets present
            dc = s.get_host("10.0.10.10")
            self.assertEqual(sorted(p.portid for p in dc.open_ports), [88, 445, 3389])
            s.close()

    def test_import_merges_and_preserves_tracking(self):
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            f = os.path.join(d, "s.gnmap")
            with open(f, "w") as fh:
                fh.write(_GNMAP)
            eng = os.path.join(d, "eng")
            def run():
                a = cli.build_arg_parser().parse_args(["import", f, "-o", eng])
                with contextlib.redirect_stdout(io.StringIO()):
                    a.func(a)
            run()
            # Tick via the `review` command (writes tracking AND regenerates the
            # workbook, the real path a tester's edits take).
            rv = cli.build_arg_parser().parse_args(
                ["review", "--host", "10.0.20.6", "--note", "manually reviewed",
                 "-o", eng])
            with contextlib.redirect_stdout(io.StringIO()):
                rv.func(rv)
            run()                                          # re-import same scan
            s = Store(os.path.join(eng, "results.sqlite"))
            h = s.get_host("10.0.20.6")
            self.assertEqual(len(h.open_ports), 3)         # not duplicated
            self.assertEqual(s.get_tracking().get("host:10.0.20.6"),
                             (True, "manually reviewed"))  # tick preserved
            s.close()


class IngestParserTest(unittest.TestCase):
    def test_parse_findings_and_skip_howto(self):
        from recce import ingest
        p = ingest.parse_loot(_LOOT_LINUX)
        self.assertTrue(p["is_recce"])
        self.assertEqual(p["hostname"], "web01")
        self.assertEqual(p["os"], "linux")
        texts = [f["text"] for f in p["findings"]]
        self.assertEqual(len(texts), 5)                       # 6 [!] lines, 1 in how-to
        self.assertTrue(all("MUST NOT BE INGESTED" not in t for t in texts))
        cats = {f["category"] for f in p["findings"]}
        self.assertTrue({"sudo", "suid", "kernel", "writable"} <= cats)

    def test_windows_detection(self):
        from recce import ingest
        p = ingest.parse_loot(_LOOT_WIN)
        self.assertEqual(p["os"], "windows")
        self.assertEqual(p["hostname"], "DBSRV01")

    def test_strips_ansi_colour(self):
        from recce import ingest
        coloured = ("recce-enum  host=h1\n\x1b[1;36m==== Sudo ====\x1b[0m\n"
                    "\x1b[1;33m[!] NOPASSWD sudo entries present\x1b[0m\n")
        p = ingest.parse_loot(coloured)
        self.assertEqual(len(p["findings"]), 1)
        self.assertEqual(p["findings"][0]["text"], "NOPASSWD sudo entries present")

    def test_dedup(self):
        from recce import ingest
        dupe = ("recce-enum host=h\n==== Sudo ====\n[!] same\n[!] same\n")
        self.assertEqual(len(ingest.parse_loot(dupe)["findings"]), 1)

    def test_empty_and_garbage_input_no_crash(self):
        from recce import ingest
        for blob in ("", "\n\n\n", "not recce output at all\nrandom text\n",
                     "\x00\x01\x02 binary-ish \xff\xfe", "=" * 5000,
                     "[!] finding with no banner and no section\n"):
            p = ingest.parse_loot(blob)          # must not raise
            self.assertIn("findings", p)
            self.assertIsInstance(p["findings"], list)

    def test_findings_without_banner_still_parse(self):
        from recce import ingest
        p = ingest.parse_loot("==== Sudo ====\n[!] NOPASSWD present\n")
        self.assertFalse(p["is_recce"])          # no banner
        self.assertEqual(len(p["findings"]), 1)  # but [!] lines still harvested

    def test_malformed_section_headers_tolerated(self):
        from recce import ingest
        # Ragged '=' fences and stray '=' in finding text must not break parsing.
        blob = ("recce-enum host=h\n=== Weird ==\n[!] a = b = c finding\n"
                "======\n[!] another\n")
        p = ingest.parse_loot(blob)
        self.assertEqual(len(p["findings"]), 2)


class IngestCommandTest(unittest.TestCase):
    def _eng(self, d, host=None):
        from recce.store import Store
        os.makedirs(os.path.join(d, "raw"), exist_ok=True)
        s = Store(os.path.join(d, "results.sqlite"))
        s.set_meta("engagement", "T")
        if host:
            s.upsert_host(host)
        s.close()

    def _ingest(self, d, loot_text, extra=None):
        from recce import cli
        loot = os.path.join(d, "loot.txt")
        with open(loot, "w") as fh:
            fh.write(loot_text)
        argv = ["ingest", loot, "-o", d] + (extra or [])
        args = cli.build_arg_parser().parse_args(argv)
        with contextlib.redirect_stdout(io.StringIO()):
            rc = args.func(args)
        return rc

    def test_ingest_matches_host_by_hostname(self):
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            self._eng(d, Host(ip="10.0.0.50", hostnames=["web01"],
                              os_family="Linux", enumerated=True))
            self.assertEqual(self._ingest(d, _LOOT_LINUX), 0)
            h = Store(os.path.join(d, "results.sqlite")).get_host("10.0.0.50")
            self.assertEqual(len(h.local_findings), 5)
            self.assertTrue(h.privesc_checked)

    def test_ingest_is_idempotent(self):
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            self._eng(d, Host(ip="10.0.0.50", hostnames=["web01"]))
            self._ingest(d, _LOOT_LINUX)
            self._ingest(d, _LOOT_LINUX)               # re-ingest same loot
            h = Store(os.path.join(d, "results.sqlite")).get_host("10.0.0.50")
            self.assertEqual(len(h.local_findings), 5)  # not doubled

    def test_ingest_synthesizes_host_when_unknown(self):
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            self._eng(d)                                # empty engagement
            self.assertEqual(self._ingest(d, _LOOT_WIN), 0)
            h = Store(os.path.join(d, "results.sqlite")).get_host("local:DBSRV01")
            self.assertIsNotNone(h)
            self.assertEqual(h.os_family, "Windows")
            self.assertEqual(len(h.local_findings), 2)

    def test_ingest_host_flag_records_hostname(self):
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            self._eng(d, Host(ip="10.0.0.50"))          # no hostname stored
            self._ingest(d, _LOOT_LINUX, extra=["--host", "10.0.0.50"])
            h = Store(os.path.join(d, "results.sqlite")).get_host("10.0.0.50")
            # hostname from the loot banner is recorded, so a later no---host
            # ingest of the same box matches this entry instead of synthesizing.
            self.assertIn("web01", h.hostnames)

    def test_ingest_dedups_incoming_rows_on_new_host(self):
        # Two sections that map to the same category with identical finding text
        # must not create duplicate rows, even on a brand-new (unmerged) host.
        from recce.store import Store
        loot = ("recce-enum host=h1\n"
                "==== SUID / SGID / capabilities ====\n[!] same finding text\n"
                "==== Capabilities ====\n[!] same finding text\n")
        with tempfile.TemporaryDirectory() as d:
            self._eng(d)                                 # empty -> synthetic host
            self._ingest(d, loot, extra=["--host", "10.0.0.1"])
            h = Store(os.path.join(d, "results.sqlite")).get_host("10.0.0.1")
            self.assertEqual(len(h.local_findings), 1)

    def test_exploitation_sheet_lists_confirmed_findings(self):
        from recce.report_excel import build_workbook
        import openpyxl
        with tempfile.TemporaryDirectory() as d:
            self._eng(d, Host(ip="10.0.0.50", hostnames=["web01"], os_family="Linux"))
            self._ingest(d, _LOOT_LINUX)      # sudo/suid/shadow -> confirmed
            from recce.store import Store
            hosts = Store(os.path.join(d, "results.sqlite")).all_hosts()
            p = os.path.join(d, "wb.xlsx")
            build_workbook(hosts, p)
            wb = openpyxl.load_workbook(p)
            self.assertIn("Exploitation", wb.sheetnames)
            ws = wb["Exploitation"]
            hdr = [c.value for c in ws[1]]
            ti = hdr.index("Existing tool")
            tools = " ".join(str(r[ti]) for r in ws.iter_rows(min_row=2, values_only=True)
                             if r[ti])
            self.assertIn("GTFOBins", tools)      # sudo / SUID findings
            self.assertGreaterEqual(ws.max_row - 1, 2)

    def test_high_signal_findings_promoted_to_vulns(self):
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            self._eng(d, Host(ip="10.0.0.50", hostnames=["web01"], os_family="Linux"))
            self._ingest(d, _LOOT_LINUX)
            h = Store(os.path.join(d, "results.sqlite")).get_host("10.0.0.50")
            local_vulns = [v for v in h.vulns if v.source == "local"]
            titles = " ".join(v.title for v in local_vulns)
            self.assertTrue(local_vulns)                    # some got promoted
            self.assertIn("Sudo misconfiguration", titles)  # NOPASSWD / ALL
            self.assertIn("Readable /etc/shadow", titles)
            # Promoted vulns are confirmed local observations with a CWE.
            self.assertTrue(all(v.confidence == "confirmed" and v.cwes
                                for v in local_vulns))

    def test_promotion_is_idempotent(self):
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            self._eng(d, Host(ip="10.0.0.50", hostnames=["web01"], os_family="Linux"))
            self._ingest(d, _LOOT_LINUX)
            self._ingest(d, _LOOT_LINUX)          # re-ingest
            h = Store(os.path.join(d, "results.sqlite")).get_host("10.0.0.50")
            local = [v for v in h.vulns if v.source == "local"]
            self.assertEqual(len(local), len({v.title for v in local}))  # no dupes

    def test_ingested_findings_appear_on_privesc_sheet(self):
        with tempfile.TemporaryDirectory() as d:
            self._eng(d, Host(ip="10.0.0.50", hostnames=["web01"], os_family="Linux"))
            self._ingest(d, _LOOT_LINUX)
            import openpyxl
            ws = openpyxl.load_workbook(os.path.join(d, "enumeration.xlsx"))["Priv-Esc"]
            hdr = [c.value for c in ws[1]]
            ti = hdr.index("Type")
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            # The 5 ingested findings each become a row verdicted as an escalation
            # path or an observation (this fresh host has no remote findings).
            on_target = sum(1 for r in rows if r[ti] in ("Escalation path", "Finding"))
            self.assertEqual(on_target, 5)
            # ...and at least some are verdicted as actual escalation paths.
            self.assertGreater(sum(1 for r in rows if r[ti] == "Escalation path"), 0)
            # The Priv-Esc tab is now findings-only: NO generic checklist rows
            # (a swept host also gets no 'run recce deploy' to-do).
            self.assertEqual(sum(1 for r in rows if r[ti] in ("Checklist", "To do")), 0)
            # The generic OS checklist lives on the separate reference sheet.
            wb = openpyxl.load_workbook(os.path.join(d, "enumeration.xlsx"))
            self.assertIn("Priv-Esc Playbook", wb.sheetnames)
            pb = wb["Priv-Esc Playbook"]
            self.assertGreater(pb.max_row, 1)         # has playbook rows

    def test_triaged_vuln_counts_toward_coverage(self):
        """Regression: the Vulnerabilities sheet's row key and the coverage
        counter's key must be identical, or ticking Triaged is never counted."""
        from recce import tracking as tr
        from recce.models import Vuln
        from recce.report_excel import _spec_vulns
        h = Host(ip="10.0.0.5", ports=[Port(portid=445, service="microsoft-ds")],
                 vulns=[Vuln(ip="10.0.0.5", port=445, protocol="tcp",
                             script_id="smb-vuln-ms17-010", title="ms17-010 RCE",
                             severity="high", source="nse")])
        sheet_key = _spec_vulns([h]).rows[0]["key"]
        # the sheet key, the canonical key, and the coverage key all agree
        self.assertEqual(sheet_key, tr.vuln_row_key(h.vulns[0]))
        self.assertIn(sheet_key, tr.item_keys([h])["vulns"])
        # untriaged -> 0/1; triaging the sheet's key -> 1/1 (was stuck at 0 before)
        self.assertEqual(tr.compute_coverage([h], {})["vulns"],
                         {"total": 1, "done": 0, "pct": 0})
        cov = tr.compute_coverage([h], {sheet_key: (True, "")})["vulns"]
        self.assertEqual(cov, {"total": 1, "done": 1, "pct": 100})

    def test_vuln_row_key_matches_store_dedup_granularity(self):
        """Regression: the workbook/coverage key must not truncate the title
        more coarsely than the store's dedup key (models.Vuln.key uses [:60]),
        or two store-distinct findings collapse to one Vulnerabilities row and
        coverage undercounts."""
        from recce import tracking as tr
        from recce.models import Vuln
        from recce.report_excel import _spec_vulns
        # Two findings identical for 40 chars, differing only at chars 41-60.
        base = "Apache httpd 2.4.49 Path Traversal RCE - "  # 41 chars
        v1 = Vuln(ip="10.0.0.5", port=443, protocol="tcp", script_id="version-db",
                  title=base + "CVE-2021-41773", severity="high", source="db")
        v2 = Vuln(ip="10.0.0.5", port=443, protocol="tcp", script_id="version-db",
                  title=base + "CVE-2021-42013", severity="high", source="db")
        # Store keeps both distinct (its key uses title[:60])...
        self.assertNotEqual(v1.key, v2.key)
        # ...so the workbook keys must also be distinct (no collapse).
        self.assertNotEqual(tr.vuln_row_key(v1), tr.vuln_row_key(v2))
        h = Host(ip="10.0.0.5", ports=[Port(portid=443, service="https")],
                 vulns=[v1, v2])
        rows = _spec_vulns([h]).rows
        self.assertEqual(len(rows), 2, "both findings must appear on the sheet")
        self.assertEqual(tr.compute_coverage([h], {})["vulns"]["total"], 2)


class ProgressAndAuthTest(unittest.TestCase):
    def test_fmt_dur(self):
        from recce import cli
        self.assertEqual(cli._fmt_dur(45), "45s")
        self.assertEqual(cli._fmt_dur(200), "3m20s")
        self.assertEqual(cli._fmt_dur(3660), "1h01m")

    def test_progress_has_pct_and_eta(self):
        from recce import cli
        import time
        s = cli._progress(2, 10, time.monotonic() - 4)
        self.assertIn("20%", s)
        self.assertIn("ETA", s)

    def test_auth_cell(self):
        from recce import cli
        self.assertEqual(cli._auth_cell(None), "-")
        self.assertEqual(cli._auth_cell({"tried": True, "auth": False}), "FAIL")
        self.assertEqual(cli._auth_cell({"tried": True, "auth": True}), "OK")
        self.assertEqual(cli._auth_cell({"tried": True, "auth": True, "admin": True}),
                         "OK (admin)")
        # A tool/connection error is ERR, not FAIL (not a credential problem).
        self.assertEqual(cli._auth_cell({"tried": True, "auth": False, "error": True}),
                         "ERR")

    def test_auth_table_prints_rows_and_flags_fail(self):
        from recce import cli
        rows = [("10.0.0.5", {"user": {"tried": True, "auth": True, "admin": True}}),
                ("10.0.0.9", {"user": {"tried": True, "auth": False}})]
        buf = io.StringIO()
        with contextlib.redirect_stdout(buf):
            cli._print_auth_table(rows)
        out = buf.getvalue()
        self.assertIn("USER ACCT", out)
        self.assertIn("OK (admin)", out)
        self.assertIn("FAIL", out)

    def test_no_yes_flag_and_ingest_present(self):
        from recce import cli
        p = cli.build_arg_parser()
        # ingest is a registered command...
        self.assertEqual(p.parse_args(["ingest", "x.txt"]).command, "ingest")
        # ...and the authorization --yes flag is gone.
        with self.assertRaises(SystemExit):
            p.parse_args(["enum", "1.2.3.4", "--yes"])


class StoreFixesTest(unittest.TestCase):
    def test_corrupt_db_raises_storeerror(self):
        from recce.store import Store, StoreError
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "results.sqlite")
            with open(p, "w") as fh:
                fh.write("not a sqlite database, just garbage")
            with self.assertRaises(StoreError):
                Store(p)

    def test_corrupt_db_gives_clean_message_not_traceback(self):
        # A carried-over corrupt DB on `report`/`status` must exit 1 with a clean
        # message, never a raw traceback (the "first command after transfer" case).
        from recce import cli
        with tempfile.TemporaryDirectory() as d:
            os.makedirs(os.path.join(d, "raw"), exist_ok=True)
            with open(os.path.join(d, "results.sqlite"), "w") as fh:
                fh.write("garbage, not a database")
            for command in ("report", "status"):
                buf = io.StringIO()
                with contextlib.redirect_stdout(buf):
                    rc = cli.main([command, "-o", d])
                self.assertEqual(rc, 1, command)
                out = buf.getvalue()
                self.assertIn("corrupt or unreadable", out)
                self.assertNotIn("Traceback", out)

    def test_distance_preserved_through_merge(self):
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            s = Store(os.path.join(d, "r.sqlite"))
            s.upsert_host(Host(ip="10.0.0.5", distance=3))
            # A second scan of the same host without distance must not zero it.
            s.upsert_host(Host(ip="10.0.0.5", os_name="Linux", os_accuracy=95))
            self.assertEqual(s.get_host("10.0.0.5").distance, 3)

    def test_rerun_does_not_duplicate_issues(self):
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            paths = cli._open_paths(d)
            s = Store(paths["db"])
            iss = [{"phase": "vuln-scan", "level": "error", "message": "nmap failed"}]
            cli._record_issues(s, paths, "10.0.0.5", iss)
            cli._record_issues(s, paths, "10.0.0.5", iss)   # re-run, same phase
            self.assertEqual(s.count_issues().get("total"), 1)  # replaced, not doubled
            s.close()


class RunbookSheetTest(unittest.TestCase):
    def test_runbook_sheet_present(self):
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "wb.xlsx")
            build_workbook([], p, meta={"subtitle": "x"})
            import openpyxl
            wb = openpyxl.load_workbook(p)
            self.assertIn("Runbook", wb.sheetnames)
            vals = [c for row in wb["Runbook"].iter_rows(values_only=True)
                    for c in row if c]
            joined = " ".join(str(v) for v in vals)
            self.assertIn("enum", joined)
            self.assertIn("ingest", joined)
            self.assertIn("--fast", joined)


class EntryPointTest(unittest.TestCase):
    def test_module_entrypoint_runs(self):
        import subprocess
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        env = dict(os.environ, PYTHONPATH=root)
        r = subprocess.run([sys.executable, "-m", "recce", "--version"],
                           capture_output=True, text=True, env=env, timeout=30)
        self.assertEqual(r.returncode, 0)
        self.assertIn("recce", (r.stdout + r.stderr).lower())


class EnumRobustnessTest(unittest.TestCase):
    """The enum phase must be robust host-by-host: one host that crashes, times
    out, or returns hostile data can never abort the run or corrupt the workbook.
    """

    def _args(self, d):
        from types import SimpleNamespace
        a = SimpleNamespace(workers=4, refresh_every=0, title="T", resume=False,
                            user=None, hash=None, domain=None, output_dir=d)
        setattr(a, "pass", None)
        for k in ("ssh_user", "ssh_pass", "ssh_key", "admin_user", "admin_pass",
                  "admin_domain", "dc_ip"):
            setattr(a, k, None)
        return a

    def test_one_bad_host_does_not_abort_run_or_corrupt_workbook(self):
        from recce import cli, scanner, xlsx
        import zipfile
        d = tempfile.mkdtemp()
        self.addCleanup(shutil.rmtree, d, ignore_errors=True)
        paths = cli._open_paths(d)
        store = cli._open_store(paths["db"])
        store.set_scope("10.0.0.0/24", 254)

        def fake_worker(ip, profile, paths, creds, port_map, subnet_map,
                        active_probe=True):
            if ip == "10.0.0.11":            # worker raises
                raise RuntimeError("boom")
            if ip == "10.0.0.12":            # timed out -> None + issue
                return None, [{"phase": "enum", "level": "error",
                               "message": "host timeout"}]
            if ip == "10.0.0.13":            # hostile data: control chars, many ports
                return Host(ip=ip, subnet="10.0.0.0/24", enumerated=True,
                            hostnames=["odd\x01\x1f"], ports=[
                                Port(portid=n, service="x\x02y", state="open")
                                for n in range(1, 60)]), []
            return Host(ip=ip, subnet="10.0.0.0/24", enumerated=True,
                        ports=[Port(portid=445, service="microsoft-ds",
                                    state="open")]), []

        orig = cli._enum_worker
        cli._enum_worker = fake_worker
        try:
            live = ["10.0.0.10", "10.0.0.11", "10.0.0.12", "10.0.0.13", "10.0.0.14"]
            with contextlib.redirect_stdout(io.StringIO()):
                cli._phase_enum(store, paths, self._args(d),
                                scanner.PROFILES["standard"],
                                {"10.0.0.10": "10.0.0.0/24"}, live,
                                {i: [] for i in live})
                cli._generate_reports(store, paths, "T", quiet=True)
        finally:
            cli._enum_worker = orig

        ips = {h.ip for h in store.all_hosts()}
        # good + hostile-but-valid hosts persist; crashed/timed-out do not
        self.assertEqual(ips, {"10.0.0.10", "10.0.0.13", "10.0.0.14"})
        issue_ips = {i["ip"] for i in store.get_issues()}
        self.assertTrue({"10.0.0.11", "10.0.0.12"} <= issue_ips)
        # the workbook is valid despite the failures + control chars
        self.assertTrue(zipfile.is_zipfile(paths["xlsx"]))
        self.assertIn("Checklist", xlsx.read_sheets(paths["xlsx"]))
        store.close()

    def test_persist_failure_on_one_host_does_not_abort_the_phase(self):
        from recce import cli, scanner
        d = tempfile.mkdtemp()
        self.addCleanup(shutil.rmtree, d, ignore_errors=True)
        paths = cli._open_paths(d)
        store = cli._open_store(paths["db"])
        store.set_scope("10.0.0.0/24", 254)

        def good_worker(ip, *a, **k):
            return (Host(ip=ip, subnet="10.0.0.0/24", enumerated=True,
                         ports=[Port(portid=445, service="microsoft-ds",
                                     state="open")]), [])
        # The datastore rejects exactly one host's write (a lock outlasting the
        # busy_timeout, say); every other host must still persist.
        real_upsert = store.upsert_host

        def flaky_upsert(host):
            if host.ip == "10.0.0.12":
                raise RuntimeError("database is locked")
            return real_upsert(host)
        store.upsert_host = flaky_upsert

        orig = cli._enum_worker
        cli._enum_worker = good_worker
        try:
            live = ["10.0.0.10", "10.0.0.11", "10.0.0.12", "10.0.0.13"]
            with contextlib.redirect_stdout(io.StringIO()):
                cli._phase_enum(store, paths, self._args(d),
                                scanner.PROFILES["standard"],
                                {"10.0.0.10": "10.0.0.0/24"}, live,
                                {i: [] for i in live})
        finally:
            cli._enum_worker = orig
        store.upsert_host = real_upsert
        ips = {h.ip for h in store.all_hosts()}
        self.assertEqual(ips, {"10.0.0.10", "10.0.0.11", "10.0.0.13"})
        self.assertIn("10.0.0.12", {i["ip"] for i in store.get_issues()})
        store.close()

    def test_zero_port_host_gets_verification_rescan(self):
        """A host the fast pass found 0 ports on is re-verified with an adaptive
        re-scan (discovered-live always; -Pn only with --verify-all), so a missed
        sweep isn't silently trusted as 'no ports'."""
        from recce import cli, scanner
        from recce.models import Host
        calls = {"verify": 0}
        saved = (cli._ports_for_host, cli._fold_host, scanner.full_port_scan,
                 scanner.verify_port_scan, scanner.enum_scan, cli.np.parse_nmap_xml)

        def fake_ports(path, ip):
            return [80, 443] if "verify" in path else []   # fast=0, verify finds some

        def fake_verify(ip, out, profile):
            calls["verify"] += 1
            return out, None
        cli._ports_for_host = fake_ports
        cli._fold_host = lambda ip, parsed, sm: Host(ip=ip, subnet="s")
        scanner.full_port_scan = lambda ip, out, profile: (out, None)
        scanner.verify_port_scan = fake_verify
        scanner.enum_scan = lambda ip, ports, out, profile, creds=None: (out, None)
        cli.np.parse_nmap_xml = lambda p: []
        try:
            d = tempfile.mkdtemp()
            self.addCleanup(shutil.rmtree, d, ignore_errors=True)
            paths = {"raw": d}
            # discovered-live 0-port host -> verified
            cli._enum_worker("1.2.3.4", scanner.ScanProfile(ping_discovery=True),
                             paths, None, None, {})
            self.assertEqual(calls["verify"], 1)
            # -Pn (assume-up) without --verify-all -> NOT re-scanning every dead IP
            calls["verify"] = 0
            cli._enum_worker("1.2.3.5",
                             scanner.ScanProfile(ping_discovery=False, verify_all=False),
                             paths, None, None, {})
            self.assertEqual(calls["verify"], 0)
            # -Pn WITH --verify-all -> verified
            cli._enum_worker("1.2.3.6",
                             scanner.ScanProfile(ping_discovery=False, verify_all=True),
                             paths, None, None, {})
            self.assertEqual(calls["verify"], 1)
            # verify disabled -> never
            calls["verify"] = 0
            cli._enum_worker("1.2.3.7",
                             scanner.ScanProfile(ping_discovery=True, verify=False),
                             paths, None, None, {})
            self.assertEqual(calls["verify"], 0)
        finally:
            (cli._ports_for_host, cli._fold_host, scanner.full_port_scan,
             scanner.verify_port_scan, scanner.enum_scan,
             cli.np.parse_nmap_xml) = saved

    def test_truncated_sweep_incomplete_flag_round_trips_and_clears_on_recompletion(self):
        """A truncated sweep flags the host incomplete_scan (persisted); a later
        complete sweep clears it, and ports union across the two scans."""
        from recce.store import Store
        from recce.models import Host, Port
        d = tempfile.mkdtemp()
        self.addCleanup(shutil.rmtree, d, ignore_errors=True)
        st = Store(os.path.join(d, "s.sqlite"))
        st.upsert_host(Host(ip="1.2.3.4", incomplete_scan=True,
                            ports=[Port(portid=80, service="http", state="open")]))
        self.assertTrue(st.all_hosts()[0].incomplete_scan)   # persisted
        # a later, complete sweep of the same host
        st.upsert_host(Host(ip="1.2.3.4", incomplete_scan=False,
                            ports=[Port(portid=443, service="https", state="open")]))
        h = st.all_hosts()[0]
        self.assertFalse(h.incomplete_scan)                  # complete once either finished
        self.assertEqual({p.portid for p in h.open_ports}, {80, 443})   # union
        st.close()

    def test_corrupt_existing_workbook_is_regenerated_not_fatal(self):
        from recce.report_excel import update_workbook
        from recce import xlsx
        import zipfile
        d = tempfile.mkdtemp()
        self.addCleanup(shutil.rmtree, d, ignore_errors=True)
        p = os.path.join(d, "enumeration.xlsx")
        with open(p, "wb") as f:            # a truncated / non-xlsx file
            f.write(b"PK\x03\x04 not a real workbook \xff\x00")
        hosts = [Host(ip="10.0.0.5", subnet="10.0.0.0/24", enumerated=True,
                      ports=[Port(portid=80, service="http", state="open")])]
        update_workbook(p, hosts, meta={"subtitle": "t"}, tracking={})
        self.assertTrue(zipfile.is_zipfile(p))
        self.assertIn("Checklist", xlsx.read_sheets(p))


if __name__ == "__main__":
    unittest.main()
