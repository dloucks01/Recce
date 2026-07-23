"""Offline tests for the enumeration pipeline (no network / nmap needed)."""

import contextlib
import io
import os
import stat
import sys
import tempfile
import unittest
from types import SimpleNamespace

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from recce import ad, exploits, parser, scanner
from recce import tracking as tr
from recce import xlsx
from recce.models import Account, Host, Port, Script, Vuln
from recce.report_excel import (build_workbook, read_workbook_tracking,
                                       update_workbook)
from recce.store import Store
from recce.targets import apply_exclusions, load_targets

SAMPLE = os.path.join(os.path.dirname(parser.__file__), "sample_scan.xml")


def header_index(rows, *must_have):
    """Row index of the real column-header row (first row holding every token in
    must_have). The Checklist puts a legend line above its header, so callers must
    locate it rather than assume row 0."""
    for i, r in enumerate(rows):
        if all(tok in r for tok in must_have):
            return i
    return 0


class TargetsTest(unittest.TestCase):
    def test_cidr_and_range(self):
        hosts, sm = load_targets(["10.0.0.0/30", "192.168.1.5-8"])
        self.assertEqual(hosts, ["10.0.0.1", "10.0.0.2", "192.168.1.5",
                                 "192.168.1.6", "192.168.1.7", "192.168.1.8"])
        # A CIDR token becomes the subnet label for its hosts.
        self.assertEqual(sm["10.0.0.1"], "10.0.0.0/30")
        # A bare range falls back to a /24 label.
        self.assertEqual(sm["192.168.1.5"], "192.168.1.0/24")

    def test_exclusions(self):
        hosts, _ = load_targets(["192.168.1.0/29"])
        kept = apply_exclusions(hosts, ["192.168.1.1", "192.168.1.2-3"])
        self.assertNotIn("192.168.1.1", kept)
        self.assertNotIn("192.168.1.3", kept)
        self.assertIn("192.168.1.4", kept)

    def test_dedup(self):
        hosts, _ = load_targets(["10.0.0.1", "10.0.0.1", "10.0.0.0/30"])
        self.assertEqual(hosts.count("10.0.0.1"), 1)

    def test_range_drops_network_and_broadcast(self):
        # A full-octet range means "the subnet", not "scan .0 and .255".
        hosts, _ = load_targets(["10.200.37.0-255"])
        self.assertNotIn("10.200.37.0", hosts)
        self.assertNotIn("10.200.37.255", hosts)
        self.assertIn("10.200.37.1", hosts)
        self.assertIn("10.200.37.254", hosts)

    def test_explicit_single_dot_zero_is_respected(self):
        # An explicitly-typed single address is kept (the user asked for it).
        hosts, _ = load_targets(["10.200.37.0"])
        self.assertEqual(hosts, ["10.200.37.0"])


class ParserTest(unittest.TestCase):
    def setUp(self):
        self.hosts = parser.parse_nmap_xml(SAMPLE)

    def test_host_count(self):
        self.assertEqual(len(self.hosts), 4)

    def test_hostnames_and_os(self):
        dc = next(h for h in self.hosts if h.ip == "10.0.10.10")
        self.assertEqual(dc.hostname, "dc01.corp.local")
        self.assertEqual(dc.os_family, "Windows")
        self.assertEqual(len(dc.open_ports), 4)

    def test_vuln_severity(self):
        # ms17-010 (CVSSv2 9.3) -> critical
        dc = next(h for h in self.hosts if h.ip == "10.0.10.10")
        sev = {v.script_id: v.severity for v in dc.vulns}
        self.assertEqual(sev["smb-vuln-ms17-010"], "critical")

    def test_vulners_score_parsed(self):
        # vulners line "CVE-2021-42013 9.8" -> critical
        web = next(h for h in self.hosts if h.ip == "10.0.20.5")
        self.assertTrue(any(v.severity == "critical" for v in web.vulns))

    def test_cvss_vector_not_misread_as_score(self):
        """Regression: a CVSS vector string ('CVSS:3.1/AV:N/...') must not be
        read as base score 3.1 (which downgraded criticals to 'low'); the
        'Base Score' phrasing must be recognized."""
        from recce.parser import _classify_vuln
        from recce.models import Script, Port
        p = Port(portid=443, protocol="tcp", service="https")
        # Vector + explicit base score 9.8 -> must classify critical, not low.
        out = ("VULNERABLE\nCVE-2021-44228\n"
               "CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:H/I:H/A:H\n"
               "CVSS Base Score: 9.8\n")
        v = _classify_vuln("10.0.0.5", p, Script(id="vuln-log4shell", output=out))
        self.assertEqual(v.severity, "critical")
        # Vector ONLY (no numeric score) must not become 'low' via the 3.1.
        v2 = _classify_vuln("10.0.0.5", p, Script(
            id="vuln-x", output="VULNERABLE\nCVE-2021-1\nCVSS:3.1/AV:N/AC:L\n"))
        self.assertNotEqual(v2.severity, "low")

    def test_ad_users_extracted(self):
        dc = next(h for h in self.hosts if h.ip == "10.0.10.10")
        users = [a.name for a in dc.accounts if a.kind == "user"]
        self.assertIn("Administrator", users)
        self.assertIn("svc_sql", users)


class ProductGroupingTest(unittest.TestCase):
    def test_same_version_groups(self):
        hosts = parser.parse_nmap_xml(SAMPLE)
        keys = {}
        for h in hosts:
            for p in h.open_ports:
                keys.setdefault(p.product_version_key, []).append(h.ip)
        apache = next(k for k in keys if k.startswith("Apache httpd|2.4.41"))
        self.assertEqual(len(keys[apache]), 3)


class StoreMergeTest(unittest.TestCase):
    def test_merge_upsert(self):
        with tempfile.TemporaryDirectory() as d:
            db = os.path.join(d, "t.sqlite")
            store = Store(db)
            h1 = Host(ip="1.2.3.4", ports=[Port(portid=80, service="http")])
            store.upsert_host(h1)
            # Second scan adds a port and enriches OS.
            h2 = Host(ip="1.2.3.4", os_name="Linux", os_accuracy=95,
                      ports=[Port(portid=443, service="https")],
                      accounts=[Account(ip="1.2.3.4", source="smb", name="bob")])
            store.upsert_host(h2)
            merged = store.get_host("1.2.3.4")
            self.assertEqual({p.portid for p in merged.ports}, {80, 443})
            self.assertEqual(merged.os_name, "Linux")
            self.assertEqual(len(merged.accounts), 1)
            store.close()


class ADAnalysisTest(unittest.TestCase):
    def setUp(self):
        self.hosts = parser.parse_nmap_xml(SAMPLE)
        ad.analyze_hosts(self.hosts)

    def test_dc_identified(self):
        dcs = ad.domain_controllers(self.hosts)
        self.assertEqual([h.ip for h in dcs], ["10.0.10.10"])

    def test_dc_signing_from_hostscript(self):
        dc = next(h for h in self.hosts if h.ip == "10.0.10.10")
        self.assertEqual(dc.smb_signing, "required")

    def test_relay_target_from_portscript(self):
        relay = ad.relay_targets(self.hosts)
        self.assertEqual([h.ip for h in relay], ["10.0.10.25"])

    def test_password_policy_parsed(self):
        doms = ad.derive_domains(self.hosts)
        corp = next(d for d in doms if d.name == "corp.local")
        self.assertEqual(corp.password_policy.get("min_length"), 7)
        self.assertEqual(corp.password_policy.get("lockout_threshold"), 0)
        self.assertIn("10.0.10.10", corp.dc_ips)

    def test_ntlm_domain_facts(self):
        ws = next(h for h in self.hosts if h.ip == "10.0.10.25")
        self.assertEqual(ws.ntlm.get("netbios_domain"), "CORP")
        self.assertEqual(ws.ntlm.get("dns_domain"), "corp.local")


class ADTargetListTest(unittest.TestCase):
    """LDAP-derived findings via synthetic accounts (no live DC needed)."""

    def _dc_with(self, *accounts):
        h = Host(ip="10.0.10.10", roles=["Domain Controller"])
        h.accounts.extend(accounts)
        return [h]

    def test_kerberoastable(self):
        hosts = self._dc_with(
            Account(ip="10.0.10.10", source="ldap", kind="user", name="svc_sql",
                    domain="corp.local", attrs={"spn": "MSSQLSvc/db01"}),
            Account(ip="10.0.10.10", source="ldap", kind="user", name="krbtgt",
                    domain="corp.local", attrs={"spn": "kadmin/changepw"}),
        )
        kerb = ad.kerberoastable(hosts)
        self.assertEqual([a.name for a in kerb], ["svc_sql"])  # krbtgt excluded

    def test_asrep_and_delegation(self):
        hosts = self._dc_with(
            Account(ip="10.0.10.10", source="ldap", kind="user", name="alice",
                    attrs={"asrep_roastable": "yes"}),
            Account(ip="10.0.10.10", source="ldap", kind="computer", name="SRV$",
                    attrs={"delegation": "unconstrained"}),
        )
        self.assertEqual([a.name for a in ad.asrep_roastable(hosts)], ["alice"])
        self.assertEqual([a.name for a in ad.delegation_accounts(hosts)], ["SRV$"])

    def test_privileged(self):
        hosts = self._dc_with(
            Account(ip="10.0.10.10", source="ldap", kind="user", name="admin",
                    attrs={"memberof": "Domain Admins; IT"}),
            Account(ip="10.0.10.10", source="ldap", kind="user", name="bob"),
        )
        self.assertEqual([a.name for a in ad.privileged_accounts(hosts)], ["admin"])

    def test_uac_flag_decoding(self):
        # DONT_REQ_PREAUTH (0x400000) + ACCOUNTDISABLE (0x2)
        flags = ad._uac_flags(0x400002)
        self.assertIn("DONT_REQ_PREAUTH", flags)
        self.assertIn("ACCOUNTDISABLE", flags)

    def test_ldap_available_is_bool(self):
        self.assertIsInstance(ad.ldap_available(), bool)


class CoverageTest(unittest.TestCase):
    def setUp(self):
        from recce.targets import _subnet_of
        self.hosts = parser.parse_nmap_xml(SAMPLE)
        for h in self.hosts:
            h.subnet = _subnet_of(h.ip)
        ad.analyze_hosts(self.hosts)

    def test_item_keys_categories(self):
        keys = tr.item_keys(self.hosts)
        self.assertEqual(len(keys["hosts"]), 4)
        self.assertEqual(len(keys["services"]), 14)
        self.assertTrue(keys["quick_wins"])  # DC + relay + smbv1

    def test_coverage_counts(self):
        tracking = {tr.host_key("10.0.10.10"): (True, "done")}
        cov = tr.compute_coverage(self.hosts, tracking)
        self.assertEqual(cov["hosts"]["done"], 1)
        self.assertEqual(cov["hosts"]["total"], 4)
        self.assertEqual(cov["overall"]["done"], 1)

    def test_subnet_coverage(self):
        tracking = {tr.host_key("10.0.10.10"): (True, "")}
        sc = tr.subnet_coverage(self.hosts, tracking)
        self.assertEqual(sc["10.0.10.0/24"]["done"], 1)
        self.assertEqual(sc["10.0.10.0/24"]["total"], 2)


class TrackingRoundTripTest(unittest.TestCase):
    def test_prefill_and_readback(self):
        hosts = parser.parse_nmap_xml(SAMPLE)
        ad.analyze_hosts(hosts)
        tracking = {
            tr.host_key("10.0.10.10"): (True, "DC reviewed"),
            tr.svc_key("10.0.20.5", "tcp", 80): (True, ""),
        }
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(hosts, out, tracking=tracking)
            back = read_workbook_tracking(out)
        self.assertTrue(back[tr.host_key("10.0.10.10")][0])
        self.assertEqual(back[tr.host_key("10.0.10.10")][1], "DC reviewed")
        self.assertTrue(back[tr.svc_key("10.0.20.5", "tcp", 80)][0])
        self.assertFalse(back[tr.host_key("10.0.20.6")][0])


class WorkbookFlowTest(unittest.TestCase):
    """The sheet order follows the engagement flow, with the service deep-dive band
    grouped and the AD cluster kept contiguous."""

    def _build(self):
        from recce.models import Domain, Vuln
        from recce import xlsx

        def V(ip, port, sid, title, sev, src):
            return Vuln(ip=ip, port=port, protocol="tcp", script_id=sid, title=title,
                        severity=sev, source=src, state="finding")
        hosts = [
            Host(ip="10.0.0.11", os_family="Windows", roles=["Domain Controller"],
                 enumerated=True,
                 ports=[Port(portid=445, state="open", service="microsoft-ds"),
                        Port(portid=1433, state="open", service="ms-sql-s",
                             product="Microsoft SQL Server")],
                 vulns=[V("10.0.0.11", 445, "smb:x", "SMB signing not required",
                          "medium", "smb"),
                        V("10.0.0.11", 1433, "mssql:x", "MSSQL sysadmin",
                          "critical", "mssql")]),
            Host(ip="10.0.0.22", os_family="Linux", enumerated=True,
                 ports=[Port(portid=2375, state="open", service="docker")],
                 vulns=[V("10.0.0.22", 2375, "docker:x",
                          "Docker Engine API exposed without authentication",
                          "critical", "docker")]),
        ]
        m = {"targets": [{"ip": "x", "port": 1}], "findings": [], "runbooks": []}
        meta = {"subtitle": "Flow", "mssql": m, "smb": m, "docker": m,
                "ad_bloodhound": {"findings": [{"severity": "high",
                    "title": "Kerberoastable", "principal": "svc", "target": "",
                    "detail": "d", "tool": "t", "command": "c", "remediation": "r",
                    "category": "kerberoast", "cwes": ["CWE-262"]}],
                    "paths": [{"start": "a", "target": "DA", "chain": "x", "length": 1,
                               "who": "a", "steps": [], "any_user": False}],
                    "kerberos": [], "domains": [{"name": "corp.local"}], "stats": {}}}
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(hosts, out, meta=meta,
                           domains=[Domain(name="corp.local", dc_ips=["10.0.0.11"])])
            return list(xlsx.read_sheets(out).keys()), xlsx.read_sheets(out)

    def test_service_band_and_ad_cluster_are_grouped(self):
        tabs, _ = self._build()
        pos = {t: i for i, t in enumerate(tabs)}
        # Service deep-dive band is contiguous and sits right after Databases.
        self.assertLess(pos["Databases"], pos["MSSQL"])
        self.assertEqual(pos["SMB"], pos["MSSQL"] + 1)
        self.assertEqual(pos["Docker"], pos["SMB"] + 1)
        # The whole service band precedes the AD cluster.
        self.assertLess(pos["Docker"], pos["Active Directory"])
        # AD cluster is contiguous: inventory -> quick wins -> findings -> paths.
        self.assertEqual(pos["AD Quick Wins"], pos["Active Directory"] + 1)
        self.assertEqual(pos["AD Findings"], pos["AD Quick Wins"] + 1)
        self.assertEqual(pos["AD Attack Paths"], pos["AD Findings"] + 1)
        # Exploit/chain the foothold BEFORE post-ex priv-esc.
        self.assertLess(pos["Exploitation"], pos["Priv-Esc"])
        self.assertLess(pos["Attack Path"], pos["Priv-Esc"])

    def test_overview_shows_confirmed_metric_and_nav_matches_order(self):
        tabs, sheets = self._build()
        ov = ["|".join(str(c) for c in r) for r in sheets["Overview"]]
        self.assertTrue(any("Confirmed by recce (prove engine)" in t for t in ov))
        # The jump-bar lists the tabs in the same left-to-right order they appear.
        navrow = next((r for r in sheets["Overview"]
                       if any(str(c).strip() == "Checklist" for c in r)), None)
        self.assertIsNotNone(navrow)
        nav = [str(c).strip() for c in navrow if str(c).strip()
               and str(c).strip() != "Jump to"]
        nav_in_tabs = [t for t in nav if t in tabs]
        self.assertEqual(nav_in_tabs, sorted(nav_in_tabs, key=lambda t: tabs.index(t)))


class PortStatusTest(unittest.TestCase):
    """Per-port tri-state work status on the Services sheet."""

    def _host(self):
        return Host(ip="10.0.0.5", subnet="10.0.0.0/24", enumerated=True,
                    ports=[Port(portid=80, service="http", state="open"),
                           Port(portid=443, service="https", state="open")])

    def test_services_sheet_has_status_column_and_dropdown(self):
        from recce.report_excel import (build_workbook, STATUS_VALUES,
                                         STATUS_TODO)
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook([self._host()], out)
            rows = xlsx.read_sheets(out)["Services"]
            hdr = rows[0]
            self.assertIn("Status", hdr)
            self.assertIn("Notes", hdr)
            si = hdr.index("Status")
            ki = hdr.index("Key")
            # Every port DATA row defaults to "Not started" (skip the collapsible
            # host-header band rows, which carry no Key).
            data_rows = [r for r in rows[1:] if len(r) > ki and r[ki]]
            self.assertTrue(data_rows)
            for r in data_rows:
                self.assertEqual(r[si], STATUS_TODO)
            # The dropdown offers all three states (find the sheet whose
            # data-validation lists them, not merely any sheet mentioning them).
            import zipfile
            listing = ",".join(STATUS_VALUES)
            with zipfile.ZipFile(out) as z:
                xmls = [z.read(n).decode() for n in z.namelist()
                        if "worksheets/sheet" in n]
            self.assertTrue(any(f'<formula1>"{listing}"</formula1>' in x
                                for x in xmls),
                            "Services Status dropdown not found")

    def test_status_roundtrip_and_reviewed_mapping(self):
        from recce.report_excel import (build_workbook, read_workbook_edits,
                                         STATUS_WIP, STATUS_DONE)
        with tempfile.TemporaryDirectory() as d:
            store = Store(os.path.join(d, "t.sqlite"))
            store.upsert_host(self._host())
            paths = {"xlsx": os.path.join(d, "wb.xlsx")}
            k80 = tr.svc_key("10.0.0.5", "tcp", 80)
            k443 = tr.svc_key("10.0.0.5", "tcp", 443)
            # Persist an in-progress port and a done port.
            store.bulk_set_status({k80: (STATUS_WIP, False, "poking at it"),
                                   k443: (STATUS_DONE, True, "")})
            # Regenerate from the store, then read the sheet back.
            build_workbook(store.all_hosts(), paths["xlsx"],
                           tracking=store.get_tracking(),
                           statuses=store.get_statuses())
            edits, statuses = read_workbook_edits(paths["xlsx"])
            self.assertEqual(statuses[k80], STATUS_WIP)
            self.assertEqual(statuses[k443], STATUS_DONE)
            # In-progress is not "reviewed"; done is.
            self.assertFalse(edits[k80][0])
            self.assertTrue(edits[k443][0])
            self.assertEqual(edits[k80][1], "poking at it")
            # Coverage counts only the done port.
            cov = tr.compute_coverage(store.all_hosts(), store.get_tracking())
            self.assertEqual(cov["services"]["done"], 1)
            store.close()

    def test_status_column_not_misread_as_checkbox(self):
        # The Status column sits at index 0 (where a checkbox used to be); a
        # "Not started" cell must not be read as reviewed=True.
        from recce.report_excel import build_workbook, read_workbook_edits
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook([self._host()], out)
            edits, _ = read_workbook_edits(out)
            self.assertFalse(edits[tr.svc_key("10.0.0.5", "tcp", 80)][0])

    def test_status_survives_store_migration(self):
        # A datastore created before the status column still gains it.
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "old.sqlite")
            import sqlite3
            con = sqlite3.connect(path)
            con.executescript(
                "CREATE TABLE tracking (key TEXT PRIMARY KEY, reviewed INTEGER "
                "DEFAULT 0, notes TEXT DEFAULT '', updated TEXT DEFAULT '');")
            con.commit(); con.close()
            store = Store(path)   # __init__ migrates
            store.bulk_set_status({"svc:x": ("◐ In progress", False, "")})
            self.assertEqual(store.get_statuses()["svc:x"], "◐ In progress")
            store.close()


class InPlaceUpdateTest(unittest.TestCase):
    def _hosts(self, ips):
        out = []
        for ip in ips:
            h = Host(ip=ip, subnet="10.0.0.0/24", ports=[Port(portid=80, service="http")])
            out.append(h)
        return out

    def test_new_ip_appended_order_preserved(self):
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            # First generation, mark the first host reviewed.
            build_workbook(self._hosts(["10.0.0.10", "10.0.0.20"]), out,
                           tracking={tr.host_key("10.0.0.10"): (True, "done")})
            # A new IP that would sort to the TOP if re-sorted.
            update_workbook(out, self._hosts(["10.0.0.10", "10.0.0.20", "10.0.0.1"]),
                            tracking={tr.host_key("10.0.0.10"): (True, "done")})
            rows = xlsx.read_sheets(out)["Checklist"]
            hidx = header_index(rows, "IP")
            hdr = rows[hidx]
            ipc = hdr.index("IP")
            # Skip the collapsible subnet band rows (empty Reviewed checkbox cell).
            data_rows = [r for r in rows[hidx + 1:]
                         if r[0] in (xlsx.CHECK_ON, xlsx.CHECK_OFF)]
            ips = [r[ipc] for r in data_rows]
        # Existing order kept; new IP appended last (not sorted in).
        self.assertEqual(ips, ["10.0.0.10", "10.0.0.20", "10.0.0.1"])
        self.assertEqual(data_rows[0][0], xlsx.CHECK_ON)  # first host still reviewed


class XlsxEngineTest(unittest.TestCase):
    def test_write_read_roundtrip(self):
        wb = xlsx.Workbook()
        sh = wb.add_sheet("S")
        sh.write([("H1", "header"), ("Key", "header")])
        sh.write([("val,with&special<chars>", None), "k1"])
        sh.write([(42, None), "k2"])
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "x.xlsx")
            wb.save(p)
            rows = xlsx.read_sheets(p)["S"]
        self.assertEqual(rows[1][0], "val,with&special<chars>")
        self.assertEqual(rows[2][0], "42")

    def test_col_letter(self):
        self.assertEqual(xlsx.col_letter(1), "A")
        self.assertEqual(xlsx.col_letter(27), "AA")


class LdifParseTest(unittest.TestCase):
    def test_parse_entries_and_base64(self):
        import base64 as b64
        enc = b64.b64encode("héllo".encode()).decode()
        ldif = (
            "dn: CN=svc_sql,DC=corp,DC=local\n"
            "sAMAccountName: svc_sql\n"
            "servicePrincipalName: MSSQLSvc/db01.corp.local:1433\n"
            "userAccountControl: 66048\n"
            f"description:: {enc}\n"
            "\n"
            "dn: CN=alice,DC=corp,DC=local\n"
            "sAMAccountName: alice\n"
            "userAccountControl: 4260352\n"
        )
        entries = ad._parse_ldif(ldif)
        self.assertEqual(len(entries), 2)
        self.assertEqual(entries[0]["sAMAccountName"], ["svc_sql"])
        # internal colon in SPN preserved
        self.assertEqual(entries[0]["servicePrincipalName"], ["MSSQLSvc/db01.corp.local:1433"])
        self.assertEqual(entries[0]["description"], ["héllo"])
        # AS-REP flag (0x400000) set on alice
        acc = ad._acc_from_ldif(entries[1], "10.0.0.1", "corp.local", "user")
        self.assertEqual(acc.attrs.get("asrep_roastable"), "yes")


class WeakConfigFindingTest(unittest.TestCase):
    def setUp(self):
        self.hosts = {h.ip: h for h in parser.parse_nmap_xml(SAMPLE)}

    def _find(self, ip, script_id):
        return next((v for v in self.hosts[ip].vulns if v.script_id == script_id), None)

    def test_ftp_anon_medium(self):
        v = self._find("10.0.20.6", "ftp-anon")
        self.assertIsNotNone(v)
        self.assertEqual(v.severity, "medium")
        self.assertTrue(v.cwes)                     # weak-config carries CWEs
        self.assertEqual(v.source, "config")

    def test_weak_tls_medium(self):
        v = self._find("10.0.20.5", "ssl-enum-ciphers")
        self.assertEqual(v.severity, "medium")

    def test_expired_cert_low(self):
        v = self._find("10.0.20.5", "ssl-cert")
        self.assertEqual(v.severity, "low")

    def test_risky_methods_low(self):
        v = self._find("10.0.20.5", "http-methods")
        self.assertEqual(v.severity, "low")

    def test_cve_still_takes_precedence(self):
        # smb-vuln-ms17-010 stays a CVE finding, not reclassified.
        v = self._find("10.0.10.10", "smb-vuln-ms17-010")
        self.assertEqual(v.severity, "critical")


def _docx_text(path):
    import zipfile
    import xml.etree.ElementTree as ET
    W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    with zipfile.ZipFile(path) as z:
        for n in z.namelist():          # every xml part must be well-formed
            if n.endswith((".xml", ".rels")):
                ET.fromstring(z.read(n))
        root = ET.fromstring(z.read("word/document.xml"))
        parts = z.namelist()
    return "\n".join("".join(t.text or "" for t in p.iter(f"{W}t"))
                     for p in root.iter(f"{W}p")), parts


class DocxWriterTest(unittest.TestCase):
    def test_writer_parts_and_text(self):
        from recce.docx import Document
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "t.docx")
            doc = Document()
            doc.title("Hello")
            doc.heading("Section")
            doc.field("Severity", "HIGH")
            doc.placeholder("do this")
            doc.save(out)
            text, parts = _docx_text(out)
        self.assertIn("[Content_Types].xml", parts)
        self.assertIn("word/document.xml", parts)
        self.assertIn("Hello", text)
        self.assertIn("Severity: HIGH", text)
        self.assertIn("[TESTER: do this]", text)

    def test_design_language_styling(self):
        """Teal accent, coloured/mono field values, teal-tinted evidence block."""
        import zipfile
        from recce.docx import Document
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "s.docx")
            doc = Document()
            doc.title("T")
            doc.field("Severity", "CRITICAL", value_color="C00000")
            doc.field("CVE / References", "CVE-2021-41773", mono=True)
            doc.mono_block("raw evidence line")
            doc.save(out)
            with zipfile.ZipFile(out) as z:
                body = z.read("word/document.xml").decode()
                styles = z.read("word/styles.xml").decode()
        self.assertIn('w:color w:val="0E6E67"', styles)      # teal accent in headings
        self.assertIn('w:color w:val="C00000"', body)        # severity value coloured
        self.assertIn('w:ascii="Consolas"', body)            # mono CVE + evidence
        self.assertIn('w:fill="EDF6F4"', body)               # teal-tinted evidence

    def test_image_embed(self):
        import struct
        import binascii
        import zlib
        from recce.docx import Document, _png_size
        sig = b"\x89PNG\r\n\x1a\n"

        def chunk(t, dat):
            return (struct.pack(">I", len(dat)) + t + dat
                    + struct.pack(">I", binascii.crc32(t + dat) & 0xffffffff))
        png = (sig + chunk(b"IHDR", struct.pack(">IIBBBBB", 640, 480, 8, 2, 0, 0, 0))
               + chunk(b"IDAT", zlib.compress(b"\x00" + b"\xff\x00\x00" * 640))
               + chunk(b"IEND", b""))
        self.assertEqual(_png_size(png), (640, 480))
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "img.docx")
            doc = Document()
            doc.image(png, caption="cap")
            doc.save(out)
            _text, parts = _docx_text(out)
            import zipfile
            with zipfile.ZipFile(out) as z:
                rels = z.read("word/_rels/document.xml.rels").decode()
                body = z.read("word/document.xml").decode()
        self.assertIn("word/media/image1.png", parts)
        self.assertIn("/image", rels)
        self.assertIn("r:embed", body)


class WriteupTest(unittest.TestCase):
    def _hosts(self):
        from recce.models import Vuln
        h1 = Host(ip="10.0.20.5", hostnames=["web01"],
                  ports=[Port(portid=443, service="https")],
                  vulns=[Vuln(ip="10.0.20.5", port=443, protocol="tcp",
                              script_id="ssl-enum-ciphers",
                              title="Weak SSL/TLS ciphers or protocols",
                              severity="medium", source="config",
                              cwes=["CWE-327"], remediation="Disable weak ciphers.",
                              output="TLSv1.0 offered")])
        h2 = Host(ip="10.0.20.9", hostnames=["web02"],
                  ports=[Port(portid=443, service="https")],
                  vulns=[Vuln(ip="10.0.20.9", port=443, protocol="tcp",
                              script_id="ssl-enum-ciphers",
                              title="Weak SSL/TLS ciphers or protocols",
                              severity="medium", source="config", cwes=["CWE-327"],
                              output="RC4 offered"),
                         Vuln(ip="10.0.20.9", port=21, protocol="tcp",
                              script_id="version-db",
                              title="vsftpd 2.3.4 backdoor", severity="critical",
                              source="version-db", ids=["CVE-2011-2523"],
                              cwes=["CWE-78"], remediation="Upgrade vsftpd.")])
        return [h1, h2]

    def test_grouping_across_hosts(self):
        from recce.report_docx import group_findings
        findings = group_findings(self._hosts())
        # 2 distinct findings; critical sorts first.
        self.assertEqual([f.severity for f in findings], ["critical", "medium"])
        tls = next(f for f in findings if "SSL" in f.title)
        self.assertEqual(sorted(a[0] for a in tls.affected),
                         ["10.0.20.5", "10.0.20.9"])   # spans both hosts

    def test_build_writeups_and_no_overwrite(self):
        from recce.report_docx import build_writeups
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "writeups")
            summary = build_writeups(self._hosts(), out)
            self.assertEqual(summary["total"], 2)
            self.assertEqual(len(summary["written"]), 2)
            f_crit = next(p for p in os.listdir(out) if p.startswith("F-001"))
            text, _ = _docx_text(os.path.join(out, f_crit))
            for expect in ("F-001", "Affected systems:", "CWE-78",
                           "CVE-2011-2523", "Recommendations", "Evidence",
                           "Mission Risk and Impact", "[TESTER:"):
                self.assertIn(expect, text)
            # Re-run: existing files are kept, not overwritten.
            again = build_writeups(self._hosts(), out)
            self.assertEqual(len(again["written"]), 0)
            self.assertEqual(len(again["skipped"]), 2)

    def test_min_severity_filter(self):
        from recce.report_docx import build_writeups
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "w")
            summary = build_writeups(self._hosts(), out, min_severity="high")
            self.assertEqual(summary["total"], 1)   # only the critical

    def _hosts_potential_and_loot(self):
        from recce.models import Vuln
        return [Host(ip="10.0.30.5", hostnames=["box"],
                     ports=[Port(portid=23, service="telnet"),
                            Port(portid=445, service="microsoft-ds")],
                     local_findings=[{"section": "Sudo", "category": "sudo",
                                      "vector": "NOPASSWD find",
                                      "text": "NOPASSWD sudo: /usr/bin/find",
                                      "source": "recce-enum"}],
                     vulns=[
                         Vuln(ip="10.0.30.5", port=23, protocol="tcp",
                              script_id="version-db", title="Telnet cleartext",
                              severity="medium", source="version-db",
                              confidence="potential", cwes=["CWE-319"]),
                         Vuln(ip="10.0.30.5", port=445, protocol="tcp",
                              script_id="smb-vuln-ms17-010", title="smb-vuln-ms17-010",
                              severity="high", source="nse", ids=["CVE-2017-0143"],
                              output="VULNERABLE"),
                     ])]

    def test_potential_excluded_by_default_included_on_flag(self):
        from recce.report_docx import build_writeups
        hosts = self._hosts_potential_and_loot()
        with tempfile.TemporaryDirectory() as d:
            real = build_writeups(hosts, os.path.join(d, "r"))
            self.assertEqual(real["total"], 1)                 # only the nse ms17-010
            self.assertEqual(real["dropped_potential"], 1)     # telnet guess skipped
            allf = build_writeups(hosts, os.path.join(d, "a"), include_potential=True)
            self.assertEqual(allf["total"], 2)                 # both

    def test_list_findings_flags_real(self):
        from recce.report_docx import list_findings
        rows = list_findings(self._hosts_potential_and_loot())
        by_title = {r["title"]: r for r in rows}
        self.assertFalse(by_title["Telnet cleartext"]["real"])
        self.assertTrue(by_title["smb-vuln-ms17-010"]["real"])
        # stable ids: high sorts before the medium
        self.assertEqual(by_title["smb-vuln-ms17-010"]["id"], "F-001")

    def test_single_writeup_prefills_looted(self):
        from recce.report_docx import build_one_writeup
        hosts = self._hosts_potential_and_loot()
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "writeups")
            res = build_one_writeup(hosts, out, "ms17")
            self.assertTrue(res["written"])
            self.assertEqual(res["looted"], 1)
            text, _ = _docx_text(res["written"])
            self.assertIn("Obtained Access / Looted Evidence", text)
            self.assertIn("NOPASSWD sudo: /usr/bin/find", text)

    def test_single_writeup_selectors(self):
        from recce.report_docx import build_one_writeup
        hosts = self._hosts_potential_and_loot()
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "w")
            # by F-id
            self.assertTrue(build_one_writeup(hosts, out, "F-001")["written"])
            # by CVE
            self.assertTrue(build_one_writeup(hosts, out, "CVE-2017-0143",
                                              overwrite=True)["written"])
            # ambiguous IP -> lists candidates, writes nothing
            amb = build_one_writeup(hosts, out, "10.0.30.5")
            self.assertIsNone(amb["written"])
            self.assertEqual(amb["reason"], "ambiguous")
            self.assertEqual(len(amb["matched"]), 2)
            # unmatched
            none = build_one_writeup(hosts, out, "zzz-nope")
            self.assertIsNone(none["written"])
            self.assertEqual(none["reason"], "none")

    def test_auto_walkthrough_steps(self):
        from recce.report_docx import group_findings, _walkthrough_steps
        findings = group_findings(self._hosts())
        tls = next(f for f in findings if "SSL" in f.title)
        steps = _walkthrough_steps(tls)
        self.assertTrue(steps)
        joined = " ".join(steps)
        self.assertIn("nmap -sV", joined)         # discovery step
        self.assertIn("ssl-enum-ciphers", joined)  # tailored confirmation step

    def test_narrative_is_multi_paragraph_and_context_aware(self):
        from recce.models import Vuln
        from recce.report_docx import group_findings, _narrative
        # A likely (version-matched) web finding.
        web = Host(ip="10.0.20.5", hostnames=["web01"],
                   ports=[Port(portid=80, service="http", product="Apache httpd",
                               version="2.4.49")],
                   vulns=[Vuln(ip="10.0.20.5", port=80, protocol="tcp",
                               script_id="version-db", title="Apache path traversal",
                               severity="critical", source="version-db",
                               confidence="likely", cwes=["CWE-22"],
                               ids=["CVE-2021-41773"])])
        f = group_findings([web])[0]
        paras = _narrative(f)
        self.assertEqual(len(paras), 3)                       # context / finding / impact
        blob = " ".join(paras).lower()
        self.assertIn("web service", blob)                    # service context
        self.assertIn("apache httpd 2.4.49", blob)            # detected product
        self.assertIn("10.0.20.5", blob)                      # affected host named
        self.assertIn("cve-2021-41773", blob)                 # CVE woven in
        self.assertIn("critical-risk", blob)                  # severity framing
        self.assertIn("read files outside", blob)             # CWE-22 plain impact
        self.assertIn("range known to be affected", blob)     # likely-confidence note

        # A potential (advisory) finding gets the "confirm by hand" caveat instead.
        adv = Host(ip="10.0.10.10", hostnames=["dc01"], os_family="Windows",
                   ports=[Port(portid=445, service="microsoft-ds",
                               product="Windows Server 2019")],
                   vulns=[Vuln(ip="10.0.10.10", port=445, protocol="tcp",
                               script_id="version-db", title="verify ZeroLogon",
                               severity="critical", source="version-db",
                               confidence="potential", cwes=["CWE-330"])])
        pa = " ".join(_narrative(group_findings([adv])[0])).lower()
        self.assertIn("smb", pa)                              # SMB service context
        self.assertIn("confirmed through hands-on", pa)       # potential caveat

    def test_every_cwe_is_classified_named_and_has_an_impact(self):
        """Guarantee: every CWE recce can emit maps to a type + a name, and every
        type has a plain-language impact - so no finding drops to a blank type."""
        import glob
        import re
        from recce.report_docx import _CWE_TYPE, _CWE_NAME, _TYPE_IMPACT
        used = set()
        for fn in glob.glob(os.path.join(os.path.dirname(os.path.dirname(
                os.path.abspath(__file__))), "recce", "*.py")):
            if fn.endswith("report_docx.py"):
                continue
            with open(fn) as fh:
                used |= set(re.findall(r"CWE-\d+", fh.read()))
        self.assertTrue(used)
        typed = set()
        for keys, _label, _cia in _CWE_TYPE:
            typed |= set(keys)
        self.assertEqual(used - typed, set(), "CWEs with no vulnerability type")
        self.assertEqual(used - set(_CWE_NAME), set(), "CWEs with no reference name")
        for _keys, label, _cia in _CWE_TYPE:
            self.assertIn(label, _TYPE_IMPACT, f"type '{label}' has no impact wording")
        # CWEs the NSE-script mapper can assign must also be named + typed.
        from recce.report_docx import _NSE_CWE
        nse_cwes = {c for cs in _NSE_CWE.values() for c in cs}
        self.assertEqual(nse_cwes - typed, set(), "NSE-mapped CWEs with no type")
        self.assertEqual(nse_cwes - set(_CWE_NAME), set(), "NSE-mapped CWEs with no name")

    def test_nse_scripts_auto_map_to_cwe_and_cve(self):
        from recce.models import Vuln
        from recce.report_docx import group_findings

        def finding_for(script_id, title=None):
            h = Host(ip="10.0.0.9", ports=[Port(portid=445, service="microsoft-ds")],
                     vulns=[Vuln(ip="10.0.0.9", port=445, protocol="tcp",
                                 script_id=script_id, title=title or script_id,
                                 severity="high", source="nse")])
            return group_findings([h])[0]

        # ms17-010 (no CVE in the id) -> mapped CVE + CWE.
        f = finding_for("smb-vuln-ms17-010")
        self.assertIn("CVE-2017-0144", f.cves)
        self.assertIn("CWE-787", f.cwes)
        # http-vuln-cveYYYY-N -> CVE parsed from the id + CWE mapped.
        f = finding_for("http-vuln-cve2021-41773")
        self.assertIn("CVE-2021-41773", f.cves)
        self.assertIn("CWE-22", f.cwes)
        # Heartbleed TLS script -> its CVE + CWE.
        f = finding_for("ssl-heartbleed")
        self.assertIn("CVE-2014-0160", f.cves)
        self.assertIn("CWE-125", f.cwes)
        # A version-db finding that already has CWE/CVE is NOT overridden.
        h = Host(ip="10.0.0.9", ports=[Port(portid=80, service="http")],
                 vulns=[Vuln(ip="10.0.0.9", port=80, protocol="tcp",
                             script_id="version-db", title="Apache thing",
                             severity="high", source="version-db",
                             cwes=["CWE-22"], ids=["CVE-2021-41773"])])
        f = group_findings([h])[0]
        self.assertEqual(f.cwes, ["CWE-22"])

    def test_marquee_vulns_get_specific_impact(self):
        from recce.models import Vuln
        from recce.report_docx import group_findings, _narrative
        cases = [
            (["CVE-2020-1472"], "verify zerologon", "ZeroLogon"),
            (["CVE-2021-34527"], "printnightmare", "Print Spooler"),
            ([], "smb-vuln-ms17-010", "EternalBlue"),        # NSE hit, no CVE
            (["CVE-2020-0796"], "smbghost", "SMBv3"),
        ]
        for cves, title, needle in cases:
            h = Host(ip="10.0.0.9", os_family="Windows",
                     ports=[Port(portid=445, service="microsoft-ds")],
                     vulns=[Vuln(ip="10.0.0.9", port=445, protocol="tcp",
                                 script_id=title, title=title, severity="critical",
                                 source="nse", ids=cves)])
            blob = " ".join(_narrative(group_findings([h])[0]))
            self.assertIn(needle, blob, f"{title} missing marquee wording")

    def test_reports_exclude_informational_by_default(self):
        from recce.models import Vuln
        from recce.report_docx import build_writeups
        h = Host(ip="10.0.0.9", ports=[Port(portid=25, service="smtp")],
                 vulns=[
                     Vuln(ip="10.0.0.9", port=25, protocol="tcp", script_id="a",
                          title="SMTP server exposed", severity="info", source="version-db"),
                     Vuln(ip="10.0.0.9", port=25, protocol="tcp", script_id="b",
                          title="Weak TLS on SMTP", severity="medium", source="probe"),
                 ])
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "w")
            summary = build_writeups([h], out)               # default = findings only
            self.assertEqual(summary["total"], 1)            # the medium, not the info
            names = os.listdir(out)
            self.assertFalse(any("_info_" in n for n in names))
            # Opting in re-includes informational items.
            summary2 = build_writeups([h], os.path.join(d, "w2"), min_severity="info")
            self.assertEqual(summary2["total"], 2)

    def test_walkthrough_uses_searchsploit_exploit(self):
        from recce.models import Vuln, Exploit
        from recce.report_docx import group_findings, _walkthrough_steps
        h = Host(ip="10.0.20.6", ports=[Port(portid=21, service="ftp",
                 product="vsftpd", version="2.3.4")],
                 vulns=[Vuln(ip="10.0.20.6", port=21, protocol="tcp",
                             script_id="version-db", title="vsftpd 2.3.4 backdoor",
                             severity="critical", source="version-db",
                             ids=["CVE-2011-2523"])],
                 exploits=[Exploit(ip="10.0.20.6", port=21, edb_id="17491",
                                   title="vsftpd 2.3.4 backdoor")])
        f = group_findings([h])[0]
        self.assertIn("17491", " ".join(_walkthrough_steps(f)))

    def test_walkthrough_only_cites_proven_exploits(self):
        from recce.models import Vuln
        from recce.report_docx import group_findings, _walkthrough_steps

        def steps(title, conf, cves, source="version-db", svc="http", port=80):
            h = Host(ip="1.1.1.1", ports=[Port(portid=port, service=svc)],
                     vulns=[Vuln(ip="1.1.1.1", port=port, protocol="tcp",
                                 script_id=source, title=title, severity="high",
                                 source=source, confidence=conf, ids=cves)])
            return " ".join(_walkthrough_steps(group_findings([h])[0]))

        # Proven exploit (curated) on a version-matched finding -> cited concretely.
        s = steps("Apache path traversal", "likely", ["CVE-2021-41773"])
        self.assertIn("Metasploit", s)
        self.assertIn("apache_normalize_path_rce", s)
        # NSE-confirmed ms17-010 -> proven EternalBlue exploit cited.
        self.assertIn("eternalblue", steps("smb-vuln-ms17-010", "", [],
                                            source="nse", svc="microsoft-ds", port=445).lower())
        # Advisory/potential finding -> NO exploit line, even with a famous CVE.
        s = steps("Windows DC - verify ZeroLogon", "potential", ["CVE-2020-1472"],
                  svc="microsoft-ds", port=445)
        self.assertNotIn("Metasploit", s)
        self.assertNotIn("exploit", s.lower())
        # A version match with no proven exploit known -> no speculative "research" line.
        s = steps("OpenSSH username enumeration", "likely", ["CVE-2018-15473"],
                  svc="ssh", port=22)
        self.assertNotIn("Metasploit", s)
        self.assertNotIn("Research a working exploit", s)

    def test_combined_report(self):
        from recce.report_docx import build_combined
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "combined.docx")
            res = build_combined(self._hosts(), out, title="Test Engagement")
            self.assertEqual(res["total"], 2)
            text, parts = _docx_text(out)
            import zipfile
            with zipfile.ZipFile(out) as z:
                body = z.read("word/document.xml").decode()
            self.assertIn("<w:tbl>", body)              # has tables
            self.assertIn("Test Engagement", text)      # title
            self.assertIn("Summary", text)
            self.assertIn("F-001", text)                # findings numbered
            self.assertIn("vsftpd 2.3.4 backdoor", text)

    def test_screenshot_url_classification(self):
        from recce import screenshot
        self.assertTrue(screenshot._web_url(Port(portid=443, service="https")))
        self.assertTrue(screenshot._web_url(Port(portid=8080, service="http-proxy")))
        self.assertIsNone(screenshot._web_url(Port(portid=22, service="ssh")))
        # No browser in the test env -> capture is a no-op, never raises.
        h = Host(ip="1.2.3.4", ports=[Port(portid=80, service="http")])
        if not screenshot.available():
            self.assertEqual(screenshot.capture_for_host(h), [])

    def _fake_browser(self, name):
        """Create a fake executable and point RECCE_BROWSER at it."""
        d = tempfile.mkdtemp()
        self.addCleanup(lambda: __import__("shutil").rmtree(d, ignore_errors=True))
        path = os.path.join(d, name)
        with open(path, "w") as fh:
            fh.write("#!/bin/sh\n")
        os.chmod(path, 0o755)
        return path

    def test_browser_found_off_path(self):
        """Regression: a browser installed but not on PATH (sudo secure_path,
        snap, /opt) must still be found via the absolute-path fallback scan."""
        import shutil as _sh
        from recce import screenshot as s
        d = tempfile.mkdtemp()
        self.addCleanup(lambda: _sh.rmtree(d, ignore_errors=True))
        # a browser in a bin dir + one nested under an /opt-style dir
        bind = os.path.join(d, "bin"); os.makedirs(bind)
        optd = os.path.join(d, "opt", "vendor"); os.makedirs(optd)
        chromium = os.path.join(bind, "chromium")
        firefox = os.path.join(optd, "firefox")
        for p in (chromium, firefox):
            with open(p, "w") as fh:
                fh.write("#!/bin/sh\n")
            os.chmod(p, 0o755)
        orig_dirs, orig_globs = s._SCAN_DIRS, s._OPT_GLOBS
        orig_path = os.environ.get("PATH", "")
        os.environ.pop("RECCE_BROWSER", None)
        try:
            os.environ["PATH"] = "/nonexistent-xyz"   # nothing resolvable on PATH
            # 1) scan-dir fallback finds the bin-dir chromium
            s._SCAN_DIRS = [bind]; s._OPT_GLOBS = []
            self.assertEqual(s.browser_tool(), chromium)
            self.assertTrue(s.available())
            # 2) /opt-style glob finds the nested firefox
            s._SCAN_DIRS = []
            s._OPT_GLOBS = [os.path.join(d, "opt", "*/{n}")]
            self.assertEqual(s.browser_tool(), firefox)
        finally:
            s._SCAN_DIRS, s._OPT_GLOBS = orig_dirs, orig_globs
            os.environ["PATH"] = orig_path

    def test_firefox_detection_and_command(self):
        from recce import screenshot
        ff = self._fake_browser("firefox")
        os.environ["RECCE_BROWSER"] = ff
        self.addCleanup(lambda: os.environ.pop("RECCE_BROWSER", None))
        try:
            self.assertEqual(screenshot.browser_tool(), ff)
            self.assertTrue(screenshot._is_firefox(ff))
            self.assertTrue(screenshot.available())

            captured = {}

            def fake_run(cmd, **kw):
                captured["cmd"] = cmd
                # Emulate Firefox writing the screenshot: -screenshot <out> URL
                out = cmd[cmd.index("--screenshot") + 1]
                with open(out, "wb") as fh:
                    fh.write(b"\x89PNG\r\n\x1a\n" + b"\0" * 32)
                return SimpleNamespace(returncode=0, stdout=b"", stderr=b"")

            import subprocess as _sp
            orig = _sp.run
            _sp.run = fake_run
            try:
                png = screenshot.capture("http://1.2.3.4:80/")
            finally:
                _sp.run = orig

            self.assertIsNotNone(png)
            self.assertTrue(png.startswith(b"\x89PNG"))
            cmd = captured["cmd"]
            self.assertEqual(os.path.basename(cmd[0]), "firefox")
            self.assertIn("--headless", cmd)
            self.assertIn("-profile", cmd)
            # Screenshot path is a positional arg (no `=` form), URL is last.
            self.assertEqual(cmd[-1], "http://1.2.3.4:80/")
            self.assertNotIn("--ignore-certificate-errors", cmd)
        finally:
            os.environ.pop("RECCE_BROWSER", None)

    def test_chrome_detection_and_command(self):
        from recce import screenshot
        ch = self._fake_browser("chromium")
        os.environ["RECCE_BROWSER"] = ch
        self.addCleanup(lambda: os.environ.pop("RECCE_BROWSER", None))
        try:
            self.assertFalse(screenshot._is_firefox(ch))
            captured = {}

            def fake_run(cmd, **kw):
                captured["cmd"] = cmd
                out = cmd[-2].split("=", 1)[1]
                with open(out, "wb") as fh:
                    fh.write(b"\x89PNG\r\n\x1a\n" + b"\0" * 32)
                return SimpleNamespace(returncode=0, stdout=b"", stderr=b"")

            import subprocess as _sp
            orig = _sp.run
            _sp.run = fake_run
            try:
                png = screenshot.capture("https://1.2.3.4:443/")
            finally:
                _sp.run = orig

            self.assertIsNotNone(png)
            cmd = captured["cmd"]
            self.assertIn("--headless", cmd)
            self.assertIn("--ignore-certificate-errors", cmd)
            self.assertTrue(cmd[-2].startswith("--screenshot="))
            self.assertEqual(cmd[-1], "https://1.2.3.4:443/")
        finally:
            os.environ.pop("RECCE_BROWSER", None)


class CredEnumTest(unittest.TestCase):
    NXC = (
        r"SMB  10.0.0.10  445  DC01  [*] Windows Server 2019 Build 17763 "
        r"(name:DC01) (domain:corp.local) (signing:True)" "\n"
        r"SMB  10.0.0.10  445  DC01  [+] corp.local\admin:Pw (Pwn3d!)" "\n"
        r"SMB  10.0.0.10  445  DC01  [*] Enumerated shares" "\n"
        r"SMB  10.0.0.10  445  DC01  Share    Permissions   Remark" "\n"
        r"SMB  10.0.0.10  445  DC01  -----    -----------   ------" "\n"
        r"SMB  10.0.0.10  445  DC01  ADMIN$   READ,WRITE    Remote Admin" "\n"
        r"SMB  10.0.0.10  445  DC01  [*] Enumerated domain user(s)" "\n"
        r"SMB  10.0.0.10  445  DC01  corp.local\Administrator  badpwdcount: 0" "\n"
        r"SMB  10.0.0.10  445  DC01  [+] Dumping password info for domain: CORP" "\n"
        r"SMB  10.0.0.10  445  DC01  Account lockout threshold: None"
    )

    def test_parse_nxc_smb(self):
        from recce import credenum as c
        d = c.parse_nxc_smb(self.NXC)
        self.assertTrue(d["admin"])
        self.assertIn("ADMIN$", [s["name"] for s in d["shares"]])
        self.assertIn("Administrator", [u["name"] for u in d["users"]])
        self.assertEqual(d["passpol"]["account lockout threshold"], "none")

    def test_parse_roasting(self):
        from recce import credenum as c
        spns = c.parse_getuserspns(
            "MSSQL/dc.corp.local  sqlsvc  Domain Users  2020\n"
            "$krb5tgs$23$*sqlsvc$CORP.LOCAL$MSSQL*$deadbeef")
        self.assertEqual(spns[0]["name"], "sqlsvc")
        self.assertTrue(spns[0]["hash"].startswith("$krb5tgs$"))
        asrep = c.parse_getnpusers("$krb5asrep$23$svc-web@CORP.LOCAL:abcd")
        self.assertEqual(asrep[0]["name"], "svc-web")

    def test_parse_secretsdump_and_ssh(self):
        from recce import credenum as c
        sd = c.parse_secretsdump(
            "Administrator:500:aad3b435b51404eeaad3b435b51404ee:"
            "31d6cfe0d16ae931b73c59d7e0c089c0:::")
        self.assertEqual(sd[0]["name"], "Administrator")
        self.assertEqual(sd[0]["nt"], "31d6cfe0d16ae931b73c59d7e0c089c0")
        ssh = c.parse_ssh_enum(
            "===ID===\nuid=0(root)\n===SUDO===\n(ALL) NOPASSWD: ALL\n"
            "===SUID===\n/usr/bin/find\n/usr/bin/sudo")
        self.assertIn("uid=0(root)", ssh["id"])
        self.assertTrue(ssh["sudo"])
        self.assertIn("/usr/bin/find", ssh["suid"])

    def test_fold_into_host_and_quickwins(self):
        from recce import credenum as c
        d = c.parse_nxc_smb(self.NXC)
        h = Host(ip="10.0.0.10", os_family="Windows", roles=["Domain Controller"],
                 ports=[Port(portid=445, state="open"),
                        Port(portid=389, state="open")])
        c._fold_nxc(h, d)
        c._fold_roast(h, [{"name": "sqlsvc", "spn": "MSSQL/dc", "hash": "$krb5tgs$x"}],
                      [{"name": "svc-web", "hash": "$krb5asrep$x"}], "corp.local")
        srcs = {a.source for a in h.accounts}
        self.assertEqual(srcs, {"netexec", "impacket"})
        titles = [v.title for v in h.vulns]
        self.assertTrue(any("Local admin" in t for t in titles))
        # Roasted accounts flow into the AD quick-wins.
        self.assertIn("sqlsvc", [a.name for a in ad.kerberoastable([h])])
        self.assertIn("svc-web", [a.name for a in ad.asrep_roastable([h])])

    def test_dual_account_user_enumerates_admin_dumps(self):
        """Low-priv account enumerates; privileged account does the admin-only
        power moves (confirm admin reach + secretsdump), labelled per account."""
        from recce import credenum as c
        used = []

        def fake_nxc(ip, creds):
            # Only the privileged account is local admin here.
            return ({"admin": creds["username"] == "da", "host_info": "corp",
                     "shares": [{"name": "C$", "perms": "READ"}],
                     "users": [{"name": "bob", "domain": "corp"}],
                     "loggedon": [], "passpol": {}}, None)

        def fake_dump(ip, creds):
            used.append(("secretsdump", creds["username"]))
            return ([{"name": "krbtgt", "rid": "502", "nt": "abc"}], None)

        onx, osd, odc = c.run_nxc_smb, c.run_secretsdump, c._is_dc
        c.run_nxc_smb, c.run_secretsdump, c._is_dc = fake_nxc, fake_dump, lambda h: False
        try:
            h = Host(ip="10.0.0.5", os_family="Windows",
                     ports=[Port(portid=445, state="open")])
            c.enrich_host(h, {"username": "bob", "password": "x", "domain": "corp"},
                          None, aggressive=False,
                          admin_creds={"username": "da", "password": "y", "domain": "corp"})
        finally:
            c.run_nxc_smb, c.run_secretsdump, c._is_dc = onx, osd, odc
        # secretsdump ran with the PRIVILEGED account, never the user account.
        self.assertEqual(used, [("secretsdump", "da")])
        titles = " ".join(v.title for v in h.vulns)
        self.assertIn("Local admin confirmed - privileged account", titles)
        self.assertIn("Credential hashes dumped", titles)
        # User enumeration still folded shares/users (once, not duplicated).
        self.assertEqual(sum(1 for a in h.accounts if a.kind == "share"), 1)

    def test_missing_tool_is_not_reported_as_auth_fail(self):
        """A missing netexec (run_nxc_smb -> (None, None)) must NOT record a FAIL
        cell nor attempt secretsdump - it's a tooling gap, not a bad credential."""
        from recce import credenum as c
        dumped = []
        onx, osd = c.run_nxc_smb, c.run_secretsdump
        c.run_nxc_smb = lambda ip, creds: (None, None)          # tool absent
        c.run_secretsdump = lambda ip, creds: (dumped.append(ip) or ([], None))
        try:
            h = Host(ip="10.0.0.5", os_family="Windows",
                     ports=[Port(portid=445, state="open")])
            issues, auth = c.enrich_host(
                h, {"username": "u", "password": "p", "domain": "d"}, None,
                admin_creds={"username": "a", "password": "p", "domain": "d"})
        finally:
            c.run_nxc_smb, c.run_secretsdump = onx, osd
        self.assertEqual(auth, {})           # nothing recorded -> cells show "-"
        self.assertEqual(dumped, [])         # no doomed secretsdump

    def test_secretsdump_skipped_when_admin_auth_rejected(self):
        """secretsdump must not run where the admin bind was rejected."""
        from recce import credenum as c
        dumped = []
        onx, osd, odc = c.run_nxc_smb, c.run_secretsdump, c._is_dc
        # Both accounts authenticate but neither is admin (auth True, admin False).
        c.run_nxc_smb = lambda ip, creds: (
            {"auth": True, "admin": False, "host_info": "", "shares": [],
             "users": [], "loggedon": [], "passpol": {}}, None)
        c.run_secretsdump = lambda ip, creds: (dumped.append(ip) or ([], None))
        c._is_dc = lambda h: False
        try:
            h = Host(ip="10.0.0.9", os_family="Windows",
                     ports=[Port(portid=445, state="open")])
            # Rejected admin: auth False for the admin account.
            c.run_nxc_smb = lambda ip, creds: (
                {"auth": creds["username"] == "u", "admin": False, "host_info": "",
                 "shares": [], "users": [], "loggedon": [], "passpol": {}}, None)
            issues, auth = c.enrich_host(
                h, {"username": "u", "password": "p", "domain": "d"}, None,
                admin_creds={"username": "adm", "password": "bad", "domain": "d"})
        finally:
            c.run_nxc_smb, c.run_secretsdump, c._is_dc = onx, osd, odc
        self.assertFalse(auth["admin"]["auth"])   # admin bind rejected
        self.assertEqual(dumped, [])              # so no secretsdump

    def test_smb_error_records_err_not_fail(self):
        """A tool/connection error (None, err) is ERR, distinct from a FAIL."""
        from recce import credenum as c
        onx = c.run_nxc_smb
        c.run_nxc_smb = lambda ip, creds: (None, "connection refused")
        try:
            h = Host(ip="10.0.0.7", os_family="Windows",
                     ports=[Port(portid=445, state="open")])
            _, auth = c.enrich_host(h, {"username": "u", "password": "p"}, None)
        finally:
            c.run_nxc_smb = onx
        self.assertTrue(auth["user"]["error"])
        self.assertFalse(auth["user"]["auth"])

    def test_ssh_finding_and_facts_recorded(self):
        from recce import credenum as c
        h = Host(ip="10.0.0.5", ports=[Port(portid=22, state="open")])
        c._fold_ssh(h, {"id": "uid=0(root)", "kernel": "Linux 5.4", "os": "Ubuntu",
                        "sudo": ["(ALL) NOPASSWD: ALL"], "suid": ["/opt/weird"]})
        self.assertTrue(any(s.id == "ssh-local-enum" for s in h.host_scripts))
        titles = [v.title for v in h.vulns]
        self.assertTrue(any("Sudo" in t for t in titles))
        self.assertTrue(any("SUID" in t for t in titles))

    def test_tool_gating_no_crash_when_absent(self):
        # With no external tools present, runners return (None/[], None) - no raise.
        from recce import credenum as c
        h = Host(ip="10.0.0.9", os_family="Windows",
                 ports=[Port(portid=445, state="open")])
        issues, auth = c.enrich_host(h, {"username": "u", "password": "p"}, None)
        self.assertTrue(h.cred_enumerated)
        self.assertIsInstance(issues, list)
        self.assertIsInstance(auth, dict)


class RobustnessTest(unittest.TestCase):
    """Field-crash guards: bad tool output / unexpected errors must not crash."""

    def test_run_survives_non_utf8_tool_output(self):
        # A service banner with raw non-UTF-8 bytes must not raise
        # UnicodeDecodeError mid-scan (errors='replace' on the runner).
        outcome = scanner._run(
            ["python3", "-c",
             "import sys; sys.stdout.buffer.write(b'open \\xff\\xfe port\\n')"])
        self.assertEqual(outcome.returncode, 0)
        self.assertIn("open", outcome.stdout)          # decoded, not crashed
        self.assertFalse(outcome.missing)

    def test_run_missing_tool_is_marked_not_raised(self):
        outcome = scanner._run(["definitely-not-a-real-binary-xyz", "--x"])
        self.assertTrue(outcome.missing)
        self.assertEqual(outcome.returncode, 127)

    def test_credenum_run_survives_non_utf8(self):
        from recce import credenum
        out, err = credenum._run(
            ["python3", "-c",
             "import sys; sys.stdout.buffer.write(b'\\xff\\xfe done')"])
        self.assertIsNone(err)
        self.assertIn("done", out)

    def test_parse_nmap_xml_never_raises_on_bad_files(self):
        with tempfile.TemporaryDirectory() as d:
            missing = os.path.join(d, "nope.xml")
            self.assertEqual(parser.parse_nmap_xml(missing), [])   # absent file
            for name, content in [
                ("empty.xml", ""),
                ("garbage.xml", "\x00\x01 not xml at all \xff"),
                ("trunc.xml", '<?xml version="1.0"?><nmaprun start="1"><host>'),
                ("partial.xml",
                 '<?xml version="1.0"?><nmaprun><host><status state="up"/>'
                 '<address addr="10.0.0.1" addrtype="ipv4"/></host>'),  # no close
            ]:
                p = os.path.join(d, name)
                with open(p, "w") as fh:
                    fh.write(content)
                out = parser.parse_nmap_xml(p)      # must not raise
                self.assertIsInstance(out, list)

    def test_xlsx_survives_control_chars_in_cells(self):
        # NSE/banner output with XML-illegal control bytes must not corrupt the
        # workbook - it must strip them and still read back.
        wb = xlsx.Workbook()
        sh = wb.add_sheet("S")
        sh.write([("banner \x00\x01\x08 with \x1f control bytes", "default")])
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "w.xlsx")
            wb.save(out)
            sheets = xlsx.read_sheets(out)          # must NOT raise ParseError
            flat = " ".join(str(c) for row in sheets["S"] for c in row)
            self.assertIn("banner", flat)
            self.assertNotIn("\x00", flat)          # control bytes stripped

    def test_docx_survives_control_chars(self):
        import xml.etree.ElementTree as ET
        import zipfile
        from recce.docx import Document
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "t.docx")
            doc = Document()
            doc.title("title \x00\x08")
            doc.mono_block("evidence \x00\x01\x1f bytes")
            doc.save(p)
            self.assertIsNone(zipfile.ZipFile(p).testzip())
            with zipfile.ZipFile(p) as z:            # Word-openable = well-formed XML
                ET.fromstring(z.read("word/document.xml"))

    def test_store_raises_clean_error_on_corrupt_db(self):
        from recce.store import Store, StoreError
        with tempfile.TemporaryDirectory() as d:
            bad = os.path.join(d, "results.sqlite")
            with open(bad, "wb") as fh:
                fh.write(b"this is not a sqlite database at all\x00\x01")
            with self.assertRaises(StoreError):
                Store(bad)

    def test_invalid_targets_exit_clean(self):
        # A bad CIDR/range must yield a clean "Invalid targets" message + a None
        # result (caller exits 1), not a traceback. Exercised via _discover so the
        # test doesn't depend on nmap being installed.
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            paths = cli._open_paths(d)
            store = Store(paths["db"])
            args = SimpleNamespace(targets=["10.0.0.0/99"], exclude=[], fast=False)
            profile = scanner.PROFILES["standard"]
            buf = io.StringIO()
            with contextlib.redirect_stdout(buf):
                result = cli._discover(args, profile, store, paths)
            store.close()
            self.assertEqual(result, (None, [], None))
            self.assertIn("Invalid targets", buf.getvalue())

    def test_main_top_level_guard_returns_clean_on_crash(self):
        # An unexpected error inside a command must become a clean exit 1, not a
        # traceback dumped at the tester.
        import argparse
        from recce import cli

        def boom(args):
            raise RuntimeError("simulated deep crash")

        class _P:
            def parse_args(self, _a):
                ns = argparse.Namespace()
                ns.command = "boom"      # non-None so main() dispatches to func
                ns.func = boom
                return ns

        orig = cli.build_arg_parser
        cli.build_arg_parser = lambda: _P()
        try:
            buf = io.StringIO()
            with contextlib.redirect_stdout(buf):
                rc = cli.main([])
        finally:
            cli.build_arg_parser = orig
        self.assertEqual(rc, 1)
        out = buf.getvalue()
        self.assertIn("unexpected error", out)
        self.assertNotIn("Traceback", out)             # no raw traceback by default


class ScanHardeningTest(unittest.TestCase):
    def test_timeout_and_version_args(self):
        p = scanner.PROFILES["standard"]
        args, kill = scanner._timeout_args(p)
        self.assertEqual(args, ["--host-timeout", f"{p.host_timeout}m"])
        self.assertEqual(kill, p.host_timeout * 60 + 120)
        # 0 disables both.
        self.assertEqual(scanner._timeout_args(p, 0), ([], None))
        # Service detection: explicit intensity vs --version-all.
        self.assertIn("--version-intensity", scanner._version_args(p))
        self.assertIn("--version-all", scanner._version_args(scanner.PROFILES["thorough"]))

    def test_issue_classification(self):
        s = scanner
        self.assertEqual(
            s._issue_from(s.RunOutcome(timed_out=True), "/no", "enum", 20).level,
            "error")
        self.assertEqual(
            s._issue_from(s.RunOutcome(missing=True), "/no", "enum", 20).level,
            "error")
        ht = s.RunOutcome(returncode=0, stdout="Skipping host X due to host timeout")
        self.assertEqual(s._issue_from(ht, "/no", "port-sweep", 20).level, "warning")
        # A clean run against a real (existing, non-empty) file -> no issue.
        self.assertIsNone(s._issue_from(s.RunOutcome(returncode=0), SAMPLE, "enum", 20))

    def test_store_issue_log(self):
        with tempfile.TemporaryDirectory() as d:
            store = Store(os.path.join(d, "s.sqlite"))
            store.add_issue("10.0.0.5", "port-sweep", "warning", "host timed out")
            store.add_issue("10.0.0.9", "enum", "error", "nmap unresponsive")
            self.assertEqual(store.count_issues(),
                             {"warning": 1, "error": 1, "total": 2})
            issues = store.get_issues()
            self.assertEqual(len(issues), 2)
            self.assertEqual(issues[0]["ip"], "10.0.0.9")   # newest first
            store.close()

    def test_overview_surfaces_issues(self):
        issues = [{"ts": "t", "ip": "10.0.0.9", "phase": "enum", "level": "error",
                   "message": "hard-timed-out"}]
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook([Host(ip="10.0.0.5", subnet="10.0.0.0/24")], out,
                           issues=issues)
            flat = [str(c) for r in xlsx.read_sheets(out)["Overview"] for c in r]
            self.assertTrue(any("SCAN ISSUES" in c for c in flat))
            self.assertTrue(any("hard-timed-out" in c for c in flat))

    def test_migration_adds_issues_table(self):
        # A datastore created before the issues table still gains it.
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "old.sqlite")
            import sqlite3
            con = sqlite3.connect(path)
            con.execute("CREATE TABLE meta (key TEXT PRIMARY KEY, value TEXT)")
            con.commit(); con.close()
            store = Store(path)
            store.add_issue("1.2.3.4", "enum", "error", "boom")
            self.assertEqual(store.count_issues()["total"], 1)
            store.close()


class ProbesTest(unittest.TestCase):
    def test_port_classification(self):
        from recce import probes
        self.assertTrue(probes._is_tls(Port(portid=443, service="https")))
        self.assertTrue(probes._is_tls(Port(portid=8443, service="http", tunnel="ssl")))
        self.assertFalse(probes._is_tls(Port(portid=80, service="http")))
        self.assertTrue(probes._is_http(Port(portid=8080, service="http-proxy")))
        self.assertTrue(probes._is_http(Port(portid=443, service="https")))
        self.assertFalse(probes._is_http(Port(portid=22, service="ssh")))

    def test_http_header_findings_flag_missing_headers(self):
        from recce import probes
        port = Port(portid=80, service="http")
        # Server present with a version, but security headers absent.
        headers = {"server": "Apache/2.4.41", "content-type": "text/html"}
        orig = probes._fetch_headers
        probes._fetch_headers = lambda ip, p, tls: (200, headers)
        try:
            findings = probes.http_findings("10.0.0.9", port)
        finally:
            probes._fetch_headers = orig
        titles = {f.title for f in findings}
        self.assertIn("Missing X-Frame-Options / frame-ancestors (clickjacking)", titles)
        self.assertIn("Missing X-Content-Type-Options header (MIME sniffing)", titles)
        self.assertTrue(any("banner discloses" in t for t in titles))
        # No HSTS finding over plain HTTP.
        self.assertNotIn("Missing HSTS header", titles)
        for f in findings:
            self.assertEqual(f.source, "probe")
            if "banner" not in f.title:
                self.assertTrue(f.cwes)

    def test_http_findings_none_when_unreachable(self):
        from recce import probes
        orig = probes._fetch_headers
        probes._fetch_headers = lambda ip, p, tls: None
        try:
            self.assertEqual(probes.http_findings("10.0.0.9", Port(portid=80)), [])
        finally:
            probes._fetch_headers = orig

    def test_parse_cert_time(self):
        from recce import probes
        epoch = probes._parse_cert_time("Jun  1 12:00:00 2030 GMT")
        self.assertIsNotNone(epoch)
        self.assertIsNone(probes._parse_cert_time("not a date"))

    def test_probe_host_dedups(self):
        from recce import probes
        h = Host(ip="10.0.0.9", ports=[Port(portid=80, service="http")])
        headers = {"server": "nginx"}
        orig = probes._fetch_headers
        probes._fetch_headers = lambda ip, p, tls: (200, headers)
        try:
            first = probes.probe_host(h)
            second = probes.probe_host(h)   # idempotent re-run
        finally:
            probes._fetch_headers = orig
        self.assertGreater(first, 0)
        self.assertEqual(second, 0)


class ExploitsTest(unittest.TestCase):
    SS_JSON = ('{"RESULTS_EXPLOIT": ['
               '{"Title": "vsftpd 2.3.4 - Backdoor Command Execution",'
               ' "EDB-ID": "17491", "Type": "remote", "Path": "unix/remote/17491.rb",'
               ' "Codes": "CVE-2011-2523"}]}')

    def test_parse_json(self):
        recs = exploits.parse_searchsploit_json(self.SS_JSON)
        self.assertEqual(len(recs), 1)
        self.assertEqual(recs[0]["EDB-ID"], "17491")

    def test_record_to_exploit_extracts_cve(self):
        rec = exploits.parse_searchsploit_json(self.SS_JSON)[0]
        e = exploits._record_to_exploit(rec, "10.0.0.9", 21, "vsftpd", "2.3.4")
        self.assertEqual(e.edb_id, "17491")
        self.assertIn("CVE-2011-2523", e.cves)
        self.assertEqual(e.type, "remote")

    def test_clean_version(self):
        self.assertEqual(exploits._clean_version("8.2p1 Ubuntu 4ubuntu0.5"), "8.2p1")
        self.assertEqual(exploits._clean_version("2.4.41"), "2.4.41")

    def test_query_terms_trims_vendor(self):
        self.assertEqual(exploits._query_terms("Apache httpd", "2.4.41"), "httpd 2.4.41")

    def test_exploit_tracking_key_and_coverage(self):
        h = Host(ip="10.0.0.9", subnet="10.0.0.0/24")
        from recce.models import Exploit
        h.exploits = [Exploit(ip="10.0.0.9", port=21, edb_id="17491")]
        keys = tr.item_keys([h])
        self.assertIn(tr.exploit_key("10.0.0.9", 21, "17491"), keys["exploits"])
        self.assertIn("exploits", tr.COVERAGE_CATEGORIES)


class SubnetCoverageTest(unittest.TestCase):
    def test_overview_includes_empty_scope_subnet(self):
        from recce.report_excel import build_workbook
        from recce import xlsx
        hosts = [Host(ip="10.0.10.5", subnet="10.0.10.0/24", enumerated=True)]
        scope = {"10.0.10.0/24": 254, "10.0.99.0/24": 254}  # 2nd has no live hosts
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(hosts, out, scope=scope)
            rows = xlsx.read_sheets(out)["Overview"]
        subnets = [r[0] for r in rows if r and r[0].startswith("10.0.")]
        self.assertIn("10.0.99.0/24", subnets)   # empty subnet still accounted for
        self.assertIn("10.0.10.0/24", subnets)

    def test_checklist_grouped_by_subnet(self):
        from recce.report_excel import _spec_checklist
        # up_reason set: a real discovery reply keeps a 0-port host on the list (the
        # Checklist shows only confirmed-up hosts).
        hosts = [Host(ip="10.0.20.9", subnet="10.0.20.0/24", up_reason="syn-ack"),
                 Host(ip="10.0.10.5", subnet="10.0.10.0/24", up_reason="echo-reply")]
        spec = _spec_checklist(hosts)
        rows = spec.rows
        # Sorted by subnet then IP -> 10.0.10.x before 10.0.20.x.
        self.assertEqual([r["data"]["IP"] for r in rows], ["10.0.10.5", "10.0.20.9"])
        # Subnet now lives in the collapsible group band, not a per-row column.
        self.assertEqual(spec.group_by, "Subnet")
        self.assertEqual(rows[0]["group"], "10.0.10.0/24")
        self.assertNotIn("Subnet", rows[0]["data"])

    def test_checklist_collapsible_band_rollup_and_risk_sort(self):
        from recce.report_excel import build_workbook
        from recce.models import Vuln
        from recce import xlsx as _x, tracking as _tr
        crit = Host(ip="10.0.10.9", subnet="10.0.10.0/24", state="up", enumerated=True,
                    ports=[Port(portid=445, service="smb")],
                    vulns=[Vuln(ip="10.0.10.9", port=445, protocol="tcp", script_id="x",
                                title="f", severity="critical", source="nse",
                                state="finding")])
        clean = Host(ip="10.0.10.1", subnet="10.0.10.0/24", state="up", enumerated=True,
                     ports=[Port(portid=22, service="ssh")])
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook([clean, crit], out,
                           tracking={_tr.host_key("10.0.10.1"): (True, "")})
            rows = xlsx.read_sheets(out)["Checklist"]
        hidx = header_index(rows, "IP")
        ipc = rows[hidx].index("IP")
        band = next(r for r in rows[hidx + 1:] if str(r[ipc]).startswith("10.0.10.0/24"))
        self.assertIn("2 hosts", str(band[ipc]))
        self.assertIn("1/2 reviewed", str(band[ipc]))
        self.assertIn("high/crit", str(band[ipc]))                # the critical host
        # Risk-first: the critical host sorts above the clean host within the subnet.
        host_ips = [r[ipc] for r in rows[hidx + 1:]
                    if r[0] in (_x.CHECK_ON, _x.CHECK_OFF)]
        self.assertEqual(host_ips, ["10.0.10.9", "10.0.10.1"])

    def test_store_scope_roundtrip(self):
        with tempfile.TemporaryDirectory() as d:
            store = Store(os.path.join(d, "s.sqlite"))
            store.set_scope("10.0.0.0/24", 254)
            store.set_scope("10.0.0.0/24", 100)  # keeps the larger
            self.assertEqual(store.get_scope()["10.0.0.0/24"], 254)
            store.close()


class HostUpCertaintyTest(unittest.TestCase):
    """The Checklist shows only hosts we can PROVE are up - but is never allowed to
    write a live host off as down. is_up is the single source of that judgement."""

    def test_is_up_only_on_positive_evidence(self):
        from recce.models import Vuln
        # An open port is unambiguous proof.
        self.assertTrue(Host(ip="1.1.1.1",
                             ports=[Port(portid=22, state="open")]).is_up)
        # A finding means a service actually responded.
        self.assertTrue(Host(ip="1.1.1.1",
                             vulns=[Vuln(ip="1.1.1.1", port=0, protocol="tcp",
                                         script_id="x", title="t", severity="low",
                                         source="nse", state="finding")]).is_up)
        # `enumerated` alone is NOT proof: the pipeline sets it on every host it tries,
        # including a dead -Pn IP that answered nothing.
        self.assertFalse(Host(ip="1.1.1.1", enumerated=True).is_up)
        # A real nmap discovery reply (not the -Pn assume-up).
        self.assertTrue(Host(ip="1.1.1.1", up_reason="echo-reply").is_up)
        self.assertTrue(Host(ip="1.1.1.1", up_reason="arp-response").is_up)
        # DNS / ARP / OS evidence => it answered something.
        self.assertTrue(Host(ip="1.1.1.1", mac="00:11:22:33:44:55").is_up)
        self.assertTrue(Host(ip="1.1.1.1", hostnames=["dc01"]).is_up)
        # A closed/filtered-only port is NOT an open port.
        self.assertFalse(Host(ip="1.1.1.1",
                              ports=[Port(portid=22, state="filtered")]).is_up)
        # The -Pn blanket assume-up ("user-set") is NOT proof, and a bare host isn't.
        self.assertFalse(Host(ip="1.1.1.1", state="up", up_reason="user-set").is_up)
        self.assertFalse(Host(ip="1.1.1.1").is_up)

    def test_checklist_hides_unconfirmed_keeps_confirmed(self):
        from recce.report_excel import build_workbook
        confirmed = Host(ip="10.0.0.5", subnet="10.0.0.0/24", state="up",
                         up_reason="syn-ack", ports=[Port(portid=445, state="open")])
        phantom = Host(ip="10.0.0.6", subnet="10.0.0.0/24", state="up",
                       up_reason="user-set")     # -Pn assume-up, no proof of life
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook([confirmed, phantom], out)
            rows = xlsx.read_sheets(out)["Checklist"]
        hidx = header_index(rows, "IP")
        ipc = rows[hidx].index("IP")
        ips = {str(r[ipc]) for r in rows[hidx + 1:]}
        self.assertIn("10.0.0.5", ips)            # confirmed-up host shown
        self.assertNotIn("10.0.0.6", ips)         # unconfirmed phantom hidden

    def test_checklist_carries_legend_above_header_and_round_trips(self):
        from recce.report_excel import (build_workbook, read_workbook_tracking,
                                         CHECKLIST_TITLE)
        h = Host(ip="10.0.0.5", subnet="10.0.0.0/24", state="up", enumerated=True,
                 ports=[Port(portid=445, service="smb", state="open", vuln_scanned=True)])
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook([h], out)
            rows = xlsx.read_sheets(out)["Checklist"]
            # A legend line precedes the header row (row 0 is not the header).
            hidx = header_index(rows, "IP")
            self.assertGreater(hidx, 0)
            self.assertIn("Legend", str(rows[0][0]))
            self.assertIn("confirmed UP", str(rows[0][0]))
            # Tracking still round-trips despite the shifted header: an auto-ticked
            # vuln step reads back True.
            back = read_workbook_tracking(out)
            self.assertTrue(back[tr.step_key("vuln", "10.0.0.5")][0])

    def test_overview_tallies_unconfirmed_hosts(self):
        from recce.report_excel import build_workbook
        confirmed = Host(ip="10.0.0.5", subnet="10.0.0.0/24",
                         ports=[Port(portid=445, state="open")])
        phantom = Host(ip="10.0.0.6", subnet="10.0.0.0/24", up_reason="user-set")
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook([confirmed, phantom], out)
            rows = xlsx.read_sheets(out)["Overview"]
        blob = "\n".join(" ".join(str(c) for c in r) for r in rows)
        self.assertIn("Scanned, not confirmed up", blob)
        self.assertIn("Hosts confirmed up", blob)

    def test_udp_fallback_flips_silent_pn_host_to_up(self):
        # A -Pn host silent on TCP gets a UDP liveness ping; a reply confirms it up.
        from recce import cli, scanner
        saved = (cli._ports_for_host, cli._fold_host, scanner.full_port_scan,
                 scanner.verify_port_scan, scanner.udp_liveness_probe,
                 scanner.enum_scan, cli.np.parse_nmap_xml)
        udp_calls = {"n": 0}

        def fake_parse(path):
            # Only the UDP-liveness XML reports a live host; TCP/enum XMLs are empty.
            if "udpalive" in path:
                return [Host(ip="10.0.0.9", up_reason="udp-response")]
            return []

        def fake_udp(ip, out, profile):
            udp_calls["n"] += 1
            return out, None
        cli._ports_for_host = lambda path, ip: []          # silent on TCP
        cli._fold_host = lambda ip, parsed, sm: Host(ip=ip, subnet="10.0.0.0/24")
        scanner.full_port_scan = lambda ip, out, profile: (out, None)
        scanner.verify_port_scan = lambda ip, out, profile: (out, None)
        scanner.udp_liveness_probe = fake_udp
        scanner.enum_scan = lambda ip, ports, out, profile, creds=None: (out, None)
        cli.np.parse_nmap_xml = fake_parse
        try:
            with tempfile.TemporaryDirectory() as d:
                prof = scanner.ScanProfile(ping_discovery=False, assume_up=True)
                host, _ = cli._enum_worker("10.0.0.9", prof, {"raw": d}, None, None,
                                           {"10.0.0.9": "10.0.0.0/24"})
            self.assertEqual(udp_calls["n"], 1)            # UDP fallback fired
            self.assertEqual(host.up_reason, "udp-response")
            self.assertTrue(host.is_up)                    # up despite 0 open TCP ports
        finally:
            (cli._ports_for_host, cli._fold_host, scanner.full_port_scan,
             scanner.verify_port_scan, scanner.udp_liveness_probe,
             scanner.enum_scan, cli.np.parse_nmap_xml) = saved

    def test_discovery_reply_reason_propagates_and_skips_udp(self):
        # A host discovered live carries its real reply reason into the stored host,
        # and the UDP fallback is NOT wasted on a host we already proved is up.
        from recce import cli, scanner
        saved = (cli._ports_for_host, cli._fold_host, scanner.full_port_scan,
                 scanner.verify_port_scan, scanner.udp_liveness_probe,
                 scanner.enum_scan, cli.np.parse_nmap_xml)
        udp_calls = {"n": 0}
        cli._ports_for_host = lambda path, ip: []          # silent on TCP
        cli._fold_host = lambda ip, parsed, sm: Host(ip=ip, subnet="10.0.0.0/24")
        scanner.full_port_scan = lambda ip, out, profile: (out, None)
        scanner.verify_port_scan = lambda ip, out, profile: (out, None)
        scanner.udp_liveness_probe = lambda ip, out, profile: (
            udp_calls.__setitem__("n", udp_calls["n"] + 1), (out, None))[1]
        scanner.enum_scan = lambda ip, ports, out, profile, creds=None: (out, None)
        cli.np.parse_nmap_xml = lambda path: []
        try:
            with tempfile.TemporaryDirectory() as d:
                prof = scanner.ScanProfile(ping_discovery=True)
                host, _ = cli._enum_worker("10.0.0.9", prof, {"raw": d}, None, None,
                                           {"10.0.0.9": "10.0.0.0/24"},
                                           disc_reason="echo-reply")
            self.assertEqual(host.up_reason, "echo-reply")
            self.assertTrue(host.is_up)
            self.assertEqual(udp_calls["n"], 0)            # already proven up -> no UDP
        finally:
            (cli._ports_for_host, cli._fold_host, scanner.full_port_scan,
             scanner.verify_port_scan, scanner.udp_liveness_probe,
             scanner.enum_scan, cli.np.parse_nmap_xml) = saved

    def test_merge_never_downgrades_proof_of_life(self):
        # A real reply must survive a later -Pn re-scan that only knows "user-set".
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            store = Store(os.path.join(d, "s.sqlite"))
            store.upsert_host(Host(ip="10.0.0.5", up_reason="echo-reply"))
            store.upsert_host(Host(ip="10.0.0.5", up_reason="user-set"))
            got = store.get_host("10.0.0.5")
            self.assertEqual(got.up_reason, "echo-reply")
            self.assertTrue(got.is_up)
            store.close()


class VulnDbTest(unittest.TestCase):
    def test_version_comparator(self):
        from recce import vulndb
        self.assertLess(vulndb._cmp("2.4.41", "2.4.53"), 0)
        self.assertGreater(vulndb._cmp("2.4.50", "2.4.49"), 0)
        self.assertLess(vulndb._cmp("8.2p1", "8.5"), 0)
        self.assertGreater(vulndb._cmp("1.0.2k", "1.0.2"), 0)
        self.assertEqual(vulndb._cmp("2.3.4", "2.3.4"), 0)

    def test_mariadb_handshake_prefix_not_read_as_eol_mysql(self):
        """Regression: MariaDB 10.x announces '5.5.5-10.x.y-MariaDB'. The
        leading 5.5.5 must not be read as the version, or a patched MariaDB gets
        a bogus EOL medium + high CVE-2012-2122."""
        from recce import vulndb
        self.assertEqual(vulndb._clean_version("5.5.5-10.11.6-MariaDB-0+deb12u1"),
                         "10.11.6-MariaDB-0+deb12u1")
        h = Host(ip="10.0.0.5", ports=[Port(portid=3306, service="mysql",
                 product="MySQL", version="5.5.5-10.11.6-MariaDB-0+deb12u1")])
        vulndb.assess_host_inplace(h)
        titles = {v.title for v in h.vulns}
        self.assertFalse(any("MySQL" in t for t in titles), titles)
        # A genuine old MySQL 5.5.40 (no handshake prefix) is still flagged.
        h2 = Host(ip="10.0.0.6", ports=[Port(portid=3306, service="mysql",
                  product="MySQL", version="5.5.40")])
        vulndb.assess_host_inplace(h2)
        self.assertTrue(any("End-of-life MySQL" in v.title for v in h2.vulns))

    def test_product_advisory_reported_on_every_matching_port(self):
        """Regression: a product-only advisory exposed on two ports must yield a
        finding per port (was deduped by title, dropping all but the first)."""
        from recce import vulndb
        from recce.report_docx import group_findings
        h = Host(ip="10.0.0.5", ports=[
            Port(portid=8090, service="http", product="Atlassian Confluence", version=""),
            Port(portid=8091, service="http", product="Atlassian Confluence", version="")])
        vulndb.assess_host_inplace(h)
        conf = [v for v in h.vulns if "Confluence" in v.title]
        self.assertEqual(sorted(v.port for v in conf), [8090, 8091])
        # The grouped write-up lists both affected ports.
        f = next(f for f in group_findings([h]) if "Confluence" in f.title)
        self.assertEqual(sorted({a[1] for a in f.affected}), [8090, 8091])

    def test_exact_and_range_matches(self):
        from recce import vulndb
        h = Host(ip="10.0.0.9", os_name="Linux", ports=[
            Port(portid=21, service="ftp", product="vsftpd", version="2.3.4"),
            Port(portid=80, service="http", product="Apache httpd", version="2.4.41"),
            Port(portid=3306, service="mysql", product="MySQL", version="5.7.38"),
        ])
        vulndb.assess_host_inplace(h)
        titles = {v.title for v in h.vulns}
        self.assertTrue(any("vsftpd 2.3.4 backdoor" in t for t in titles))   # exact
        self.assertTrue(any("Apache" in t for t in titles))                  # range
        # MySQL 5.7.38 is >= 5.7 -> not flagged as EOL (< 5.7).
        self.assertFalse(any("End-of-life MySQL" in t for t in titles))

    def test_findings_carry_remediation_and_source(self):
        from recce import vulndb
        h = Host(ip="10.0.0.9", ports=[Port(portid=21, service="ftp",
                 product="vsftpd", version="2.3.4")])
        vulndb.assess_host_inplace(h)
        v = h.vulns[0]
        self.assertEqual(v.source, "version-db")
        self.assertEqual(v.severity, "critical")
        self.assertIn("CVE-2011-2523", v.ids)
        self.assertTrue(v.remediation)

    def test_multiple_findings_per_port_have_distinct_keys(self):
        from recce.models import Vuln
        a = Vuln(ip="1.1.1.1", port=80, protocol="tcp", script_id="version-db",
                 title="Finding A")
        b = Vuln(ip="1.1.1.1", port=80, protocol="tcp", script_id="version-db",
                 title="Finding B")
        self.assertNotEqual(a.key, b.key)

    def test_no_version_no_false_positive(self):
        from recce import vulndb
        # product matches but no version -> a version-gated sig must not fire.
        h = Host(ip="10.0.0.9", ports=[Port(portid=80, service="http",
                 product="Apache httpd", version="")])
        n = vulndb.assess_host_inplace(h)
        self.assertEqual(n, 0)

    def test_signature_database_is_large(self):
        from recce import vulndb
        self.assertGreaterEqual(vulndb.signature_count(), 80)

    def test_new_signature_categories_match(self):
        from recce import vulndb
        cases = {
            "ActiveMQ OpenWire transport": "ActiveMQ",
            "Oracle WebLogic admin httpd": "WebLogic",
            "Docker": "Docker Engine API",
            "Apache Solr": "Solr",
            "Zabbix": "Zabbix",
            "JetBrains TeamCity": "TeamCity",
            "VMware ESXi": "ESXi",
            "Apache CouchDB": "CouchDB",
            "Ivanti Connect Secure": "Ivanti",
            "F5 BIG-IP": "BIG-IP",
            "MikroTik RouterOS": "MikroTik",
            "Cisco ASA": "Cisco ASA",
        }
        for product, expect in cases.items():
            h = Host(ip="1.1.1.1", ports=[Port(portid=8080, service="http",
                     product=product, state="open")])
            vulndb.assess_host_inplace(h)
            self.assertTrue(any(expect in v.title for v in h.vulns),
                            f"{product} -> expected a '{expect}' finding")

    def test_windows_advisories_are_os_gated(self):
        from recce import vulndb
        # A non-DC Windows host gets the Windows SMB advisories, but NOT ZeroLogon
        # (which attacks a domain controller's Netlogon only).
        win = Host(ip="1.1.1.1", os_family="Windows", os_name="Windows Server 2019",
                   ports=[Port(portid=445, service="microsoft-ds",
                               product="Microsoft Windows Server 2019", state="open")])
        vulndb.assess_host_inplace(win)
        titles = " ".join(v.title for v in win.vulns)
        for expect in ("SMBGhost", "PrintNightmare"):
            self.assertIn(expect, titles)
        self.assertNotIn("ZeroLogon", titles)          # DC-only -> not on a member
        # A Linux/Samba SMB host must NOT get the Windows-only advisories.
        lin = Host(ip="1.1.1.2", os_family="Linux", os_name="Linux",
                   ports=[Port(portid=445, service="microsoft-ds",
                               product="Samba smbd", version="4.13.0", state="open")])
        vulndb.assess_host_inplace(lin)
        self.assertFalse(any(w in " ".join(v.title for v in lin.vulns)
                             for w in ("SMBGhost", "PrintNightmare", "ZeroLogon")))

    def test_iis_mssql_seimpersonate_potato_advisories(self):
        from recce import vulndb
        h = Host(ip="10.0.10.50", os_family="Windows", os_name="Windows 11",
                 ports=[Port(portid=80, service="http",
                             product="Microsoft IIS httpd", version="10.0"),
                        Port(portid=1433, service="ms-sql-s",
                             product="Microsoft SQL Server", version="15.0")])
        vulndb.assess_host_inplace(h)
        titles = " ".join(v.title for v in h.vulns)
        self.assertIn("IIS AppPool - SeImpersonate", titles)
        self.assertIn("MSSQL service account - SeImpersonate", titles)
        potato = [v for v in h.vulns if "SeImpersonate" in v.title]
        for v in potato:
            self.assertEqual(v.confidence, "potential")       # advisory
            self.assertIn("CWE-269", v.cwes)
            self.assertIn("GodPotato", v.output + v.remediation or "")

    def test_zerologon_is_dc_only(self):
        from recce import vulndb
        # A real DC (Kerberos 88 + LDAP 389 + SMB 445) DOES get ZeroLogon.
        dc = Host(ip="10.0.10.10", os_family="Windows", os_name="Windows Server 2019",
                  ports=[Port(portid=88, service="kerberos-sec", state="open"),
                         Port(portid=389, service="ldap", state="open"),
                         Port(portid=445, service="microsoft-ds",
                              product="Windows Server 2019", state="open")])
        vulndb.assess_host_inplace(dc)
        self.assertIn("ZeroLogon", " ".join(v.title for v in dc.vulns))
        # Role-tagged DC with only SMB visible still matches via the role.
        dc2 = Host(ip="10.0.10.11", os_family="Windows", roles=["Domain Controller"],
                   ports=[Port(portid=445, service="microsoft-ds",
                               product="Windows Server", state="open")])
        vulndb.assess_host_inplace(dc2)
        self.assertIn("ZeroLogon", " ".join(v.title for v in dc2.vulns))

    def test_jetty_version_gate(self):
        from recce import vulndb
        for ver, should in [("9.4.30.v20200611", True), ("9.4.50", False)]:
            h = Host(ip="1.1.1.1", ports=[Port(portid=8080, service="http",
                     product="Jetty", version=ver, state="open")])
            vulndb.assess_host_inplace(h)
            hit = any("Jetty" in v.title for v in h.vulns)
            self.assertEqual(hit, should, f"Jetty {ver}")

    def test_findings_carry_cwes(self):
        from recce import vulndb
        h = Host(ip="10.0.0.9", ports=[Port(portid=21, service="ftp",
                 product="vsftpd", version="2.3.4")])
        vulndb.assess_host_inplace(h)
        v = h.vulns[0]
        self.assertTrue(v.cwes)
        self.assertTrue(all(c.startswith("CWE-") for c in v.cwes))

    def test_advisory_signature_is_product_only_and_potential(self):
        from recce import vulndb
        # A product-only advisory (no version) should still fire, tagged potential.
        h = Host(ip="10.0.0.9", ports=[Port(portid=8080, service="http",
                 product="Apache Tomcat", version="")])
        vulndb.assess_host_inplace(h)
        adv = [v for v in h.vulns if "default credentials" in v.title]
        self.assertTrue(adv)
        self.assertEqual(adv[0].confidence, "potential")
        self.assertTrue(adv[0].cwes)

    def test_every_signature_has_cwe_field(self):
        from recce import vulndb
        for sig in vulndb.SIGNATURES:
            self.assertIn("cwe", sig, f"{sig['title']} missing cwe")
            self.assertTrue(sig["cwe"], f"{sig['title']} empty cwe")


class PhaseModelTest(unittest.TestCase):
    def _host(self, ip="10.0.0.5", scanned=None):
        h = Host(ip=ip, subnet="10.0.0.0/24", ports=[
            Port(portid=80, service="http"), Port(portid=445, service="microsoft-ds")])
        if scanned:
            for p in h.ports:
                if p.portid in scanned:
                    p.vuln_scanned = True
        return h

    def test_status_transitions(self):
        h = self._host()
        self.assertEqual(h.status, "discovered")
        h.enumerated = True
        self.assertEqual(h.status, "enumerated")
        h.ports[0].vuln_scanned = True
        self.assertEqual(h.status, "vuln-scanned 1/2")
        h.ports[1].vuln_scanned = True
        self.assertEqual(h.status, "vuln-scanned")

    def test_vuln_targets_only_and_unscanned(self):
        from recce import cli
        h = self._host(scanned={80})
        h.enumerated = True
        # --only http -> just port 80
        ns = SimpleNamespace(only=["http"], subnet=None, host=None, unscanned=False)
        tgt = cli._vuln_targets([h], ns)
        self.assertEqual(tgt, [(h, [80])])
        # --unscanned -> only port 445 (80 already scanned)
        ns = SimpleNamespace(only=None, subnet=None, host=None, unscanned=True)
        self.assertEqual(cli._vuln_targets([h], ns), [(h, [445])])
        # --only by port number
        ns = SimpleNamespace(only=["445"], subnet=None, host=None, unscanned=False)
        self.assertEqual(cli._vuln_targets([h], ns), [(h, [445])])

    def test_vuln_targets_subnet_and_host_filter(self):
        from recce import cli
        a = self._host("10.0.0.5"); b = self._host("10.0.1.9")
        b.subnet = "10.0.1.0/24"
        ns = SimpleNamespace(only=None, subnet=["10.0.0.0/24"], host=None, unscanned=False)
        got = cli._vuln_targets([a, b], ns)
        self.assertEqual([h.ip for h, _ in got], ["10.0.0.5"])

    def test_merge_vuln_results(self):
        from recce import cli
        from recce.models import Vuln
        h = self._host()
        parsed = Host(ip="10.0.0.5", ports=[Port(portid=80, service="http",
                      scripts=[Script(id="http-git", output="x")])],
                      vulns=[Vuln(ip="10.0.0.5", port=80, protocol="tcp",
                                  script_id="http-git", severity="medium")])
        cli._merge_vuln_results(h, [parsed])
        self.assertEqual(len(h.vulns), 1)
        self.assertTrue(any(s.id == "http-git" for s in h.ports[0].scripts))


class TargetingTest(unittest.TestCase):
    def test_ip_matcher(self):
        from recce.targets import ip_matcher
        m = ip_matcher(["10.0.0.5", "10.0.1.0/24", "192.168.1.10-12"])
        self.assertTrue(m("10.0.0.5"))       # exact ip
        self.assertTrue(m("10.0.1.99"))      # in cidr
        self.assertTrue(m("192.168.1.11"))   # in range
        self.assertFalse(m("10.0.0.6"))
        self.assertFalse(m("172.16.0.1"))

    def test_empty_matches_all(self):
        from recce.targets import ip_matcher
        m = ip_matcher([])
        self.assertTrue(m("1.2.3.4"))

    def test_selected_hosts(self):
        from recce import cli
        a = Host(ip="10.0.0.5", subnet="10.0.0.0/24")
        b = Host(ip="10.0.9.9", subnet="10.0.9.0/24")
        ns = SimpleNamespace(targets=["10.0.0.0/24"], host=None, subnet=None)
        self.assertEqual([h.ip for h in cli._selected_hosts([a, b], ns)], ["10.0.0.5"])


class DatabaseModuleTest(unittest.TestCase):
    def test_engine_detection(self):
        from recce import db
        self.assertEqual(db.engine_for(Port(portid=3306)), "mysql")
        self.assertEqual(db.engine_for(Port(portid=1433)), "mssql")
        self.assertEqual(db.engine_for(Port(portid=9999, service="postgresql")), "postgresql")
        self.assertIsNone(db.engine_for(Port(portid=80, service="http")))

    def test_db_instances(self):
        from recce import db
        from recce.models import Vuln
        h = Host(ip="10.0.0.9", ports=[Port(portid=3306, service="mysql",
                 product="MySQL", version="5.7.38")])
        h.vulns = [Vuln(ip="10.0.0.9", port=3306, protocol="tcp",
                        script_id="mysql-empty-password", title="Database account "
                        "with empty password", severity="high")]
        inst = db.db_instances([h])
        self.assertEqual(len(inst), 1)
        self.assertEqual(inst[0]["engine"], "mysql")
        self.assertEqual(inst[0]["auth"], "EMPTY PASSWORD")

    def test_script_selection_aggressive(self):
        from recce import db
        safe = db.script_selection(False)
        aggr = db.script_selection(True)
        self.assertIn("mysql-info", safe)
        self.assertNotIn("mysql-brute", safe)
        self.assertIn("mysql-brute", aggr)


class PrivescModuleTest(unittest.TestCase):
    def test_windows_playbook(self):
        # The generic OS checklist now lives on the separate reference sheet
        # (playbook_rows), scoped to the OSes present in the engagement.
        from recce import privesc
        h = Host(ip="10.0.0.5", os_family="Windows",
                 ports=[Port(portid=445, service="microsoft-ds")])
        oses = {r["os"] for r in privesc.playbook_rows([h])}
        self.assertEqual(oses, {"windows"})

    def test_linux_playbook(self):
        from recce import privesc
        h = Host(ip="10.0.0.6", os_family="Linux",
                 ports=[Port(portid=22, service="ssh")])
        oses = {r["os"] for r in privesc.playbook_rows([h])}
        self.assertEqual(oses, {"linux"})

    def test_playbook_shows_both_oses_for_mixed_or_unknown_scope(self):
        from recce import privesc
        mixed = [Host(ip="10.0.0.5", os_family="Windows"),
                 Host(ip="10.0.0.6", os_family="Linux")]
        self.assertEqual({r["os"] for r in privesc.playbook_rows(mixed)},
                         {"windows", "linux"})
        unknown = [Host(ip="10.0.0.9")]
        self.assertEqual({r["os"] for r in privesc.playbook_rows(unknown)},
                         {"windows", "linux"})

    def test_remote_finding_from_vuln(self):
        from recce import privesc
        from recce.models import Vuln
        h = Host(ip="10.0.0.5", os_family="Windows")
        h.vulns = [Vuln(ip="10.0.0.5", port=445, protocol="tcp",
                        script_id="smb-vuln-ms17-010", title="ms17-010",
                        severity="critical")]
        findings = [r for r in privesc.plan(h) if r["category"] == "finding"]
        self.assertTrue(any("MS17-010" in r["vector"] for r in findings))

    def test_current_potato_playbook_and_service_hints(self):
        from recce import privesc
        h = Host(ip="10.0.10.50", os_family="Windows", os_name="Windows 11",
                 ports=[Port(portid=80, service="http",
                             product="Microsoft IIS httpd", version="10.0"),
                        Port(portid=1433, service="ms-sql-s",
                             product="Microsoft SQL Server")])
        # The Potato playbook is reference material (playbook sheet)...
        pb_blob = " ".join(f"{r['vector']} {r['howto']} {r['note']}"
                           for r in privesc.playbook_rows([h]))
        for tool in ("GodPotato", "PrintSpoofer", "EfsPotato", "JuicyPotatoNG",
                     "RoguePotato", "LocalPotato"):
            self.assertIn(tool, pb_blob)
        self.assertIn("CVE-2023-21746", pb_blob)              # LocalPotato CVE
        self.assertIn("SeImpersonate", pb_blob)               # precondition named
        # ...but recce flags the opportunity remotely from the IIS + MSSQL services
        # as real findings on the Priv-Esc tab.
        findings = [r for r in privesc.plan(h) if r["category"] == "finding"]
        self.assertTrue(any("IIS" in r["vector"] for r in findings))
        self.assertTrue(any("MSSQL" in r["vector"] for r in findings))


class StepCheckboxTest(unittest.TestCase):
    def _host(self, **kw):
        h = Host(ip="10.0.0.5", subnet="10.0.0.0/24",
                 ports=[Port(portid=80, service="http"), Port(portid=3306, service="mysql")])
        for k, v in kw.items():
            setattr(h, k, v)
        return h

    def test_step_auto(self):
        h = self._host()
        self.assertFalse(tr.step_auto(h, "enum"))
        h.enumerated = True
        self.assertTrue(tr.step_auto(h, "enum"))
        self.assertFalse(tr.step_auto(h, "vuln"))   # ports not scanned
        for p in h.ports:
            p.vuln_scanned = True
        self.assertTrue(tr.step_auto(h, "vuln"))
        self.assertFalse(tr.step_auto(h, "db"))     # has mysql, not db_scanned
        h.db_scanned = True
        self.assertTrue(tr.step_auto(h, "db"))

    def test_db_not_applicable_when_no_db(self):
        # An SSH-only Linux host: DB, Web and AD steps simply don't apply.
        h = Host(ip="10.0.0.6", os_family="Linux", enumerated=True,
                 ports=[Port(portid=22, service="ssh")])
        self.assertFalse(tr.step_applies(h, "db"))
        self.assertFalse(tr.step_applies(h, "web"))
        self.assertFalse(tr.step_applies(h, "ad"))
        # Universal steps still apply.
        self.assertTrue(tr.step_applies(h, "enum"))
        self.assertTrue(tr.step_applies(h, "vuln"))

    def test_step_applicability_by_surface(self):
        web = Host(ip="10.0.0.7", os_family="Linux",
                   ports=[Port(portid=443, service="https")])
        self.assertTrue(tr.step_applies(web, "web"))
        self.assertFalse(tr.step_applies(web, "ad"))   # Linux web, not a DC
        self.assertFalse(tr.step_applies(web, "db"))

        # A plain SMB file server is NOT an AD host (SMB is tracked on Services).
        smb = Host(ip="10.0.0.8", os_family="Windows",
                   ports=[Port(portid=445, service="microsoft-ds")])
        self.assertFalse(tr.step_applies(smb, "ad"))

        # A DC (LDAP/Kerberos) is an AD host.
        dc = Host(ip="10.0.0.10", os_family="Windows",
                  ports=[Port(portid=389, service="ldap"),
                         Port(portid=88, service="kerberos-sec")])
        self.assertTrue(tr.step_applies(dc, "ad"))

        # Kill-chain markers apply to anything with an open port.
        for step in ("access", "creds", "lateral"):
            self.assertTrue(tr.step_applies(web, step))
        dead = Host(ip="10.0.0.11", state="up", ports=[])
        for step in ("access", "creds", "lateral", "vuln"):
            self.assertFalse(tr.step_applies(dead, step))

        # Priv-esc only applies once the phase has run (a foothold exists).
        self.assertFalse(tr.step_applies(dc, "privesc"))
        dc.privesc_checked = True
        self.assertTrue(tr.step_applies(dc, "privesc"))

    def test_manual_steps_never_auto_check(self):
        # AD review + kill-chain markers are operator sign-offs: applicable but
        # never auto-completed by the tool, even after enumeration.
        dc = Host(ip="10.0.0.10", os_family="Windows", enumerated=True,
                  roles=["Domain Controller"],
                  ports=[Port(portid=389, service="ldap"),
                         Port(portid=88, service="kerberos-sec")])
        for step in ("ad", "access", "creds", "lateral"):
            self.assertTrue(tr.step_applies(dc, step))
            self.assertFalse(tr.step_auto(dc, step))

    def test_manual_marker_ticks_persist(self):
        # Ticking a manual kill-chain box is recorded as an override and, unlike
        # auto steps, no phase clears it.
        from recce import cli
        with tempfile.TemporaryDirectory() as d:
            store = Store(os.path.join(d, "t.sqlite"))
            store.upsert_host(Host(ip="10.0.0.5", subnet="10.0.0.0/24",
                                   enumerated=True,
                                   ports=[Port(portid=80, service="http")]))
            akey = tr.step_key("access", "10.0.0.5")
            cli._reconcile_steps(store, {akey: (True, "")})   # tester ticked it
            self.assertTrue(store.get_tracking()[akey][0])
            # Unticking matches the auto default (False) -> override cleared.
            cli._reconcile_steps(store, {akey: (False, "")})
            self.assertNotIn(akey, store.get_tracking())
            store.close()

    def test_web_step_auto_done_when_web_ports_scanned(self):
        h = Host(ip="10.0.0.9", enumerated=True,
                 ports=[Port(portid=80, service="http"), Port(portid=22, service="ssh")])
        self.assertFalse(tr.step_auto(h, "web"))
        h.ports[0].vuln_scanned = True    # the web port got probed
        self.assertTrue(tr.step_auto(h, "web"))

    def test_na_step_renders_dash_and_no_override(self):
        # A Linux SSH box: the DB/Web/AD columns show N/A, not a checkbox, and
        # reading the workbook back records no step override for them.
        h = Host(ip="10.0.0.6", os_family="Linux", enumerated=True,
                 ports=[Port(portid=22, service="ssh")])
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook([h], out)
            rows = xlsx.read_sheets(out)["Checklist"]
            hidx = header_index(rows, "IP")
            header = rows[hidx]
            # The row after the header is the collapsible subnet band; take the first host row.
            row = next(r for r in rows[hidx + 1:] if r[0] in (xlsx.CHECK_ON, xlsx.CHECK_OFF))
            for col in ("DB", "Web", "AD"):
                self.assertEqual(row[header.index(col)], tr.STEP_NA)
            back = read_workbook_tracking(out)
            self.assertNotIn(tr.step_key("db", "10.0.0.6"), back)
            self.assertNotIn(tr.step_key("web", "10.0.0.6"), back)
            self.assertNotIn(tr.step_key("ad", "10.0.0.6"), back)
            # Universal steps (enum + kill-chain, host has an open port) are tracked.
            self.assertIn(tr.step_key("enum", "10.0.0.6"), back)
            self.assertIn(tr.step_key("access", "10.0.0.6"), back)

    def test_checkbox_reflects_auto_then_override(self):
        h = self._host(enumerated=True)
        for p in h.ports:
            p.vuln_scanned = True
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            # No override -> follows auto (vuln done -> TRUE).
            build_workbook([h], out)
            back = read_workbook_tracking(out)
            self.assertTrue(back[tr.step_key("vuln", "10.0.0.5")][0])
            # Override FALSE -> checkbox shows FALSE despite auto TRUE.
            build_workbook([h], out, tracking={tr.step_key("vuln", "10.0.0.5"): (False, "")})
            back = read_workbook_tracking(out)
            self.assertFalse(back[tr.step_key("vuln", "10.0.0.5")][0])

    def test_services_vulnscan_not_read_as_step(self):
        # The Services sheet also has a "Vuln-scan" column; it must NOT pollute steps.
        h = self._host(enumerated=True)  # ports NOT vuln_scanned -> Services shows "pending"
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook([h], out)
            back = read_workbook_tracking(out)
        # vuln step comes from the Checklist (auto = pending = False), and there's
        # exactly one value - not overwritten to False by the Services rows.
        self.assertIn(tr.step_key("vuln", "10.0.0.5"), back)

    def test_reconcile_records_and_clears_override(self):
        from recce import cli
        with tempfile.TemporaryDirectory() as d:
            store = Store(os.path.join(d, "t.sqlite"))
            h = self._host(enumerated=True)
            for p in h.ports:
                p.vuln_scanned = True     # vuln auto = True
            store.upsert_host(h)
            key = tr.step_key("vuln", "10.0.0.5")
            # Shown FALSE but auto TRUE -> record override.
            cli._reconcile_steps(store, {key: (False, "")})
            self.assertIn(key, store.get_tracking())
            # Shown TRUE matches auto -> clear override.
            cli._reconcile_steps(store, {key: (True, "")})
            self.assertNotIn(key, store.get_tracking())
            store.close()


class CliSmokeTest(unittest.TestCase):
    def test_arg_parser_has_all_commands(self):
        from recce import cli
        p = cli.build_arg_parser()
        # Parse a representative invocation of each command without executing.
        for argv in (["enum", "10.0.0.1"], ["vulns", "10.0.0.0/24", "--fast"],
                     ["db", "-o", "x"], ["privesc", "--scan"], ["scan", "10.0.0.1"],
                     ["credenum", "-u", "a", "-p", "b", "-d", "corp.local"],
                     ["writeups", "--min-severity", "high", "--no-screenshots"],
                     ["writeups", "--include-potential"],
                     ["writeup", "F-007", "-o", "eng"],
                     ["services", "-o", "eng", "-a"],
                     ["exploitplan", "-o", "eng", "--lhost", "10.0.0.1", "--run"],
                     ["attackpath", "-o", "eng"],
                     ["creds", "--add", "CORP\\alice:Pw!", "--plan", "-o", "eng"],
                     ["ingest", "loot.txt", "--host", "1.2.3.4"],
                     ["import", "scan.xml", "-o", "eng"],
                     ["report"], ["status"], ["review", "--host", "1.2.3.4"],
                     ["demo"], ["doctor", "--no-self-scan"]):
            ns = p.parse_args(argv)
            self.assertTrue(callable(ns.func))

    def test_doctor_runs_without_crashing(self):
        from recce import cli
        rc = cli.cmd_doctor(SimpleNamespace(no_self_scan=True))
        self.assertIn(rc, (0, 1))  # 0 if nmap present, 1 if not - never raises

    def test_doctor_ldap_uses_capability_gate_not_just_binary(self):
        """Regression: doctor reports LDAP via ad.ldap_available() (ldapsearch OR
        the ldap3 package), not a raw which('ldapsearch') - else a box with only
        the ldap3 package is falsely told LDAP is missing, and the detail line
        and the summary disagree."""
        import io
        import contextlib
        import shutil
        from recce import cli, ad
        orig_avail, orig_which = ad.ldap_available, shutil.which

        def run():
            buf = io.StringIO()
            with contextlib.redirect_stdout(buf):
                cli.cmd_doctor(SimpleNamespace(no_self_scan=True))
            return buf.getvalue()
        try:
            # ldap3 present, ldapsearch binary absent -> capability IS available
            ad.ldap_available = lambda: True
            shutil.which = lambda n: None if n == "ldapsearch" else orig_which(n)
            out = run()
            self.assertTrue(any(l.strip().startswith("ldap") and "OK" in l
                                for l in out.splitlines()), out)
            missing = next((l for l in out.splitlines()
                            if "Optional tools missing" in l), "")
            self.assertNotIn("ldap", missing)
            # neither backend present -> reported missing (detail + summary agree)
            ad.ldap_available = lambda: False
            out = run()
            self.assertTrue(any(l.strip().startswith("ldap") and "-   (optional)" in l
                                for l in out.splitlines()), out)
        finally:
            ad.ldap_available, shutil.which = orig_avail, orig_which


class ExploitPlanTest(unittest.TestCase):
    @staticmethod
    def _read(*parts):
        with open(os.path.join(*parts)) as fh:
            return fh.read()

    def _hosts(self):
        from recce.models import Vuln, Account
        dc = Host(ip="10.0.10.5", hostnames=["dc01"], os_family="Windows",
                  roles=["Domain Controller"], smb_signing="not required",
                  accounts=[Account(ip="10.0.10.5", source="nse", kind="domain",
                                    domain="CORP")],
                  ports=[Port(portid=445, service="microsoft-ds")],
                  vulns=[Vuln(ip="10.0.10.5", port=445, protocol="tcp",
                              script_id="smb-vuln-ms17-010", title="smb-vuln-ms17-010",
                              severity="high", source="nse", ids=["CVE-2017-0143"],
                              output="VULNERABLE")])
        ftp = Host(ip="10.0.10.30", os_family="Linux",
                   ports=[Port(portid=21, service="ftp")],
                   vulns=[Vuln(ip="10.0.10.30", port=21, protocol="tcp",
                               script_id="version-db", title="vsftpd 2.3.4 backdoor",
                               severity="critical", source="version-db",
                               confidence="likely", ids=["CVE-2011-2523"]),
                          # a 'potential' guess must NOT get a plan
                          Vuln(ip="10.0.10.30", port=23, protocol="tcp",
                               script_id="version-db", title="Telnet cleartext",
                               severity="medium", source="version-db",
                               confidence="potential")])
        return [dc, ftp]

    def test_msf_mapping(self):
        from recce import exploitplan as ep
        self.assertEqual(ep._msf_for("smb-vuln-ms17-010 CVE-2017-0143")["module"],
                         "exploit/windows/smb/ms17_010_eternalblue")
        self.assertEqual(ep._msf_for("vsftpd 2.3.4 backdoor")["module"],
                         "exploit/unix/ftp/vsftpd_234_backdoor")
        self.assertIsNone(ep._msf_for("telnet cleartext credentials"))

    def test_build_plan_safe_default(self):
        from recce import exploitplan as ep
        with tempfile.TemporaryDirectory() as d:
            s = ep.build_plan(self._hosts(), d, lhost="10.9.9.9")
            self.assertEqual(sorted(s["plans"]), ["10.0.10.30", "10.0.10.5"])
            self.assertEqual(s["rc_files"], 2)          # ms17-010 + vsftpd
            pd = s["dir"]
            eb = next(f for f in os.listdir(pd) if "eternalblue" in f)
            rc = self._read(pd, eb)
            self.assertIn("set RHOSTS 10.0.10.5", rc)
            self.assertIn("set LHOST 10.9.9.9", rc)
            self.assertIn("check", rc)
            self.assertIn("# exploit -j", rc)           # launch commented (safe)
            # DC gets AS-REP + Kerberoast + relay actions with the domain filled in.
            dc_sh = self._read(pd, "10.0.10.5.sh")
            self.assertIn("impacket-GetNPUsers CORP/", dc_sh)
            self.assertIn("impacket-GetUserSPNs CORP/", dc_sh)
            self.assertIn("ntlmrelayx", dc_sh)

    def test_run_arms_launch(self):
        from recce import exploitplan as ep
        with tempfile.TemporaryDirectory() as d:
            s = ep.build_plan(self._hosts(), d, lhost="10.9.9.9", run=True)
            eb = next(f for f in os.listdir(s["dir"]) if "eternalblue" in f)
            rc = self._read(s["dir"], eb)
            self.assertRegex(rc, r"(?m)^exploit -j$")   # active, not commented

    def test_potential_findings_get_no_plan(self):
        from recce import exploitplan as ep
        from recce.models import Vuln
        h = Host(ip="10.0.0.9", os_family="Linux",
                 ports=[Port(portid=23, service="telnet")],
                 vulns=[Vuln(ip="10.0.0.9", port=23, protocol="tcp",
                             script_id="version-db", title="Telnet cleartext",
                             severity="medium", source="version-db",
                             confidence="potential")])
        with tempfile.TemporaryDirectory() as d:
            s = ep.build_plan([h], d)
            self.assertEqual(s["plans"], [])            # nothing confirmed -> no plan

    def test_actions_for_host_structured(self):
        from recce import exploitplan as ep
        dc = self._hosts()[0]                            # DC with ms17-010 + signing off
        acts = ep.actions_for_host(dc, lhost="10.9.9.9")
        kinds = {a["kind"] for a in acts}
        self.assertIn("remote-msf", kinds)
        self.assertIn("remote-tool", kinds)             # AS-REP/Kerberoast/relay
        msf = next(a for a in acts if a["kind"] == "remote-msf")
        self.assertIn("ms17_010_eternalblue", msf["cmd"])
        self.assertIn("10.9.9.9", msf["cmd"])           # LHOST filled in

    def test_exploitation_sheet_unifies_actions(self):
        from recce.report_excel import _spec_exploitation
        spec = _spec_exploitation(self._hosts())
        types = {r["data"]["Type"] for r in spec.rows}
        self.assertIn("remote (msf)", types)

    def test_services_sheet_has_enum_command(self):
        from recce.report_excel import _spec_services
        spec = _spec_services(self._hosts())
        self.assertIn("Enum command", [c[0] for c in spec.cols])
        cmds = [r["data"].get("Enum command", "") for r in spec.rows]
        self.assertTrue(any("recce-service.sh smb" in c for c in cmds))


class IngestServiceTest(unittest.TestCase):
    OUT = ("\n==== SMB  ->  10.0.0.5:445 ====\n"
           "[+] 445/tcp is open\n"
           "[!] SMB signing NOT required -> NTLM relay target\n"
           "[!] Null session lists shares -> anonymous SMB access\n"
           "[!] Test BlueKeep CVE-2019-0708 on legacy Windows\n"
           "==== SNMP  ->  10.0.0.9:161 ====\n"
           "[!] SNMP community string works: 'public' (v2c)\n")

    def test_parse_service_output(self):
        from recce import ingest
        p = ingest.parse_service_output(self.OUT)
        self.assertTrue(p["is_service"])
        self.assertEqual(len(p["findings"]), 4)
        self.assertEqual({f["ip"] for f in p["findings"]}, {"10.0.0.5", "10.0.0.9"})
        smb = [f for f in p["findings"] if f["ip"] == "10.0.0.5"]
        self.assertTrue(all(f["port"] == 445 for f in smb))

    def test_service_vulns_confidence_and_source(self):
        from recce import ingest
        vulns = ingest.service_findings_to_vulns(ingest.parse_service_output(self.OUT))
        adv = next(v for v in vulns if v.title.startswith("Test BlueKeep"))
        self.assertEqual(adv.confidence, "potential")   # advisory -> off writeups
        sign = next(v for v in vulns if "signing" in v.title)
        self.assertEqual(sign.confidence, "")           # observed -> real
        self.assertEqual(sign.source, "service-enum")
        self.assertEqual(sign.port, 445)
        self.assertEqual(sign.severity, "high")

    def test_ingest_service_output_into_store(self):
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            db = os.path.join(d, "results.sqlite")
            st = Store(db)
            st.upsert_host(Host(ip="10.0.0.5", subnet="10.0.0.0/24",
                                ports=[Port(portid=445, service="microsoft-ds")]))
            st.close()
            loot = os.path.join(d, "svc.txt")
            with open(loot, "w") as fh:
                fh.write(self.OUT)
            rc = cli.cmd_ingest(SimpleNamespace(
                output_dir=d, loot=loot, host=None, title="t"))
            self.assertEqual(rc, 0)
            st = Store(db)
            hosts = {h.ip: h for h in st.all_hosts()}
            st.close()
            self.assertIn("10.0.0.9", hosts)            # new host created from output
            titles = [v.title for v in hosts["10.0.0.5"].vulns]
            self.assertTrue(any("signing" in t for t in titles))


class CheckboxPersistenceTest(unittest.TestCase):
    def test_every_checkbox_header_round_trips(self):
        """Every column with the checkbox role must be recognised by the read-back
        (CHECKBOX_HEADERS), or the operator's ticks are silently lost on regen."""
        from recce import report_excel as rx
        from recce.models import Vuln, Credential
        hosts = [Host(ip="10.0.0.5", os_family="Windows", roles=["Domain Controller"],
                      local_findings=[{"category": "sudo",
                                       "vector": "NOPASSWD sudo: /usr/bin/find",
                                       "section": "Sudo", "source": "recce-enum"}],
                      accounts=[__import__("recce.models", fromlist=["Account"]).Account(
                          ip="10.0.0.5", source="nse", kind="domain", domain="CORP")],
                      ports=[Port(portid=445, service="microsoft-ds")],
                      vulns=[Vuln(ip="10.0.0.5", port=445, protocol="tcp",
                                  script_id="smb-vuln-ms17-010", title="ms17-010",
                                  severity="high", source="nse", ids=["CVE-2017-0143"],
                                  output="VULNERABLE")])]
        creds = [Credential(username="alice", secret="Pw!", domain="CORP")]
        pre, ad_specs, tail = rx._ordered_specs(hosts, None, creds)
        for spec in pre + list(ad_specs.values()) + tail:
            cb = [h for h, role, _w in spec.cols if role == "checkbox"]
            for header in cb:
                self.assertIn(header, rx.CHECKBOX_HEADERS,
                              f"{spec.title}: checkbox column {header!r} not in "
                              "CHECKBOX_HEADERS -> ticks won't persist")


class VersionTupleTest(unittest.TestCase):
    def test_openssh_patch_level_preserved(self):
        """Regression: greedy [a-z]* used to swallow the 'p', collapsing 9.3p1 and
        9.3p2 to the same tuple and losing the OpenSSH < 9.3p2 finding."""
        from recce.vulndb import _ver_tuple, _cmp
        self.assertEqual(_ver_tuple("8.2p1"), (8, 2, 1))      # docstring example
        self.assertEqual(_ver_tuple("9.3p1"), (9, 3, 1))
        self.assertEqual(_ver_tuple("9.3p2"), (9, 3, 2))
        self.assertEqual(_cmp("9.3p1", "9.3p2"), -1)          # p1 sorts below p2
        self.assertEqual(_ver_tuple("1.0.2k"), (1, 0, 2, 11))  # letter suffix intact
        self.assertEqual(_cmp("2.3.4", "2.3.4a"), -1)          # ...still < a-suffix

    def test_openssh_9_3p1_flags_double_free(self):
        from recce import vulndb
        h = Host(ip="10.0.0.9", os_family="Linux",
                 ports=[Port(portid=22, service="ssh", product="OpenSSH",
                             version="9.3p1")])
        vulndb.assess_host_inplace(h)
        self.assertTrue(any("double-free" in v.title for v in h.vulns))
        # 9.3p2 (patched) must NOT flag it
        h2 = Host(ip="10.0.0.10", os_family="Linux",
                  ports=[Port(portid=22, service="ssh", product="OpenSSH",
                              version="9.3p2")])
        vulndb.assess_host_inplace(h2)
        self.assertFalse(any("double-free" in v.title for v in h2.vulns))


class HtmlReportTest(unittest.TestCase):
    def _hosts(self):
        from recce.models import Vuln
        return [Host(ip="10.0.0.5", hostnames=["dc01"], os_family="Windows",
                     roles=["Domain Controller"], defenses=["EDR/AV: CSFalcon (process)"],
                     ports=[Port(portid=445, service="microsoft-ds")],
                     vulns=[Vuln(ip="10.0.0.5", port=445, protocol="tcp",
                                 script_id="smb-vuln-ms17-010",
                                 title="smb-vuln-ms17-010 <x>", severity="high",
                                 source="nse", ids=["CVE-2017-0143"],
                                 output="VULNERABLE")])]

    def test_self_contained_and_escaped(self):
        from recce import report_html
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "report.html")
            report_html.build_html(self._hosts(), p, title="Client X",
                                   generated="2026-01-01")
            with open(p, encoding="utf-8") as fh:
                html = fh.read()
        self.assertIn("<!doctype html>", html)
        # self-contained: no external resources at all.
        for bad in ("http://", "https://", "src=", "<link"):
            self.assertNotIn(bad, html)
        self.assertIn("Client X", html)
        for section in ("Executive summary", "Findings by severity", "Attack path",
                        "Hosts", "CVE-2017-0143"):
            self.assertIn(section, html)
        self.assertIn("smb-vuln-ms17-010 &lt;x&gt;", html)   # HTML-escaped title
        # Attack-path graph is embedded (Mermaid), copyable and offline.
        self.assertIn('class="mermaid', html)
        self.assertIn("flowchart LR", html)

    def test_detailed_findings_section(self):
        from recce import report_html
        from recce.models import Vuln
        h = Host(ip="10.0.0.6", os_family="Linux",
                 ports=[Port(portid=21, service="ftp")],
                 vulns=[Vuln(ip="10.0.0.6", port=21, protocol="tcp",
                             script_id="vsftpd-backdoor", title="vsFTPd backdoor",
                             severity="critical", source="nse",
                             ids=["CVE-2011-2523"],
                             remediation="Upgrade vsFTPd to a patched release.",
                             output="Backdoor shell on 6200 confirmed")])
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "r.html")
            report_html.build_html([h], p)
            with open(p, encoding="utf-8") as fh:
                html = fh.read()
        self.assertIn("Finding details", html)
        self.assertIn("class=\"fcard\"", html)
        self.assertIn("Upgrade vsFTPd", html)                 # remediation card
        self.assertIn("Backdoor shell on 6200 confirmed", html)  # evidence excerpt
        self.assertIn("10.0.0.6:21", html)                    # affected system

    def test_empty_hosts_ok(self):
        from recce import report_html
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "r.html")
            report_html.build_html([], p, title="Empty")
            self.assertTrue(os.path.exists(p))


class PrivEscVerdictTest(unittest.TestCase):
    def test_verdict_orders_and_classifies(self):
        from recce import privesc as pe
        h = Host(ip="10.0.0.5", os_family="Linux", local_findings=[
            {"category": "sudo",
             "vector": "NOPASSWD sudo: /usr/bin/find -> GTFOBins 'find'",
             "section": "Sudo", "source": "recce-enum"},
            {"category": "local",
             "vector": "recently modified config /opt/app/settings.conf",
             "section": "Files", "source": "recce-enum"}])
        rows = pe.plan(h)
        # escalation sorts first; the unmappable observation is a finding. The
        # generic checklist is NOT here anymore (it's the Playbook sheet), and a
        # swept host gets no 'run recce deploy' to-do.
        self.assertEqual(rows[0]["type"], "escalation")
        self.assertIn("GTFOBins", rows[0]["howto"])       # verdict shows the tool
        types = [r["type"] for r in rows]
        self.assertIn("finding", types)                   # the unmappable observation
        self.assertNotIn("checklist", types)
        self.assertNotIn("action", types)                 # already swept -> no to-do
        order = {"escalation": 0, "finding": 1, "action": 2}
        idx = [order[t] for t in types]
        self.assertEqual(idx, sorted(idx))

    def test_unswept_host_with_ports_gets_a_deploy_todo_not_a_checklist(self):
        from recce import privesc as pe
        rows = pe.plan(Host(ip="10.0.0.6", os_family="Windows",
                            ports=[Port(portid=445, service="microsoft-ds")]))
        self.assertEqual([r["type"] for r in rows], ["action"])
        self.assertIn("recce deploy", rows[0]["howto"])

    def test_dead_ip_produces_no_privesc_rows(self):
        # A host with no open ports and nothing observed (e.g. a network/broadcast
        # address that slipped into scope) must not fabricate privesc entries.
        from recce import privesc as pe
        self.assertEqual(pe.plan(Host(ip="10.200.37.0")), [])
        self.assertEqual(pe.all_rows([Host(ip="10.200.37.0")]), [])


class LocalEnumEnrichmentTest(unittest.TestCase):
    """The lateral-movement / shell-escape / persistence additions to the on-target
    scripts must flow through parsing, categorization, promotion and the playbook."""

    def test_new_sections_categorize(self):
        from recce import ingest
        loot = (
            "recce-enum  host=WEB01  user=svc  now\n"
            "==== Lateral movement & pivoting ====\n"
            "[!] ssh-agent socket live (/tmp/ssh-x/agent.1) -> hijack to SSH onward\n"
            "[!] Kerberoastable accounts (SPN set): svc_sql, svc_web\n"
            "==== Restricted shell & shell escape ====\n"
            "[!] Restricted shell (/bin/rbash) -> escape via an allowed interpreter\n"
            "==== Persistence footholds (writable login/boot hooks) ====\n"
            "[!] Writable login-time file: /etc/profile.d/init.sh\n")
        parsed = ingest.parse_loot(loot)
        cats = {f["text"][:12]: f["category"] for f in parsed["findings"]}
        self.assertEqual(cats["ssh-agent so"], "lateral")
        self.assertEqual(cats["Kerberoastab"], "lateral")
        self.assertEqual(cats["Restricted s"], "escape")
        self.assertEqual(cats["Writable log"], "persistence")

    def test_high_value_lateral_findings_promote(self):
        from recce import ingest
        findings = [
            {"vector": "Unconstrained-delegation hosts: SRV01 -> coerce auth + capture a TGT"},
            {"vector": "Kerberoastable accounts (SPN set): svc_sql"},
            {"vector": "Kubernetes service-account token readable (/var/run/secrets...)"},
        ]
        titles = {v.title for v in ingest.promote_to_vulns("10.0.0.9", findings)}
        self.assertTrue(any("Unconstrained delegation" in t for t in titles))
        self.assertTrue(any("Kerberoastable" in t for t in titles))
        self.assertTrue(any("Kubernetes" in t for t in titles))

    def test_playbook_maps_new_vectors(self):
        from recce import playbook as pb
        self.assertEqual(pb.for_text("Kerberoastable accounts (SPN set): svc",
                                     "windows")["id"], "win-kerberoast")
        self.assertEqual(pb.for_text("Unconstrained-delegation hosts: SRV01",
                                     "windows")["id"], "win-delegation")
        p = pb.for_text("ssh-agent socket live (/tmp/ssh-x/agent.1) -> hijack", "linux")
        self.assertEqual(p["id"], "lin-ssh-agent")
        self.assertIn("/tmp/ssh-x/agent.1", p["cmd"])          # {X} filled in
        self.assertEqual(pb.for_text("Restricted shell (/bin/rbash) -> escape",
                                     "linux")["id"], "lin-restricted-shell")

    def test_shipped_linux_script_parses(self):
        import shutil
        import subprocess
        bash = shutil.which("bash")
        if not bash:
            self.skipTest("bash not available")
        script = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                              "recce", "local", "recce-enum.sh")
        r = subprocess.run([bash, "-n", script], capture_output=True, text=True)
        self.assertEqual(r.returncode, 0, r.stderr)

    def test_gtfobins_lite_returns_exact_technique(self):
        # The script's embedded GTFOBins-lite must resolve a specific binary to a
        # precise command (this is the "dive deeper into the exact exploit" logic).
        import shutil
        import subprocess
        bash = shutil.which("bash")
        if not bash:
            self.skipTest("bash not available")
        script = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                              "recce", "local", "recce-enum.sh")
        harness = (
            f"source <(sed -n '/^gtfo_suid()/,/^}}/p;/^gtfo_sudo()/,/^}}/p' {script})\n"
            "echo SUID_FIND:$(gtfo_suid /usr/bin/find)\n"
            "echo SUID_PY:$(gtfo_suid /usr/bin/python3)\n"
            "echo SUDO_VIM:$(gtfo_sudo /usr/bin/vim)\n"
            "echo UNKNOWN:[$(gtfo_suid /usr/bin/nope)]\n")
        r = subprocess.run([bash, "-c", harness], capture_output=True, text=True)
        self.assertEqual(r.returncode, 0, r.stderr)
        self.assertIn("SUID_FIND:/usr/bin/find . -exec /bin/sh -p", r.stdout)
        self.assertIn("os.setuid(0)", r.stdout)                 # python technique
        self.assertIn("SUDO_VIM:sudo /usr/bin/vim -c ':!/bin/sh'", r.stdout)
        self.assertIn("UNKNOWN:[]", r.stdout)                   # no false technique

    def test_suid_static_analysis_and_secret_phrasings_map(self):
        from recce import ingest, playbook as pb
        # The static-analysis SUID findings promote and map to a play.
        promoted = ingest.promote_to_vulns("10.0.0.5", [
            {"vector": "SUID PATH-hijack candidate: /usr/bin/foo invokes bare "
                       "command(s) [backup] -> plant a malicious binary earlier in PATH"}])
        self.assertTrue(any("Custom SUID" in v.title for v in promoted))
        play = pb.for_text("SUID PATH-hijack candidate: /usr/bin/foo invokes bare "
                           "command(s) [backup] -> plant", "linux")
        self.assertEqual(play["id"], "lin-suid-pathhijack")
        self.assertIn("backup", play["cmd"])                    # {X} filled in
        # Encrypted vs ready-to-use private keys are both surfaced as SSH_KEY-ish
        # credential findings and categorized under creds.
        parsed = ingest.parse_loot(
            "recce-enum  host=h  user=u  now\n"
            "==== Credential & secret hunting ====\n"
            "[!] Private key (UNENCRYPTED, ready): /home/u/.ssh/id_rsa\n")
        self.assertEqual(parsed["findings"][0]["category"], "creds")

    def test_windows_exact_exploit_findings_map_and_promote(self):
        from recce import ingest, playbook as pb
        unq = ("Unquoted service path EXPLOITABLE: service 'Foo' runs as LocalSystem "
               "-> plant your payload at  C:\\Program Files\\Sub.exe  (dir 'C:\\Program "
               "Files' is writable), then: sc stop Foo & sc start Foo")
        p = pb.for_text(unq, "windows")
        self.assertEqual(p["id"], "win-unquoted")
        self.assertIn("C:\\Program Files\\Sub.exe", p["cmd"])       # exact intercept
        dll = ("Writable app dir (DLL hijack): C:\\Program Files\\App -> exe(s): app.exe. "
               "The app dir is searched FIRST, so drop a DLL...")
        self.assertEqual(pb.for_text(dll, "windows")["id"], "win-dll-hijack")
        titles = {v.title for v in ingest.promote_to_vulns("10.0.0.5", [
            {"vector": unq}, {"vector": dll},
            {"vector": "Writable service binary EXPLOITABLE: C:\\svc\\a.exe (service X)"}])}
        self.assertTrue(any("Unquoted service path" in t for t in titles))
        self.assertTrue(any("DLL hijack" in t for t in titles))
        self.assertTrue(any("Writable service binary/registry" in t for t in titles))


class WebPutProofTest(unittest.TestCase):
    """Gap-2: the dangerous-methods finding is PROVEN by a PUT round-trip."""
    @classmethod
    def setUpClass(cls):
        import http.server
        import threading
        cls.store = {}

        class H(http.server.BaseHTTPRequestHandler):
            def log_message(self, *a):
                pass

            def do_OPTIONS(self):
                self.send_response(200)
                self.send_header("Allow", "GET, PUT, DELETE, POST")
                self.end_headers()

            def do_PUT(self):
                n = int(self.headers.get("Content-Length", 0))
                cls.store[self.path] = self.rfile.read(n)
                self.send_response(201)
                self.end_headers()

            def do_GET(self):
                b = cls.store.get(self.path, b"index")
                self.send_response(200)
                self.send_header("Content-Length", str(len(b)))
                self.end_headers()
                self.wfile.write(b)

            def do_DELETE(self):
                cls.store.pop(self.path, None)
                self.send_response(204)
                self.end_headers()
        cls.httpd = http.server.ThreadingHTTPServer(("127.0.0.1", 0), H)
        cls.port = cls.httpd.server_address[1]
        cls.thread = threading.Thread(target=cls.httpd.serve_forever, daemon=True)
        cls.thread.start()

    @classmethod
    def tearDownClass(cls):
        cls.httpd.shutdown()

    def test_put_write_is_proven_and_reverted(self):
        from recce import web, proofs
        _profile, vulns = web.scan_endpoint("127.0.0.1",
                                            Port(portid=self.port, service="http",
                                                 state="open"), active=True)
        proven = [v for v in vulns if v.script_id == "web-methods"
                  and "proven" in v.title.lower()]
        self.assertTrue(proven, "PUT write should be proven")
        self.assertEqual(proven[0].confidence, "confirmed")
        self.assertIn("returned the uploaded marker", proven[0].output)
        self.assertNotIn("/recce_put_probe.txt", self.store)     # cleaned up
        # The prove engine reads it as CONFIRMED, not LIKELY.
        h = Host(ip="127.0.0.1", ports=[Port(portid=self.port, state="open")],
                 vulns=[proven[0]])
        self.assertEqual(proofs.verify_host(h)[0]["verdict"], proofs.CONFIRMED)


class WebJwtNoneProofTest(unittest.TestCase):
    """Gap-3: alg:none is PROVEN by forging an unsigned token and replaying it."""
    @classmethod
    def setUpClass(cls):
        import http.server
        import threading
        import base64 as _b64
        import json as _json
        import re as _re

        def _b64u(obj):
            return _b64.urlsafe_b64encode(
                _json.dumps(obj).encode()).rstrip(b"=").decode()
        # A session token the app issues with the insecure alg:none.
        cls.valid = f'{_b64u({"alg": "none", "typ": "JWT"})}.{_b64u({"user": "admin"})}.'

        def _decode(seg):
            return _b64.urlsafe_b64decode(seg + "=" * (-len(seg) % 4))

        class H(http.server.BaseHTTPRequestHandler):
            def log_message(self, *a):
                pass

            def do_GET(self):
                authed = False
                m = _re.search(r"session=([A-Za-z0-9_\-.]+)",
                               self.headers.get("Cookie", ""))
                if m:
                    parts = m.group(1).split(".")
                    if len(parts) >= 2:
                        try:                    # the bug: trust claims, never verify sig
                            hdr = _json.loads(_decode(parts[0]))
                            pl = _json.loads(_decode(parts[1]))
                            if str(hdr.get("alg", "")).lower() == "none" and pl.get("user"):
                                authed = True
                        except Exception:
                            pass
                body = (b"WELCOME ADMIN - secret dashboard: users, billing, settings, logs "
                        b"and much more privileged content here" if authed
                        else b"please log in")
                self.send_response(200)
                self.send_header("Set-Cookie", f"session={cls.valid}; Path=/")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
        cls.httpd = http.server.ThreadingHTTPServer(("127.0.0.1", 0), H)
        cls.port = cls.httpd.server_address[1]
        cls.thread = threading.Thread(target=cls.httpd.serve_forever, daemon=True)
        cls.thread.start()

    @classmethod
    def tearDownClass(cls):
        cls.httpd.shutdown()

    def test_alg_none_is_proven_by_replay(self):
        from recce import web, proofs
        _profile, vulns = web.scan_endpoint("127.0.0.1",
                                            Port(portid=self.port, service="http",
                                                 state="open"), active=True)
        proven = [v for v in vulns if v.script_id == "web-jwt"
                  and "proven" in v.title.lower()]
        self.assertTrue(proven, "alg:none acceptance should be proven")
        self.assertEqual(proven[0].confidence, "confirmed")
        self.assertEqual(proven[0].severity, "high")
        self.assertIn("same authenticated response", proven[0].output)
        h = Host(ip="127.0.0.1", ports=[Port(portid=self.port, state="open")],
                 vulns=[proven[0]])
        self.assertEqual(proofs.verify_host(h)[0]["verdict"], proofs.CONFIRMED)

    def test_alg_none_rejected_is_not_a_finding(self):
        # A server that ignores the forged token (rejects unsigned) must NOT be flagged
        # as exploitable - the forge-and-replay downgrades it.
        from recce import web
        import http.server
        import threading
        import base64 as _b64
        import json as _json

        def _b64u(obj):
            return _b64.urlsafe_b64encode(_json.dumps(obj).encode()).rstrip(b"=").decode()
        tok = f'{_b64u({"alg": "none"})}.{_b64u({"user": "admin"})}.'

        class H(http.server.BaseHTTPRequestHandler):
            def log_message(self, *a):
                pass

            def do_GET(self):
                # Always the same page regardless of token -> not gated / rejects it.
                body = b"please log in"
                self.send_response(200)
                self.send_header("Set-Cookie", f"session={tok}; Path=/")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
        httpd = http.server.ThreadingHTTPServer(("127.0.0.1", 0), H)
        port = httpd.server_address[1]
        threading.Thread(target=httpd.serve_forever, daemon=True).start()
        try:
            _p, vulns = web.scan_endpoint("127.0.0.1",
                                          Port(portid=port, service="http", state="open"),
                                          active=True)
        finally:
            httpd.shutdown()
        proven = [v for v in vulns if v.script_id == "web-jwt" and "proven" in v.title.lower()]
        self.assertFalse(proven, "an ungated server must not be reported as proven")


class WebModuleTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        import http.server
        import threading

        class H(http.server.BaseHTTPRequestHandler):
            def log_message(self, *a):
                pass

            def _send(self, code, body=b"", extra=None):
                self.send_response(code)
                self.send_header("Server", "Apache/2.4.49 (Unix)")
                for k, v in (extra or {}).items():
                    self.send_header(k, v)
                self.end_headers()
                if body:
                    self.wfile.write(body)

            def do_HEAD(self):
                self._send(200)

            def do_OPTIONS(self):
                self._send(200, extra={"Allow": "GET, POST, PUT, OPTIONS"})

            def do_POST(self):
                length = int(self.headers.get("Content-Length", 0) or 0)
                body = self.rfile.read(length) if length else b""
                if self.path == "/graphql" and b"__schema" in body:
                    return self._send(200, b'{"data":{"__schema":{"queryType":{"name":"Query"}}}}')
                return self._send(404, b"nope")

            def do_GET(self):
                if self.path == "/reflect":
                    return self._send(200, (self.headers.get("X-Test", "none")).encode())
                if self.path == "/metrics":
                    return self._send(200, b"# HELP go_gc_duration_seconds ...\n# TYPE x gauge\n")
                if self.path == "/crossdomain.xml":
                    return self._send(200, b'<cross-domain-policy><allow-access-from domain="*"/></cross-domain-policy>')
                if self.path == "/actuator":
                    return self._send(200, b'{"_links":{"env":{"href":"/actuator/env"}}}')
                if self.path == "/actuator/env":
                    return self._send(200, b'{"propertySources":[{"properties":{"db.password":{"value":"S3cr3tPass"}}}]}')
                if self.path == "/actuator/heapdump":
                    return self._send(200, b"JAVA PROFILE 1.0.2\x00" + b"\x00" * 32,
                                      extra={"Content-Type": "application/octet-stream"})
                if self.path == "/.git/config":
                    return self._send(200, b"[core]\n\trepositoryformatversion = 0\n")
                if self.path == "/backup.sql":
                    return self._send(200, b"-- MySQL dump\nCREATE TABLE users (id int);\n")
                if self.path == "/private":
                    if "Authorization" not in self.headers:
                        return self._send(401, b"auth", extra={"WWW-Authenticate": 'Basic realm="x"'})
                    return self._send(200, b"secret area")
                if self.path.startswith("/?rc="):
                    from urllib.parse import unquote
                    val = unquote(self.path.split("=", 1)[1])
                    # Evaluate {{7*7}} like a vulnerable template engine.
                    rendered = val.replace("{{7*7}}", "49")
                    return self._send(200, ("<html>" + rendered + "</html>").encode())
                if self.path == "/app.js":
                    return self._send(200, b"var k='AIzaSyA1234567890abcdefghijklmnopqrstuvw';")
                if self.path == "/readme.html":
                    return self._send(200, b"<h1>WordPress</h1> Version 6.4.2")
                if self.path == "/wp-content/plugins/woocommerce/readme.txt":
                    return self._send(200, b"=== WooCommerce ===\nStable tag: 8.3.1\n")
                if self.path == "/jwt":
                    return self._send(200, b"token=eyJhbGciOiJub25lIn0.eyJ1c2VyIjoiYSJ9.")
                if self.path == "/.git/HEAD":
                    return self._send(200, b"ref: refs/heads/main\n")
                if self.path == "/.env":
                    return self._send(200, b"APP_KEY=base64:x\nDB_PASSWORD=secret\n")
                if self.path == "/":
                    body = (b"<html><head><title>My Site</title>"
                            b"<script src=\"/app.js\"></script></head><body>"
                            b"<a href=\"/page2?q=1\">next</a>"
                            b"<form method=post action=/login>"
                            b"<input type=text name=user><input type=password name=pw></form>"
                            b"Directory listing for /  wp-content/themes</body></html>")
                    return self._send(200, body, extra={"Set-Cookie": "PHPSESSID=abc; path=/"})
                if self.path.startswith("/page2"):
                    return self._send(200, b"<html><body>page two</body></html>")
                return self._send(404, b"nope")

        cls.httpd = http.server.ThreadingHTTPServer(("127.0.0.1", 0), H)
        cls.port = cls.httpd.server_address[1]
        cls.thread = threading.Thread(target=cls.httpd.serve_forever, daemon=True)
        cls.thread.start()

    @classmethod
    def tearDownClass(cls):
        cls.httpd.shutdown()

    def _port(self):
        return Port(portid=self.port, service="http", state="open")

    def test_fingerprint_from_headers_and_body(self):
        from recce import web
        fp = web.fingerprint({"server": "nginx", "set-cookie": "JSESSIONID=1"},
                             "<title>Home</title> wp-content")
        self.assertIn("server=nginx", fp["tech"])
        self.assertIn("Java/Servlet", fp["tech"])
        self.assertIn("WordPress", fp["tech"])
        self.assertEqual(fp["title"], "Home")

    def test_deep_scan_finds_git_env_listing_methods_cookie(self):
        from recce import web
        profile, findings = web.scan_endpoint("127.0.0.1", self._port(), active=True)
        sids = {v.script_id for v in findings}
        self.assertIn("web-git", sids)          # exposed .git
        self.assertIn("web-dotenv", sids)       # exposed .env
        self.assertIn("web-dirlisting", sids)   # directory listing
        self.assertIn("web-methods", sids)      # PUT advertised
        self.assertIn("web-cookie", sids)       # no HttpOnly
        self.assertIn("WordPress", profile["tech"])
        # The .git finding is high severity and carries the exact URL.
        git = next(v for v in findings if v.script_id == "web-git")
        self.assertEqual(git.severity, "high")
        self.assertIn("/.git/HEAD", git.output)

    def test_high_value_exposures(self):
        from recce import web
        _, findings = web.scan_endpoint("127.0.0.1", self._port(), active=True)
        sids = {v.script_id for v in findings}
        self.assertIn("web-metrics", sids)        # Prometheus /metrics
        self.assertIn("web-crossdomain", sids)    # permissive crossdomain.xml
        self.assertIn("web-graphql", sids)        # GraphQL introspection (POST)

    def test_deep_actuator_backup_gitconfig_and_secret_extraction(self):
        from recce import web
        _, findings = web.scan_endpoint("127.0.0.1", self._port(), active=True)
        by = {v.script_id: v for v in findings}
        self.assertIn("web-actuator", by)               # actuator index
        self.assertIn("web-actuator-env", by)           # /env
        self.assertIn("web-actuator-heapdump", by)      # downloadable heapdump
        self.assertIn("web-gitconfig", by)              # .git/config
        self.assertIn("web-backup", by)                 # backup.sql
        # /env leaked secret is surfaced REDACTED (not the raw value).
        self.assertIn("db.password=", by["web-actuator-env"].output)
        self.assertNotIn("S3cr3tPass", by["web-actuator-env"].output)

    def test_default_creds_probe_opt_in(self):
        from recce import web
        # Without creds=True, the Basic-auth endpoint isn't brute-tried.
        _, f0 = web.scan_endpoint("127.0.0.1", self._port(), active=True, creds=False)
        self.assertNotIn("web-default-creds", {v.script_id for v in f0})
        # The bounded default list finds admin:admin on /private... but our server
        # only 200s WITH any Authorization header, so admin:admin (first try) works.
        found = web._basic_auth_defaults("127.0.0.1", self._port(),
                                         web.url_for("127.0.0.1", self._port()), ["/private"])
        self.assertTrue(any(v.script_id == "web-default-creds" for v in found))

    def test_product_version_fingerprint(self):
        from recce import web
        self.assertEqual(web.product_version({"x-jenkins": "2.401.1"}, ""), ("Jenkins", "2.401.1"))
        prod, ver = web.product_version({}, '<meta name="generator" content="WordPress 6.4.2">')
        self.assertEqual(prod, "WordPress")
        self.assertEqual(ver, "6.4.2")

    def test_ssti_js_secret_and_wordpress_enum(self):
        from recce import web
        _, findings = web.scan_endpoint("127.0.0.1", self._port(), active=True)
        sids = {v.script_id for v in findings}
        self.assertIn("web-ssti", sids)          # {{7*7}} -> 49 in the reflected page
        self.assertIn("web-js-secret", sids)     # AIza… key in /app.js
        self.assertIn("web-wp-version", sids)    # readme.html -> 6.4.2
        self.assertIn("web-wp-plugin", sids)     # woocommerce readme
        js = next(v for v in findings if v.script_id == "web-js-secret")
        self.assertIn("Google API key", js.title)

    def test_authenticated_crawl_discovers_pages_forms_and_params(self):
        from recce import web
        cres = web.crawl("127.0.0.1", self._port(), auth={"Cookie": "PHPSESSID=abc"})
        paths = {p["path"] for p in cres["pages"]}
        self.assertIn("/page2?q=1", paths)                     # followed the link
        self.assertIn(("/page2", "q"), cres["params"])         # captured the param
        self.assertTrue(any(f["password"] for f in cres["forms"]))   # parsed the login form

    def test_crawl_flags_cleartext_login_and_reflected_param(self):
        from recce import web
        cres = web.crawl("127.0.0.1", self._port())
        fs = web._crawl_findings("127.0.0.1", self._port(), cres)
        self.assertIn("web-cleartext-login", {v.script_id for v in fs})   # pw form over HTTP
        # discovered-param reflection: the server evaluates {{7*7}} on ?rc=
        ref = web._reflect_param("127.0.0.1", self._port(), "/", "rc", None)
        self.assertEqual(ref[0].script_id, "web-ssti")

    def test_jwt_alg_none_detected(self):
        from recce import web
        # A header.payload.sig where header = {"alg":"none"}.
        findings = web._scan_jwts("1.1.1.1", Port(portid=443, service="https"),
                                  {"set-cookie": "t=eyJhbGciOiJub25lIn0.eyJ1IjoiYSJ9."}, "")
        self.assertTrue(findings)
        self.assertEqual(findings[0].severity, "high")
        self.assertIn("alg:none", findings[0].title.lower())
        # Proof engine renders a verdict + jwt_tool step.
        from recce import proofs
        r = proofs.recipe_for(findings[0])
        self.assertEqual(r["id"], "web-jwt")

    def test_passive_mode_skips_path_probes(self):
        from recce import web
        _, findings = web.scan_endpoint("127.0.0.1", self._port(), active=False)
        self.assertNotIn("web-git", {v.script_id for v in findings})

    def test_web_endpoints_categorization_and_bridge(self):
        from recce import web
        h = Host(ip="127.0.0.1", ports=[self._port(),
                                        Port(portid=445, service="microsoft-ds", state="open")])
        h.vulns = web_git = []
        eps = web.web_endpoints([h])
        self.assertEqual(len(eps), 1)                    # only the http port
        self.assertIn("whatweb", eps[0]["commands"])
        self.assertIn("nikto", eps[0]["commands"])

    def test_auth_headers_are_sent(self):
        from recce import web
        r = web._fetch("127.0.0.1", self._port(), "/reflect", auth={"X-Test": "hello"})
        self.assertIsNotNone(r)
        self.assertEqual(r[2], "hello")            # server echoed our auth header
        r2 = web._fetch("127.0.0.1", self._port(), "/reflect")
        self.assertEqual(r2[2], "none")            # no header without auth

    def test_non_http_port_skips_active_probes(self):
        from recce import web
        # A closed/non-HTTP port: root fetch fails -> no path probes, no crash.
        dead = Port(portid=1, service="https", state="open")     # nothing listening
        profile, findings = web.scan_endpoint("127.0.0.1", dead, active=True)
        self.assertIsNone(profile["status"])
        self.assertNotIn("web-git", {v.script_id for v in findings})

    def test_web_proof_and_poc_wiring(self):
        from recce import proofs, poc
        v = Vuln(ip="1.1.1.1", port=80, protocol="tcp", script_id="web-git",
                 title="Exposed Git repository (.git) - source/secret disclosure",
                 output="GET http://1.1.1.1/.git/HEAD -> HTTP 200", source="web")
        r = proofs.recipe_for(v)
        self.assertEqual(r["id"], "web-exposure")
        self.assertEqual(r["fn"](Host(ip="1.1.1.1"), None, v)[0], proofs.CONFIRMED)
        self.assertEqual(poc.recipe_key_for(v.title), "web")


class PocRecipeTest(unittest.TestCase):
    def test_finding_text_selects_the_right_recipe(self):
        from recce import poc
        cases = {
            "SUID env-injection candidate: /usr/bin/foo reads LD_PRELOAD": "ld_preload",
            "/etc/passwd is WRITABLE -> add a UID 0 user": "linux_passwd",
            "SUID PATH-hijack candidate: /usr/bin/foo invokes bare command(s) [backup]": "linux_root_job",
            "Unquoted service path EXPLOITABLE: service 'Foo' runs as LocalSystem": "win_service_exe",
            "Writable app dir (DLL hijack): C:\\Program Files\\App": "win_dll",
            "AlwaysInstallElevated = 1 (HKLM+HKCU)": "win_msi",
        }
        for text, key in cases.items():
            self.assertEqual(poc.recipe_key_for(text), key, text)

    def test_select_for_host_covers_confirmed_findings(self):
        from recce import poc
        h = Host(ip="10.0.0.5", local_findings=[
            {"category": "suid", "vector": "SUID env-injection candidate: /x reads LD_PRELOAD"},
            {"category": "writable", "vector": "/etc/passwd is WRITABLE -> add a UID 0 user"}])
        keys = set(poc.select_for_host(h))
        self.assertEqual(keys, {"ld_preload", "linux_passwd"})

    def test_write_files_and_plan_lines(self):
        from recce import poc
        with tempfile.TemporaryDirectory() as d:
            recipes = {k: poc.RECIPES[k] for k in ("ld_preload", "win_dll")}
            written = poc.write_files(d, recipes)
            self.assertTrue(any(p.endswith("recce_poc_preload.c") for p in written))
            self.assertTrue(any(p.endswith("recce_poc_dll.c") for p in written))
            block = "\n".join(poc.plan_lines(recipes))
            self.assertIn("PoC BUILD RECIPES", block)
            self.assertIn("gcc -fPIC -shared", block)
            self.assertIn("msfvenom", block)

    def test_web_pocs_per_finding(self):
        from recce import poc
        from recce.models import Host, Vuln

        def v(sid, out):
            return Vuln(ip="10.0.0.5", port=443, protocol="tcp", script_id=sid,
                        title=sid, output=out, source="web")
        h = Host(ip="10.0.0.5", vulns=[
            v("web-git", "GET https://10.0.0.5/.git/HEAD -> HTTP 200"),
            v("web-cors", "Origin: … -> ACAO https://10.0.0.5"),
            v("web-jwt", "alg=none token"),
            v("web-ssti", "GET https://10.0.0.5/?rc=… -> 49"),
            v("web-graphql", "POST https://10.0.0.5/graphql"),
            v("web-actuator-heapdump", "GET https://10.0.0.5/actuator/heapdump"),
        ])
        pocs = {f: (c, n) for f, c, n in poc.web_pocs_for_host(h)}
        # one artifact per finding, URL filled in, right extension.
        self.assertTrue(any(f.startswith("poc_web-git_") and f.endswith(".sh") for f in pocs))
        self.assertTrue(any(f.startswith("poc_web-cors_") and f.endswith(".html") for f in pocs))
        jwt_f = next(f for f in pocs if f.startswith("poc_web-jwt_"))
        self.assertTrue(jwt_f.endswith(".py"))
        # the generated Python + shell must be valid.
        compile(pocs[jwt_f][0], jwt_f, "exec")             # JWT forge PoC parses
        cors = next(pocs[f][0] for f in pocs if "cors" in f)
        self.assertIn("credentials:'include'", cors)
        self.assertIn("https://10.0.0.5", cors)             # target URL embedded
        # Every PoC states an unambiguous PROVEN verdict + the ROE hand-off marker.
        for fname, (content, _note) in pocs.items():
            self.assertIn("ROE:", content, fname)
            self.assertIn("PROVEN", content, fname)
        # The JWT PoC actually replays the forged token (accepted-vs-denied).
        jwt_src = pocs[jwt_f][0]
        self.assertIn("forged  status", jwt_src)
        self.assertIn("urllib.request", jwt_src)
        import shutil
        import subprocess
        if shutil.which("sh"):
            for f, (content, _) in pocs.items():
                if f.endswith(".sh"):
                    r = subprocess.run(["sh", "-n", "/dev/stdin"], input=content,
                                       capture_output=True, text=True)
                    self.assertEqual(r.returncode, 0, f"{f}: {r.stderr}")

    def test_exploitplan_writes_web_pocs(self):
        from recce import exploitplan
        from recce.models import Host, Vuln
        h = Host(ip="10.0.0.5", os_family="Linux", vulns=[
            Vuln(ip="10.0.0.5", port=80, protocol="tcp", script_id="web-git",
                 title="Exposed Git repository (.git)", output="GET http://10.0.0.5/.git/HEAD",
                 source="web", confidence="confirmed")])
        with tempfile.TemporaryDirectory() as d:
            summary = exploitplan.build_plan([h], d)
            poc_dir = os.path.join(summary["dir"], "poc")
            files = os.listdir(poc_dir) if os.path.isdir(poc_dir) else []
            self.assertTrue(any(f.startswith("poc_web-git_") for f in files))

    def test_ld_preload_poc_source_actually_compiles(self):
        # The emitted .so source must be valid C that builds - proves it's real,
        # not pseudo-code. (Skipped where gcc is unavailable.)
        import shutil
        import subprocess
        gcc = shutil.which("gcc")
        if not gcc:
            self.skipTest("gcc not available")
        from recce import poc
        with tempfile.TemporaryDirectory() as d:
            poc.write_files(d, {"ld_preload": poc.RECIPES["ld_preload"]})
            src = os.path.join(d, "recce_poc_preload.c")
            so = os.path.join(d, "recce_poc.so")
            r = subprocess.run([gcc, "-fPIC", "-shared", "-nostartfiles", "-o", so, src],
                               capture_output=True, text=True)
            self.assertEqual(r.returncode, 0, r.stderr)
            self.assertTrue(os.path.exists(so))

    def test_exploitplan_writes_poc_files(self):
        from recce import exploitplan
        from recce.models import Vuln
        h = Host(ip="10.0.0.5", os_family="Linux",
                 local_findings=[{"category": "writable",
                                  "vector": "/etc/passwd is WRITABLE -> add a UID 0 user"}])
        h.vulns = [Vuln(ip="10.0.0.5", port=None, protocol="tcp", script_id="local-enum",
                        title="Writable /etc/passwd (add a UID 0 user)", source="local",
                        confidence="confirmed")]
        with tempfile.TemporaryDirectory() as d:
            summary = exploitplan.build_plan([h], d)
            self.assertGreaterEqual(summary.get("poc_files", 0), 0)
            script = os.path.join(summary["dir"], "10.0.0.5.sh")
            self.assertTrue(os.path.exists(script))
            with open(script) as fh:
                self.assertIn("PoC BUILD RECIPES", fh.read())


class ProofEngineTest(unittest.TestCase):
    def _vuln(self, **kw):
        base = dict(ip="10.0.0.5", port=None, protocol="tcp", script_id="s",
                    title="", output="", source="nse")
        base.update(kw)
        return Vuln(**base)

    def test_activemq_patched_is_false_positive(self):
        from recce import proofs
        h = Host(ip="10.0.0.5", ports=[Port(portid=61616, service="activemq",
                                            product="Apache ActiveMQ", version="5.18.3",
                                            state="open")])
        h.vulns = [self._vuln(port=61616, title="Apache ActiveMQ 5.18.3",
                              ids=["CVE-2023-46604"])]
        r = proofs.verify_host(h)[0]
        self.assertEqual(r["verdict"], proofs.FALSE_POSITIVE)

    def test_activemq_old_with_openwire_is_likely(self):
        from recce import proofs
        h = Host(ip="10.0.0.5", ports=[Port(portid=61616, service="activemq",
                                            product="Apache ActiveMQ", version="5.17.1",
                                            state="open")])
        h.vulns = [self._vuln(port=61616, title="Apache ActiveMQ 5.17.1",
                              ids=["CVE-2023-46604"])]
        r = proofs.verify_host(h)[0]
        self.assertEqual(r["verdict"], proofs.LIKELY)
        self.assertTrue(any("61616 is OPEN" in e for e in r["evidence"]))

    def _ver_host(self, portid, prod, ver, title, ids=None):
        h = Host(ip="10.0.0.5", ports=[Port(portid=portid, service="x", product=prod,
                                            version=ver, state="open")])
        h.vulns = [self._vuln(port=portid, title=title, source="version-db",
                              ids=ids or [])]
        return h

    def test_version_cve_findings_now_get_a_verdict(self):
        # Gap-1: version->CVE matches that previously had NO prove path.
        from recce import proofs
        # regreSSHion: patched build -> FALSE POSITIVE (catches the over-flag).
        h = self._ver_host(22, "OpenSSH", "9.8p1", "OpenSSH regreSSHion pre-auth RCE",
                           ["CVE-2024-6387"])
        self.assertEqual(proofs.verify_host(h)[0]["verdict"], proofs.FALSE_POSITIVE)
        # regreSSHion: affected window -> LIKELY.
        h = self._ver_host(22, "OpenSSH", "9.2p1", "OpenSSH regreSSHion", ["CVE-2024-6387"])
        r = proofs.verify_host(h)[0]
        self.assertEqual(r["verdict"], proofs.LIKELY)
        self.assertTrue(any("backport" in e.lower() or "glibc" in e.lower()
                            for e in r["evidence"]))
        # Apache smuggling: version match -> LIKELY with the backport caveat.
        h = self._ver_host(80, "Apache httpd", "2.4.52",
                           "Apache httpd < 2.4.59 mod_proxy SSRF / smuggling")
        self.assertEqual(proofs.verify_host(h)[0]["verdict"], proofs.LIKELY)
        # EOL software: the version fact IS the proof -> CONFIRMED.
        for prod, ver, title in [("MySQL", "5.6.0", "End-of-life MySQL (< 5.7) exposed"),
                                 ("MongoDB", "3.4", "Legacy MongoDB (< 3.6) exposure"),
                                 ("Microsoft IIS", "6.0", "Legacy Microsoft IIS - unsupported")]:
            h = self._ver_host(3306, prod, ver, title)
            self.assertEqual(proofs.verify_host(h)[0]["verdict"], proofs.CONFIRMED, title)

    def test_smb_signing_confirmed_vs_false_positive(self):
        from recce import proofs
        h = Host(ip="10.0.0.5", smb_signing="not required",
                 ports=[Port(portid=445, service="microsoft-ds", state="open")])
        h.vulns = [self._vuln(port=445, title="SMB signing not required",
                              script_id="smb2-security-mode")]
        self.assertEqual(proofs.verify_host(h)[0]["verdict"], proofs.CONFIRMED)
        h.smb_signing = "required"
        self.assertEqual(proofs.verify_host(h)[0]["verdict"], proofs.FALSE_POSITIVE)

    def test_ms17_010_nse_state_drives_verdict(self):
        from recce import proofs
        h = Host(ip="10.0.0.5", ports=[Port(portid=445, service="microsoft-ds", state="open")])
        h.vulns = [self._vuln(port=445, script_id="smb-vuln-ms17-010",
                              title="ms17-010", state="VULNERABLE", source="nse")]
        self.assertEqual(proofs.verify_host(h)[0]["verdict"], proofs.CONFIRMED)
        h.vulns[0].state = "NOT VULNERABLE"
        h.vulns[0].output = "NOT VULNERABLE (patched)"
        self.assertEqual(proofs.verify_host(h)[0]["verdict"], proofs.FALSE_POSITIVE)

    def test_seimpersonate_enabled_confirms_but_remote_only_inconclusive(self):
        from recce import proofs
        # On-target enum says Enabled -> CONFIRMED.
        h = Host(ip="10.0.0.5", os_family="Windows",
                 local_findings=[{"category": "token",
                                  "vector": "SeImpersonate / SeAssignPrimaryToken held (Enabled) -> Potato"}])
        h.vulns = [self._vuln(port=None, title="SeImpersonate -> Potato -> SYSTEM")]
        self.assertEqual(proofs.verify_host(h)[0]["verdict"], proofs.CONFIRMED)
        # Remote inference only (no on-target confirmation) -> INCONCLUSIVE.
        h2 = Host(ip="10.0.0.6", os_family="Windows",
                  ports=[Port(portid=1433, service="ms-sql-s", state="open")])
        h2.vulns = [self._vuln(ip="10.0.0.6", port=1433,
                               title="MSSQL service - likely holds SeImpersonate")]
        self.assertEqual(proofs.verify_host(h2)[0]["verdict"], proofs.INCONCLUSIVE)

    def test_confirmed_sorts_before_false_positive(self):
        from recce import proofs
        h = Host(ip="10.0.0.5", smb_signing="not required",
                 ports=[Port(portid=445, state="open"),
                        Port(portid=61616, product="Apache ActiveMQ", version="5.18.5",
                             state="open")])
        h.vulns = [self._vuln(port=61616, title="ActiveMQ 5.18.5", ids=["CVE-2023-46604"]),
                   self._vuln(port=445, title="SMB signing not required",
                              script_id="smb2-security-mode")]
        verdicts = [r["verdict"] for r in proofs.verify_host(h)]
        self.assertEqual(verdicts[0], proofs.CONFIRMED)
        self.assertEqual(verdicts[-1], proofs.FALSE_POSITIVE)

    def test_printnightmare_verdicts(self):
        from recce import proofs
        # On-target LPE precondition present -> LIKELY.
        h = Host(ip="10.0.0.5", os_family="Windows", local_findings=[{"category": "hardening",
                 "vector": "PrintNightmare surface: Spooler running + PointAndPrint "
                           "NoWarningNoElevationOnInstall=1 (CVE-2021-34527)"}])
        h.vulns = [self._vuln(title="PrintNightmare surface", script_id="local")]
        self.assertEqual(proofs.verify_host(h)[0]["verdict"], proofs.LIKELY)
        # Non-Windows host flagged -> FALSE POSITIVE.
        h2 = Host(ip="10.0.0.6", os_family="Linux")
        h2.vulns = [self._vuln(ip="10.0.0.6", title="printnightmare (CVE-2021-34527)")]
        self.assertEqual(proofs.verify_host(h2)[0]["verdict"], proofs.FALSE_POSITIVE)

    def test_bluekeep_os_gating(self):
        from recce import proofs
        old = Host(ip="10.0.0.5", os_name="Windows 7 Professional",
                   ports=[Port(portid=3389, service="ms-wbt-server", state="open")])
        old.vulns = [self._vuln(port=3389, title="BlueKeep", ids=["CVE-2019-0708"])]
        self.assertEqual(proofs.verify_host(old)[0]["verdict"], proofs.LIKELY)
        new = Host(ip="10.0.0.6", os_name="Windows Server 2019",
                   ports=[Port(portid=3389, service="ms-wbt-server", state="open")])
        new.vulns = [self._vuln(ip="10.0.0.6", port=3389, title="BlueKeep",
                                ids=["CVE-2019-0708"])]
        self.assertEqual(proofs.verify_host(new)[0]["verdict"], proofs.FALSE_POSITIVE)

    def test_zerologon_only_on_dcs(self):
        from recce import proofs
        dc = Host(ip="10.0.0.5", os_family="Windows",
                  ports=[Port(portid=88, service="kerberos", state="open"),
                         Port(portid=389, service="ldap", state="open")])
        dc.vulns = [self._vuln(port=None, title="ZeroLogon", ids=["CVE-2020-1472"])]
        self.assertEqual(proofs.verify_host(dc)[0]["verdict"], proofs.LIKELY)
        member = Host(ip="10.0.0.6", os_family="Windows",
                      ports=[Port(portid=445, service="microsoft-ds", state="open")])
        member.vulns = [self._vuln(ip="10.0.0.6", title="ZeroLogon", ids=["CVE-2020-1472"])]
        self.assertEqual(proofs.verify_host(member)[0]["verdict"], proofs.FALSE_POSITIVE)

    def test_heartbleed_and_kerberoast(self):
        from recce import proofs
        h = Host(ip="10.0.0.5", ports=[Port(portid=443, service="https", state="open")])
        h.vulns = [self._vuln(port=443, script_id="ssl-heartbleed", title="heartbleed",
                              state="VULNERABLE", source="nse")]
        self.assertEqual(proofs.verify_host(h)[0]["verdict"], proofs.CONFIRMED)
        k = Host(ip="10.0.0.7", os_family="Windows",
                 local_findings=[{"category": "lateral",
                                  "vector": "Kerberoastable accounts (SPN set): svc_sql"}])
        k.vulns = [self._vuln(ip="10.0.0.7", title="Kerberoastable accounts (SPN set): svc_sql")]
        self.assertEqual(proofs.verify_host(k)[0]["verdict"], proofs.CONFIRMED)

    def test_verification_sheet_builds(self):
        from recce.report_excel import _spec_verification
        h = Host(ip="10.0.0.5", smb_signing="not required",
                 ports=[Port(portid=445, state="open")])
        h.vulns = [self._vuln(port=445, title="SMB signing not required",
                              script_id="smb2-security-mode")]
        spec = _spec_verification([h])
        self.assertEqual(spec.title, "Verification")
        self.assertTrue(spec.rows)
        self.assertIn("Verdict", [c[0] for c in spec.cols])


class CredentialsTest(unittest.TestCase):
    def _hosts(self):
        return [Host(ip="10.0.10.5", subnet="10.0.10.0/24", os_family="Windows",
                     ports=[Port(portid=445, service="microsoft-ds"),
                            Port(portid=5985, service="http"),
                            Port(portid=389, service="ldap")]),
                Host(ip="10.0.20.9", subnet="10.0.20.0/24", os_family="Linux",
                     ports=[Port(portid=22, service="ssh")])]

    def test_parse_and_stack_dedupe(self):
        from recce import cli, credentials as cr
        a = cli._parse_cred_spec("CORP\\alice:Passw0rd!")
        self.assertEqual((a.domain, a.username, a.kind), ("CORP", "alice", "password"))
        b = cli._parse_cred_spec("administrator:aad3b435b51404eeaad3b435b51404ee")
        self.assertEqual(b.kind, "nthash")             # 32-hex -> hash
        c = cli._parse_cred_spec("bob")
        self.assertEqual(c.kind, "blank")
        stacked = cr.stack([], [a, b, a])              # duplicate a collapses
        self.assertEqual(len(stacked), 2)

    def test_spray_plan_files_and_commands(self):
        from recce import credentials as cr
        from recce.models import Credential
        creds = [Credential(username="alice", secret="Pw!", kind="password", domain="CORP"),
                 Credential(username="administrator",
                            secret="aad3b435b51404eeaad3b435b51404ee", kind="nthash")]
        with tempfile.TemporaryDirectory() as d:
            s = cr.build_spray(creds, self._hosts(), d)
            self.assertIn("users.txt", s["files"])
            self.assertIn("passwords.txt", s["files"])
            self.assertIn("nthashes.txt", s["files"])
            cmds = "\n".join(s["commands"])
            self.assertIn("netexec smb 10.0.10.0/24", cmds)
            self.assertIn("-H nthashes.txt", cmds)     # pass-the-hash line
            self.assertIn("netexec ssh 10.0.20.0/24", cmds)
            self.assertNotIn("netexec ssh 10.0.20.0/24 -u users.txt -H", cmds)  # no PtH over ssh

    def test_harvest_from_accounts(self):
        from recce import credentials as cr
        from recce.models import Account
        h = Host(ip="10.0.0.5", accounts=[
            Account(ip="10.0.0.5", source="secretsdump", kind="user", name="svc",
                    domain="CORP", attrs={"password": "S3cret"})])
        got = cr.harvest([h])
        self.assertEqual(len(got), 1)
        self.assertEqual((got[0].username, got[0].secret), ("svc", "S3cret"))

    def test_creds_add_list_plan_via_cli(self):
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            db = os.path.join(d, "results.sqlite")
            st = Store(db)
            st.upsert_host(Host(ip="10.0.10.5", subnet="10.0.10.0/24",
                                ports=[Port(portid=445, service="microsoft-ds")]))
            st.close()
            def ns(**kw):
                base = dict(output_dir=d, targets=[], host=[], subnet=[], add=None,
                            user=None, password=None, hash=None, domain=None,
                            plan=False, title="t")
                base.update(kw)
                return SimpleNamespace(**base)
            self.assertEqual(cli.cmd_creds(ns(add=["CORP\\alice:Pw!"])), 0)
            st = Store(db)
            self.assertEqual(len(st.all_credentials()), 1)
            st.close()
            self.assertEqual(cli.cmd_creds(ns(plan=True)), 0)
            self.assertTrue(os.path.exists(os.path.join(d, "creds", "users.txt")))


class KubernetesTest(unittest.TestCase):
    def _host(self):
        return Host(ip="10.0.0.90", os_family="Linux",
                    ports=[Port(portid=10250, state="open"),
                           Port(portid=2379, state="open"),
                           Port(portid=6443, state="open")])

    def test_findings_all_surfaces(self):
        from recce import kubernetes as k8s
        from recce.report_docx import _vuln_type
        pr = {("10.0.0.90", 10250): {"role": "kubelet", "anon_pods": True, "pod_count": 7},
              ("10.0.0.90", 2379): {"role": "etcd", "v2_readable": True,
                                    "etcd_version": "3.5.9"},
              ("10.0.0.90", 6443): {"role": "apiserver", "version": "v1.28",
                                    "anon_list": True, "anon_secrets": True,
                                    "anon_status": 200}}
        fs = k8s.findings([self._host()], pr)
        titles = " | ".join(f["title"] for f in fs)
        self.assertIn("Kubelet allows anonymous", titles)
        self.assertIn("etcd exposed", titles)
        self.assertIn("anonymous resource listing", titles)
        by = k8s.findings_to_vulns(fs)
        for v in by["10.0.0.90"]:
            vt, _ = _vuln_type(v.cwes)
            self.assertTrue(vt, v.cwes)

    def test_prove_engine_confirms_and_downgrades(self):
        from recce import kubernetes as k8s, proofs
        pr = {("10.0.0.90", 10250): {"role": "kubelet", "anon_pods": True, "pod_count": 3},
              ("10.0.0.90", 6443): {"role": "apiserver", "version": "v1.28",
                                    "anon_list": False, "anon_status": 403}}
        h = Host(ip="10.0.0.90", ports=[Port(portid=10250, state="open"),
                                        Port(portid=6443, state="open")])
        h.vulns = k8s.findings_to_vulns(k8s.findings([h], pr))["10.0.0.90"]
        verdicts = [r["verdict"] for r in proofs.verify_host(h)]
        self.assertIn(proofs.CONFIRMED, verdicts)                  # kubelet read
        self.assertIn(proofs.LIKELY, verdicts)                     # anonymous-auth 403

    def test_v3_etcd_is_flagged(self):
        # Modern etcd disables the v2 keys API; a readable v3 gateway must still fire.
        from recce import kubernetes as k8s
        pr = {("10.0.0.90", 2379): {"role": "etcd", "v2_readable": False,
                                    "v3_readable": True, "etcd_version": "3.5.9"}}
        h = Host(ip="10.0.0.90", ports=[Port(portid=2379, state="open")])
        fs = k8s.findings([h], pr)
        self.assertTrue(any("etcd exposed" in f["title"] for f in fs))
        self.assertIn("v3", " ".join(f["detail"] for f in fs))

    def test_8080_is_not_auto_selected_as_apiserver(self):
        from recce import kubernetes as k8s
        self.assertEqual(k8s.role(8080), "unknown")
        self.assertFalse(k8s.is_k8s(Port(portid=8080, state="open", service="http")))
        # but a service explicitly named kube-apiserver is still caught
        self.assertTrue(k8s.is_k8s(Port(portid=8080, state="open",
                                        service="kube-apiserver")))

    def test_probe_parsers(self):
        from recce import kubernetes as k8s
        self.assertTrue(k8s._is_podlist({"kind": "PodList", "items": [1, 2]}))
        self.assertEqual(k8s._pod_count({"items": [1, 2, 3]}), 3)
        self.assertTrue(k8s._is_list({"kind": "NamespaceList", "items": []}))
        self.assertEqual(k8s._etcd_version({"etcdserver": "3.5.9"}), "3.5.9")
        self.assertEqual(k8s.role(10250), "kubelet")
        self.assertEqual(k8s.role(2379), "etcd")

    def test_cmd_kubernetes_end_to_end(self):
        from recce import cli, xlsx, kubernetes as k8s
        from recce.store import Store
        import http.server
        import threading
        import json as _json

        class H(http.server.BaseHTTPRequestHandler):
            def log_message(self, *a):
                pass

            def do_GET(self):
                if self.path == "/pods":
                    b = _json.dumps({"kind": "PodList",
                                     "items": [{"m": 1}, {"m": 2}]}).encode()
                    self.send_response(200)
                else:
                    b = b"{}"
                    self.send_response(404)
                self.send_header("Content-Length", str(len(b)))
                self.end_headers()
                self.wfile.write(b)
        httpd = http.server.ThreadingHTTPServer(("127.0.0.1", 0), H)
        port = httpd.server_address[1]
        threading.Thread(target=httpd.serve_forever, daemon=True).start()
        orig_role, orig_is = k8s.role, k8s.is_k8s
        k8s.role = lambda p: "kubelet-ro" if p == port else orig_role(p)
        k8s.is_k8s = lambda p: (p.state == "open" and (p.portid == port or orig_is(p)))
        try:
            with tempfile.TemporaryDirectory() as d:
                out = os.path.join(d, "eng")
                os.makedirs(out)
                st = Store(os.path.join(out, "results.sqlite"))
                st.upsert_host(Host(ip="127.0.0.1",
                                    ports=[Port(portid=port, state="open",
                                                service="kubelet")]))
                st.close()
                rc = cli.main(["k8s", "-o", out])
                self.assertEqual(rc, 0)
                sheets = xlsx.read_sheets(os.path.join(out, "enumeration.xlsx"))
                self.assertIn("Kubernetes", sheets)
                vtxt = "\n".join(" ".join(map(str, r))
                                 for r in sheets["Vulnerabilities"])
                self.assertIn("Kubelet", vtxt)
                st = Store(os.path.join(out, "results.sqlite"))
                h = st.get_host("127.0.0.1")
                st.close()
                self.assertTrue([v for v in h.vulns if v.source == "kubernetes"])
        finally:
            httpd.shutdown()
            k8s.role, k8s.is_k8s = orig_role, orig_is

    def test_no_endpoints_is_graceful(self):
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "eng")
            os.makedirs(out)
            st = Store(os.path.join(out, "results.sqlite"))
            st.upsert_host(Host(ip="10.0.0.7", ports=[Port(portid=80, service="http")]))
            st.close()
            self.assertEqual(cli.main(["kubernetes", "-o", out, "--no-probe"]), 0)


class CapabilityAutoCheckTest(unittest.TestCase):
    """Running a deep-service capability auto-marks the Checklist boxes for the
    ports it assessed (no manual ticking)."""

    def test_mark_capability_scanned_flags_ports_and_db(self):
        from recce import cli, tracking as tr
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            st = Store(os.path.join(d, "s.sqlite"))
            st.upsert_host(Host(ip="10.0.0.5", subnet="10.0.0.0/24", enumerated=True,
                                ports=[Port(portid=445, state="open", service="smb"),
                                       Port(portid=1433, state="open",
                                            service="ms-sql-s")]))
            # An SMB run assessed only 445 -> that port is scanned, 1433 is not, and a
            # host with an un-scanned port is NOT yet 'vuln-scanned' overall.
            cli._mark_capability_scanned(st, [{"ip": "10.0.0.5", "port": 445}])
            h = st.get_host("10.0.0.5")
            self.assertTrue(next(p for p in h.ports if p.portid == 445).vuln_scanned)
            self.assertFalse(next(p for p in h.ports if p.portid == 1433).vuln_scanned)
            self.assertFalse(tr.step_auto(h, "vuln"))
            self.assertFalse(h.db_scanned)
            # An MSSQL run assesses 1433 AND flags the host db-scanned -> now every port
            # is covered, so Vuln-scan and DB both auto-tick.
            cli._mark_capability_scanned(st, [{"ip": "10.0.0.5", "port": 1433}], db=True)
            h = st.get_host("10.0.0.5")
            self.assertTrue(tr.step_auto(h, "vuln"))
            self.assertTrue(tr.step_auto(h, "db"))
            st.close()


class DockerTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        import http.server
        import threading
        import json as _json

        class H(http.server.BaseHTTPRequestHandler):
            def log_message(self, *a):
                pass

            def do_GET(self):
                m = {"/version": {"Version": "24.0.5", "ApiVersion": "1.43",
                                  "Os": "linux", "KernelVersion": "6.1"},
                     "/info": {"Name": "node1", "Containers": 2, "ContainersRunning": 1,
                               "Images": 5, "ServerVersion": "24.0.5"},
                     "/containers/json": [{"Image": "nginx", "Names": ["/web"],
                                           "Command": "nginx", "State": "running"}],
                     "/images/json": [{"RepoTags": ["nginx:latest", "app:1.2"]}]}
                b = _json.dumps(m.get(self.path, {})).encode()
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(b)))
                self.end_headers()
                self.wfile.write(b)
        cls.httpd = http.server.ThreadingHTTPServer(("127.0.0.1", 0), H)
        cls.port = cls.httpd.server_address[1]
        threading.Thread(target=cls.httpd.serve_forever, daemon=True).start()

    @classmethod
    def tearDownClass(cls):
        cls.httpd.shutdown()

    def test_probe_and_findings(self):
        from recce import docker
        from recce.report_docx import _vuln_type
        pr = docker.probe("127.0.0.1", self.port)
        self.assertTrue(pr and pr["exposed"])
        self.assertEqual(pr["server_version"], "24.0.5")
        # findings() needs is_docker(port) True; the test server is on a random port,
        # so exercise the finding path on a canonical 2375 host with the same probe.
        h2 = Host(ip="127.0.0.1", ports=[Port(portid=2375, state="open")])
        fs = docker.findings([h2], {("127.0.0.1", 2375): pr})
        titles = " ".join(f["title"] for f in fs)
        self.assertIn("exposed without authentication", titles)
        by = docker.findings_to_vulns(fs)
        for v in by["127.0.0.1"]:
            vt, _ = _vuln_type(v.cwes)
            self.assertTrue(vt, v.cwes)

    def test_prove_engine_confirms_exposure(self):
        from recce import docker, proofs
        pr = docker.probe("127.0.0.1", self.port)
        h = Host(ip="127.0.0.1", ports=[Port(portid=2375, state="open")])
        h.vulns = docker.findings_to_vulns(
            docker.findings([h], {("127.0.0.1", 2375): pr}))["127.0.0.1"]
        verdicts = [r["verdict"] for r in proofs.verify_host(h)]
        self.assertIn(proofs.CONFIRMED, verdicts)

    def test_probed_but_not_exposed_marks_false(self):
        # A Docker port that answers TCP but whose API read fails (TLS-locked/auth) must
        # come back exposed=False + probed=True, not an unset 'not probed'.
        from recce import docker
        h = Host(ip="127.0.0.1", ports=[Port(portid=2375, state="open")])
        an = docker.analyze([h], active=True)   # nothing is listening on 2375 here
        t = an["targets"][0]
        self.assertFalse(t.get("exposed"))
        self.assertTrue(t.get("probed"))

    def test_cmd_docker_end_to_end(self):
        from recce import cli, xlsx, docker
        from recce.store import Store
        orig = docker.is_docker
        docker.is_docker = lambda p: (p.state == "open"
                                      and (p.portid == self.port or orig(p)))
        try:
            with tempfile.TemporaryDirectory() as d:
                out = os.path.join(d, "eng")
                os.makedirs(out)
                st = Store(os.path.join(out, "results.sqlite"))
                st.upsert_host(Host(ip="127.0.0.1",
                                    ports=[Port(portid=self.port, state="open",
                                                service="docker")]))
                st.close()
                rc = cli.main(["docker", "-o", out])
                self.assertEqual(rc, 0)
                sheets = xlsx.read_sheets(os.path.join(out, "enumeration.xlsx"))
                self.assertIn("Docker", sheets)
                vtxt = "\n".join(" ".join(map(str, r))
                                 for r in sheets["Vulnerabilities"])
                self.assertIn("Docker Engine API", vtxt)
                st = Store(os.path.join(out, "results.sqlite"))
                h = st.get_host("127.0.0.1")
                st.close()
                self.assertTrue([v for v in h.vulns if v.source == "docker"])
        finally:
            docker.is_docker = orig

    def test_no_endpoints_is_graceful(self):
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "eng")
            os.makedirs(out)
            st = Store(os.path.join(out, "results.sqlite"))
            st.upsert_host(Host(ip="10.0.0.7", ports=[Port(portid=22, service="ssh")]))
            st.close()
            self.assertEqual(cli.main(["docker", "-o", out, "--no-probe"]), 0)


class FtpTest(unittest.TestCase):
    def _host(self):
        return Host(ip="10.0.0.80", subnet="10.0.0.0/24", hostnames=["FTP01"],
                    os_family="Linux", enumerated=True,
                    ports=[Port(portid=21, service="ftp", product="vsftpd",
                                version="2.3.4", state="open")])

    def test_findings_from_probe(self):
        from recce import ftp
        pr = {("10.0.0.80", 21): {"banner": "(vsFTPd 2.3.4)", "anonymous": True,
                                  "auth_tls": False}}
        fs = ftp.findings([self._host()], pr)
        titles = " ".join(f["title"] for f in fs)
        self.assertIn("backdoor", titles.lower())                  # vsftpd 2.3.4
        self.assertIn("Anonymous FTP login", titles)
        self.assertIn("cleartext", titles.lower())
        self.assertTrue(all(f.get("narrative") for f in fs))

    def test_findings_to_vulns_have_classified_cwes(self):
        from recce import ftp
        from recce.report_docx import _vuln_type
        pr = {("10.0.0.80", 21): {"banner": "(vsFTPd 2.3.4)", "anonymous": True,
                                  "auth_tls": False}}
        by_ip = ftp.findings_to_vulns(ftp.findings([self._host()], pr))
        self.assertIn("10.0.0.80", by_ip)
        for v in by_ip["10.0.0.80"]:
            vt, _ = _vuln_type(v.cwes)
            self.assertTrue(vt, v.cwes)

    def test_prove_engine_adjudicates_ftp(self):
        from recce import ftp, proofs
        pr = {("10.0.0.80", 21): {"banner": "(vsFTPd 2.3.4)", "anonymous": True,
                                  "auth_tls": False}}
        h = self._host()
        h.vulns = ftp.findings_to_vulns(ftp.findings([h], pr))["10.0.0.80"]
        verdicts = {r["vuln"]: r["verdict"] for r in proofs.verify_host(h)}
        anon = next(v for k, v in verdicts.items() if "Anonymous FTP" in k)
        back = next(v for k, v in verdicts.items() if "Backdoor" in k or "RCE FTP" in k)
        self.assertEqual(anon, proofs.CONFIRMED)                   # 230 observed
        self.assertEqual(back, proofs.LIKELY)                      # banner-based

    def test_write_proof_finding(self):
        from recce import ftp
        f = ftp.write_proof_finding("10.0.0.80", 21,
                                    {"writable": True, "evidence": "STOR ok\nDELE ok"},
                                    None)
        self.assertIsNotNone(f)
        self.assertIn("proven", f["title"].lower())
        self.assertIn("CWE-732", f["cwes"])
        self.assertIsNone(ftp.write_proof_finding("10.0.0.80", 21,
                                                  {"writable": False}, None))

    def test_multiline_220_banner_reaches_backdoor_match(self):
        # A ProFTPD version on the SECOND 220 line must still be captured + matched.
        import socket
        import threading
        from recce import ftp
        srv = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        srv.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        srv.bind(("127.0.0.1", 0))
        srv.listen(1)
        port = srv.getsockname()[1]

        def serve():
            try:
                c, _ = srv.accept()
                c.sendall(b"220-Welcome to ACME FTP\r\n220 ProFTPD 1.3.5 Server ready\r\n")
                while True:
                    data = c.recv(1024)
                    if not data:
                        break
                    cmd = data.decode("latin-1", "replace").upper()
                    if cmd.startswith("FEAT"):
                        c.sendall(b"211-Features:\r\n AUTH TLS\r\n211 End\r\n")
                    elif cmd.startswith("USER"):
                        c.sendall(b"331 password please\r\n")
                    elif cmd.startswith("PASS"):
                        c.sendall(b"530 login incorrect\r\n")
                    elif cmd.startswith("SYST"):
                        c.sendall(b"215 UNIX Type: L8\r\n")
                    elif cmd.startswith("QUIT"):
                        c.sendall(b"221 bye\r\n")
                        break
                c.close()
            except OSError:
                pass
        threading.Thread(target=serve, daemon=True).start()
        try:
            pr = ftp.probe("127.0.0.1", port, timeout=3.0)
        finally:
            srv.close()
        self.assertIsNotNone(pr)
        self.assertIn("ProFTPD 1.3.5", pr["banner"])              # 2nd line captured
        h = Host(ip="127.0.0.1", ports=[Port(portid=21, state="open", service="ftp")])
        fs = ftp.findings([h], {("127.0.0.1", 21): pr})
        self.assertTrue(any("mod_copy" in f["title"].lower() for f in fs))

    def test_cmd_ftp_end_to_end(self):
        from recce import cli, xlsx
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "eng")
            os.makedirs(out)
            st = Store(os.path.join(out, "results.sqlite"))
            st.upsert_host(self._host())
            st.close()
            rc = cli.main(["ftp", "-o", out, "--no-run", "--no-probe"])
            self.assertEqual(rc, 0)
            sheets = xlsx.read_sheets(os.path.join(out, "enumeration.xlsx"))
            self.assertIn("FTP", sheets)
            mtxt = "\n".join(" ".join(map(str, r)) for r in sheets["FTP"])
            self.assertIn("10.0.0.80:21", mtxt)
            vtxt = "\n".join(" ".join(map(str, r)) for r in sheets["Vulnerabilities"])
            self.assertIn("backdoor", vtxt.lower())                # folded into totals
            st = Store(os.path.join(out, "results.sqlite"))
            h = st.get_host("10.0.0.80")
            st.close()
            self.assertTrue([v for v in h.vulns if v.source == "ftp"])

    def test_no_endpoints_is_graceful(self):
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "eng")
            os.makedirs(out)
            st = Store(os.path.join(out, "results.sqlite"))
            st.upsert_host(Host(ip="10.0.0.7", ports=[Port(portid=22, service="ssh")]))
            st.close()
            self.assertEqual(cli.main(["ftp", "-o", out, "--no-probe", "--no-run"]), 0)


class SmbTest(unittest.TestCase):
    def _host(self):
        return Host(ip="10.0.0.60", subnet="10.0.0.0/24", hostnames=["FS01"],
                    os_family="Windows", enumerated=True,
                    ports=[Port(portid=445, service="microsoft-ds",
                                product="Windows Server 2019", state="open"),
                           Port(portid=80, service="http", state="open")])

    def test_smb2_negotiate_roundtrips(self):
        import struct
        from recce import smb
        req = smb._build_smb2_negotiate()
        self.assertEqual(req[4:8], b"\xfeSMB")
        self.assertEqual(struct.unpack(">I", req[:4])[0], len(req) - 4)
        # Synthetic SMB2 negotiate response: 3.1.1, signing NOT required.
        hdr = smb._smb2_header(0x0000, flags=0x00000001)
        body = (struct.pack("<H", 65) + struct.pack("<H", 0x01)   # signing enabled only
                + struct.pack("<H", 0x0311) + struct.pack("<H", 0) + b"\x11" * 16
                + struct.pack("<I", 7) + struct.pack("<I", 0x800000) * 3)
        resp = struct.pack(">I", len(hdr + body)) + hdr + body
        p = smb.parse_smb2_negotiate(resp)
        self.assertEqual(p["dialect_name"], "SMB 3.1.1")
        self.assertFalse(p["signing_required"])
        self.assertTrue(p["signing_enabled"])

    def test_smb1_negotiate_detection(self):
        import struct
        from recce import smb
        req = smb._build_smb1_negotiate()
        self.assertEqual(req[4:8], b"\xffSMB")
        # SMBv1 answer with a selected dialect index -> enabled.
        hdr = (b"\xffSMB" + b"\x72" + b"\x00\x00\x00\x00" + b"\x98" + b"\x01\x28"
               + b"\x00\x00" + b"\x00" * 8 + b"\x00\x00" + b"\x00\x00" + b"\x2f\x4b"
               + b"\x00\x08" + b"\xc5\x5e")
        body = struct.pack("<B", 17) + struct.pack("<H", 5) + b"\x00" * 30
        resp = struct.pack(">I", len(hdr + body)) + hdr + body
        self.assertTrue(smb.parse_smb1_negotiate(resp)["smbv1"])
        # A server answering SMB2 to the SMB1 negotiate -> SMBv1 off.
        self.assertFalse(smb.parse_smb1_negotiate(
            struct.pack(">I", 8) + b"\xfeSMB" + b"\x00" * 4)["smbv1"])

    def test_findings_from_probe(self):
        from recce import smb
        pr = {("10.0.0.60", 445): {"smbv1": True, "signing_required": False,
                                   "dialect_name": "SMB 3.1.1"}}
        fs = smb.findings([self._host()], pr)
        titles = " ".join(f["title"] for f in fs)
        self.assertIn("SMBv1", titles)
        self.assertIn("signing not required", titles.lower())
        self.assertTrue(all(f.get("narrative") for f in fs))       # narratives attached

    def test_findings_to_vulns_have_classified_cwes(self):
        from recce import smb
        from recce.report_docx import _vuln_type
        pr = {("10.0.0.60", 445): {"smbv1": True, "signing_required": False,
                                   "dialect_name": "SMB 3.1.1"}}
        by_ip = smb.findings_to_vulns(smb.findings([self._host()], pr))
        self.assertIn("10.0.0.60", by_ip)
        for v in by_ip["10.0.0.60"]:
            vt, _ = _vuln_type(v.cwes)
            self.assertTrue(vt, v.cwes)                             # every CWE classifies

    def test_prove_engine_adjudicates_smb(self):
        from recce import smb, proofs
        pr = {("10.0.0.60", 445): {"smbv1": True, "signing_required": False,
                                   "dialect_name": "SMB 3.1.1"}}
        h = self._host()
        h.vulns = smb.findings_to_vulns(smb.findings([h], pr))["10.0.0.60"]
        h.smb_signing = "not required"
        verdicts = {r["vuln"]: r["verdict"] for r in proofs.verify_host(h)}
        smbv1 = next(v for k, v in verdicts.items() if "SMBv1" in k)
        signing = next(v for k, v in verdicts.items() if "signing" in k.lower())
        self.assertEqual(smbv1, proofs.CONFIRMED)
        self.assertEqual(signing, proofs.CONFIRMED)

    def test_writable_shares_do_not_collapse(self):
        # Two writable shares on one host must survive as two distinct findings/Vulns.
        from recce import smb
        f1 = smb.write_proof_finding("10.0.0.60", 445, "data",
                                     {"writable": True, "evidence": "e"}, None)
        f2 = smb.write_proof_finding("10.0.0.60", 445, "backups",
                                     {"writable": True, "evidence": "e"}, None)
        self.assertNotEqual(f1["title"], f2["title"])
        vulns = smb.findings_to_vulns([f1, f2])["10.0.0.60"]
        self.assertEqual(len({v.key for v in vulns}), 2)          # both survive dedup

    def test_writable_share_confirmed_by_prove_engine(self):
        from recce import smb, proofs
        f = smb.write_proof_finding("10.0.0.60", 445, "data",
                                    {"writable": True, "evidence": "e"}, None)
        h = Host(ip="10.0.0.60", ports=[Port(portid=445, state="open")],
                 vulns=smb.findings_to_vulns([f])["10.0.0.60"])
        verdicts = [r["verdict"] for r in proofs.verify_host(h)]
        self.assertIn(proofs.CONFIRMED, verdicts)                 # dedicated recipe

    def test_prove_writable_judges_the_put_and_cleans_up(self):
        from recce import smb
        orig_tool, orig_run = smb.smbclient_tool, smb._run
        smb.smbclient_tool = lambda: "/usr/bin/smbclient"
        try:
            # Success: put lands, delete is silent -> writable + cleaned up.
            smb._run = lambda cmd, timeout=60: (
                "putting file /tmp/x as \\recce_smb_probe.txt (10.0 kb/s)\n"
                "  recce_smb_probe.txt\n", None)
            r = smb.prove_writable("1.2.3.4", "data", None)
            self.assertTrue(r["writable"])
            self.assertTrue(r["cleanup_ok"])
            # Write lands but the in-script delete is DENIED -> still writable, and a
            # second explicit delete is attempted (cleanup retried).
            calls = []

            def run_deldenied(cmd, timeout=60):
                calls.append(cmd)
                if len(calls) == 1:
                    return ("putting file /tmp/x as \\recce_smb_probe.txt (9 kb/s)\n"
                            "NT_STATUS_ACCESS_DENIED deleting remote file "
                            "\\recce_smb_probe.txt\n", None)
                return ("", None)                                 # explicit retry del
            smb._run = run_deldenied
            r = smb.prove_writable("1.2.3.4", "data", None)
            self.assertTrue(r["writable"])
            self.assertEqual(len(calls), 2)                       # cleanup retried
            # Put refused -> not writable (no false positive from a trailing marker).
            smb._run = lambda cmd, timeout=60: (
                "NT_STATUS_ACCESS_DENIED opening remote file "
                "\\recce_smb_probe.txt\n", None)
            self.assertFalse(smb.prove_writable("1.2.3.4", "data", None)["writable"])
        finally:
            smb.smbclient_tool, smb._run = orig_tool, orig_run

    def test_null_session_findings(self):
        from recce import smb
        session = {"ran": True, "error": None,
                   "shares": [{"name": "backups", "perms": "READ"},
                              {"name": "IPC$", "perms": "READ"}],
                   "users": [{"domain": "CORP", "name": "alice"}]}
        fs = smb.null_session_findings("10.0.0.60", 445, session)
        titles = " ".join(f["title"] for f in fs)
        self.assertIn("null / anonymous session", titles)
        self.assertIn("readable without credentials", titles)       # backups (not IPC$)

    def test_cmd_smb_end_to_end(self):
        from recce import cli, xlsx
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "eng")
            os.makedirs(out)
            st = Store(os.path.join(out, "results.sqlite"))
            h = self._host()
            h.ports[0].scripts = [Script(id="smb-vuln-ms17-010",
                                         output="VULNERABLE: MS17-010")]
            st.upsert_host(h)
            st.close()
            # --no-probe (no live socket in CI); feed a synthetic probe via meta? No -
            # instead assert the sheet renders and the runbook is creds-filled.
            rc = cli.main(["smb", "-o", out, "--no-run", "--no-probe",
                           "-u", "alice", "-p", "P@ss", "-d", "corp.local"])
            self.assertEqual(rc, 0)
            sheets = xlsx.read_sheets(os.path.join(out, "enumeration.xlsx"))
            self.assertIn("SMB", sheets)
            mtxt = "\n".join(" ".join(map(str, r)) for r in sheets["SMB"])
            self.assertIn("10.0.0.60:445", mtxt)
            self.assertIn("corp.local", mtxt)                       # runbook creds-filled

    def test_no_endpoints_is_graceful(self):
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "eng")
            os.makedirs(out)
            st = Store(os.path.join(out, "results.sqlite"))
            st.upsert_host(Host(ip="10.0.0.7", ports=[Port(portid=25, service="smtp")]))
            st.close()
            self.assertEqual(cli.main(["smb", "-o", out, "--no-probe", "--no-run"]), 0)


class MssqlTest(unittest.TestCase):
    def _host(self):
        from recce.models import Vuln
        return Host(ip="10.0.0.50", subnet="10.0.0.0/24", hostnames=["SQL01"],
                    os_family="Windows", enumerated=True,
                    ports=[Port(portid=1433, service="ms-sql-s",
                                product="Microsoft SQL Server", version="12.0.2000",
                                state="open",
                                scripts=[Script(id="ms-sql-ntlm-info",
                                                output="NetBIOS_Domain_Name: CORP")])],
                    vulns=[Vuln(ip="10.0.0.50", port=1433, protocol="tcp",
                                script_id="ms-sql-empty-password",
                                title="MSSQL sa empty password", severity="critical",
                                source="nse")])

    def test_sql_browser_parse(self):
        from recce import mssql
        insts = mssql._parse_browser(
            "ServerName;WINSQL;InstanceName;SQLEXPRESS;IsClustered;No;"
            "Version;15.0.2000.5;tcp;1433;;")
        self.assertEqual(insts[0]["instance"], "SQLEXPRESS")
        self.assertEqual(insts[0]["tcp"], "1433")
        self.assertEqual(insts[0]["version"], "15.0.2000.5")

    def test_prelogin_request_is_wellformed_and_response_parses(self):
        import struct
        from recce import mssql
        req = mssql._build_prelogin()
        self.assertEqual(req[0], 0x12)                              # PRELOGIN type
        self.assertEqual(struct.unpack(">H", req[2:4])[0], len(req))  # length field
        # Synthetic response: SQL 2019, encryption required.
        table = struct.pack(">BHH", 0x00, 11, 6) + struct.pack(">BHH", 0x01, 17, 1) + b"\xff"
        data = bytes([15, 0]) + struct.pack(">H", 2000) + b"\x00\x00" + bytes([3])
        payload = table + data
        resp = struct.pack(">BBHHBB", 0x04, 0x01, 8 + len(payload), 0, 0, 0) + payload
        p = mssql._parse_prelogin(resp)
        self.assertEqual(p["version"], "15.0.2000")
        self.assertEqual(p["encryption"], "required")
        self.assertIn("SQL Server 2019", mssql.version_name(p["version"]))

    def test_findings_from_nse_and_version(self):
        from recce import mssql
        fs = mssql.findings([self._host()])
        titles = " ".join(f["title"] for f in fs)
        self.assertIn("blank password", titles)                    # ms-sql-empty-password
        self.assertIn("End-of-life", titles)                       # 12.x = 2014
        self.assertIn("NetBIOS", titles)                           # ms-sql-ntlm-info
        blank = next(f for f in fs if "blank password" in f["title"])
        self.assertEqual(blank["severity"], "critical")
        self.assertIn("impacket-mssqlclient", blank["command"])

    def test_runbook_commands_prefilled_with_creds(self):
        from recce import mssql
        an = mssql.analyze([self._host()], creds={"user": "alice", "secret": "P@ss",
                           "domain": "corp.local", "dc_ip": "10.0.0.9"}, active=False)
        cmds = " ".join(s["cmd"] for s in an["runbooks"][0]["credentialed"])
        self.assertIn("nxc mssql 10.0.0.50 -u alice -p P@ss", cmds)
        self.assertIn("corp.local/alice:P@ss@10.0.0.50", cmds)     # mssqlclient target
        self.assertIn("OPENQUERY", cmds)                           # linked-server chain
        self.assertIn("EXECUTE AS LOGIN", cmds)                    # impersonation chain

    def test_parse_nxc_mssql(self):
        from recce import mssql
        r = mssql.parse_nxc_mssql("MSSQL 10.0.0.50 1433 SQL01 [+] CORP\\alice:P@ss (Pwn3d!)")
        self.assertTrue(r["access"] and r["admin"])
        r2 = mssql.parse_nxc_mssql("MSSQL 10.0.0.50 1433 SQL01 [-] CORP\\bob:x")
        self.assertFalse(r2["access"])

    def test_findings_to_vulns_have_classified_cwes(self):
        from recce import mssql
        from recce.report_docx import _vuln_type
        fs = mssql.findings([self._host()])
        by_ip = mssql.findings_to_vulns(fs)
        self.assertIn("10.0.0.50", by_ip)
        for v in by_ip["10.0.0.50"]:
            vt, _ = _vuln_type(v.cwes)
            self.assertTrue(vt, v.cwes)                            # every CWE classifies

    def test_cmd_mssql_end_to_end(self):
        from recce import cli, xlsx
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "eng")
            os.makedirs(out)
            st = Store(os.path.join(out, "results.sqlite"))
            st.upsert_host(self._host())
            st.close()
            rc = cli.main(["mssql", "-o", out, "--no-run", "--no-probe",
                           "-u", "alice", "-p", "P@ss", "-d", "corp.local",
                           "--lhost", "10.0.0.9"])
            self.assertEqual(rc, 0)
            sheets = xlsx.read_sheets(os.path.join(out, "enumeration.xlsx"))
            self.assertIn("MSSQL", sheets)
            mtxt = "\n".join(" ".join(map(str, r)) for r in sheets["MSSQL"])
            self.assertIn("10.0.0.50:1433", mtxt)
            self.assertIn("corp.local/alice:P@ss", mtxt)           # runbook creds-filled
            vtxt = "\n".join(" ".join(map(str, r)) for r in sheets["Vulnerabilities"])
            self.assertIn("blank password", vtxt)                  # folded into main totals
            st = Store(os.path.join(out, "results.sqlite"))
            h = st.get_host("10.0.0.50")
            st.close()
            self.assertTrue([v for v in h.vulns if v.source == "mssql"])

    def test_no_endpoints_is_graceful(self):
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "eng")
            os.makedirs(out)
            st = Store(os.path.join(out, "results.sqlite"))
            st.upsert_host(Host(ip="10.0.0.7", ports=[Port(portid=80, service="http")]))
            st.close()
            self.assertEqual(cli.main(["mssql", "-o", out, "--no-probe"]), 0)

    _LIVE = ("SQL (CORP\\alice guest@master)>\n"
             "@@B:server\nSQL01|CORP\\alice|0|1|15.0.2000.5\n@@E:server\n"
             "@@B:logins\nsa|1\nCORP\\alice|0\n@@E:logins\n"
             "@@B:databases\nmaster|0|sa\npayroll|1|sa\nappdb|0|CORP\\svc\n@@E:databases\n"
             "@@B:links\nDW01|SQL Server|dw01.corp.local\n@@E:links\n"
             "@@B:impersonate\nsa|1\n@@E:impersonate\n"
             "@@B:config\nxp_cmdshell|1\n@@E:config\n"
             "@@B:hashes\nsa|0x0200ABCD\n@@E:hashes\n")

    def test_parse_enum_extracts_sections(self):
        from recce import mssql
        e = mssql.parse_enum(self._LIVE)
        self.assertEqual(e["server"][0][0], "SQL01")
        self.assertEqual([r for r in e["logins"] if r[1] == "1"][0][0], "sa")
        self.assertEqual([r[0] for r in e["databases"] if r[1] == "1"], ["payroll"])
        self.assertEqual(e["links"][0][0], "DW01")

    def test_build_enum_script_wraps_sentinels(self):
        from recce import mssql
        script = mssql.build_enum_script()
        self.assertIn("@@B:databases", script)
        self.assertIn("@@E:impersonate", script)
        self.assertTrue(script.strip().endswith("exit"))

    def test_chains_from_enum_detects_concrete_chain(self):
        from recce import mssql
        e = mssql.parse_enum(self._LIVE)
        t = {"ip": "10.0.0.50", "port": 1433}
        fs, chain, summary = mssql.chains_from_enum(
            t, e, {"user": "alice", "secret": "P@ss", "domain": "corp.local"})
        joined = " -> ".join(chain)
        self.assertIn("impersonate sysadmin login 'sa'", joined)
        self.assertIn("TRUSTWORTHY db 'payroll'", joined)
        self.assertIn("hop linked server(s) DW01", joined)
        titles = " ".join(f["title"] for f in fs)
        self.assertIn("Impersonatable sysadmin login", titles)
        self.assertIn("TRUSTWORTHY database owned by a sysadmin", titles)
        self.assertIn("linked server(s) reachable", titles)
        self.assertIn("SQL login password hash", titles)
        # A concrete command with the real db name filled in.
        tw = next(f for f in fs if "TRUSTWORTHY database" in f["title"])
        self.assertIn("USE [payroll]", tw["command"])

    def test_chains_direct_sysadmin(self):
        from recce import mssql
        e = mssql.parse_enum(
            "@@B:server\nSQL01|sa|1|0|15.0.2000.5\n@@E:server\n"
            "@@B:logins\nsa|1\n@@E:logins\n")
        t = {"ip": "10.0.0.50", "port": 1433}
        fs, chain, summary = mssql.chains_from_enum(t, e, {"user": "sa", "secret": "x"})
        self.assertTrue(summary["is_sysadmin"])
        self.assertTrue(t["admin"])
        self.assertIn("already sysadmin", chain[0])
        self.assertTrue(any("sysadmin on this MSSQL" in f["title"] for f in fs))

    def test_nested_exec_at_quote_doubling(self):
        from recce import mssql
        self.assertEqual(mssql._nested_at(["DW01"], "SELECT 1"),
                         "EXEC ('SELECT 1') AT [DW01]")
        d2 = mssql._nested_at(["DW01", "DW02"], "SELECT x+'|'+y")
        # inner quotes double once per hop (quadruple at depth 2).
        self.assertEqual(d2, "EXEC ('EXEC (''SELECT x+''''|''''+y'') AT [DW02]') AT [DW01]")
        self.assertEqual(mssql._nested_at([], "SELECT 1"), "SELECT 1")

    def test_walk_links_bfs_with_cycle(self):
        from recce import mssql
        calls = {"n": 0}

        def fake(script):
            calls["n"] += 1
            if calls["n"] == 1:                                 # entry -> DW01 (not sa)
                return "@@L:0\nDW01SRV|CORP\\svc|0|DW02\n@@LE:0\n"
            if calls["n"] == 2:                                 # DW01 -> DW02 (sa), loops back
                return "@@L:0\nDW02SRV|sa|1|DW01\n@@LE:0\n"
            return ""
        nodes = mssql.walk_links(["DW01"], fake, max_depth=5)
        self.assertEqual(len(nodes), 2)
        self.assertEqual(nodes[0]["server"], "DW01SRV")
        self.assertTrue(nodes[1]["sysadmin"])
        self.assertEqual(nodes[1]["path"], ["DW01", "DW02"])
        self.assertEqual(calls["n"], 2)                         # cycle stopped the walk

    def test_walk_links_respects_depth_bound(self):
        from recce import mssql
        calls = {"n": 0}

        def fake(script):                                       # each hop leads to a NEW server
            calls["n"] += 1
            n = calls["n"]
            return f"@@L:0\nSRV{n}|u|0|L{n + 1}\n@@LE:0\n"
        nodes = mssql.walk_links(["L1"], fake, max_depth=3)
        self.assertEqual(max(n["depth"] for n in nodes), 3)     # walked exactly to the bound
        self.assertEqual(calls["n"], 3)                         # and stopped there

    def test_link_findings_flag_sysadmin_node_with_rce(self):
        from recce import mssql
        nodes = [{"path": ["DW01"], "depth": 1, "server": "DW01SRV",
                  "login": "CORP\\svc", "sysadmin": False, "links": ["DW02"]},
                 {"path": ["DW01", "DW02"], "depth": 2, "server": "DW02SRV",
                  "login": "sa", "sysadmin": True, "links": []}]
        t = {"ip": "10.0.0.50", "port": 1433, "live_login": "CORP\\alice"}
        fs, chain = mssql.link_findings(t, nodes, {"user": "alice"})
        crit = next(f for f in fs if "SYSADMIN on DW02SRV" in f["title"])
        self.assertEqual(crit["severity"], "critical")
        self.assertIn("xp_cmdshell", crit["command"])           # nested RCE command
        self.assertIn("AT [DW01]", crit["command"])             # walks through the chain
        self.assertIn("DW01 -> DW02", " ".join(chain))

    def test_linked_server_walk_flows_into_sheet_and_totals(self):
        from unittest import mock
        from recce import cli, mssql, xlsx
        from recce.store import Store
        enum = mssql.parse_enum(
            "@@B:server\nSQL01|CORP\\alice|0|1|15.0.2000.5\n@@E:server\n"
            "@@B:logins\nsa|1\n@@E:logins\n@@B:databases\nmaster|0|sa\n@@E:databases\n"
            "@@B:links\nDW01|SQL Server|dw01\n@@E:links\n"
            "@@B:impersonate\n@@E:impersonate\n@@B:config\nxp_cmdshell|0\n@@E:config\n"
            "@@B:hashes\n@@E:hashes\n")
        lvl = {"n": 0}

        def runner_factory(*a, **k):
            def run(script):
                lvl["n"] += 1
                if lvl["n"] == 1:
                    return "@@L:0\nDW01SRV|CORP\\svc|0|DW02\n@@LE:0\n"
                if lvl["n"] == 2:
                    return "@@L:0\nDW02SRV|sa|1|DW01\n@@LE:0\n"
                return ""
            return run
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "eng")
            os.makedirs(out)
            st = Store(os.path.join(out, "results.sqlite"))
            st.upsert_host(self._host())
            st.close()
            with mock.patch.object(mssql, "mssqlclient_tool", return_value="x"), \
                    mock.patch.object(mssql, "nxc_tool", return_value=None), \
                    mock.patch.object(mssql, "run_mssqlclient", return_value=(enum, None)), \
                    mock.patch.object(mssql, "link_runner", side_effect=runner_factory):
                rc = cli.main(["mssql", "-o", out, "--no-probe",
                               "-u", "alice", "-p", "P@ss", "-d", "corp.local"])
            self.assertEqual(rc, 0)
            sheets = xlsx.read_sheets(os.path.join(out, "enumeration.xlsx"))
            m = "\n".join(" ".join(map(str, r)) for r in sheets["MSSQL"])
            self.assertIn("Linked-server chain:", m)
            self.assertIn("Linked-server graph", m)
            self.assertIn("DW02SRV", m)
            v = "\n".join(" ".join(map(str, r)) for r in sheets["Vulnerabilities"])
            self.assertIn("chain to SYSADMIN on DW02SRV", v)     # in the main totals

    def test_verify_dbowner_confirms_and_guards_context(self):
        from recce import mssql
        # db_owner=1 and DB_NAME() matches -> confirmed.
        ok = mssql.parse_dbowner("@@DBO:0\n1|payroll\n@@DBOE:0\n", ["payroll"])
        self.assertTrue(ok["payroll"])
        # db_owner=1 but a failed USE left us in another db -> NOT confirmed.
        bad = mssql.parse_dbowner("@@DBO:0\n1|master\n@@DBOE:0\n", ["payroll"])
        self.assertFalse(bad["payroll"])
        # db_owner=0 -> not confirmed.
        no = mssql.parse_dbowner("@@DBO:0\n0|payroll\n@@DBOE:0\n", ["payroll"])
        self.assertFalse(no["payroll"])

    def test_trustworthy_chain_confirmed_vs_candidate(self):
        from recce import mssql
        enum = mssql.parse_enum(
            "@@B:server\nSQL01|CORP\\alice|0|1|15.0.2000.5\n@@E:server\n"
            "@@B:logins\nsa|1\n@@E:logins\n"
            "@@B:databases\npayroll|1|sa\n@@E:databases\n"
            "@@B:links\n@@E:links\n@@B:impersonate\n@@E:impersonate\n"
            "@@B:config\n@@E:config\n@@B:hashes\n@@E:hashes\n")
        t = {"ip": "10.0.0.50", "port": 1433}
        # Candidate (no verification): high.
        fs, _c, _s = mssql.chains_from_enum(t, enum, {"user": "alice"})
        tw = next(f for f in fs if "TRUSTWORTHY" in f["title"])
        self.assertEqual(tw["severity"], "high")
        # Verified db_owner: critical + CONFIRMED wording.
        fs2, _c2, s2 = mssql.chains_from_enum(t, enum, {"user": "alice"},
                                              dbo_map={"payroll": True})
        tw2 = next(f for f in fs2 if "CONFIRMED privesc" in f["title"])
        self.assertEqual(tw2["severity"], "critical")
        self.assertEqual(s2["dbowner_confirmed"], ["payroll"])
        # Verified NOT db_owner: no trustworthy finding at all.
        fs3, _c3, _s3 = mssql.chains_from_enum(t, enum, {"user": "alice"},
                                               dbo_map={"payroll": False})
        self.assertFalse([f for f in fs3 if "TRUSTWORTHY" in f["title"]
                          or "CONFIRMED" in f["title"]])

    def test_server_level_deep_checks(self):
        from recce import mssql
        enum = mssql.parse_enum(
            "@@B:server\nSQL01|CORP\\alice|0|0|15.0.2000.5\n@@E:server\n"
            "@@B:logins\nsa|1\n@@E:logins\n@@B:databases\nmaster|0|sa\n@@E:databases\n"
            "@@B:links\n@@E:links\n@@B:impersonate\n@@E:impersonate\n@@B:config\n@@E:config\n"
            "@@B:hashes\n@@E:hashes\n@@B:credentials\n@@E:credentials\n@@B:proxies\n@@E:proxies\n"
            "@@B:linkedlogins\n@@E:linkedlogins\n"
            "@@B:serverperms\nCONNECT SQL|GRANT\nIMPERSONATE ANY LOGIN|GRANT\n@@E:serverperms\n"
            "@@B:publicserver\nALTER ANY LOGIN|GRANT\n@@E:publicserver\n"
            "@@B:startup\nsp_backdoor|startup\n@@E:startup\n")
        fs, chain, summary = mssql.chains_from_enum(
            {"ip": "10.0.0.50", "port": 1433}, enum, {"user": "alice"})
        kinds = {f["kind"] for f in fs}
        self.assertIn("mixed_mode", kinds)                       # IsIntegratedSecurityOnly=0
        self.assertIn("server_perms", kinds)                     # IMPERSONATE ANY LOGIN
        self.assertIn("public_role", kinds)                      # ALTER ANY LOGIN to public
        self.assertIn("startup_proc", kinds)                     # sp_backdoor
        sp = next(f for f in fs if f["kind"] == "server_perms")
        self.assertEqual(sp["severity"], "high")
        self.assertIn("IMPERSONATE ANY LOGIN", sp["detail"])
        self.assertIn("abuse server permission(s) IMPERSONATE ANY LOGIN", " -> ".join(chain))
        # public ALTER ANY LOGIN is a dangerous perm -> high.
        self.assertEqual(next(f for f in fs if f["kind"] == "public_role")["severity"], "high")
        self.assertTrue(summary["mixed_mode"])
        self.assertEqual(summary["public_server"], ["ALTER ANY LOGIN"])

    def test_permission_mining_guest_and_public_grants(self):
        from recce import mssql
        dbs = ["master", "payroll", "hr"]
        script = mssql.build_permmine_script(dbs)
        self.assertIn("USE [payroll]", script)
        self.assertIn("guest", script.lower())
        out = ("@@GST:1\npayroll|guest_enabled\n@@GSTE:1\n"
               "@@PBP:1\npayroll|public|SELECT|dbo.Salaries\npayroll|guest|EXECUTE|dbo.sp_Pay\n@@PBPE:1\n"
               "@@GST:2\n@@GSTE:2\n@@PBP:2\nhr|public|SELECT|dbo.Employees\n@@PBPE:2\n")
        perms = mssql.parse_permmine(out, dbs)
        self.assertTrue(perms["payroll"]["guest"])
        self.assertFalse(perms["hr"]["guest"])
        self.assertIn(("guest", "EXECUTE", "dbo.sp_Pay"), perms["payroll"]["grants"])
        fs = mssql.permmine_findings({"ip": "10.0.0.50", "port": 1433}, perms, {"user": "a"})
        kinds = {f["kind"] for f in fs}
        self.assertIn("guest_enabled", kinds)
        obj = next(f for f in fs if f["kind"] == "object_perms")
        self.assertEqual(obj["severity"], "high")                # EXECUTE = write/execute
        self.assertIn("dbo.sp_Pay", obj["detail"])

    def test_permmine_context_guard(self):
        from recce import mssql
        # Rows tagged with the wrong DB_NAME (failed USE) are rejected.
        out = "@@GST:0\nmaster|guest_enabled\n@@GSTE:0\n@@PBP:0\n@@PBPE:0\n"
        perms = mssql.parse_permmine(out, ["payroll"])
        self.assertFalse(perms["payroll"]["guest"])

    def test_proof_screenshot_html_and_gating(self):
        from recce import mssql, cli
        from types import SimpleNamespace
        html = mssql.proof_html(["EXEC xp_cmdshell 'whoami'"], "nt service\\mssql <b>x</b>",
                                banner="impacket-mssqlclient alice@10.0.0.50")
        # A faithful terminal render: the real command at a SQL> prompt, verbatim
        # output, and NO recce branding/badge.
        self.assertIn("SQL&gt;", html)
        self.assertIn("EXEC xp_cmdshell", html)
        self.assertIn("impacket-mssqlclient alice@10.0.0.50", html)  # console banner
        self.assertIn("&lt;b&gt;", html)                         # output is HTML-escaped
        self.assertNotIn("PROOF", html)                          # no manufactured badge
        self.assertNotIn(">recce<", html)                        # unbranded
        # Multiple command lines each get a prompt.
        multi = mssql.proof_html(["CREATE TABLE ##x (...)", "DROP TABLE ##x"], "")
        self.assertEqual(multi.count("SQL&gt;"), 3)              # 2 commands + trailing prompt
        # _mssql_shot is a no-op unless --screenshots is set.
        self.assertIsNone(cli._mssql_shot(
            SimpleNamespace(screenshots=False), "10.0.0.50", "n", "b", "c", "o"))

    def test_datamine_finds_tables_and_sensitive_columns(self):
        from recce import mssql
        dbs = ["master", "payroll", "appdb"]
        script = mssql.build_datamine_script(dbs)
        self.assertIn("USE [payroll]", script)
        self.assertIn("c.name LIKE '%ssn%'", script)             # interesting-column filter
        out = ("@@TBL:1\npayroll|dbo.Employees|1240\npayroll|dbo.Salaries|1240\n@@TBLE:1\n"
               "@@COL:1\npayroll|dbo.Employees.ssn\npayroll|dbo.Employees.email\n@@COLE:1\n"
               "@@TBL:2\nappdb|dbo.Users|55\n@@TBLE:2\n"
               "@@COL:2\nappdb|dbo.Users.password_hash\n@@COLE:2\n")
        mined = mssql.parse_datamine(out, dbs)
        self.assertEqual(mined["payroll"]["tables"],
                         [("dbo.Employees", "1240"), ("dbo.Salaries", "1240")])
        self.assertIn("dbo.Employees.ssn", mined["payroll"]["interesting"])
        fs = mssql.datamine_findings({"ip": "10.0.0.50", "port": 1433}, mined,
                                     {"user": "alice"})
        f = fs[0]
        self.assertEqual(f["severity"], "high")                  # sensitive data present
        self.assertEqual(f["kind"], "data_at_rest")
        self.assertIn("ssn", f["detail"])
        self.assertIn("Users", f["detail"])                      # interesting table name
        self.assertGreater(len(f["narrative"]), 120)

    def test_datamine_context_guard_rejects_wrong_db(self):
        from recce import mssql
        # A failed USE leaves rows tagged with the wrong DB_NAME -> not attributed.
        out = "@@TBL:0\nmaster|dbo.x|1\n@@TBLE:0\n@@COL:0\n@@COLE:0\n"
        mined = mssql.parse_datamine(out, ["payroll"])
        self.assertEqual(mined["payroll"]["tables"], [])         # 'master' != 'payroll'

    def test_write_proof_is_reversible_and_evidenced(self):
        from recce import mssql
        s = mssql.build_write_proof_script("ab12cd")
        # Proves create/insert/update AND reverts everything.
        self.assertIn("CREATE TABLE ##recce_ab12cd", s)
        self.assertIn("UPDATE ##recce_ab12cd SET note='MODIFIED_ab12cd'", s)
        self.assertIn("DROP TABLE ##recce_ab12cd", s)            # reverted
        self.assertIn("ALTER SERVER ROLE dbcreator ADD MEMBER recce_ab12cd", s)
        self.assertIn("DROP LOGIN recce_ab12cd", s)              # reverted
        ev = mssql.parse_write_proof(
            "@@W:begin\nINSERT|before\nUPDATE|MODIFIED_ab12cd\nPERM|1\n@@W:end\n")
        self.assertEqual(ev["update"], "MODIFIED_ab12cd")
        self.assertEqual(ev["perm"], "1")
        f = mssql.write_proof_finding({"ip": "10.0.0.50", "port": 1433}, ev, {"user": "a"})
        self.assertEqual(f["severity"], "critical")
        self.assertIn("reverted", f["detail"])
        self.assertIn("role", f["detail"])

    def test_write_proof_requires_actual_modification(self):
        from recce import mssql
        from unittest import mock
        # If UPDATE didn't round-trip, prove_write reports failure (no false claim).
        with mock.patch.object(mssql, "_mssqlclient_cmd", return_value=["x"]), \
                mock.patch.object(mssql, "_run_stdin", return_value=("@@W:begin\n@@W:end\n", None)):
            ev, err = mssql.prove_write("10.0.0.50", {"user": "a", "secret": "b"}, "tok")
        self.assertIsNone(ev)
        self.assertIn("not proven", err)

    def test_data_and_prove_write_flow_into_totals(self):
        from unittest import mock
        from recce import cli, mssql, xlsx
        from recce.store import Store
        enum = mssql.parse_enum(
            "@@B:server\nSQL01|sa|1|0|15.0.2000.5\n@@E:server\n@@B:logins\nsa|1\n@@E:logins\n"
            "@@B:databases\nmaster|0|sa\npayroll|0|sa\n@@E:databases\n@@B:links\n@@E:links\n"
            "@@B:impersonate\n@@E:impersonate\n@@B:config\n@@E:config\n@@B:hashes\n@@E:hashes\n"
            "@@B:credentials\n@@E:credentials\n@@B:proxies\n@@E:proxies\n"
            "@@B:linkedlogins\n@@E:linkedlogins\n")

        def runner_factory(*a, **k):
            def run(script):
                if "@@TBL:" in script:                          # dbs order: [master, payroll]
                    return ("@@TBL:0\nmaster|dbo.spt_values|1\n@@TBLE:0\n@@COL:0\n@@COLE:0\n"
                            "@@TBL:1\npayroll|dbo.Employees|1240\n@@TBLE:1\n"
                            "@@COL:1\npayroll|dbo.Employees.ssn\n@@COLE:1\n")
                return ""
            return run
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "eng")
            os.makedirs(out)
            st = Store(os.path.join(out, "results.sqlite"))
            st.upsert_host(self._host())
            st.close()
            with mock.patch.object(mssql, "mssqlclient_tool", return_value="x"), \
                    mock.patch.object(mssql, "nxc_tool", return_value=None), \
                    mock.patch.object(mssql, "run_mssqlclient", return_value=(enum, None)), \
                    mock.patch.object(mssql, "link_runner", side_effect=runner_factory), \
                    mock.patch.object(mssql, "prove_write",
                                      return_value=({"insert": "before", "update": "MODIFIED_x",
                                                     "perm": "1"}, None)):
                rc = cli.main(["mssql", "-o", out, "--no-probe", "--no-links",
                               "--data", "--prove-write",
                               "-u", "alice", "-p", "P@ss", "-d", "corp.local"])
            self.assertEqual(rc, 0)
            sheets = xlsx.read_sheets(os.path.join(out, "enumeration.xlsx"))
            v = "\n".join(" ".join(map(str, r)) for r in sheets["Vulnerabilities"])
            self.assertIn("Sensitive data accessible", v)
            self.assertIn("Proved write + permission-modify", v)
            m = "\n".join(" ".join(map(str, r)) for r in sheets["MSSQL"])
            self.assertIn("SENSITIVE COLUMNS", m)

    def test_findings_carry_detailed_narratives(self):
        from recce import mssql
        # A rich enum that exercises many finding kinds.
        enum = mssql.parse_enum(
            "@@B:server\nSQL01|CORP\\alice|0|1|12.0.2000.5\n@@E:server\n"
            "@@B:logins\nsa|1\n@@E:logins\n@@B:databases\npayroll|1|sa\n@@E:databases\n"
            "@@B:links\nDW01|SQL Server|dw01\n@@E:links\n@@B:impersonate\nsa|1\n@@E:impersonate\n"
            "@@B:config\nxp_cmdshell|1\n@@E:config\n@@B:hashes\nsa|0x0200AB\n@@E:hashes\n"
            "@@B:credentials\nAppCred|CORP\\svc\n@@E:credentials\n@@B:proxies\n@@E:proxies\n"
            "@@B:linkedlogins\nDW01|sa|0\n@@E:linkedlogins\n")
        fs, _c, _s = mssql.chains_from_enum(
            t := {"ip": "10.0.0.50", "port": 1433}, enum,
            {"user": "alice", "secret": "P@ss", "domain": "corp.local"})
        _ = t
        # Every finding must carry a substantial narrative and a kind.
        for f in fs:
            self.assertTrue(f.get("kind"), f["title"])
            self.assertGreater(len(f.get("narrative", "")), 120, f["title"])
        # The xp_cmdshell narrative explains its real capability in detail.
        xp = next(f for f in fs if f["kind"] == "xp_cmdshell")
        for phrase in ("service account", "SeImpersonate", "SYSTEM", "LSASS"):
            self.assertIn(phrase, xp["narrative"])

    def test_narrative_folds_into_vuln_evidence(self):
        from recce import mssql
        fs = mssql.findings([self._host()])
        by_ip = mssql.findings_to_vulns(fs)
        blob = "\n".join(v.output for v in by_ip["10.0.0.50"])
        self.assertIn("What this enables", blob)                # narrative in evidence

    def test_testing_methodology_narrative(self):
        from recce import mssql
        phases = [p for p, _t in mssql.TESTING_NARRATIVE]
        self.assertTrue(any("Discovery" in p for p in phases))
        self.assertTrue(any("Escalation" in p for p in phases))
        self.assertEqual(len(mssql.TESTING_NARRATIVE), 6)
        # Each phase has a real explanation.
        for _p, text in mssql.TESTING_NARRATIVE:
            self.assertGreater(len(text), 100)

    def test_credential_and_linked_login_secret_extraction(self):
        from recce import mssql
        from recce.report_docx import _vuln_type
        enum = mssql.parse_enum(
            "@@B:server\nSQL01|CORP\\alice|0|1|15.0.2000.5\n@@E:server\n"
            "@@B:logins\nsa|1\n@@E:logins\n@@B:databases\nmaster|0|sa\n@@E:databases\n"
            "@@B:links\nDW01|SQL Server|dw01\n@@E:links\n@@B:impersonate\n@@E:impersonate\n"
            "@@B:config\n@@E:config\n@@B:hashes\n@@E:hashes\n"
            "@@B:credentials\nAppCred|CORP\\svc_backup\n@@E:credentials\n"
            "@@B:proxies\nDeployProxy|CORP\\svc_deploy\n@@E:proxies\n"
            "@@B:linkedlogins\nDW01|sa|0\nRPT01|reader|1\n@@E:linkedlogins\n")
        t = {"ip": "10.0.0.50", "port": 1433}
        fs, chain, summary = mssql.chains_from_enum(
            t, enum, {"user": "alice", "secret": "P@ss", "domain": "corp.local"})
        cred = next(f for f in fs if "stored SQL credential" in f["title"])
        self.assertIn("CORP\\svc_backup", cred["detail"])              # the stored account
        self.assertIn("DeployProxy", cred["detail"])                   # agent proxy shown
        self.assertIn("Get-SQLCredential", cred["command"])            # extraction command
        # Fixed linked login mapping to sa -> critical + decrypt command.
        link = next(f for f in fs if "stored fixed login" in f["title"])
        self.assertEqual(link["severity"], "critical")                # maps to sa
        self.assertIn("Get-SQLServerLinkedServerLogin", link["command"])
        self.assertIn("DW01->sa", " ".join(chain))
        # Self-mapping (uses_self_credential=1) is NOT flagged as a stored secret.
        self.assertEqual(summary["linkedlogins"], ["DW01->sa [fixed]", "RPT01->reader"])
        # CWEs classify (keeps the coverage test green + gives writeups a type).
        for f in (cred, link):
            vt, _ = _vuln_type(f["cwes"])
            self.assertTrue(vt, f["cwes"])

    def test_exec_script_builders_per_method(self):
        from recce import mssql
        xp = mssql.build_exec_script("whoami", "xp")
        self.assertIn("EXEC xp_cmdshell 'whoami'", xp)
        self.assertIn("sp_configure 'xp_cmdshell',1", xp)
        ole = mssql.build_exec_script("whoami", "ole")
        self.assertIn("sp_OACreate 'WScript.Shell'", ole)
        self.assertIn("OPENROWSET(BULK", ole)              # reads output back
        agent = mssql.build_exec_script("whoami", "agent")
        self.assertIn("sp_add_job", agent)
        self.assertIn("@subsystem='CmdExec'", agent)
        self.assertIn("sp_delete_job", agent)              # cleans up after itself
        self.assertIsNone(mssql.build_exec_script("x", "clr"))
        # A single quote in the command is doubled for the T-SQL literal.
        self.assertIn("echo ''hi''", mssql.build_exec_script("echo 'hi'", "ole"))

    def test_parse_exec_strips_chrome(self):
        from recce import mssql
        out = mssql.parse_exec("SQL>\n@@X:out\n--------\noutput\ncorp\\alice\nNULL\n@@XE:out\n")
        self.assertEqual(out, "corp\\alice")

    def test_exec_command_clr_is_a_handoff_not_executed(self):
        from recce import mssql
        o, e, ref = mssql.exec_command("10.0.0.50",
                                       {"user": "alice", "secret": "P@ss", "domain": "corp.local"},
                                       "whoami", method="clr")
        self.assertIsNone(o)
        self.assertIsNone(e)
        self.assertIn("mssqlpwner", ref)                   # delegates, never loads a DLL
        self.assertIn("custom-asm", ref)

    def test_exec_rce_flows_into_totals(self):
        from unittest import mock
        from recce import cli, mssql, xlsx
        from recce.store import Store
        enum = mssql.parse_enum(
            "@@B:server\nSQL01|sa|1|0|15.0.2000.5\n@@E:server\n@@B:logins\nsa|1\n@@E:logins\n"
            "@@B:databases\nmaster|0|sa\n@@E:databases\n@@B:links\n@@E:links\n"
            "@@B:impersonate\n@@E:impersonate\n@@B:config\nxp_cmdshell|1\n@@E:config\n"
            "@@B:hashes\n@@E:hashes\n")
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "eng")
            os.makedirs(out)
            st = Store(os.path.join(out, "results.sqlite"))
            st.upsert_host(self._host())
            st.close()
            with mock.patch.object(mssql, "mssqlclient_tool", return_value="x"), \
                    mock.patch.object(mssql, "nxc_tool", return_value=None), \
                    mock.patch.object(mssql, "run_mssqlclient", return_value=(enum, None)), \
                    mock.patch.object(mssql, "link_runner",
                                      side_effect=lambda *a, **k: (lambda s: "")), \
                    mock.patch.object(mssql, "exec_command",
                                      return_value=("nt service\\mssqlserver", None, None)):
                rc = cli.main(["mssql", "-o", out, "--no-probe", "--no-links",
                               "-u", "alice", "-p", "P@ss", "-d", "corp.local",
                               "--exec", "whoami", "--method", "agent"])
            self.assertEqual(rc, 0)
            sheets = xlsx.read_sheets(os.path.join(out, "enumeration.xlsx"))
            v = "\n".join(" ".join(map(str, r)) for r in sheets["Vulnerabilities"])
            self.assertIn("Confirmed OS command execution via agent", v)
            self.assertIn("nt service\\mssqlserver", v)    # captured output

    def test_relay_targets_and_finding(self):
        from recce import mssql
        hosts = [
            Host(ip="10.0.0.50", ports=[Port(portid=1433, service="ms-sql-s")]),
            Host(ip="10.0.0.9", roles=["Domain Controller"],
                 ports=[Port(portid=389, service="ldap")]),
            Host(ip="10.0.0.20", smb_signing="not required",
                 ports=[Port(portid=445, service="microsoft-ds")]),
            Host(ip="10.0.0.60", ports=[Port(portid=1433, service="ms-sql-s")]),
        ]
        rt = mssql.relay_targets(hosts, "10.0.0.50")
        kinds = {r["kind"] for r in rt}
        self.assertEqual(kinds, {"ldap", "mssql", "smb"})
        self.assertTrue(any(r["target"] == "10.0.0.9" for r in rt))       # DC ldap
        self.assertTrue(any(r["target"] == "10.0.0.60:1433" for r in rt))  # other mssql
        self.assertFalse(any("10.0.0.50" in r["target"] for r in rt))      # not itself
        f = mssql.relay_finding({"ip": "10.0.0.50", "port": 1433}, rt, "10.10.14.5",
                                {"user": "alice"})
        self.assertIn("ntlmrelayx", f["command"])
        self.assertIn("xp_dirtree", f["command"])
        self.assertIn("10.10.14.5", f["command"])                          # lhost filled

    def test_live_enum_flows_into_sheet_and_totals(self):
        from unittest import mock
        from recce import cli, mssql, xlsx
        from recce.store import Store
        enum = mssql.parse_enum(self._LIVE)
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "eng")
            os.makedirs(out)
            st = Store(os.path.join(out, "results.sqlite"))
            st.upsert_host(self._host())
            st.close()
            with mock.patch.object(mssql, "mssqlclient_tool", return_value="impacket-mssqlclient"), \
                    mock.patch.object(mssql, "nxc_tool", return_value=None), \
                    mock.patch.object(mssql, "run_mssqlclient",
                                      return_value=(enum, None)):
                rc = cli.main(["mssql", "-o", out, "--no-probe", "--no-links",
                               "-u", "alice", "-p", "P@ss", "-d", "corp.local"])
            self.assertEqual(rc, 0)
            sheets = xlsx.read_sheets(os.path.join(out, "enumeration.xlsx"))
            m = "\n".join(" ".join(map(str, r)) for r in sheets["MSSQL"])
            self.assertIn("Live chain:", m)
            self.assertIn("payroll", m)                        # TRUSTWORTHY db named
            self.assertIn("DW01", m)                           # linked server named
            self.assertIn("Live enumeration (impacket", m)
            v = "\n".join(" ".join(map(str, r)) for r in sheets["Vulnerabilities"])
            self.assertIn("Impersonatable sysadmin", v)        # chain finding in totals
            self.assertIn("TRUSTWORTHY database", v)


class BloodHoundTest(unittest.TestCase):
    BASE = "S-1-5-21-1-2-3"

    def _collection(self, d):
        """Write a synthetic SharpHound collection into dir `d`. Encodes:
        BOB(user, Domain Users) --GenericAll--> HELPDESK --MemberOf--> Domain Admins,
        BOB has DCSync on the domain, is kerberoastable, and has a pwd in its
        description; ALICE is AS-REP roastable; SVC has unconstrained delegation."""
        B = self.BASE
        users = {"meta": {"type": "users", "count": 3}, "data": [
            {"ObjectIdentifier": f"{B}-1001",
             "Properties": {"name": "BOB@CORP.LOCAL", "domain": "CORP.LOCAL",
                            "enabled": True, "hasspn": True,
                            "serviceprincipalnames": ["MSSQL/db.corp.local"],
                            "description": "svc account pwd=Summer2024!"},
             "Aces": []},
            {"ObjectIdentifier": f"{B}-1002",
             "Properties": {"name": "ALICE@CORP.LOCAL", "enabled": True,
                            "dontreqpreauth": True}, "Aces": []},
            {"ObjectIdentifier": f"{B}-1003",
             "Properties": {"name": "SVC@CORP.LOCAL", "enabled": True,
                            "unconstraineddelegation": True}, "Aces": []},
        ]}
        groups = {"meta": {"type": "groups", "count": 3}, "data": [
            {"ObjectIdentifier": f"{B}-512",
             "Properties": {"name": "DOMAIN ADMINS@CORP.LOCAL", "highvalue": True},
             "Members": [{"ObjectIdentifier": f"{B}-1105", "ObjectType": "Group"}],
             "Aces": []},
            {"ObjectIdentifier": f"{B}-513",
             "Properties": {"name": "DOMAIN USERS@CORP.LOCAL"},
             "Members": [{"ObjectIdentifier": f"{B}-1001", "ObjectType": "User"}],
             "Aces": []},
            {"ObjectIdentifier": f"{B}-1105",
             "Properties": {"name": "HELPDESK@CORP.LOCAL"}, "Members": [],
             "Aces": [{"PrincipalSID": f"{B}-1001", "PrincipalType": "User",
                       "RightName": "GenericAll"}]},
        ]}
        domains = {"meta": {"type": "domains", "count": 1}, "data": [
            {"ObjectIdentifier": B,
             "Properties": {"name": "CORP.LOCAL", "functionallevel": "2016",
                            "machineaccountquota": 10},
             "Trusts": [],
             "Aces": [{"PrincipalSID": f"{B}-1001", "RightName": "GetChanges"},
                      {"PrincipalSID": f"{B}-1001", "RightName": "GetChangesAll"}]},
        ]}
        import json as _json
        for name, blob in (("users", users), ("groups", groups), ("domains", domains)):
            with open(os.path.join(d, f"2026_{name}.json"), "w") as fh:
                fh.write(_json.dumps(blob))

    def test_load_graph_builds_nodes_and_edges(self):
        from recce import bloodhound as bh
        with tempfile.TemporaryDirectory() as d:
            self._collection(d)
            g = bh.load_graph(d)
        self.assertEqual(g["nodes"][f"{self.BASE}-1001"]["type"], "User")
        # GenericAll ACE -> edge; group Members -> MemberOf; GetChanges*2 -> DCSync.
        labels = {(s.split("-")[-1], lbl, dd.split("-")[-1]) for s, lbl, dd in g["edges"]}
        self.assertIn(("1001", "GenericAll", "1105"), labels)
        self.assertIn(("1105", "MemberOf", "512"), labels)
        self.assertTrue(any(lbl == "DCSync" for _s, lbl, _d in g["edges"]))

    def test_is_sharphound_detects_collection(self):
        from recce import bloodhound as bh
        with tempfile.TemporaryDirectory() as d:
            self._collection(d)
            self.assertTrue(bh.is_sharphound(d))
        with tempfile.TemporaryDirectory() as d2:
            with open(os.path.join(d2, "x.json"), "w") as fh:
                fh.write('{"nope": 1}')
            self.assertFalse(bh.is_sharphound(d2))

    def test_findings_cover_the_classics(self):
        from recce import bloodhound as bh
        with tempfile.TemporaryDirectory() as d:
            self._collection(d)
            fs = bh.findings(bh.load_graph(d))
        cats = {f["category"] for f in fs}
        for expect in ("kerberoast", "asrep", "delegation", "dcsync", "hygiene", "creds"):
            self.assertIn(expect, cats)
        dcsync = next(f for f in fs if f["category"] == "dcsync")
        self.assertEqual(dcsync["severity"], "critical")
        self.assertIn("secretsdump", dcsync["command"])

    def test_attack_path_owned_user_to_domain_admin(self):
        from recce import bloodhound as bh
        with tempfile.TemporaryDirectory() as d:
            self._collection(d)
            g = bh.load_graph(d)
        paths = bh.attack_paths(g, owned={"BOB@CORP.LOCAL"})
        da = next((p for p in paths if "DOMAIN ADMINS" in p["target"].upper()), None)
        self.assertIsNotNone(da)
        self.assertEqual(da["length"], 2)                       # BOB->HELPDESK->DA
        self.assertEqual([s["label"] for s in da["steps"]], ["GenericAll", "MemberOf"])
        self.assertIn("GenericAll", da["chain"])
        # The domain object is reachable in one DCSync hop.
        self.assertTrue(any(p["length"] == 1 and s["label"] == "DCSync"
                            for p in paths for s in p["steps"]))

    def test_kerberos_actions_with_hash(self):
        from recce import bloodhound as bh
        with tempfile.TemporaryDirectory() as d:
            self._collection(d)
            g = bh.load_graph(d)
        acts = bh.kerberos_actions(g, {"domain": "CORP.LOCAL", "user": "bob",
                                       "secret": "aad3b...:31d6c...", "is_hash": True,
                                       "dc_ip": "10.0.0.1"})
        titles = " ".join(a["title"] for a in acts)
        self.assertIn("Kerberoast", titles)
        self.assertIn("AS-REP", titles)
        self.assertTrue(any("-hashes :" in a["command"] for a in acts))

    def test_live_kerberos_parsers(self):
        from recce import bloodhound as bh
        tgs = ("[*] Getting TGS for svc_sql\n"
               "$krb5tgs$23$*svc_sql$CORP.LOCAL$MSSQLSvc/db.corp.local:1433*$"
               "a1b2c3d4e5f6a7b8c9d0e1f2$deadbeef" * 1 + "\n"
               "$krb5tgs$23$*svc_web$CORP.LOCAL$HTTP/web.corp.local*$00112233$cafebabe\n")
        rows = bh.parse_tgs(tgs)
        self.assertEqual([r["user"] for r in rows], ["svc_sql", "svc_web"])
        self.assertEqual(rows[0]["spn"], "MSSQLSvc/db.corp.local:1433")
        asrep = ("[*] AS-REP for jdoe\n"
                 "$krb5asrep$23$jdoe@CORP.LOCAL:aabbcc$ddeeff001122\n")
        ar = bh.parse_asrep(asrep)
        self.assertEqual(ar[0]["user"], "jdoe")
        dump = ("Administrator:500:aad3b435b51404eeaad3b435b51404ee:"
                "31d6cfe0d16ae931b73c59d7e0c089c0:::\n"
                "CORP.LOCAL\\krbtgt:502:aad3b435b51404eeaad3b435b51404ee:"
                "1a2b3c4d5e6f70819293a4b5c6d7e8f9:::\n")
        sd = bh.parse_secretsdump(dump)
        self.assertEqual(len(sd), 2)
        krb = [h for h in sd if h["krbtgt"]]
        self.assertEqual(len(krb), 1)
        self.assertEqual(krb[0]["nt"], "1a2b3c4d5e6f70819293a4b5c6d7e8f9")

    def test_live_kerberos_toolmissing_is_clean(self):
        # No impacket installed in CI -> each runner reports the missing tool, never
        # raises, and produces no findings.
        from recce import bloodhound as bh
        creds = {"domain": "CORP.LOCAL", "user": "bob", "secret": "Pw",
                 "is_hash": False, "dc_ip": "10.0.0.1"}
        res = bh.live_kerberos(creds, None, do_roast=True, do_asrep=True, do_dcsync=True)
        self.assertEqual(res["findings"], [])
        self.assertEqual(len(res["errors"]), 3)
        self.assertTrue(all("not installed" in e for e in res["errors"]))

    def test_live_capture_findings_fold_into_vulns(self):
        # A captured TGS -> a proven 'roasted' finding -> a confirmed Vuln that reaches
        # the main totals with the real hash as evidence and the right CWE.
        from recce import bloodhound as bh
        out = bh.parse_tgs("$krb5tgs$23$*svc_sql$CORP.LOCAL$MSSQLSvc/db*$aa$bb\n")
        # Simulate a successful capture by exercising the finding-builder path.
        creds = {"user": "bob", "domain": "CORP.LOCAL"}
        fs = []
        for h in out:
            fs.append(bh._finding(
                "roasted", "high", "Kerberoast hash captured (proven)", h["user"], "",
                f"Captured a live TGS-REP for SPN '{h['spn']}'.\n\n{h['hash']}",
                "hashcat", "hashcat -m 13100 kerberoast.hash rockyou.txt", "rotate"))
        an = {"findings": fs}
        vulns = bh.findings_to_vulns(an, "10.0.0.9", "CORP.LOCAL")
        self.assertEqual(len(vulns), 1)
        v = vulns[0]
        self.assertEqual(v.confidence, "confirmed")
        self.assertIn("CWE-262", v.cwes)
        self.assertIn("$krb5tgs$", v.output)
        _ = creds

    def test_analyze_is_json_serialisable(self):
        from recce import bloodhound as bh
        import json as _json
        with tempfile.TemporaryDirectory() as d:
            self._collection(d)
            an = bh.analyze(d, owned={"BOB@CORP.LOCAL"})
        _json.dumps(an)                                          # must round-trip
        self.assertEqual(an["stats"]["nodes"], 7)
        self.assertTrue(an["stats"]["findings"] >= 6)
        self.assertTrue(an["paths"])

    def test_report_sheets_render(self):
        from recce import bloodhound as bh, report_excel, xlsx
        with tempfile.TemporaryDirectory() as d:
            self._collection(d)
            an = bh.analyze(d, owned={"BOB@CORP.LOCAL"},
                            creds={"domain": "CORP.LOCAL", "user": "bob",
                                   "secret": "x", "is_hash": False, "dc_ip": "1.2.3.4"})
            p = os.path.join(d, "wb.xlsx")
            report_excel.build_workbook([], p, meta={"subtitle": "T", "ad_bloodhound": an})
            sheets = xlsx.read_sheets(p)
        self.assertIn("AD Findings", sheets)
        self.assertIn("AD Attack Paths", sheets)
        findings_txt = "\n".join(" ".join(map(str, r)) for r in sheets["AD Findings"])
        self.assertIn("DCSync", findings_txt)
        self.assertIn("secretsdump", findings_txt)               # the prove command
        paths_txt = "\n".join(" ".join(map(str, r)) for r in sheets["AD Attack Paths"])
        self.assertIn("DOMAIN ADMINS", paths_txt)
        self.assertIn("MemberOf", paths_txt)                     # the edge chain
        self.assertIn("GetUserSPNs", paths_txt)                  # kerberos action

    def test_cmd_bloodhound_end_to_end(self):
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            self._collection(d)
            out = os.path.join(d, "eng")
            rc = cli.cmd_bloodhound(SimpleNamespace(
                paths=[d], username=None, password=None, domain=None,
                owned=["BOB@CORP.LOCAL"], creds=None, dc_ip=None,
                output_dir=out, title="T"))
            self.assertEqual(rc, 0)
            self.assertTrue(os.path.exists(os.path.join(out, "enumeration.xlsx")))
            st = Store(os.path.join(out, "results.sqlite"))
            blob = st.get_meta("ad_bloodhound")
            doms = {dm.name: dm for dm in st.all_domains()}
            st.close()
            self.assertTrue(blob)                                # analysis persisted
            self.assertIn("corp.local", doms)                    # domain merged in
            self.assertIn("bloodhound", doms["corp.local"].sources)

    def test_findings_to_vulns_feed_main_findings_and_writeups(self):
        from recce import bloodhound as bh
        from recce.report_docx import group_findings, list_findings, _vuln_type
        with tempfile.TemporaryDirectory() as d:
            self._collection(d)
            an = bh.analyze(d, owned={"BOB@CORP.LOCAL"})
        vulns = bh.findings_to_vulns(an, "10.0.0.9", "CORP.LOCAL")
        self.assertTrue(vulns)
        h = Host(ip="10.0.0.9", os_family="Windows", roles=["Domain Controller"],
                 vulns=vulns)
        # group_findings powers the severity rollup, Vulnerabilities sheet + writeups.
        groups = group_findings([h])
        titles = {f.title for f in groups}
        self.assertIn("DCSync rights held off tier-0", titles)
        dcsync = next(f for f in groups if f.title == "DCSync rights held off tier-0")
        self.assertEqual(dcsync.severity, "critical")
        self.assertTrue(dcsync.remediation)                      # remediation carried
        self.assertIn("secretsdump", " ".join(o for _i, _p, o in dcsync.evidence))
        # Every AD CWE must classify (keeps the CWE-coverage test green) + have a type.
        for f in groups:
            vt, _cia = _vuln_type(f.cwes)
            self.assertTrue(vt, f.cwes)
        # list_findings (the appendix/HTML feed) includes them with severity.
        lf = list_findings([h], min_severity="info")
        self.assertTrue(any("DCSync" in x["title"] for x in lf))

    def test_ad_findings_reach_vulnerabilities_sheet_e2e(self):
        from recce import cli, xlsx
        with tempfile.TemporaryDirectory() as d:
            self._collection(d)
            out = os.path.join(d, "eng")
            cli.cmd_bloodhound(SimpleNamespace(
                paths=[d], username="alice", password="Passw0rd!", domain="corp.local",
                owned=None, creds=None, dc_ip="10.0.0.9", output_dir=out, title="T"))
            sheets = xlsx.read_sheets(os.path.join(out, "enumeration.xlsx"))
        vtxt = "\n".join(" ".join(map(str, r)) for r in sheets.get("Vulnerabilities", []))
        self.assertIn("DCSync", vtxt)                            # in the MAIN vuln sheet
        self.assertIn("Kerberoastable", vtxt)

    def _collection_small(self, d):
        """A reduced collection - only ALICE (AS-REP roastable) - simulating a
        follow-up import after the other findings were remediated."""
        B = self.BASE
        import json as _json
        users = {"meta": {"type": "users"}, "data": [
            {"ObjectIdentifier": f"{B}-1002",
             "Properties": {"name": "ALICE@CORP.LOCAL", "enabled": True,
                            "dontreqpreauth": True}, "Aces": []}]}
        domains = {"meta": {"type": "domains"}, "data": [
            {"ObjectIdentifier": B, "Properties": {"name": "CORP.LOCAL"},
             "Trusts": [], "Aces": []}]}
        for n, b in (("users", users), ("domains", domains)):
            with open(os.path.join(d, f"2026_{n}.json"), "w") as fh:
                fh.write(_json.dumps(b))

    def test_replace_ad_clears_remediated_findings_but_keeps_scan_vulns(self):
        from recce import cli
        from recce.models import Host, Vuln
        from recce.store import Store

        def db(out):
            st = Store(os.path.join(out, "results.sqlite"))
            h = st.get_host("10.0.0.9")
            st.close()
            return {v.title for v in h.vulns}
        with tempfile.TemporaryDirectory() as d:
            full = os.path.join(d, "full")
            small = os.path.join(d, "small")
            out = os.path.join(d, "eng")
            os.makedirs(full)
            os.makedirs(small)
            os.makedirs(out)
            self._collection(full)
            self._collection_small(small)
            # Pre-seed the DC host with a scan-sourced vuln that must SURVIVE replace.
            st = Store(os.path.join(out, "results.sqlite"))
            st.upsert_host(Host(ip="10.0.0.9",
                                ports=[Port(portid=445, service="microsoft-ds")],
                                vulns=[Vuln(ip="10.0.0.9", port=445, protocol="tcp",
                                            script_id="smb-vuln-ms17-010", title="MS17-010",
                                            severity="critical", source="nse")]))
            st.close()
            base = dict(username="alice", password="p", domain="corp.local", owned=None,
                        creds=None, dc_ip="10.0.0.9", output_dir=out, title="T")
            cli.cmd_bloodhound(SimpleNamespace(paths=[full], replace_ad=False, **base))
            t1 = db(out)
            self.assertIn("Kerberoastable account", t1)
            self.assertIn("MS17-010", t1)
            # Re-import the remediated (smaller) collection WITH --replace-ad.
            cli.cmd_bloodhound(SimpleNamespace(paths=[small], replace_ad=True, **base))
            t2 = db(out)
            self.assertNotIn("Kerberoastable account", t2)       # remediated -> gone
            self.assertNotIn("DCSync rights held off tier-0", t2)
            self.assertIn("AS-REP roastable account (no Kerberos pre-auth)", t2)  # kept
            self.assertIn("MS17-010", t2)                        # scan vuln survived

    def test_distinct_findings_are_not_deduped_in_main_totals(self):
        # Two kerberoastable users share the generic title but must produce TWO
        # Vulns (distinct keys) so the main severity totals aren't undercounted.
        from recce import bloodhound as bh
        B = self.BASE
        analysis = {"findings": [
            {"category": "kerberoast", "severity": "medium",
             "title": "Kerberoastable account", "principal": "SVC1@C", "target": "",
             "detail": "", "command": "x", "remediation": "y"},
            {"category": "kerberoast", "severity": "medium",
             "title": "Kerberoastable account", "principal": "SVC2@C", "target": "",
             "detail": "", "command": "x", "remediation": "y"}]}
        vulns = bh.findings_to_vulns(analysis, "10.0.0.9", "C")
        self.assertEqual(len({v.key for v in vulns}), 2)         # distinct, not collapsed
        _ = B  # silence

    def test_domain_controller_not_flagged_for_unconstrained_delegation(self):
        from recce import bloodhound as bh
        B = self.BASE
        with tempfile.TemporaryDirectory() as d:
            # DC01 has unconstrained delegation AND is a member of Domain Controllers
            # (RID 516). It must NOT be reported (that's normal for a DC).
            comps = {"meta": {"type": "computers"}, "data": [
                {"ObjectIdentifier": f"{B}-1000",
                 "Properties": {"name": "DC01.CORP.LOCAL", "enabled": True,
                                "unconstraineddelegation": True}, "Aces": []}]}
            groups = {"meta": {"type": "groups"}, "data": [
                {"ObjectIdentifier": f"{B}-516",
                 "Properties": {"name": "DOMAIN CONTROLLERS@CORP.LOCAL"},
                 "Members": [{"ObjectIdentifier": f"{B}-1000", "ObjectType": "Computer"}],
                 "Aces": []}]}
            import json as _json
            for n, b in (("computers", comps), ("groups", groups)):
                with open(os.path.join(d, f"{n}.json"), "w") as fh:
                    fh.write(_json.dumps(b))
            fs = bh.findings(bh.load_graph(d))
        self.assertFalse([f for f in fs if f["category"] == "delegation"])
        # A non-DC computer with unconstrained delegation IS flagged.
        with tempfile.TemporaryDirectory() as d:
            comps = {"meta": {"type": "computers"}, "data": [
                {"ObjectIdentifier": f"{B}-1001",
                 "Properties": {"name": "APP01.CORP.LOCAL", "enabled": True,
                                "unconstraineddelegation": True}, "Aces": []}]}
            import json as _json
            with open(os.path.join(d, "computers.json"), "w") as fh:
                fh.write(_json.dumps(comps))
            fs = bh.findings(bh.load_graph(d))
        self.assertTrue([f for f in fs if f["category"] == "delegation"])

    def test_enabled_null_is_treated_as_enabled(self):
        from recce import bloodhound as bh
        B = self.BASE
        with tempfile.TemporaryDirectory() as d:
            users = {"meta": {"type": "users"}, "data": [
                {"ObjectIdentifier": f"{B}-1001",
                 "Properties": {"name": "SVC@C", "enabled": None, "hasspn": True,
                                "serviceprincipalnames": ["x/y"]}, "Aces": []}]}
            import json as _json
            with open(os.path.join(d, "users.json"), "w") as fh:
                fh.write(_json.dumps(users))
            fs = bh.findings(bh.load_graph(d))
        self.assertTrue([f for f in fs if f["category"] == "kerberoast"])

    def test_bare_string_members_do_not_crash(self):
        from recce import bloodhound as bh
        B = self.BASE
        with tempfile.TemporaryDirectory() as d:
            # Members / LocalAdmins as bare SID strings (older SharpHound).
            groups = {"meta": {"type": "groups"}, "data": [
                {"ObjectIdentifier": f"{B}-512", "Properties": {"name": "DA@C"},
                 "Members": [f"{B}-1001"], "Aces": []}]}
            comps = {"meta": {"type": "computers"}, "data": [
                {"ObjectIdentifier": f"{B}-1000", "Properties": {"name": "WS@C"},
                 "LocalAdmins": [f"{B}-1001"], "Aces": []}]}
            import json as _json
            for n, b in (("groups", groups), ("computers", comps)):
                with open(os.path.join(d, f"{n}.json"), "w") as fh:
                    fh.write(_json.dumps(b))
            g = bh.load_graph(d)                                 # must not raise
        labels = {(s.split("-")[-1], lbl, dd.split("-")[-1]) for s, lbl, dd in g["edges"]}
        self.assertIn(("1001", "MemberOf", "512"), labels)
        self.assertIn(("1001", "AdminTo", "1000"), labels)

    def test_fill_creds_password_containing_a_token_is_safe(self):
        from recce import bloodhound as bh
        an = {"findings": [{"command": "run <DOMAIN>/<user>:<pass> against <dc>"}],
              "kerberos": [], "paths": []}
        # Password literally contains "<dc>" - must NOT be re-substituted.
        bh.fill_creds(an, {"domain": "corp.local", "user": "alice",
                           "secret": "p<dc>w", "is_hash": False, "dc_ip": "10.0.0.1"})
        cmd = an["findings"][0]["command"]
        self.assertIn("corp.local/alice:p<dc>w", cmd)            # password intact
        self.assertTrue(cmd.endswith("against 10.0.0.1"))        # real <dc> filled

    def test_fill_creds_makes_commands_copy_paste_ready(self):
        from recce import bloodhound as bh
        with tempfile.TemporaryDirectory() as d:
            self._collection(d)
            an = bh.analyze(d, owned={"BOB@CORP.LOCAL"})
        bh.fill_creds(an, {"domain": "corp.local", "user": "alice",
                           "secret": "Passw0rd!", "is_hash": False, "dc_ip": "10.0.0.1"})
        cmds = " ".join(f["command"] for f in an["findings"])
        self.assertIn("corp.local/alice:Passw0rd!", cmds)        # DOMAIN/user:pass filled
        self.assertIn("10.0.0.1", cmds)                          # dc-ip filled
        self.assertNotIn("<dc>", cmds)
        self.assertNotIn("<DOMAIN>", cmds)

    def test_simple_credentialed_run_defaults_owned_to_you(self):
        # -u alice with no --owned: paths must start from ALICE (the simple UX).
        from recce import cli
        with tempfile.TemporaryDirectory() as d:
            # ALICE has GenericAll on HELPDESK instead of BOB.
            self._collection(d)
            import json as _json
            groups_path = os.path.join(d, "2026_groups.json")
            g = _json.loads(open(groups_path).read())
            for obj in g["data"]:
                if obj["ObjectIdentifier"].endswith("-1105"):
                    obj["Aces"] = [{"PrincipalSID": f"{self.BASE}-1002",
                                    "RightName": "GenericAll"}]
            open(groups_path, "w").write(_json.dumps(g))
            out = os.path.join(d, "eng")
            rc = cli.cmd_bloodhound(SimpleNamespace(
                paths=[d], username="alice", password="Passw0rd!", domain="CORP.LOCAL",
                owned=None, creds=None, dc_ip="10.0.0.1", output_dir=out, title="T"))
            self.assertEqual(rc, 0)
            from recce import xlsx
            sheets = xlsx.read_sheets(os.path.join(out, "enumeration.xlsx"))
            paths_txt = "\n".join(" ".join(map(str, r))
                                  for r in sheets.get("AD Attack Paths", []))
            self.assertIn("ALICE", paths_txt.upper())            # path starts from ALICE
            # Kerberos command carries the real creds, not placeholders.
            self.assertIn("CORP.LOCAL/alice", paths_txt)


class AdcsCertipyTest(unittest.TestCase):
    def _certipy(self, path):
        data = {
            "Certificate Authorities": {
                "0": {"CA Name": "CORP-CA", "DNS Name": "ca.corp.local",
                      "Web Enrollment": "Enabled",
                      "[!] Vulnerabilities": {
                          "ESC8": "Web Enrollment is enabled and Request Disposition is Issue"}}},
            "Certificate Templates": {
                "0": {"Template Name": "VulnUser", "Enabled": True,
                      "Client Authentication": True, "Enrollee Supplies Subject": True,
                      "Certificate Authorities": ["CORP-CA"],
                      "Permissions": {"Enrollment Permissions": {
                          "Enrollment Rights": ["CORP.LOCAL\\Domain Users"]}},
                      "[!] Vulnerabilities": {
                          "ESC1": "'CORP.LOCAL\\Domain Users' can enroll and supply a SAN"}},
                "1": {"Template Name": "Boring", "Enabled": True,
                      "[!] Vulnerabilities": {}}}}
        import json as _json
        with open(path, "w") as fh:
            fh.write(_json.dumps(data))

    def test_is_certipy_detects_file(self):
        from recce import adcs
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "20260101_Certipy.json")
            self._certipy(p)
            self.assertTrue(adcs.is_certipy(p))
            other = os.path.join(d, "sh.json")
            with open(other, "w") as fh:
                fh.write('{"meta": {"type": "users"}, "data": []}')
            self.assertFalse(adcs.is_certipy(other))

    def test_findings_map_esc_to_exact_commands(self):
        from recce import adcs
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "c.json")
            self._certipy(p)
            fs = adcs.findings(p)
        cats = {f["category"] for f in fs}
        self.assertIn("adcs-esc1", cats)
        self.assertIn("adcs-esc8", cats)
        esc1 = next(f for f in fs if f["category"] == "adcs-esc1")
        self.assertEqual(esc1["severity"], "critical")
        self.assertIn("certipy req", esc1["command"])
        self.assertIn("VulnUser", esc1["command"])               # real template name
        self.assertIn("Domain Users", esc1["principal"])         # who can enroll

    def test_enrollment_rights_as_dict_does_not_crash(self):
        from recce import adcs
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "c.json")
            data = {"Certificate Templates": {"0": {
                "Template Name": "VulnUser",
                "Permissions": {"Enrollment Permissions": {
                    "Enrollment Rights": {"CORP\\Domain Users": "Enroll"}}},  # dict form
                "[!] Vulnerabilities": {"ESC1": "x"}}}}
            import json as _json
            with open(p, "w") as fh:
                fh.write(_json.dumps(data))
            fs = adcs.findings(p)                                # must not raise
        self.assertTrue(fs)
        self.assertIn("Domain Users", fs[0]["principal"])

    def test_certipy_flows_into_workbook_with_creds(self):
        from recce import cli, xlsx
        with tempfile.TemporaryDirectory() as d:
            cp = os.path.join(d, "certipy.json")
            self._certipy(cp)
            out = os.path.join(d, "eng")
            rc = cli.cmd_bloodhound(SimpleNamespace(
                paths=[cp], username="alice", password="Passw0rd!",
                domain="corp.local", owned=None, creds=None, dc_ip="10.0.0.1",
                output_dir=out, title="T"))
            self.assertEqual(rc, 0)
            sheets = xlsx.read_sheets(os.path.join(out, "enumeration.xlsx"))
            txt = "\n".join(" ".join(map(str, r)) for r in sheets["AD Findings"])
            self.assertIn("ESC1", txt)
            self.assertIn("VulnUser", txt)
            self.assertIn("alice@corp.local", txt)               # creds pre-filled
            self.assertIn("10.0.0.1", txt)


class AttackPathTest(unittest.TestCase):
    def _hosts(self):
        from recce.models import Vuln, Account
        dc = Host(ip="10.0.10.5", hostnames=["dc01"], os_family="Windows",
                  roles=["Domain Controller"], smb_signing="not required",
                  accounts=[Account(ip="10.0.10.5", source="nse", kind="domain",
                                    domain="CORP")],
                  ports=[Port(portid=445, service="microsoft-ds"),
                         Port(portid=5985, service="http")],
                  vulns=[Vuln(ip="10.0.10.5", port=445, protocol="tcp",
                              script_id="smb-vuln-ms17-010", title="smb-vuln-ms17-010",
                              severity="high", source="nse", ids=["CVE-2017-0143"],
                              output="VULNERABLE"),
                         Vuln(ip="10.0.10.5", port=0, protocol="tcp",
                              script_id="local-enum", title="SeImpersonate -> Potato",
                              severity="high", source="local", confidence="confirmed",
                              output="SeImpersonate held")])
        return [dc]

    def test_stages_and_ordering(self):
        from recce import attackpath as ap
        steps = ap.build(self._hosts())
        stages = [s["stage"] for s in steps]
        # ordered by STAGE_ORDER
        idx = [ap.STAGE_ORDER.index(s) for s in stages]
        self.assertEqual(idx, sorted(idx))
        self.assertIn("Initial Access", stages)          # ms17-010
        self.assertIn("Privilege Escalation", stages)    # SeImpersonate/Potato
        self.assertIn("Domain Dominance", stages)        # AS-REP/Kerberoast/relay on DC
        self.assertIn("Lateral Movement", stages)        # SMB/WinRM present

    def test_narrative_grounded(self):
        from recce import attackpath as ap
        hosts = self._hosts()
        text = " ".join(ap.narrative(hosts))
        self.assertIn("Likely path", text)
        self.assertIn("10.0.10.5", text)                 # names the real host

    def test_attackpath_sheet(self):
        from recce.report_excel import _spec_attackpath
        spec = _spec_attackpath(self._hosts())
        self.assertEqual(spec.title, "Attack Path")
        self.assertIn("Stage", [c[0] for c in spec.cols])
        self.assertTrue(spec.rows)

    def test_empty_when_no_confirmed(self):
        from recce import attackpath as ap
        h = Host(ip="10.0.0.1", os_family="Linux",
                 ports=[Port(portid=23, service="telnet")])
        self.assertEqual(ap.build([h]), [])              # no confirmed findings

    def test_mermaid_graph(self):
        from recce import attackpath as ap
        hosts = self._hosts()
        mmd = ap.mermaid(hosts)
        self.assertTrue(mmd.startswith("flowchart LR"))
        self.assertIn('subgraph S0["Initial Access"]', mmd)
        self.assertIn("10.0.10.5", mmd)                  # real host in a node
        self.assertIn("-->", mmd)                        # stage-to-stage flow
        # Same host walks stages -> a dashed continuity edge.
        self.assertIn("same host", mmd)

    def test_dot_graph(self):
        from recce import attackpath as ap
        hosts = self._hosts()
        dot = ap.dot(hosts)
        self.assertTrue(dot.startswith("digraph attack_path {"))
        self.assertIn("rankdir=LR", dot)
        self.assertIn("cluster_0", dot)
        self.assertIn("10.0.10.5", dot)
        self.assertTrue(dot.rstrip().endswith("}"))

    def test_graph_empty_is_valid(self):
        from recce import attackpath as ap
        h = Host(ip="10.0.0.1", os_family="Linux",
                 ports=[Port(portid=23, service="telnet")])
        self.assertIn("flowchart LR", ap.mermaid([h]))
        self.assertIn("digraph", ap.dot([h]))

    def test_cmd_writes_graph_files(self):
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as dd:
            db = os.path.join(dd, "results.sqlite")
            st = Store(db)
            for h in self._hosts():
                st.upsert_host(h)
            st.close()
            rc = cli.cmd_attackpath(SimpleNamespace(output_dir=dd, targets=[]))
            self.assertEqual(rc, 0)
            self.assertTrue(os.path.exists(os.path.join(dd, "attack_path.mmd")))
            self.assertTrue(os.path.exists(os.path.join(dd, "attack_path.dot")))


class ListenerBackfillTest(unittest.TestCase):
    LOOT = (
        "recce-enum  host=web01  user=root  now\n"
        "==== Network ====\n"
        "    listening-service inventory (proto/port/process/binary):\n"
        "    LISTEN proto=tcp addr=0.0.0.0 port=80 pid=1337 proc=nginx bin=/usr/sbin/nginx\n"
        "    LISTEN proto=tcp addr=127.0.0.1 port=6379 pid=990 proc=redis-server bin=/usr/bin/redis-server\n"
        "    LISTEN proto=tcp addr=0.0.0.0 port=5985 pid=1200 proc=svchost svc=WinRM bin=C:\\Windows\\svchost.exe\n"
        "    LISTEN proto=udp addr=[::] port=53 pid=800 proc=named bin=/usr/sbin/named\n")

    def test_parse_listeners_linux_and_windows_lines(self):
        from recce import ingest
        ls = {(x["proto"], x["port"]): x for x in ingest.parse_listeners(self.LOOT)}
        self.assertEqual(ls[("tcp", 80)]["bin"], "/usr/sbin/nginx")
        self.assertFalse(ls[("tcp", 80)]["loopback"])
        self.assertTrue(ls[("tcp", 6379)]["loopback"])          # 127.0.0.1
        self.assertEqual(ls[("tcp", 5985)]["svc"], "WinRM")     # windows svc= field
        self.assertEqual(ls[("udp", 53)]["port"], 53)
        # No listener lines -> empty (older loot degrades gracefully).
        self.assertEqual(ingest.parse_listeners("recce-enum host=x\n[!] a finding"), [])

    def test_backfill_enriches_and_adds_ports(self):
        from recce import ingest
        h = Host(ip="10.0.0.9", ports=[
            Port(portid=80, protocol="tcp", service="http", product="nginx",
                 detect_source="nmap", state="open")])
        added, enriched = ingest.backfill_ports(h, ingest.parse_listeners(self.LOOT))
        self.assertEqual((added, enriched), (3, 1))
        idx = {(p.protocol, p.portid): p for p in h.ports}
        # Existing nmap port keeps its service; only gains the backing binary.
        self.assertEqual(idx[("tcp", 80)].service, "http")
        self.assertEqual(idx[("tcp", 80)].detect_source, "nmap")
        self.assertEqual(idx[("tcp", 80)].binary, "/usr/sbin/nginx")
        # Loopback-only service the network scan never saw is now on the host.
        self.assertIn(("tcp", 6379), idx)
        self.assertEqual(idx[("tcp", 6379)].detect_source, "local")
        self.assertIn("loopback", idx[("tcp", 6379)].extrainfo)
        # Windows svc name becomes the service label + noted in extra info.
        self.assertEqual(idx[("tcp", 5985)].service, "WinRM")
        self.assertEqual(idx[("udp", 53)].service, "named")

    def test_fold_loot_backfills_ports_end_to_end(self):
        from recce import cli
        h = Host(ip="10.0.0.9", os_family="Linux",
                 ports=[Port(portid=80, protocol="tcp", service="http",
                             detect_source="nmap", state="open")])
        cli._fold_loot(h, self.LOOT, "loot.txt")
        idx = {(p.protocol, p.portid): p for p in h.ports}
        self.assertEqual(idx[("tcp", 80)].binary, "/usr/sbin/nginx")
        self.assertIn(("tcp", 6379), idx)               # loopback service added

    def test_backfill_survives_store_round_trip(self):
        from recce import ingest
        from recce.store import Store
        h = Host(ip="10.0.0.9", subnet="10.0.0.0/24", ports=[])
        ingest.backfill_ports(h, ingest.parse_listeners(self.LOOT))
        with tempfile.TemporaryDirectory() as d:
            st = Store(os.path.join(d, "r.sqlite"))
            st.upsert_host(h)
            back = st.get_host("10.0.0.9")
            st.close()
        binp = {(p.protocol, p.portid): p.binary for p in back.ports}
        self.assertEqual(binp[("tcp", 80)], "/usr/sbin/nginx")


class EngagementPermsTest(unittest.TestCase):
    def test_relax_perms_makes_tree_777(self):
        from recce import cli
        with tempfile.TemporaryDirectory() as d:
            sub = os.path.join(d, "raw")
            os.makedirs(sub)
            f1 = os.path.join(d, "report.html")
            f2 = os.path.join(sub, "10.0.0.5.xml")
            for f in (f1, f2):
                with open(f, "w") as fh:
                    fh.write("x")
                os.chmod(f, 0o600)          # simulate a root-created, locked file
            os.chmod(sub, 0o700)
            cli._relax_perms(d)
            for p in (d, sub, f1, f2):
                self.assertEqual(stat.S_IMODE(os.stat(p).st_mode), 0o777, p)

    def test_relax_perms_is_best_effort_on_missing_dir(self):
        from recce import cli
        cli._relax_perms("/nonexistent/path/xyz")      # must not raise

    def test_open_paths_relaxes_output_dir(self):
        from recce import cli
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "engagement")
            cli._open_paths(out)
            self.assertEqual(stat.S_IMODE(os.stat(out).st_mode), 0o777)
            self.assertEqual(stat.S_IMODE(os.stat(os.path.join(out, "raw")).st_mode),
                             0o777)

    def test_main_finally_relaxes_perms_even_on_early_return(self):
        from recce import cli
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "eng")
            # attackpath with no datastore returns 1 early - the finally must still
            # relax the folder that _open_paths created.
            rc = cli.main(["attackpath", "-o", out])
            self.assertEqual(rc, 1)
            self.assertEqual(stat.S_IMODE(os.stat(out).st_mode), 0o777)


class AVAwarenessTest(unittest.TestCase):
    LOOT = ("recce-enum  host=DC01  user=admin  now\n"
            "==== AV / EDR detection ====\n"
            "    AV product: Windows Defender\n"
            "[!] EDR/AV process: CSFalcon\n"
            "    EDR/AV service: CSAgent\n"
            "==== OS hardening & defences ====\n"
            "    Defender: RealTime=True Tamper=True\n"
            "    LSA protection (RunAsPPL)=1\n"
            "    Sysmon service present (activity is being logged)\n"
            "    AppLocker policy present (review allowed paths)\n")

    def test_extract_defenses(self):
        from recce import ingest
        j = " | ".join(ingest.extract_defenses(self.LOOT))
        for expect in ("AV: Windows Defender", "CSFalcon (process)",
                       "CSAgent (service)", "Defender RTP=True",
                       "LSASS protected (RunAsPPL)", "Sysmon present (logging)",
                       "AppLocker enforced"):
            self.assertIn(expect, j)

    def test_ingest_populates_defenses(self):
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as dd:
            db = os.path.join(dd, "results.sqlite")
            st = Store(db)
            st.upsert_host(Host(ip="10.0.0.5", subnet="10.0.0.0/24",
                                os_family="Windows",
                                ports=[Port(portid=445, service="microsoft-ds")]))
            st.close()
            loot = os.path.join(dd, "l.txt")
            with open(loot, "w") as fh:
                fh.write(self.LOOT)
            cli.cmd_ingest(SimpleNamespace(output_dir=dd, loot=loot,
                                           host="10.0.0.5", title="t"))
            st = Store(db)
            h = {x.ip: x for x in st.all_hosts()}["10.0.0.5"]
            st.close()
            self.assertTrue(any("CSFalcon" in x for x in h.defenses))

    def test_checklist_and_exploitation_columns(self):
        from recce.report_excel import _spec_checklist, _spec_exploitation
        from recce.models import Vuln
        h = Host(ip="10.0.0.5", os_family="Windows",
                 defenses=["EDR/AV: CSFalcon (process)"],
                 ports=[Port(portid=445, service="microsoft-ds")],
                 vulns=[Vuln(ip="10.0.0.5", port=445, protocol="tcp",
                             script_id="local-enum",
                             title="SeImpersonate -> Potato -> SYSTEM", severity="high",
                             source="local", confidence="confirmed",
                             output="SeImpersonate held")])
        cl = _spec_checklist([h])
        self.assertIn("AV / EDR", [c[0] for c in cl.cols])
        self.assertEqual(cl.rows[0]["data"]["AV / EDR"], "EDR/AV: CSFalcon (process)")
        ex = _spec_exploitation([h])
        self.assertIn("Defenses (host)", [c[0] for c in ex.cols])
        self.assertTrue(any("CSFalcon" in r["data"].get("Defenses (host)", "")
                            for r in ex.rows))

    def test_exploitplan_defenses_banner(self):
        from recce import exploitplan as ep
        from recce.models import Vuln
        h = Host(ip="10.0.0.5", os_family="Windows",
                 defenses=["EDR/AV: CSFalcon (process)"],
                 ports=[Port(portid=445, service="microsoft-ds")],
                 vulns=[Vuln(ip="10.0.0.5", port=445, protocol="tcp",
                             script_id="smb-vuln-ms17-010", title="smb-vuln-ms17-010",
                             severity="high", source="nse", ids=["CVE-2017-0143"],
                             output="VULNERABLE")])
        with tempfile.TemporaryDirectory() as dd:
            s = ep.build_plan([h], dd)
            with open(os.path.join(s["dir"], "10.0.0.5.sh")) as fh:
                sh = fh.read()
        self.assertIn("DEFENSES on 10.0.0.5", sh)
        self.assertIn("CSFalcon", sh)
        self.assertIn("does not evade AV", sh)   # coordination, not evasion


class ServiceEnumTest(unittest.TestCase):
    def test_script_mapping(self):
        from recce import serviceenum as se
        self.assertEqual(se.script_for("microsoft-ds", 445), "smb")
        self.assertEqual(se.script_for("netbios-ssn", 139), "smb")
        self.assertEqual(se.script_for("ssl/http", 8443), "http")
        self.assertEqual(se.script_for("", 6379), "redis")       # port fallback
        self.assertEqual(se.script_for("http", 5985), "winrm")   # WinRM port wins
        self.assertEqual(se.script_for("ms-wbt-server", 3389), "rdp")
        self.assertEqual(se.script_for("unknown-thing", 12345), "")

    def test_commands_and_unmapped(self):
        from recce import serviceenum as se
        h = Host(ip="10.0.0.5", hostnames=["dc"],
                 ports=[Port(portid=445, service="microsoft-ds"),
                        Port(portid=6379, service="redis"),
                        Port(portid=9999, service="weird", state="open")])
        cmds = se.commands_for_host(h)
        scripts = {c[2] for c in cmds}
        self.assertEqual(scripts, {"smb", "redis"})
        self.assertTrue(all(c[3].startswith("./scripts/recce-service.sh") for c in cmds))
        self.assertIn((9999, "weird"), se.unmapped_ports(h))

    def test_cmd_services_smoke(self):
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            db = os.path.join(d, "results.sqlite")
            st = Store(db)
            st.upsert_host(Host(ip="10.0.0.5", subnet="10.0.0.0/24",
                                ports=[Port(portid=445, service="microsoft-ds"),
                                       Port(portid=80, service="http")]))
            st.close()
            rc = cli.cmd_services(SimpleNamespace(output_dir=d, targets=[],
                                                  host=[], subnet=[], aggressive=False))
            self.assertEqual(rc, 0)


class SvcDetectTest(unittest.TestCase):
    def test_servicefp_mining_names_unknown_port(self):
        from recce import svcdetect as sd
        p = Port(portid=5900, service="unknown", servicefp="RFB 003.008\n")
        self.assertTrue(sd.enrich_port("1.1.1.1", p, active=False))
        self.assertEqual(p.service, "vnc")
        self.assertEqual(p.detect_source, "inferred")

    def test_curated_port_map_labels_windows_services(self):
        from recce import svcdetect as sd
        p = Port(portid=5040, service="unknown")
        sd.enrich_port("1.1.1.1", p, active=False)
        self.assertEqual(p.service, "cdpsvc")
        self.assertIn("CDPSvc", p.extrainfo)
        self.assertEqual(p.detect_source, "inferred")
        # Dynamic MSRPC ephemeral range.
        p2 = Port(portid=49664, service="")
        sd.enrich_port("1.1.1.1", p2, active=False)
        self.assertEqual(p2.service, "msrpc")

    def test_nmap_named_port_is_never_overwritten(self):
        from recce import svcdetect as sd
        p = Port(portid=80, service="http", detect_source="nmap")
        self.assertFalse(sd.enrich_port("1.1.1.1", p, active=False))
        self.assertEqual(p.service, "http")

    def test_banner_signature_matching(self):
        from recce import svcdetect as sd
        self.assertEqual(sd._match_signature("SSH-2.0-OpenSSH_8.9")[0], "ssh")
        self.assertEqual(sd._match_signature("HTTP/1.1 200 OK")[0], "http")
        self.assertEqual(sd._match_signature("+PONG\r\n")[0], "redis")
        self.assertEqual(sd._match_signature("\x03\x00\x00\x13")[0], "ms-wbt-server")
        self.assertIsNone(sd._match_signature("random noise"))

    def test_suggest_command_only_for_still_unknown(self):
        from recce import svcdetect as sd
        unknown = Port(portid=1234, service="unknown")
        self.assertIn("nmap -sV --version-all",
                      sd.suggest_id_command("1.1.1.1", unknown))
        named = Port(portid=1234, service="cdpsvc", detect_source="inferred")
        self.assertEqual(sd.suggest_id_command("1.1.1.1", named), "")

    def test_reprobe_upgrades_still_unknown_ports(self):
        from recce import svcdetect as sd
        host = Host(ip="10.0.0.7", ports=[
            Port(portid=8888, service="unknown", state="open"),
            Port(portid=5040, service="cdpsvc", detect_source="inferred", state="open"),
        ])
        self.assertEqual(sd.still_unknown_ports(host), [8888])
        # nmap's second-opinion parse now names 8888 concretely.
        parsed = [Host(ip="10.0.0.7", ports=[
            Port(portid=8888, service="http", product="nginx", version="1.25",
                 state="open")])]
        n = sd.apply_reprobe(host, parsed)
        self.assertEqual(n, 1)
        p = next(p for p in host.ports if p.portid == 8888)
        self.assertEqual((p.service, p.product, p.detect_source),
                         ("http", "nginx", "nmap"))
        # The inferred port nmap still can't name is left untouched.
        self.assertEqual(sd.still_unknown_ports(host), [])

    def test_reprobe_scanner_command_targets_only_leftover_ports(self):
        from recce import scanner
        seen = {}
        orig = scanner._run

        def fake_run(cmd, timeout=None):
            seen["cmd"] = cmd
            return scanner.RunOutcome(returncode=0)
        scanner._run = fake_run
        try:
            with tempfile.TemporaryDirectory() as d:
                out = os.path.join(d, "rp.xml")
                scanner.reprobe_services("10.0.0.7", [8888, 3389], out,
                                         scanner.PROFILES["standard"])
        finally:
            scanner._run = orig
        cmd = seen["cmd"]
        self.assertIn("--version-all", cmd)
        self.assertIn("3389,8888", cmd)          # ports are sorted
        # Empty leftover list -> no scan (returns an empty XML, never shells out).
        seen.clear()
        with tempfile.TemporaryDirectory() as d:
            scanner.reprobe_services("10.0.0.7", [], os.path.join(d, "e.xml"),
                                     scanner.PROFILES["standard"])
        self.assertNotIn("cmd", seen)

    def test_parse_product_version_from_banners(self):
        from recce import svcdetect as sd
        cases = {
            "SSH-2.0-OpenSSH_8.9p1 Ubuntu-3": ("OpenSSH", "8.9p1"),
            "220 (vsFTPd 3.0.3)": ("vsFTPd", "3.0.3"),
            "220 mail ESMTP Exim 4.94 Debian": ("Exim", "4.94"),
            "5.5.5-10.3.34-MariaDB-log": ("MariaDB", "10.3.34"),
            "Server: Apache/2.4.41 (Ubuntu)": ("Apache", "2.4.41"),
            "+OK Dovecot ready.": ("Dovecot", ""),
        }
        for banner, (prod, ver) in cases.items():
            got = sd.parse_product_version(banner)
            self.assertIsNotNone(got, banner)
            self.assertEqual(got[0], prod, banner)
            if ver:
                self.assertEqual(got[1], ver, banner)
        self.assertIsNone(sd.parse_product_version("just some noise"))

    def test_enrich_versions_fills_product_for_cve_mapping(self):
        from recce import svcdetect as sd
        # nmap named the service but left product blank; we hold its banner.
        host = Host(ip="10.0.0.8", ports=[
            Port(portid=22, service="ssh", detect_source="nmap", state="open",
                 banner="SSH-2.0-OpenSSH_7.4"),
            Port(portid=25, service="smtp", detect_source="nmap", state="open",
                 servicefp="220 relay ESMTP Postfix 3.4.14"),
        ])
        n = sd.enrich_versions(host)
        self.assertEqual(n, 2)
        p22 = next(p for p in host.ports if p.portid == 22)
        self.assertEqual((p22.product, p22.version), ("OpenSSH", "7.4"))
        p25 = next(p for p in host.ports if p.portid == 25)
        self.assertEqual(p25.product, "Postfix")

    def test_enrich_versions_never_overwrites_nmap_product(self):
        from recce import svcdetect as sd
        host = Host(ip="10.0.0.8", ports=[
            Port(portid=22, service="ssh", product="OpenSSH", version="9.6",
                 detect_source="nmap", state="open",
                 banner="SSH-2.0-OpenSSH_7.4")])   # stale banner must NOT win
        self.assertEqual(sd.enrich_versions(host), 0)
        self.assertEqual(host.ports[0].version, "9.6")

    def test_new_port_fields_round_trip_through_store(self):
        # servicefp / detect_source / banner must survive a datastore round-trip.
        with tempfile.TemporaryDirectory() as d:
            st = Store(os.path.join(d, "r.sqlite"))
            st.upsert_host(Host(ip="10.0.0.9", subnet="10.0.0.0/24",
                                ports=[Port(portid=5040, service="cdpsvc",
                                            detect_source="inferred",
                                            servicefp="fp", banner="b")]))
            back = st.get_host("10.0.0.9")
            st.close()
            p = back.ports[0]
            self.assertEqual((p.service, p.detect_source), ("cdpsvc", "inferred"))
            self.assertEqual((p.servicefp, p.banner), ("fp", "b"))


class ReportTest(unittest.TestCase):
    def test_workbook_builds_and_has_sheets(self):
        hosts = parser.parse_nmap_xml(SAMPLE)
        ad.analyze_hosts(hosts)
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "x.xlsx")
            build_workbook(hosts, out, meta={"subtitle": "t"})
            self.assertTrue(os.path.exists(out))
            sheets = xlsx.read_sheets(out)
        for name in ("Start Here", "Overview", "Checklist", "Services by Product",
                     "Vulnerabilities", "AD Quick Wins"):
            self.assertIn(name, sheets)

    def test_opens_in_openpyxl_if_available(self):
        # Optional: proves the stdlib-written file parses in a real xlsx engine.
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        hosts = parser.parse_nmap_xml(SAMPLE)
        ad.analyze_hosts(hosts)
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "x.xlsx")
            build_workbook(hosts, out)
            wb = load_workbook(out)
            self.assertIn("Checklist", wb.sheetnames)


class MasscanParseTest(unittest.TestCase):
    def test_parse_sweep(self):
        xml = (
            '<?xml version="1.0"?><nmaprun>'
            '<host><address addr="10.0.0.5" addrtype="ipv4"/>'
            '<ports><port protocol="tcp" portid="22"><state state="open"/></port>'
            '<port protocol="tcp" portid="443"><state state="open"/></port></ports></host>'
            '<host><address addr="10.0.0.6" addrtype="ipv4"/>'
            '<ports><port protocol="tcp" portid="3389"><state state="open"/></port></ports>'
            '</host></nmaprun>'
        )
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "m.xml")
            with open(p, "w") as fh:
                fh.write(xml)
            got = scanner.parse_masscan_sweep_xml(p)
        self.assertEqual(got["10.0.0.5"], [22, 443])
        self.assertEqual(got["10.0.0.6"], [3389])


class StoreTrackingTest(unittest.TestCase):
    def test_set_and_get(self):
        with tempfile.TemporaryDirectory() as d:
            store = Store(os.path.join(d, "t.sqlite"))
            store.set_reviewed("host:1.2.3.4", True, notes="checked")
            store.bulk_set_tracking({"svc:1.2.3.4:tcp:80": (True, "")})
            t = store.get_tracking()
            self.assertTrue(t["host:1.2.3.4"][0])
            self.assertEqual(t["host:1.2.3.4"][1], "checked")
            self.assertTrue(t["svc:1.2.3.4:tcp:80"][0])
            # Un-review preserves prior note when notes not passed.
            store.set_reviewed("host:1.2.3.4", False)
            self.assertEqual(store.get_tracking()["host:1.2.3.4"], (False, "checked"))
            store.close()


class PlaybookTest(unittest.TestCase):
    def test_host_level_finding_walkthrough_has_no_bogus_port(self):
        # A port-less (host-level) priv-esc finding must not render "nmap -p None".
        from recce.report_docx import group_findings, _walkthrough_steps
        h = Host(ip="10.0.20.5", os_family="Linux", vulns=[
            Vuln(ip="10.0.20.5", port=None, protocol="tcp", script_id="local-enum",
                 title="SUID GTFOBins escalation candidate", severity="high",
                 source="local", confidence="confirmed",
                 output="On-target enum: SUID /usr/bin/find - GTFOBins")])
        steps = " ".join(_walkthrough_steps(group_findings([h])[0]))
        self.assertNotIn("None", steps)
        self.assertNotIn("-p ,", steps)

    def test_port_less_finding_writeup_has_no_none(self):
        # The whole rendered write-up (Affected systems / Evidence / walkthrough)
        # must never show "ip:None" for a host-level finding.
        import zipfile
        from recce.report_docx import build_writeups
        h = Host(ip="10.0.20.5", os_family="Linux", vulns=[
            Vuln(ip="10.0.20.5", port=None, protocol="tcp", script_id="local-enum",
                 title="Sudo misconfiguration -> root", severity="high",
                 source="local", confidence="confirmed",
                 output="On-target enum: NOPASSWD sudo entries present")])
        with tempfile.TemporaryDirectory() as d:
            build_writeups([h], d, min_severity="low")
            import glob
            docs = glob.glob(os.path.join(d, "*.docx"))
            self.assertTrue(docs)
            for f in docs:
                t = zipfile.ZipFile(f).read("word/document.xml").decode("utf-8", "replace")
                self.assertNotIn(":None", t)
                self.assertNotIn("-p None", t)

    def test_windows_seimpersonate_maps_to_potato(self):
        from recce import playbook
        e = playbook.for_text("Token holds SeImpersonate -> SYSTEM", "Windows")
        self.assertIsNotNone(e)
        self.assertIn("GodPotato", e["tool"])
        self.assertIn("whoami", e["cmd"].lower())
        self.assertIn("SYSTEM", e["validate"])

    def test_finding_values_are_substituted_into_command(self):
        from recce import playbook
        # the SUID binary path from the finding is filled into the command
        e = playbook.for_text("SUID /usr/bin/find - GTFOBins escalation candidate",
                              "Linux")
        self.assertIn("/usr/bin/find", e["cmd"])
        # unquoted service path is extracted too
        e2 = playbook.for_text(
            r"Unquoted service path with a writable parent: C:\Program Files\X\s.exe",
            "Windows")
        self.assertIn(r"C:\Program Files\X\s.exe", e2["cmd"])

    def test_no_match_returns_none(self):
        from recce import playbook
        self.assertIsNone(playbook.for_text("some benign http banner", "Linux"))
        self.assertIsNone(playbook.for_text("", ""))

    def test_confirmed_only_advisories_excluded(self):
        # A 'potential' advisory vuln must NOT get an exploitation entry, even if
        # its text would otherwise match.
        from recce import playbook
        h = Host(ip="10.0.0.5", os_family="Windows", vulns=[
            Vuln(ip="10.0.0.5", port=445, protocol="tcp", script_id="adv",
                 title="SeImpersonate advisory", severity="high",
                 source="version-db", confidence="potential")])
        self.assertEqual(playbook.host_entries(h), [])
        h.vulns[0].confidence = "confirmed"
        self.assertEqual(len(playbook.host_entries(h)), 1)

    def test_linux_writable_service_unit_does_not_get_windows_command(self):
        # OS-distinct matching: a Linux systemd 'writable service unit' finding
        # must not resolve to the Windows sc-config play.
        from recce import playbook
        e = playbook.for_text("Writable service unit: /etc/systemd/system/x.service",
                              "Linux")
        if e is not None:
            self.assertNotIn("sc config", e["cmd"].lower())


class ExploitRefTest(unittest.TestCase):
    def test_cve_exact_match(self):
        from recce.exploitref import proven_exploit_ref
        ref = proven_exploit_ref(["CVE-2017-0144"])
        self.assertIsNotNone(ref)
        self.assertIn("eternalblue", ref.lower())

    def test_no_match_returns_none(self):
        from recce.exploitref import proven_exploit_ref
        self.assertIsNone(proven_exploit_ref(["CVE-1999-0001"]))
        self.assertIsNone(proven_exploit_ref(None))
        self.assertIsNone(proven_exploit_ref([], ""))

    def test_cve_embedded_in_nse_id_text(self):
        # A raw NSE finding carrying the CVE only in its id must resolve the same.
        from recce.exploitref import proven_exploit_ref
        ref = proven_exploit_ref([], "http-vuln-cve2021-41773")
        self.assertIsNotNone(ref)
        self.assertIn("apache", ref.lower())

    def test_keyword_fallback_when_no_cve(self):
        from recce.exploitref import proven_exploit_ref
        self.assertIn("ms17_010",
                      (proven_exploit_ref([], "SMB ms17-010 vulnerable") or "").lower())
        self.assertIn("vsftpd",
                      (proven_exploit_ref([], "vsftpd 2.3.4 backdoor") or "").lower())

    def test_explicit_cve_beats_text(self):
        from recce.exploitref import proven_exploit_ref
        # A known CVE in the list wins even if the text mentions nothing.
        self.assertEqual(proven_exploit_ref(["CVE-2014-0160"]),
                         proven_exploit_ref([], "heartbleed"))

    def test_windows_references_resolve(self):
        from recce.exploitref import proven_exploit_ref
        cases = [
            (["CVE-2008-4250"], "ms08_067"),      # MS08-067
            (["CVE-2017-0147"], "eternalblue"),   # EternalBlue variant CVE
            (["CVE-2020-1472"], "zerologon"),     # ZeroLogon (module + PoC)
            (["CVE-2020-0796"], "smbghost"),      # SMBGhost
            (["CVE-2014-6324"], "ms14-068"),      # Kerberos PAC
        ]
        for cves, needle in cases:
            self.assertIn(needle, (proven_exploit_ref(cves) or "").lower(), cves)

    def test_token_privilege_maps_to_potato_tools(self):
        # A confirmed SeImpersonate finding (no CVE) points at the existing Potato
        # tools - a reference, not generated code.
        from recce.exploitref import proven_exploit_ref
        ref = proven_exploit_ref([], "Token holds SeImpersonate -> Potato -> SYSTEM")
        self.assertIsNotNone(ref)
        self.assertIn("godpotato", ref.lower())
        self.assertIn("printspoofer", ref.lower())

    def test_keyword_table_values_are_real_exploit_entries(self):
        # Integrity: every keyword ref must be a real curated reference - either a
        # concrete CVE entry, or the (CVE-less) token-privilege Potato reference.
        # Catches a typo'd/dangling keyword value.
        from recce.exploitref import PROVEN_EXPLOIT, PROVEN_KW, _POTATO
        allowed = set(PROVEN_EXPLOIT.values()) | {_POTATO}
        self.assertTrue(set(PROVEN_KW.values()) <= allowed)
        self.assertTrue(all(v.strip() for v in PROVEN_KW.values()))


class DeployTest(unittest.TestCase):
    def _host(self, ip, os_, ports):
        return Host(ip=ip, os_family=os_,
                    ports=[Port(portid=p, state="open") for p in ports])

    def test_transport_selection(self):
        from recce import deploy
        ssh = {"username": "u", "password": "p"}
        win = {"username": "a", "password": "b"}
        self.assertEqual(deploy.transport_for(self._host("1", "Linux", [22, 80]), ssh, win), "ssh")
        self.assertEqual(deploy.transport_for(self._host("2", "Windows", [445, 5985]), ssh, win), "winrm")
        self.assertEqual(deploy.transport_for(self._host("3", "Windows", [445]), ssh, win), "smb")
        # Windows box but we only have SSH creds and it runs sshd -> ssh
        self.assertEqual(deploy.transport_for(self._host("4", "Windows", [22, 445]), ssh, None), "ssh")
        self.assertIsNone(deploy.transport_for(self._host("5", "Linux", [80]), ssh, win))   # no exec port
        self.assertIsNone(deploy.transport_for(self._host("6", "Linux", [22]), None, None))  # no creds

    def test_skip_reason_explains_why_a_host_is_unable(self):
        from recce import deploy
        ssh = {"username": "u", "password": "p"}
        win = {"username": "a", "password": "b"}
        # No remote-exec port at all.
        self.assertIn("no remote-exec port",
                      deploy.skip_reason(self._host("1", "Linux", [80]), ssh, win))
        # SSH port open but no SSH creds held.
        self.assertIn("SSH creds",
                      deploy.skip_reason(self._host("2", "Linux", [22]), None, win))
        # SMB/WinRM open but no Windows creds held.
        self.assertIn("Windows creds",
                      deploy.skip_reason(self._host("3", "Windows", [445]), ssh, None))
        # nxc precheck said none of the protocols authenticated on this host.
        amap = {"4": {"smb": False, "winrm": False, "ssh": False}}
        self.assertIn("did not authenticate",
                      deploy.skip_reason(self._host("4", "Windows", [445, 5985]),
                                         ssh, win, amap))

    def test_domain_qualified_username_is_split(self):
        from recce import cli

        class A:
            def __init__(self, u, d=None, p="Pw"):
                self.username, self.domain, self.password = u, d, p
        # NetBIOS backslash form.
        c = cli._creds_of(A("CORP\\administrator"))
        self.assertEqual((c["username"], c["domain"]), ("administrator", "CORP"))
        # UPN @ form.
        c = cli._creds_of(A("administrator@corp.local"))
        self.assertEqual((c["username"], c["domain"]), ("administrator", "corp.local"))
        # domain/user form.
        c = cli._creds_of(A("corp.local/svc"))
        self.assertEqual((c["username"], c["domain"]), ("svc", "corp.local"))
        # Explicit -d wins over an embedded NetBIOS domain.
        c = cli._creds_of(A("CORP\\administrator", d="corp.local"))
        self.assertEqual((c["username"], c["domain"]), ("administrator", "corp.local"))
        # Plain username, explicit domain - unchanged.
        c = cli._creds_of(A("administrator", d="corp.local"))
        self.assertEqual((c["username"], c["domain"]), ("administrator", "corp.local"))
        # Plain username, no domain.
        c = cli._creds_of(A("administrator"))
        self.assertEqual((c["username"], c["domain"]), ("administrator", ""))

    def test_ps_payload_is_utf16le_base64(self):
        import base64
        from recce import deploy
        b = deploy._b64_ps("Write-Host hi")
        self.assertEqual(base64.b64decode(b).decode("utf-16-le"), "Write-Host hi")

    def test_ssh_key_auth_pipes_script_no_disk_artifact(self):
        from recce import deploy
        calls = {}

        def fake_run(argv, timeout, stdin=None):
            calls["argv"], calls["stdin"] = argv, stdin
            return 0, "recce-enum host=x\n[!] finding", ""
        orig = deploy._run
        deploy._run = fake_run
        try:
            out, err = deploy.run_ssh("1.2.3.4", {"username": "u", "key": "/k"}, "SCRIPT", 60)
        finally:
            deploy._run = orig
        self.assertIsNone(err)
        self.assertEqual(calls["stdin"], "SCRIPT")            # script piped over stdin
        self.assertIn("bash -s -- -q", calls["argv"])         # not written to disk
        self.assertNotEqual(calls["argv"][0], "sshpass")      # key auth, no sshpass
        self.assertIn("/k", calls["argv"])

    def test_winrm_and_smb_run_encoded_powershell(self):
        from recce import deploy
        seen = {}

        def fake_run(argv, timeout, stdin=None):
            seen.setdefault("argvs", []).append(argv)
            return 0, "recce-enum host=x\n[!] x", ""
        o_run, o_tool = deploy._run, deploy.smb_tool
        deploy._run, deploy.smb_tool = fake_run, (lambda: "nxc")
        try:
            _, e1 = deploy.run_winrm("1.2.3.4", {"username": "a", "password": "b"}, "S", 60)
            _, e2 = deploy.run_smb("1.2.3.4", {"username": "a", "password": "b"}, "/tmp/x.ps1", 60)
        finally:
            deploy._run, deploy.smb_tool = o_run, o_tool
        self.assertIsNone(e1)
        winrm = seen["argvs"][0]
        self.assertIn("winrm", winrm)
        self.assertIn("EncodedCommand", " ".join(winrm))
        self.assertIn("--put-file", " ".join(seen["argvs"][1]))   # smb pushes the script

    def test_deploy_dry_run_executes_nothing(self):
        from recce import cli, deploy
        called = {"n": 0}
        orig = deploy.deploy_one
        deploy.deploy_one = lambda *a, **k: (called.__setitem__("n", called["n"] + 1)
                                             or ("ssh", "x", None))
        try:
            with tempfile.TemporaryDirectory() as d:
                paths = cli._open_paths(d)
                st = cli._open_store(paths["db"])
                st.upsert_host(self._host("10.0.0.5", "Linux", [22]))
                st.close()
                args = SimpleNamespace(output_dir=d, workers=2, title="t", dry_run=True,
                                       ssh_user="u", ssh_pass=None, ssh_key="/k",
                                       username=None, password=None, domain=None,
                                       hash=None, targets=[], host=None)
                with contextlib.redirect_stdout(io.StringIO()):
                    rc = cli.cmd_deploy(args)
                self.assertEqual(rc, 0)
                self.assertEqual(called["n"], 0)   # dry-run ran nothing on the target
        finally:
            deploy.deploy_one = orig

    def test_stager_serves_script_under_token_only(self):
        import urllib.request
        import urllib.error
        from recce.stager import Stager
        data = b"# recce-enum.ps1"
        with Stager("127.0.0.1", {"recce-enum.ps1": data}) as st:
            got = urllib.request.urlopen(st.url("recce-enum.ps1"), timeout=5).read()
            self.assertEqual(got, data)
            self.assertEqual(st.hits, 1)
            with self.assertRaises(urllib.error.HTTPError) as cm:
                urllib.request.urlopen(
                    f"http://127.0.0.1:{st.port}/wrong/recce-enum.ps1", timeout=5)
            self.assertEqual(cm.exception.code, 404)

    def test_nxc_auth_parse_and_authmap_selection(self):
        from recce import deploy
        rows = deploy._parse_nxc_auth(
            "SMB   10.0.0.1  445  DC  [+] d\\a:p (Pwn3d!)\n"
            "SMB   10.0.0.2  445  WS  [-] d\\a:p STATUS_LOGON_FAILURE")
        self.assertEqual(rows, [("10.0.0.1", True, True), ("10.0.0.2", False, False)])
        ssh = {"username": "u", "password": "p"}
        win = {"username": "a", "password": "b"}
        amap = {"1": {"winrm": True}, "2": {"winrm": False, "smb": True},
                "3": {"ssh": True}, "4": {"winrm": False, "smb": False}}
        self.assertEqual(deploy.transport_for(self._host("1", "Windows", [445, 5985]), ssh, win, amap), "winrm")
        self.assertEqual(deploy.transport_for(self._host("2", "Windows", [445, 5985]), ssh, win, amap), "smb")
        self.assertEqual(deploy.transport_for(self._host("3", "Linux", [22]), ssh, win, amap), "ssh")
        self.assertIsNone(deploy.transport_for(self._host("4", "Windows", [445]), ssh, win, amap))

    def test_rejected_winrm_login_not_folded_as_success(self):
        """A rejected nxc WinRM login is a bare '[-]' banner with no STATUS keyword
        and no script output - it must be reported as a failure, never folded as a
        successful run with garbage loot."""
        from recce import deploy
        o_run, o_smb = deploy._run, deploy.smb_tool
        deploy.smb_tool = lambda: "nxc"
        deploy._run = lambda argv, timeout, stdin=None: (
            0, "WINRM 10.0.0.5 5985 HOST [-] corp\\u:BadPw", "")
        try:
            out, err = deploy.run_winrm("10.0.0.5", {"username": "u", "password": "x"},
                                        "SCRIPT", 60)
        finally:
            deploy._run, deploy.smb_tool = o_run, o_smb
        self.assertIsNone(out)                          # not a success
        self.assertIn("auth", err.lower())

    def test_exploit_cell_needs_cve_match_not_just_port(self):
        from recce.report_excel import _exploit_cell, _curated_exploit
        from recce.models import Exploit
        host = Host(ip="1.1.1.1", exploits=[
            Exploit(ip="1.1.1.1", port=80, edb_id="99999", cves=["CVE-2099-9999"])])
        # unrelated port-80 finding, no shared CVE -> NO exploit attached (was the bug)
        risky = Vuln(ip="1.1.1.1", port=80, protocol="tcp", script_id="http-methods",
                     title="Risky HTTP methods enabled", severity="low", confidence="likely")
        self.assertEqual(_exploit_cell(host, risky), "")
        # a finding sharing the EDB's CVE -> a clearly-labelled CANDIDATE, not proof
        match = Vuln(ip="1.1.1.1", port=80, protocol="tcp", script_id="x", title="Some RCE",
                     severity="high", confidence="likely", ids=["CVE-2099-9999"])
        cell = _exploit_cell(host, match)
        self.assertIn("EDB-99999", cell)
        self.assertIn("candidate", cell.lower())
        # a weak-TLS finding never claims a proven exploit, even with heartbleed's CVE
        tls = Vuln(ip="1.1.1.1", port=443, protocol="tcp", script_id="ssl-enum-ciphers",
                   title="Weak SSL/TLS ciphers or protocols", severity="medium",
                   confidence="likely", ids=["CVE-2014-0160"])
        self.assertEqual(_curated_exploit(tls), "")

    def test_impacket_engine_runs_stager_cradle_when_no_nxc(self):
        """With netexec absent but impacket present, the Windows path uses
        impacket wmiexec (which pairs cleanly with --stager: runs the cradle, no
        file push)."""
        from recce import deploy
        seen = []

        def fake_run(argv, timeout, stdin=None):
            seen.append(argv[0])
            return 0, "recce-enum host=x\n[!] finding", ""

        class FS:
            def url(self, n):
                return f"http://1.2.3.4:8000/t/{n}"
        o_run, o_smb, o_imp = deploy._run, deploy.smb_tool, deploy.impacket_tool
        deploy._run = fake_run
        deploy.smb_tool = lambda: None                                  # no nxc
        deploy.impacket_tool = lambda n: "impacket-wmiexec" if n == "wmiexec" else None
        try:
            self.assertEqual(deploy.win_engine(), ("impacket", "impacket-wmiexec"))
            out, err, status = deploy.run_win_stager(
                "10.0.0.9", {"username": "a", "password": "b", "domain": "d"},
                "smb", FS(), 60)
            self.assertEqual(status, "ok")
            self.assertEqual(seen[0], "impacket-wmiexec")
            self.assertEqual(deploy._impacket_target({"username": "a", "hash": "NT"}, "1.2.3.4"),
                             "a@1.2.3.4")
        finally:
            deploy._run, deploy.smb_tool, deploy.impacket_tool = o_run, o_smb, o_imp

    def test_stager_unreachable_falls_back_to_push(self):
        from recce import deploy
        win = {"username": "a", "password": "b"}
        seen = []

        def fake_run(argv, timeout, stdin=None):
            joined = " ".join(argv)
            if "EncodedCommand" in joined:            # the stager cradle
                seen.append("stager")
                return 0, "PowerShell WebException: unable to connect", ""
            if "--put-file" in joined:                # the push fallback
                seen.append("put")
                return 0, "", ""
            return 0, "recce-enum host=x\n[!] finding", ""   # push exec

        class FakeStager:
            def url(self, n):
                return f"http://1.2.3.4:8000/tok/{n}"
        o_run, o_tool = deploy._run, deploy.smb_tool
        deploy._run, deploy.smb_tool = fake_run, (lambda: "nxc")
        try:
            t, out, err = deploy.deploy_one(
                self._host("10.0.0.9", "Windows", [445]), None, win,
                stager=FakeStager(), authmap={"10.0.0.9": {"smb": True}})
        finally:
            deploy._run, deploy.smb_tool = o_run, o_tool
        self.assertIn("stager", seen)                 # tried the stager first
        self.assertIn("put", seen)                    # then fell back to push
        self.assertTrue(out and "recce-enum" in out)  # and got output

    def test_deploy_worker_folds_recce_enum_output(self):
        from recce import cli, deploy
        sample = ("recce-enum host=web01 os=linux\n"
                  "[!] sudo: NOPASSWD entry - run a root command via sudo\n")
        orig = deploy.deploy_one
        deploy.deploy_one = lambda host, s, w, timeout, stager=None, authmap=None: (
            "ssh", sample, None)
        try:
            with tempfile.TemporaryDirectory() as d:
                host, transport, added, promoted, err = cli._deploy_worker(
                    self._host("10.0.0.5", "Linux", [22]), {"username": "u"}, None, 60, d)
                self.assertIsNone(err)
                self.assertEqual(transport, "ssh")
                self.assertGreaterEqual(added, 1)              # finding folded in
                self.assertTrue(host.local_findings)
                self.assertTrue(host.privesc_checked)
                self.assertTrue(os.path.exists(os.path.join(d, "10.0.0.5.txt")))  # loot saved
        finally:
            deploy.deploy_one = orig


if __name__ == "__main__":
    unittest.main(verbosity=2)
